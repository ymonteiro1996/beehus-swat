import io
import json
import os
import re

import openpyxl
from bson import ObjectId
from bson.errors import InvalidId
from flask import Blueprint, Response, jsonify, render_template, request
from openpyxl.styles import Alignment, Font, PatternFill

from db import db, get_company_filter, get_wallet_names

bp = Blueprint("precificacao", __name__)

SAVED_LISTS_FILE  = os.path.join(os.path.dirname(__file__), "..", "data", "precificacao_lists.json")
CONFIG_FILE       = os.path.join(os.path.dirname(__file__), "..", "data", "precificacao_config.json")

DEFAULT_LIST_NAME = "Lista padrão"

_DEFAULT_CONFIG = {
    "benchmarks": [
        {"id": "66fc4a71e88f2f542b805639", "name": "CDI"}
    ]
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _load_lists():
    """Return saved lists as `[{"name": str, "securities": [...]}, ...]`.

    Migrates older on-disk shapes the first time they are read and rewrites
    the file in place:
      - `{"securities": [...]}` (the single-template format used between
        revisions) → wrapped into one named list (DEFAULT_LIST_NAME).
      - Already a list of `{name, securities}` → kept, with duplicate
        names disambiguated and missing names backfilled."""
    if not os.path.exists(SAVED_LISTS_FILE):
        return []
    try:
        with open(SAVED_LISTS_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except (OSError, json.JSONDecodeError):
        return []

    if isinstance(raw, dict) and "securities" in raw:
        secs = list(raw.get("securities") or [])
        lists = [{"name": DEFAULT_LIST_NAME, "securities": secs}] if secs else []
        try:
            _write_lists(lists)
        except OSError:
            pass
        return lists

    if isinstance(raw, list):
        cleaned, seen = [], set()
        for entry in raw:
            if not isinstance(entry, dict):
                continue
            name = str(entry.get("name") or "").strip() or DEFAULT_LIST_NAME
            base, i = name, 2
            while name in seen:
                name = f"{base} ({i})"
                i += 1
            seen.add(name)
            cleaned.append({
                "name": name,
                "securities": list(entry.get("securities") or []),
            })
        return cleaned

    return []


def _write_lists(lists):
    with open(SAVED_LISTS_FILE, "w", encoding="utf-8") as f:
        json.dump(lists, f, indent=2, ensure_ascii=False)


def _load_config():
    if not os.path.exists(CONFIG_FILE):
        return dict(_DEFAULT_CONFIG)
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def _write_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


def _extract_hp(doc):
    """Return the historyPrice sub-document from a securityPrices doc."""
    if not doc:
        return {}
    raw = doc.get("historyPrice")
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, list) and raw:
        try:
            return max(raw, key=lambda x: str(x.get("date", "")))
        except Exception:
            return raw[0] if raw else {}
    return {}


def _extract_all_hp(doc):
    """Return a list of all historyPrice entries from a securityPrices doc."""
    if not doc:
        return []
    raw = doc.get("historyPrice")
    if isinstance(raw, dict):
        return [raw]
    if isinstance(raw, list):
        return raw
    return []


def _find_price(sec_id_str, query_extra, proj):
    """Find one price doc trying ObjectId then string for securityId."""
    try:
        oid = ObjectId(sec_id_str)
        doc = next(iter(db.securityPrices.find(
            {"securityId": oid, **query_extra}, proj
        ).sort("historyPrice.date", -1).limit(1)), None)
        if doc:
            return doc
    except Exception:
        pass
    return next(iter(db.securityPrices.find(
        {"securityId": sec_id_str, **query_extra}, proj
    ).sort("historyPrice.date", -1).limit(1)), None)


def _find_all_prices(sec_id_str, query_extra, proj, ascending=True):
    """Find all price docs trying ObjectId then string for securityId."""
    direction = 1 if ascending else -1
    try:
        oid  = ObjectId(sec_id_str)
        docs = list(db.securityPrices.find(
            {"securityId": oid, **query_extra}, proj
        ).sort("historyPrice.date", direction))
        if docs:
            return docs
    except Exception:
        pass
    return list(db.securityPrices.find(
        {"securityId": sec_id_str, **query_extra}, proj
    ).sort("historyPrice.date", direction))


def _find_price_as_of(sec_id_str, ref_date):
    """Latest `historyPrice` entry whose `date` is `<= ref_date` for this
    security. Walks every `securityPrices` doc (since `historyPrice` may
    be embedded as either a single dict or an array of entries — see
    [[securityprices_schema]]) and returns the best `{date, value, ...}`
    dict. When `ref_date` is falsy, falls back to the all-time-latest
    behaviour of `_find_price + _extract_hp` so callers can opt-in.
    Returns `{}` when no historical entry satisfies the cutoff."""
    if not ref_date:
        return _extract_hp(_find_price(sec_id_str, {}, {"historyPrice": 1}))
    docs = _find_all_prices(sec_id_str, {}, {"historyPrice": 1}, ascending=False)
    best = None
    best_dt = ""
    for d in docs:
        for hp in _extract_all_hp(d):
            dt = str(hp.get("date") or "")[:10]
            if not dt or dt > ref_date:
                continue
            if dt > best_dt:
                best_dt = dt
                best = hp
    return best or {}


def _get_most_recent_position(wallet_id, position_date=None):
    """Get processedPosition for a wallet at a given date (or most recent).
    Returns (pos_doc, position_date_str) or (None, None).
    """
    or_q = [{"walletId": wallet_id}]
    try:
        or_q.append({"walletId": ObjectId(wallet_id)})
    except Exception:
        pass

    query = {"$or": or_q}
    if position_date:
        query["positionDate"] = position_date

    pos_doc = next(iter(
        db.processedPosition.find(
            query,
            {"securities": 1, "positionDate": 1}
        ).sort("positionDate", -1).limit(1)
    ), None)

    if not pos_doc:
        return None, None
    return pos_doc, str(pos_doc.get("positionDate", ""))[:10]


# ── Page ───────────────────────────────────────────────────────────────────────

@bp.route("/precificacao")
def index():
    cfg        = _load_config()
    benchmarks = cfg.get("benchmarks", [])
    return render_template("precificacao.html", benchmarks=benchmarks)


# ── API: Companies & Wallets ──────────────────────────────────────────────────

@bp.route("/api/precificacao/companies")
def get_companies():
    companies = sorted(
        [{"id": str(c["_id"]), "name": c.get("name", "")}
         for c in db.companies.find({}, {"name": 1})],
        key=lambda c: c["name"],
    )
    cf = get_company_filter()
    if cf:
        companies = [c for c in companies if c["id"] in cf]
    return jsonify(companies)


@bp.route("/api/precificacao/wallets")
def get_wallets():
    company_id = request.args.get("companyId", "").strip()
    if not company_id:
        return jsonify([])
    wallets = list(db.wallets.find(
        {"companyId": company_id},
        {"name": 1, "accountCode": 1}
    ).sort("name", 1))
    return jsonify([
        {"id": str(w["_id"]), "name": w.get("name", ""), "accountCode": w.get("accountCode", "")}
        for w in wallets
    ])


@bp.route("/api/precificacao/latest-position-date")
def get_latest_position_date():
    """Return the most recent positionDate for a wallet."""
    wallet_id = request.args.get("walletId", "").strip()
    if not wallet_id:
        return jsonify({"date": None})
    _, pos_date = _get_most_recent_position(wallet_id)
    return jsonify({"date": pos_date})


# ── API: Wallet Securities (from processedPosition) ──────────────────────────

@bp.route("/api/precificacao/wallet-securities")
def get_wallet_securities():
    wallet_id     = request.args.get("walletId", "").strip()
    pos_date_arg  = request.args.get("positionDate", "").strip()
    if not wallet_id:
        return jsonify({"error": "walletId obrigatório"}), 400

    pos_doc, position_date = _get_most_recent_position(wallet_id, pos_date_arg or None)
    if not pos_doc:
        return jsonify({"securities": [], "positionDate": None})

    raw_secs = pos_doc.get("securities", [])

    # Batch-fetch security names from securities collection
    oid_ids = []
    for s in raw_secs:
        sid = s.get("securityId")
        if sid:
            try:
                oid_ids.append(ObjectId(str(sid)))
            except Exception:
                pass

    sec_meta = {}
    for sec in db.securities.find({"_id": {"$in": oid_ids}}, {"beehusName": 1, "mainId": 1}):
        sec_meta[str(sec["_id"])] = {
            "beehusName": sec.get("beehusName", ""),
            "mainId":     sec.get("mainId", ""),
        }

    result = []
    for s in raw_secs:
        sid = str(s.get("securityId", ""))
        qty = s.get("quantity")
        pu  = s.get("pu")
        balance = None
        if qty is not None and pu is not None:
            try:
                balance = round(float(qty) * float(pu), 2)
            except Exception:
                pass
        meta = sec_meta.get(sid, {})
        result.append({
            "securityId":  sid,
            "beehusName":  meta.get("beehusName", sid),
            "mainId":      meta.get("mainId", ""),
            "pricingType": s.get("pricingType", ""),
            "quantity":    float(qty) if qty is not None else None,
            "pu":          float(pu)  if pu  is not None else None,
            "balance":     balance,
        })

    result.sort(key=lambda x: (x["beehusName"] or "").lower())
    return jsonify({"securities": result, "positionDate": position_date})


# ── API: Security detail ─────────────────────────────────────────────────────
# Merges: securities collection (metadata) + securityPrices (lastPU)
#       + processedPosition (pricingType, quantity, pu) from most recent position

@bp.route("/api/precificacao/security/<sec_id>")
def get_security(sec_id):
    try:
        oid = ObjectId(sec_id)
    except Exception:
        return jsonify({"error": "ID inválido"}), 400

    sec = db.securities.find_one({"_id": oid}, {
        "beehusName": 1, "mainId": 1, "securityType": 1, "type": 1,
        "currency": 1, "maturityDate": 1, "emissionDate": 1, "issuer": 1,
        "indexer": 1, "indexerPercentual": 1, "yield": 1,
    })
    if not sec:
        return jsonify({"error": "Ativo não encontrado"}), 404

    # Last PU from securityPrices
    last_price = _find_price(sec_id, {}, {"historyPrice": 1})
    hp = _extract_hp(last_price)

    # Position data from processedPosition
    wallet_id    = request.args.get("walletId", "").strip()
    pos_date_arg = request.args.get("positionDate", "").strip()
    pos_pricing_type = ""
    pos_quantity     = None
    pos_pu           = None

    if wallet_id:
        pos_doc, _ = _get_most_recent_position(wallet_id, pos_date_arg or None)
        if pos_doc:
            for ps in pos_doc.get("securities", []):
                if str(ps.get("securityId", "")) == sec_id:
                    pos_pricing_type = ps.get("pricingType", "")
                    pos_quantity     = float(ps["quantity"]) if ps.get("quantity") is not None else None
                    pos_pu           = float(ps["pu"])       if ps.get("pu")       is not None else None
                    break

    return jsonify({
        "id":               str(sec["_id"]),
        "beehusName":       sec.get("beehusName", ""),
        "mainId":           sec.get("mainId", ""),
        "securityType":     sec.get("securityType", ""),
        "type":             sec.get("type", ""),
        "currency":         sec.get("currency", ""),
        "maturityDate":     str(sec.get("maturityDate", ""))[:10] if sec.get("maturityDate") else None,
        "emissionDate":     str(sec.get("emissionDate", ""))[:10] if sec.get("emissionDate") else None,
        "issuer":           sec.get("issuer", ""),
        "indexer":          sec.get("indexer", ""),
        "indexerPercentual": sec.get("indexerPercentual"),
        "yield":            sec.get("yield"),
        "lastPU":           hp.get("value"),
        "lastPUDate":       str(hp.get("date", ""))[:10] if hp.get("date") else None,
        "pricingType":      pos_pricing_type,
        "posQuantity":      pos_quantity,
        "posPU":            pos_pu,
    })


# ── API: Security search ─────────────────────────────────────────────────────

@bp.route("/api/precificacao/search")
def search_securities():
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify({"results": []})

    escaped_q = re.escape(q)
    or_clauses = [
        {"beehusName": {"$regex": escaped_q, "$options": "i"}},
        {"mainId":     {"$regex": escaped_q, "$options": "i"}},
    ]
    try:
        or_clauses.append({"_id": ObjectId(q)})
    except Exception:
        pass

    results = []
    for sec in db.securities.find({"$or": or_clauses}, {
        "beehusName": 1, "mainId": 1, "securityType": 1,
        "indexer": 1, "indexerPercentual": 1,
    }).limit(20):
        results.append({
            "id":               str(sec["_id"]),
            "beehusName":       sec.get("beehusName", ""),
            "mainId":           sec.get("mainId", ""),
            "securityType":     sec.get("securityType", ""),
            "indexer":          sec.get("indexer", ""),
            "indexerPercentual": sec.get("indexerPercentual"),
        })
    return jsonify({"results": results})


# ── API: Transactions ─────────────────────────────────────────────────────────

@bp.route("/api/precificacao/security-transactions")
def get_security_transactions():
    sec_id        = request.args.get("securityId", "").strip()
    wallet_id     = request.args.get("walletId",   "").strip()
    position_date = request.args.get("positionDate", "").strip()
    if not sec_id or not wallet_id:
        return jsonify({"transactions": []})

    # Match securityId as both ObjectId and string
    sid_q = [{"securityId": sec_id}]
    try:
        sid_q.append({"securityId": ObjectId(sec_id)})
    except Exception:
        pass

    # Match walletId as both ObjectId and string
    wid_q = [{"walletId": wallet_id}]
    try:
        wid_q.append({"walletId": ObjectId(wallet_id)})
    except Exception:
        pass

    query = {
        "$or": sid_q,
        "$and": [{"$or": wid_q}],
        "beehusTransactionType": {"$in": ["buySell", "securityTransfer"]},
    }
    if position_date:
        query["liquidationDate"] = {"$lte": position_date}

    txns = list(db.transactions.find(
        query,
        {"liquidationDate": 1, "beehusTransactionType": 1,
         "quantity": 1, "price": 1, "balance": 1, "description": 1}
    ).sort("liquidationDate", 1))

    result = []
    for txn in txns:
        qty = txn.get("quantity")
        if qty is None:
            continue
        try:
            if float(qty) == 0:
                continue
        except (ValueError, TypeError):
            continue
        prc = txn.get("price")
        q   = float(qty)
        p   = float(prc) if prc is not None else None
        result.append({
            "liquidationDate":      str(txn.get("liquidationDate", "") or "")[:10],
            "beehusTransactionType": txn.get("beehusTransactionType", ""),
            "quantity":             q,
            "price":                p,
            "balance":              round(q * p, 2) if p is not None else None,
            "description":          txn.get("description", ""),
        })
    return jsonify({"transactions": result})


# ── API: Calculate ────────────────────────────────────────────────────────────

def calculate_curva(securities_list):
    """Pure function: run the curva pricing engine over a list of
    security input dicts and return the raw `results` array (same shape
    the /calcular route returns).

    Extracted from the Flask route so other blueprints (notably
    /repetir-posicoes, which now auto-applies curva PUs during the
    preview build) can call the engine directly without going through
    HTTP. The route below is a thin wrapper that adapts the request
    body and wraps the response in jsonify."""
    if not securities_list:
        return []
    return _calculate_curva_impl(securities_list)


# Tipos de transação que reduzem o PU na curva (juros/principal pagos).
_TXN_TYPES_COUPON_AMORT = ("coupon", "amortization")


def _qty_before(wallet_id, sec_id, date_str):
    """Quantity of `sec_id` in the most recent `processedPosition` for
    `wallet_id` whose `positionDate` is strictly **before** `date_str`
    (the "dia anterior" ao evento). Matches `walletId` as ObjectId and
    string. Returns `None` when no prior position or no entry for the
    security."""
    if not wallet_id or not sec_id or not date_str:
        return None
    or_q = [{"walletId": wallet_id}]
    try:
        or_q.append({"walletId": ObjectId(wallet_id)})
    except Exception:
        pass
    doc = next(iter(db.processedPosition.find(
        {"$or": or_q, "positionDate": {"$lt": date_str}},
        {"securities": 1, "positionDate": 1},
    ).sort("positionDate", -1).limit(1)), None)
    if not doc:
        return None
    sid = str(sec_id)
    for ps in doc.get("securities", []):
        if str(ps.get("securityId", "")) == sid:
            q = ps.get("quantity")
            try:
                return float(q) if q is not None else None
            except (TypeError, ValueError):
                return None
    return None


def _event_pu_impacts(sec_id, wallet_id):
    """Per-unit PU drops for coupon/amortization events on (wallet,
    security), keyed by `liquidationDate` (`YYYY-MM-DD`).

        impacto = (Σ balance[coupon|amortization] + Σ balance[taxes]) / qtd_prev

    `taxes` (IR retido — `balance` já negativo) só entra quando há uma
    transação `taxes` com o **mesmo** `walletId`/`securityId`/`liquidationDate`
    de um cupom/amortização, baixando o degrau ao valor líquido (marcação
    caixa-neutra). `qtd_prev` vem de `_qty_before` (posição do dia
    anterior). Datas com `taxes` mas sem cupom/amort, ou com `qtd_prev`
    nula/zero, são descartadas. Retorna `{}` quando não há `wallet_id`
    (sem carteira não há como casar a transação)."""
    if not wallet_id or not sec_id:
        return {}

    sid_q = [{"securityId": sec_id}]
    try:
        sid_q.append({"securityId": ObjectId(sec_id)})
    except Exception:
        pass
    wid_q = [{"walletId": wallet_id}]
    try:
        wid_q.append({"walletId": ObjectId(wallet_id)})
    except Exception:
        pass

    cursor = db.transactions.find(
        {"$or": sid_q,
         "$and": [{"$or": wid_q}],
         "beehusTransactionType": {"$in": list(_TXN_TYPES_COUPON_AMORT) + ["taxes"]},
         "trashed": {"$ne": True}},
        {"liquidationDate": 1, "beehusTransactionType": 1, "balance": 1},
    )

    by_date = {}  # date -> {"event": float, "taxes": float}
    for t in cursor:
        dt = str(t.get("liquidationDate") or "")[:10]
        if not dt:
            continue
        try:
            bal = float(t.get("balance"))
        except (TypeError, ValueError):
            continue
        slot = by_date.setdefault(dt, {"event": 0.0, "taxes": 0.0})
        if t.get("beehusTransactionType") == "taxes":
            slot["taxes"] += bal
        else:
            slot["event"] += bal

    impacts = {}
    for dt, slot in by_date.items():
        if slot["event"] == 0.0:
            continue  # taxes-only date — sem cupom/amort, não há degrau
        qty_prev = _qty_before(wallet_id, sec_id, dt)
        if not qty_prev:
            continue  # sem posição anterior ou qtd 0 → pula (sem div/0)
        impacts[dt] = (slot["event"] + slot["taxes"]) / qty_prev
    return impacts


def _snap_impacts(impacts, emitted_dates, base_date):
    """Map raw event impacts (keyed by `liquidationDate`) onto the curve's
    emitted calendar: each impact lands on the **first emitted date ≥ its
    liquidationDate**. Impacts on/below `base_date` (already baked into the
    PU-base) or beyond the horizon are dropped. `emitted_dates` must be
    sorted ascending. Returns `{cal_date -> Σ impacto_por_unidade}`."""
    if not impacts or not emitted_dates:
        return {}
    out = {}
    for liq in sorted(impacts):
        if base_date and liq <= base_date:
            continue
        cal = next((d for d in emitted_dates if d >= liq), None)
        if cal is None:
            continue
        out[cal] = out.get(cal, 0.0) + impacts[liq]
    return out


@bp.route("/api/precificacao/calcular", methods=["POST"])
def calcular():
    data            = request.get_json() or {}
    securities_list = data.get("securities", [])

    if not securities_list:
        return jsonify({"error": "Parâmetros inválidos"}), 400

    return jsonify({"results": _calculate_curva_impl(securities_list)})


def _calculate_curva_impl(securities_list):
    # Find each security's last available PU
    sec_last = {}
    for s in securities_list:
        sec_id = s.get("id")
        if not sec_id or sec_id in sec_last:
            continue
        last_price = _find_price(sec_id, {}, {"historyPrice": 1})
        last_hp    = _extract_hp(last_price)
        if last_hp.get("value"):
            sec_last[sec_id] = (float(last_hp["value"]), str(last_hp.get("date", ""))[:10])

    # Earliest start date to limit benchmark fetch
    all_start_dates = [v[1] for v in sec_last.values() if v[1]]
    global_start    = min(all_start_dates) if all_start_dates else ""

    # Build factor map for each unique benchmark
    unique_bm_ids = {s.get("benchmarkId") for s in securities_list if s.get("benchmarkId")}
    has_curva = any(s.get("calcType") in ("pre_fixado_curva", "inflacao_curva") for s in securities_list)
    if has_curva and not unique_bm_ids:
        cfg = _load_config()
        default_bms = cfg.get("benchmarks", [])
        if default_bms:
            unique_bm_ids.add(default_bms[0]["id"])

    bm_factor_map = {}
    bm_errors     = {}

    for bm_id in unique_bm_ids:
        query_extra = {"historyPrice.date": {"$gte": global_start}} if global_start else {}
        docs = _find_all_prices(bm_id, query_extra, {"historyPrice": 1}, ascending=True)
        if not docs:
            bm_errors[bm_id] = "Preços do benchmark não encontrados"
            continue

        all_hps = []
        for d in docs:
            all_hps.extend(_extract_all_hp(d))

        seen = set()
        series = []
        for hp in sorted(all_hps, key=lambda x: str(x.get("date", ""))):
            dt = str(hp.get("date", ""))[:10]
            if dt and dt not in seen:
                seen.add(dt)
                series.append((dt, hp))

        if not series:
            bm_errors[bm_id] = "Datas do benchmark não encontradas"
            continue

        fmap = {}
        for i, (dt, hp) in enumerate(series):
            if hp.get("rentability") is not None:
                fmap[dt] = float(hp["rentability"])
            elif i > 0:
                prev_val = series[i - 1][1].get("value")
                curr_val = hp.get("value")
                if prev_val and curr_val:
                    fmap[dt] = float(curr_val) / float(prev_val) - 1
        bm_factor_map[bm_id] = fmap

    # Roll PU forward for each security
    results = []
    for sec_inp in securities_list:
        sec_id       = sec_inp.get("id")
        calc_type    = sec_inp.get("calcType", "pos_fixado")
        bm_id        = sec_inp.get("benchmarkId", "")
        bm_name      = sec_inp.get("benchmarkName", "")
        wallet_id    = sec_inp.get("walletId", "")
        wallet_name  = sec_inp.get("walletName", "")
        pricing_type = sec_inp.get("pricingType", "")

        if calc_type == "pos_fixado":
            # Priority: posPU/positionDate from wallet > securityPrices latest
            pos_pu_raw   = sec_inp.get("posPU")
            pos_date_raw = str(sec_inp.get("positionDate") or "")[:10]
            if pos_pu_raw is not None and pos_date_raw:
                try:
                    current_pu, last_pu_date = float(pos_pu_raw), pos_date_raw
                except (ValueError, TypeError):
                    if sec_id not in sec_last:
                        results.append({"securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                                        "calcType": calc_type, "error": "Sem PU disponível"})
                        continue
                    current_pu, last_pu_date = sec_last[sec_id]
            elif sec_id in sec_last:
                current_pu, last_pu_date = sec_last[sec_id]
            else:
                results.append({"securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                                "calcType": calc_type, "error": "Sem PU disponível"})
                continue
            if not bm_id or bm_id in bm_errors:
                results.append({"securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                                "benchmarkName": bm_name, "calcType": calc_type,
                                "error": bm_errors.get(bm_id, "Benchmark não configurado")})
                continue
            idx_pct = sec_inp.get("indexerPercentual")
            if idx_pct is None:
                results.append({"securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                                "calcType": calc_type, "error": "indexerPercentual ausente"})
                continue
            idx_pct_f  = float(idx_pct) / 100.0
            factor_map = bm_factor_map.get(bm_id, {})
            emitted    = sorted(dt for dt in factor_map if dt > last_pu_date)
            impacts    = _snap_impacts(_event_pu_impacts(sec_id, wallet_id),
                                       emitted, last_pu_date)
            for dt in emitted:
                factor = factor_map[dt]
                new_pu = current_pu * (1 + factor * idx_pct_f)
                row = {
                    "securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                    "calcType": calc_type, "walletId": wallet_id, "walletName": wallet_name,
                    "pricingType": pricing_type, "benchmarkName": bm_name,
                    "date": dt, "benchmarkFactor": round(1 + factor, 10),
                    "indexerPercentual": idx_pct_f,
                }
                imp = impacts.get(dt)
                if imp:
                    new_pu -= imp           # degrau de cupom/amortização (líquido de IR)
                    row["eventImpact"] = round(-imp, 8)
                row["pu"] = round(new_pu, 8)
                results.append(row)
                current_pu = new_pu
            # NOTE: `walletId` agora SEMPRE acompanha o resultado (inclusive
            # pos_fixado) — o lookup no Repetir Posições é estritamente
            # `walletId|securityId|targetDate` (sem fallback agnóstico),
            # então uma entrada `pos_fixado` sem walletId no template não
            # seria localizada na repetição.

        elif calc_type in ("pre_fixado_curva", "inflacao_curva"):
            txn_list = sec_inp.get("transactions", [])
            if not txn_list:
                results.append({"securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                                "calcType": calc_type, "walletId": wallet_id, "walletName": wallet_name,
                                "pricingType": pricing_type, "error": "Nenhuma transação informada"})
                continue

            cal_fmap = next(iter(bm_factor_map.values()), {})
            if not cal_fmap:
                results.append({"securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                                "calcType": calc_type, "walletId": wallet_id, "walletName": wallet_name,
                                "pricingType": pricing_type,
                                "error": "Nenhum calendário disponível — configure ao menos um benchmark"})
                continue

            lots = []
            for t in txn_list:
                qty, yld = t.get("quantity"), t.get("yield")
                if qty is None or yld is None:
                    continue
                try:
                    lots.append({"quantity": float(qty), "yield": float(yld)})
                except (ValueError, TypeError):
                    continue

            if not lots:
                results.append({"securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                                "calcType": calc_type, "walletId": wallet_id, "walletName": wallet_name,
                                "pricingType": pricing_type, "error": "Nenhuma transação válida"})
                continue

            # Determine initial PU — priority: posPU/positionDate > initialPU > securityPrices
            pos_pu_raw    = sec_inp.get("posPU")
            pos_date_raw  = str(sec_inp.get("positionDate") or "")[:10]
            init_pu_raw   = sec_inp.get("initialPU")
            init_date_raw = str(sec_inp.get("initialPUDate") or "")[:10]

            last_pu, last_pu_date = None, None
            if pos_pu_raw is not None and pos_date_raw:
                try:
                    last_pu, last_pu_date = float(pos_pu_raw), pos_date_raw
                except (ValueError, TypeError):
                    pass
            if (not last_pu or not last_pu_date) and init_pu_raw is not None and init_date_raw:
                try:
                    last_pu, last_pu_date = float(init_pu_raw), init_date_raw
                except (ValueError, TypeError):
                    pass
            if not last_pu or not last_pu_date:
                last_price_doc = _find_price(sec_id, {}, {"historyPrice": 1})
                last_hp = _extract_hp(last_price_doc)
                last_pu = float(last_hp["value"]) if last_hp.get("value") else None
                last_pu_date = str(last_hp.get("date", ""))[:10] if last_hp.get("date") else None

            if not last_pu or not last_pu_date:
                results.append({"securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                                "calcType": calc_type, "walletId": wallet_id, "walletName": wallet_name,
                                "pricingType": pricing_type,
                                "error": "Nenhum PU disponível — informe o PU inicial"})
                continue

            sorted_cal  = sorted(cal_fmap.keys())
            date_to_idx = {d: i for i, d in enumerate(sorted_cal)}

            if last_pu_date in date_to_idx:
                base_idx = date_to_idx[last_pu_date]
            else:
                base_idx = next((date_to_idx[d] for d in sorted_cal if d >= last_pu_date), None)
            if base_idx is None:
                results.append({"securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                                "calcType": calc_type, "walletId": wallet_id, "walletName": wallet_name,
                                "pricingType": pricing_type, "error": "Data do último PU fora do calendário"})
                continue

            active_lots = [l for l in lots if l["quantity"] > 0]
            total_qty   = sum(l["quantity"] for l in active_lots)
            if not active_lots or total_qty <= 0:
                results.append({"securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                                "calcType": calc_type, "walletId": wallet_id, "walletName": wallet_name,
                                "pricingType": pricing_type, "error": "Nenhuma transação ativa"})
                continue

            w_yield      = sum(l["quantity"] * l["yield"] for l in active_lots) / total_qty
            daily_factor = (1 + w_yield / 100) ** (1 / 252)

            # For inflacao_curva, accumulate benchmark factor on top of yield accrual
            inf_factor_map = {}
            if calc_type == "inflacao_curva" and bm_id:
                inf_factor_map = bm_factor_map.get(bm_id, {})
                if not inf_factor_map:
                    results.append({"securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                                    "calcType": calc_type, "walletId": wallet_id, "walletName": wallet_name,
                                    "pricingType": pricing_type, "benchmarkName": bm_name,
                                    "error": bm_errors.get(bm_id, "Benchmark de inflação não encontrado")})
                    continue

            # Use the default calendar benchmark for annualized yield display
            cal_bm_id = bm_id if calc_type == "inflacao_curva" else next(iter(bm_factor_map), None)
            cal_bm_fmap = bm_factor_map.get(cal_bm_id, {}) if cal_bm_id else {}

            # Roll incremental (`running_pu *= …` por dia) — idêntico à forma
            # fechada `last_pu × accum_bm × daily_factor^n` na ausência de
            # eventos, mas permite que o degrau de cupom/amortização se
            # propague para os dias seguintes (e, na amortização, o acrual
            # passe a incidir sobre a base reduzida). Ver docs/PRECIFICACAO.md.
            emitted = [dt for dt in sorted_cal if date_to_idx[dt] > base_idx]
            impacts = _snap_impacts(_event_pu_impacts(sec_id, wallet_id),
                                    emitted, last_pu_date)
            running_pu = last_pu
            for dt in sorted_cal:
                dt_idx = date_to_idx[dt]
                if dt_idx <= base_idx:
                    continue
                running_pu *= daily_factor
                if calc_type == "inflacao_curva" and dt in inf_factor_map:
                    running_pu *= (1 + inf_factor_map[dt])
                imp = impacts.get(dt)
                if imp:
                    running_pu -= imp   # degrau de cupom/amortização (líquido de IR)
                pu_val = running_pu
                # Annualize daily rentability: ((1 + daily)^252 - 1) * 100
                daily_rent = cal_bm_fmap.get(dt)
                bm_yield_ann = round(((1 + daily_rent) ** 252 - 1) * 100, 4) if daily_rent is not None else None
                # Daily factor for inflacao_curva = yield daily factor × benchmark daily factor
                if calc_type == "inflacao_curva":
                    bm_daily = inf_factor_map.get(dt, 0)
                    combined_daily = daily_factor * (1 + bm_daily)
                    display_factor = round(combined_daily, 10)
                else:
                    display_factor = round(daily_factor, 10)

                row = {
                    "securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                    "calcType": calc_type, "walletId": wallet_id, "walletName": wallet_name,
                    "pricingType": pricing_type, "date": dt,
                    "benchmarkName": bm_name if calc_type == "inflacao_curva" else "",
                    "benchmarkFactor": display_factor,
                    "benchmarkYield": bm_yield_ann,
                    "pu": round(pu_val, 8),
                }
                if imp:
                    row["eventImpact"] = round(-imp, 8)
                results.append(row)
        else:
            results.append({"securityId": sec_id, "beehusName": sec_inp.get("beehusName", ""),
                            "calcType": calc_type, "error": f"Tipo de cálculo desconhecido: {calc_type}"})

    return results


# ── Saved lists (multiple named lists) ───────────────────────────────────────
# The "Preços na curva" tab persists multiple named lists of securities.
# Each list is keyed by its `name`; saving a name that already exists
# replaces the stored securities for that list.

@bp.route("/api/precificacao/lists")
def get_lists():
    """Return saved lists as `[{name, securities[]}, ...]`."""
    return jsonify(_load_lists())


@bp.route("/api/precificacao/lists", methods=["POST"])
def save_list():
    """Create or replace a named list. Body: `{name, securities}`.
    If a list with `name` already exists, its securities are overwritten;
    otherwise a new entry is appended."""
    data = request.get_json() or {}
    name = str(data.get("name") or "").strip()
    secs = data.get("securities", [])
    if not name:
        return jsonify({"error": "name é obrigatório"}), 400
    if not isinstance(secs, list):
        return jsonify({"error": "securities deve ser uma lista"}), 400

    lists = _load_lists()
    for entry in lists:
        if entry.get("name") == name:
            entry["securities"] = secs
            break
    else:
        lists.append({"name": name, "securities": secs})
    _write_lists(lists)
    return jsonify({"ok": True, "name": name, "count": len(secs)})


@bp.route("/api/precificacao/lists", methods=["DELETE"])
def delete_list():
    """Delete a single named list (pass `?name=<list_name>`)."""
    name = request.args.get("name", "").strip()
    if not name:
        return jsonify({"error": "name é obrigatório"}), 400
    lists = _load_lists()
    new_lists = [l for l in lists if l.get("name") != name]
    if len(new_lists) == len(lists):
        return jsonify({"error": "Lista não encontrada"}), 404
    _write_lists(new_lists)
    return jsonify({"ok": True})


# ── Config (benchmarks) ──────────────────────────────────────────────────────

@bp.route("/api/precificacao/config")
def get_config():
    cfg        = _load_config()
    benchmarks = cfg.get("benchmarks", [])

    enriched = []
    for b in benchmarks:
        entry = {"id": b["id"], "name": b.get("name", "")}
        try:
            oid = ObjectId(b["id"])
            row = _find_price(b["id"], {}, {"historyPrice": 1})
            entry["lastDate"] = str(_extract_hp(row).get("date", ""))[:10] if row else None
            sec = db.securities.find_one({"_id": oid}, {"beehusName": 1, "mainId": 1})
            if sec:
                entry["beehusName"] = sec.get("beehusName", "")
                entry["mainId"]     = sec.get("mainId", "")
        except Exception:
            entry["lastDate"] = None
        enriched.append(entry)
    return jsonify({"benchmarks": enriched})


@bp.route("/api/precificacao/config", methods=["POST"])
def save_config():
    data       = request.get_json() or {}
    benchmarks = data.get("benchmarks", [])
    clean = [
        {"id": str(b["id"]).strip(), "name": str(b.get("name", "")).strip()}
        for b in benchmarks if b.get("id") and b.get("name")
    ]
    _write_config({"benchmarks": clean})
    return jsonify({"ok": True})


# ── Export results to Excel ───────────────────────────────────────────────────

@bp.route("/api/precificacao/exportar", methods=["POST"])
def exportar():
    data    = request.get_json() or {}
    results = data.get("results", [])

    CALC_LABELS = {
        "pos_fixado":       "Pós-fixado",
        "pre_fixado_curva": "Pré-fixado Curva",
        "inflacao_curva":   "Benchmark +",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Precificação"

    headers = ["Ativo", "SecurityId", "Carteira", "Tipo Cálculo",
               "Data", "Benchmark", "Fator Diário", "Yield BM (% a.a.)",
               "PU Calculado", "Ajuste Evento"]
    ws.append(headers)

    hdr_fill = PatternFill("solid", fgColor="FCE4D6")
    hdr_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        if r.get("error"):
            continue
        ws.append([
            r.get("beehusName", ""),
            r.get("securityId", ""),
            r.get("walletName", ""),
            CALC_LABELS.get(r.get("calcType", ""), r.get("calcType", "")),
            r.get("date", ""),
            r.get("benchmarkName", ""),
            r.get("benchmarkFactor"),
            r.get("benchmarkYield"),
            r.get("pu"),
            r.get("eventImpact"),
        ])

    col_widths = [50, 28, 25, 20, 12, 15, 18, 18, 18, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for cell in ws["G"][1:]:
        cell.number_format = "#,##0.0000000000"
    for cell in ws["H"][1:]:
        cell.number_format = "0.0000%"
    for cell in ws["I"][1:]:
        cell.number_format = "#,##0.00000000"
    for cell in ws["J"][1:]:
        cell.number_format = "#,##0.00000000"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="precificacao.xlsx"'},
    )


# ── Bulk import via Excel ───────────────────────────────────────────────────
# The "Importar Excel" workflow lets the operator hand-edit a spreadsheet
# of securities + HTM lots and bulk-insert it into the Step 3 list. The
# template/upload pair share the same shape:
#
#   Sheet "Ativos" — one row per security (calcType + lookup hints + params)
#   Sheet "Lotes"  — one row per HTM lot, joined back to a security row by
#                    (securityId, walletId)
#
# Internal calcType ids (pos_fixado / pre_fixado_curva / inflacao_curva)
# are used in the sheet so the parser stays stable across UI label edits.
# Example rows ship with `SUBSTITUIR_*` placeholder IDs so a forgetful
# operator gets a clean "ativo não encontrado" error rather than silently
# importing junk values.

_UPLOAD_TEMPLATE_ATIVOS_COLS = [
    ("securityId", 30), ("walletId", 30), ("calcType", 20),
    ("indexerPercentual", 18), ("benchmarkName", 16),
    ("quantity", 14), ("yield", 14),
]
# `posPU` is intentionally NOT in the default template — the engine resolves
# it from processedPosition for the given walletId, mirroring what the
# selectSecurity() UI flow does. The parser still accepts a `posPU` column
# if the operator adds one manually for override scenarios.
#
# HTM lots: one row per lot. Rows sharing (securityId, walletId, calcType)
# are merged into a single security with multiple transactions. Lot dates
# default to 1900-01-01 since the engine doesn't use them (display only).

_VALID_CALC_TYPES = ("pos_fixado", "pre_fixado_curva", "inflacao_curva")


def _style_header_row(ws, ncols):
    fill = PatternFill("solid", fgColor="FCE4D6")
    font = Font(bold=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


@bp.route("/api/precificacao/upload-template")
def upload_template():
    """Return a .xlsx template the operator can fill out and re-upload.
    Three sheets: Ativos (one row per security), Lotes (one row per HTM
    lot), Instruções (column reference)."""
    wb = openpyxl.Workbook()

    ws_a = wb.active
    ws_a.title = "Ativos"
    ws_a.append([c for c, _ in _UPLOAD_TEMPLATE_ATIVOS_COLS])
    for i, (_, w) in enumerate(_UPLOAD_TEMPLATE_ATIVOS_COLS, 1):
        ws_a.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    _style_header_row(ws_a, len(_UPLOAD_TEMPLATE_ATIVOS_COLS))

    # One example per calcType. HTM types repeat the row to add more lots
    # (rows sharing securityId+walletId+calcType are merged on the server).
    # Columns: securityId | walletId | calcType | indexerPercentual | benchmarkName | quantity | yield
    ws_a.append(["SUBSTITUIR_PELO_ID_DO_ATIVO", "SUBSTITUIR_PELO_WALLET_ID", "pos_fixado",       100,  "CDI",   "",   ""])
    ws_a.append(["SUBSTITUIR_PELO_ID_DO_ATIVO", "SUBSTITUIR_PELO_WALLET_ID", "pre_fixado_curva", "",   "",      1000, 12.5])
    ws_a.append(["SUBSTITUIR_PELO_ID_DO_ATIVO", "SUBSTITUIR_PELO_WALLET_ID", "pre_fixado_curva", "",   "",      500,  13.0])
    ws_a.append(["SUBSTITUIR_PELO_ID_DO_ATIVO", "SUBSTITUIR_PELO_WALLET_ID", "inflacao_curva",   "",   "IPCA",  1000, 6.0])
    ws_a.append(["SUBSTITUIR_PELO_ID_DO_ATIVO", "SUBSTITUIR_PELO_WALLET_ID", "inflacao_curva",   "",   "IPCA",  500,  6.5])

    ws_i = wb.create_sheet("Instruções")
    ws_i.column_dimensions["A"].width = 110
    instr = [
        "INSTRUÇÕES DE PREENCHIMENTO",
        "",
        'Aba "Ativos" — uma linha por ativo (pos_fixado) ou por lote (HTM).',
        "  • securityId        : ID do ativo (obrigatório).",
        "  • walletId          : obrigatório para tipos HTM; opcional p/ pos_fixado.",
        "  • calcType          : pos_fixado, pre_fixado_curva ou inflacao_curva.",
        "  • indexerPercentual : usado em pos_fixado (% do indexador, ex.: 98).",
        "  • benchmarkName     : usado em pos_fixado e inflacao_curva. Resolvido",
        "                        case-insensitive contra os benchmarks configurados.",
        "  • quantity / yield  : usados nos tipos HTM (pre_fixado_curva, inflacao_curva).",
        "",
        "Para HTM com múltiplos lotes, repita a linha com o mesmo securityId+walletId",
        "alterando apenas quantity e yield. O servidor agrupa pelos três campos",
        "(securityId + walletId + calcType) e calcula o yield ponderado.",
        "A data do lote é fixada em 1900-01-01 (a engine não usa a data, só qty e yield).",
        "",
        "Os IDs no template (SUBSTITUIR_*) são placeholders — substitua antes do upload.",
    ]
    for line in instr:
        ws_i.append([line])
    ws_i.cell(1, 1).font = Font(bold=True, size=12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="precificacao_template.xlsx"'},
    )


def _read_excel_cell(row, header_map, name):
    i = header_map.get(name)
    if i is None or i >= len(row):
        return None
    v = row[i]
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


@bp.route("/api/precificacao/upload", methods=["POST"])
def upload_excel():
    """Parse an uploaded .xlsx (same shape as /upload-template) and return
    `{securities: [...], errors: [...]}` ready to merge into the Step 3
    list. All data lives in the "Ativos" sheet — HTM lots live in the
    same row as the security, and rows sharing (securityId, walletId,
    calcType) are merged into a single security with multiple lots.

    Lookups: securityId → ObjectId, fallback `mainId` if that column is
    present; benchmark by id then by name (case-insensitive) against
    configured benchmarks. Walletless rows are allowed for pos_fixado but
    rejected for HTM types, since the PU/lots come from
    processedPosition + transactions for that wallet."""
    if "file" not in request.files:
        return jsonify({"error": "Arquivo não enviado"}), 400
    f = request.files["file"]
    if not f.filename or not f.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Envie um arquivo .xlsx"}), 400

    # Optional reference date — when provided, drives the position lookup
    # (`processedPosition` on that date) and the price lookup (`_find_price_as_of`
    # — latest historyPrice with date <= ref) instead of "all-time most
    # recent". Empty / malformed values fall back to the legacy behaviour.
    ref_date = (request.form.get("referenceDate") or "").strip()
    if ref_date and not re.match(r"^\d{4}-\d{2}-\d{2}$", ref_date):
        return jsonify({"error": "referenceDate inválida (use YYYY-MM-DD)"}), 400

    try:
        wb = openpyxl.load_workbook(f.stream, read_only=True, data_only=True)
    except Exception as e:
        return jsonify({"error": f"Falha ao ler o arquivo: {e}"}), 400

    if "Ativos" not in wb.sheetnames:
        return jsonify({"error": 'Planilha "Ativos" não encontrada'}), 400

    def header_map(ws):
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not row:
            return {}
        return {str(h).strip(): i for i, h in enumerate(row) if h is not None}

    ws_a = wb["Ativos"]
    a_h  = header_map(ws_a)
    if "calcType" not in a_h:
        return jsonify({"error": "Coluna obrigatória ausente em 'Ativos': calcType"}), 400

    cfg          = _load_config()
    bms          = cfg.get("benchmarks", [])
    bms_by_id    = {b["id"]: b for b in bms if b.get("id")}
    bms_by_name  = {b["name"].lower(): b for b in bms if b.get("name")}
    # Wallet name lookup — used to backfill walletName on each entry so the
    # saved list (precificacao_lists.json) ends up with the human-readable
    # name instead of an empty string. The function caches 5 min internally.
    wallet_names = get_wallet_names()

    # Group rows by (sec_id, walletId, calcType). The first row of each
    # group seeds the entry; every row contributes a lot when calcType is
    # HTM. `order` preserves the input ordering for the response.
    groups = {}
    order  = []
    errors = []

    for row_idx, row in enumerate(ws_a.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue

        calc_type = (str(_read_excel_cell(row, a_h, "calcType") or "")).strip()
        if calc_type not in _VALID_CALC_TYPES:
            errors.append(f"Linha {row_idx}: calcType inválido ({calc_type or 'vazio'})")
            continue

        sec_id_raw = _read_excel_cell(row, a_h, "securityId")
        main_id    = _read_excel_cell(row, a_h, "mainId")
        sec_id_str = str(sec_id_raw).strip() if sec_id_raw else ""

        sec_doc = None
        if sec_id_str:
            oid = _to_oid_safe(sec_id_str)
            if oid is not None:
                sec_doc = db.securities.find_one({"_id": oid}, {
                    "beehusName": 1, "mainId": 1, "securityType": 1,
                    "indexer": 1, "indexerPercentual": 1,
                })
        if not sec_doc and main_id:
            sec_doc = db.securities.find_one({"mainId": str(main_id).strip()}, {
                "beehusName": 1, "mainId": 1, "securityType": 1,
                "indexer": 1, "indexerPercentual": 1,
            })
        if not sec_doc:
            errors.append(
                f"Linha {row_idx}: ativo não encontrado "
                f"(securityId={sec_id_str or '—'}, mainId={main_id or '—'})"
            )
            continue
        sec_id = str(sec_doc["_id"])

        wallet_id    = str(_read_excel_cell(row, a_h, "walletId")   or "").strip()
        sheet_wname  = str(_read_excel_cell(row, a_h, "walletName") or "").strip()
        # Sheet value wins (lets the operator override); fall back to the
        # DB lookup so the saved JSON gets a real name even when the column
        # isn't present in the template.
        wallet_name = sheet_wname or (wallet_names.get(wallet_id, "") if wallet_id else "")
        key         = (sec_id, wallet_id, calc_type)

        if key not in groups:
            # First row of this group — seed the entry with full metadata.
            entry = {
                "id":                sec_id,
                "beehusName":        sec_doc.get("beehusName", ""),
                "mainId":            sec_doc.get("mainId", ""),
                "securityType":      sec_doc.get("securityType", ""),
                "indexer":           sec_doc.get("indexer", ""),
                "indexerPercentual": sec_doc.get("indexerPercentual"),
                "calcType":          calc_type,
                "walletId":          wallet_id,
                "walletName":        wallet_name,
                "pricingType":       "",
                "positionDate":      str(_read_excel_cell(row, a_h, "positionDate") or "")[:10],
                "posPU":             None,
                "lastPU":            None,
                "lastPUDate":        None,
                "benchmarkId":       "",
                "benchmarkName":     "",
                "transactions":      [],
                "weightedYield":     None,
            }

            bm_id_in   = _read_excel_cell(row, a_h, "benchmarkId")
            bm_name_in = _read_excel_cell(row, a_h, "benchmarkName")
            bm = None
            if bm_id_in and str(bm_id_in).strip() in bms_by_id:
                bm = bms_by_id[str(bm_id_in).strip()]
            elif bm_name_in and str(bm_name_in).strip().lower() in bms_by_name:
                bm = bms_by_name[str(bm_name_in).strip().lower()]
            if bm:
                entry["benchmarkId"]   = bm["id"]
                entry["benchmarkName"] = bm.get("name", "")

            pu_sheet = _read_excel_cell(row, a_h, "posPU")
            if pu_sheet is not None:
                try:
                    entry["posPU"] = float(pu_sheet)
                except (TypeError, ValueError):
                    pass

            # Quando o operador informou `referenceDate`, ela tem prioridade
            # sobre o `positionDate` da planilha — a busca de posição usa ela
            # exata (sem fallback para a mais recente do banco, que poderia
            # carregar PU de outro dia). Em modo legado (sem ref_date), mantém
            # o comportamento antigo de "mais recente" via _get_most_recent_position.
            pos_lookup_date = ref_date or entry["positionDate"] or None
            need_pos_lookup = (
                wallet_id
                and (entry["posPU"] is None or not entry["positionDate"])
            )
            if need_pos_lookup:
                pos_doc, posdate = _get_most_recent_position(
                    wallet_id, pos_lookup_date,
                )
                if pos_doc:
                    if not entry["positionDate"]:
                        entry["positionDate"] = posdate
                    for ps in pos_doc.get("securities", []):
                        if str(ps.get("securityId", "")) != sec_id:
                            continue
                        if entry["posPU"] is None and ps.get("pu") is not None:
                            try:
                                entry["posPU"] = float(ps["pu"])
                            except (TypeError, ValueError):
                                pass
                        if not entry["pricingType"]:
                            entry["pricingType"] = ps.get("pricingType", "")
                        break

            # Price lookup: "as of `ref_date`" when informado, senão all-time
            # most recent (legacy). `_find_price_as_of` cobre ambos os modos.
            hp = _find_price_as_of(sec_id, ref_date)
            if hp.get("value") is not None:
                try:
                    entry["lastPU"] = float(hp["value"])
                except (TypeError, ValueError):
                    pass
            if hp.get("date"):
                entry["lastPUDate"] = str(hp["date"])[:10]

            if calc_type == "pos_fixado":
                idx_in = _read_excel_cell(row, a_h, "indexerPercentual")
                if idx_in is None:
                    idx_in = sec_doc.get("indexerPercentual")
                try:
                    entry["indexerPercentual"] = float(idx_in) if idx_in is not None else None
                except (TypeError, ValueError):
                    entry["indexerPercentual"] = None

            groups[key] = entry
            order.append(key)

        # Collect lot from every HTM row (including the seeding one).
        if calc_type in ("pre_fixado_curva", "inflacao_curva"):
            qty = _read_excel_cell(row, a_h, "quantity")
            yld = _read_excel_cell(row, a_h, "yield")
            if qty is None and yld is None:
                continue  # blank lot columns — silently skip
            try:
                qf = float(qty) if qty is not None else 0.0
                yf = float(yld) if yld is not None else None
            except (TypeError, ValueError):
                errors.append(f"Linha {row_idx}: quantity/yield inválidos")
                continue
            if qf <= 0 or yf is None:
                errors.append(f"Linha {row_idx}: lote inválido (quantity > 0 e yield obrigatórios)")
                continue
            groups[key]["transactions"].append({
                "date":     "1900-01-01",
                "quantity": qf,
                "yield":    yf,
            })

    # Finalize each group: validate required fields and compute weightedYield.
    securities = []
    for key in order:
        entry = groups[key]
        ct    = entry["calcType"]
        sid   = entry["id"]

        if ct == "pos_fixado":
            if not entry["benchmarkId"]:
                errors.append(f"Ativo {sid}: benchmark obrigatório para pos_fixado")
                continue
        elif ct in ("pre_fixado_curva", "inflacao_curva"):
            if not entry["walletId"]:
                errors.append(f"Ativo {sid}: walletId obrigatório para tipos HTM")
                continue
            valid = [
                l for l in entry["transactions"]
                if l["quantity"] > 0 and l["yield"] is not None
            ]
            if not valid:
                errors.append(f"Ativo {sid}: nenhum lote válido (preencha quantity e yield)")
                continue
            entry["transactions"] = valid
            total = sum(l["quantity"] for l in valid)
            entry["weightedYield"] = (
                sum(l["quantity"] * l["yield"] for l in valid) / total
                if total else None
            )
            if ct == "inflacao_curva" and not entry["benchmarkId"]:
                errors.append(f"Ativo {sid}: benchmark obrigatório para inflacao_curva")
                continue

        securities.append(entry)

    return jsonify({"securities": securities, "errors": errors})


@bp.route("/api/precificacao/refresh-list", methods=["POST"])
def refresh_list():
    """Reapply position + price lookups on the **current Step-3 list** for a
    given reference date. Body: `{ securities: [...], referenceDate:
    "YYYY-MM-DD" }`. The endpoint walks the input list and re-fetches
    `posPU` / `positionDate` / `pricingType` (from `processedPosition` on
    that date) and `lastPU` / `lastPUDate` (from `_find_price_as_of`),
    overwriting those four/five fields on each entry while preserving
    everything else (transactions, calcType, walletId, benchmark, etc.).
    Returns `{securities: [...], errors: [...]}` — the FE merges by
    `(id, calcType, walletId)` so a refresh never duplicates rows."""
    body = request.get_json(silent=True) or {}
    securities = body.get("securities") or []
    ref_date   = (body.get("referenceDate") or "").strip()
    if not isinstance(securities, list):
        return jsonify({"error": "securities deve ser uma lista"}), 400
    if not ref_date or not re.match(r"^\d{4}-\d{2}-\d{2}$", ref_date):
        return jsonify({"error": "referenceDate obrigatória (YYYY-MM-DD)"}), 400

    out, errors = [], []
    for idx, entry_in in enumerate(securities, start=1):
        if not isinstance(entry_in, dict):
            errors.append(f"Linha {idx}: entrada inválida (não é objeto)")
            continue
        sec_id    = str(entry_in.get("id") or "").strip()
        wallet_id = str(entry_in.get("walletId") or "").strip()
        if not sec_id:
            errors.append(f"Linha {idx}: ativo sem id")
            out.append(dict(entry_in))
            continue
        # Copia o entry original e sobrescreve apenas os campos dependentes
        # da data — calcType, transactions, indexerPercentual, benchmark,
        # weightedYield, etc. permanecem intactos.
        entry = dict(entry_in)
        # Reset dos campos que serão recalculados pela data nova.
        entry["positionDate"] = ""
        entry["posPU"]        = None
        entry["pricingType"]  = ""
        entry["lastPU"]       = None
        entry["lastPUDate"]   = None

        if wallet_id:
            pos_doc, posdate = _get_most_recent_position(wallet_id, ref_date)
            if pos_doc:
                entry["positionDate"] = posdate
                for ps in pos_doc.get("securities", []):
                    if str(ps.get("securityId", "")) != sec_id:
                        continue
                    if ps.get("pu") is not None:
                        try:
                            entry["posPU"] = float(ps["pu"])
                        except (TypeError, ValueError):
                            pass
                    entry["pricingType"] = ps.get("pricingType", "")
                    break

        hp = _find_price_as_of(sec_id, ref_date)
        if hp.get("value") is not None:
            try:
                entry["lastPU"] = float(hp["value"])
            except (TypeError, ValueError):
                pass
        if hp.get("date"):
            entry["lastPUDate"] = str(hp["date"])[:10]

        out.append(entry)

    return jsonify({
        "securities":    out,
        "errors":        errors,
        "referenceDate": ref_date,
    })


def _to_oid_safe(s):
    try:
        return ObjectId(s)
    except (InvalidId, TypeError):
        return None

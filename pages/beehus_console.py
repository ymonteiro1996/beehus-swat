"""Routes for the Beehus API console page.

Layout (one logical operation per group):
    /beehus                                — page
    /api/beehus/token         GET/POST/DELETE
    /api/beehus/transactions  POST           — create
    /api/beehus/transactions/<id>  DELETE   — delete one (proxies to API)
    /api/beehus/transactions/<id>  PATCH    — update one (proxies to API)
    /api/beehus/transactions/search POST    — list local mongo by filters
    /api/beehus/identify-transactions/config GET/PUT — types-needing-security list
    /api/beehus/identify-transactions/identify POST  — suggest type/security per id
                                                       (+ executionPrice/IRRF per row)
    /api/beehus/identify-transactions/execution-extras POST
                                            — push executionPrices + create IRRF
                                              `taxes` transactions on Implementar
    /api/beehus/positions/process   POST    — trigger processed-position run
    /api/beehus/positions/delete    POST    — delete processed positions
    /api/beehus/provisions          POST    — create a provision
    /api/beehus/provisions/search   POST    — list local mongo by filters
    /api/beehus/provisions/<id>     DELETE  — delete one (proxies to API)
    /api/beehus/execution-prices    POST    — create an execution price
    /api/beehus/util/parse-dates-excel POST — extract sorted YYYY-MM-DD list
                                              from an uploaded .xlsx (no header)
    /api/beehus/util/parse-strings-excel POST — extract deduped string values
                                              from an uploaded .xlsx (no header)
    /api/beehus/nav/calculate-wallets POST  — recalculate NAV contribution
    /api/beehus/nav/explosion-proportions POST — recalculate explosion proportions
    /api/beehus/nav/calculate-groupings POST — recalculate NAV at grouping level
    /api/beehus/nav/publish         POST    — publish NAV results (PATCH upstream)
    /api/beehus/nav/unpublish       POST    — unpublish NAV results (PATCH upstream)
    /api/beehus/filters/groupings-by-publish-state
                                    GET ?companyId=&positionDate=&published=true|false
    /api/beehus/filters/companies   GET
    /api/beehus/filters/groupings   GET ?companyId=
    /api/beehus/filters/wallets     GET ?companyId=&groupingId=
    /api/beehus/filters/wallets-with-position
                                    GET ?companyId=&positionDate=
    /api/beehus/filters/grouping-return-deltas
                                    GET ?companyId=&positionDate=&published=
    /api/beehus/filters/entities    GET ?companyId=
    /api/beehus/filters/securities  GET
"""
import json
import logging
import os
import threading
from datetime import datetime

from bson import ObjectId
from bson.errors import InvalidId
import beehus_catalog
from flask import Blueprint, render_template, jsonify, make_response, request

from beehus_api import (
    BeehusAPIError,
    BeehusAuthError,
    calculate_nav_groupings,
    calculate_nav_wallets,
    create_execution_price,
    create_provision,
    create_transaction,
    delete_provision,
    delete_transaction,
    delete_processed_position,
    update_transaction,
    process_processed_position,
    proportion_explosion,
    publish_nav,
    unpublish_nav,
    set_token,
    clear_token,
    token_status,
    verify_token,
)
from db import (
    atomic_write_json,
    biz_days_between,
    business_days_before,
    db,
    company_visible,
    get_company_filter,
    get_company_names,
    get_entity_names,
    get_grouping_index,
    get_security_names,
    get_wallet_names,
    resolve_wallet,
    _cached_ttl,
)

bp = Blueprint("beehus_console", __name__)


# ── Page ──────────────────────────────────────────────────────────────────────

@bp.route("/beehus")
def index():
    return render_template("beehus_console.html")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _api_error_response(e: BeehusAPIError):
    status = 401 if isinstance(e, BeehusAuthError) else 502
    return jsonify({
        "error": str(e),
        "upstream_status": e.status,
        "upstream_body": e.body,
    }), status


def _sorted_dicts_to_list(d: dict, name_key="name"):
    return sorted(
        [{"id": k, name_key: v or k} for k, v in d.items()],
        key=lambda x: (x[name_key] or "").lower(),
    )


# ── Token ─────────────────────────────────────────────────────────────────────

@bp.route("/api/beehus/token", methods=["GET"])
def token_get():
    return jsonify(token_status())


@bp.route("/api/beehus/token", methods=["POST"])
def token_set():
    data = request.get_json(silent=True) or {}
    try:
        set_token(data.get("token", ""))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    # Validate the pasted token against the API immediately (cheap authenticated
    # GET) so an invalid/expired token is rejected here instead of failing
    # silently on every later page read. verify_token() also sets the in-process
    # `rejected` flag (via the 401 path), which the global banner picks up.
    try:
        verify_token()
    except BeehusAuthError as e:
        return jsonify({
            "error": "Token rejeitado pela API (401/403). Verifique se copiou o token de hoje por completo.",
            "upstream_status": e.status,
        }), 401
    except BeehusAPIError as e:
        # Token might be fine but the API is unreachable/erroring — don't block
        # the save on that; surface a soft warning and let the banner reflect state.
        return jsonify({**token_status(), "warning": f"Não foi possível validar agora: {e}"})
    return jsonify(token_status())


@bp.route("/api/beehus/token", methods=["DELETE"])
def token_clear():
    clear_token()
    return jsonify(token_status())


# ── Filters (local Mongo) ─────────────────────────────────────────────────────

def _companies_empty_reason():
    """Por que `/filters/companies` veio vazio.

    Sonda `list_companies()` direto (sem cache) — só no caminho vazio, então não
    pesa no fluxo normal — para distinguir token ausente/expirado
    (`BeehusAuthError`, levantado tanto sem token quanto em 401/403) de falha de
    rede (`BeehusAPIError`) ou de uma resposta legitimamente vazia."""
    from beehus_api import list_companies
    try:
        list_companies()
    except BeehusAuthError:
        return "Token expirado ou ausente. Cole o token atual na página /beehus."
    except BeehusAPIError:
        return "Falha ao consultar a API de empresas. Tente novamente em instantes."
    except Exception:  # noqa: BLE001
        return "Não foi possível carregar as empresas."
    return ""  # API respondeu OK, porém sem empresas


@bp.route("/api/beehus/filters/companies")
def filter_companies():
    cf = get_company_filter()
    names = get_company_names()
    items = [{"id": cid, "name": name or cid}
             for cid, name in names.items()
             if not cf or cid in cf]
    items.sort(key=lambda x: x["name"].lower())
    if items:
        return jsonify(items)
    # Vazio: SEMPRE devolve o array (mantém compat com todos os consumidores que
    # fazem `r.body.map(...)` — ferramentas do beehus_console). O MOTIVO vai num
    # header, lido apenas pelo dropdown que quer mostrá-lo (Painel → Day-trade).
    reason = (
        # API trouxe empresas, mas o filtro de visibilidade (Configurações)
        # excluiu todas — não é problema de token.
        "Nenhuma empresa no filtro de visibilidade. Ajuste em Configurações."
        if names else _companies_empty_reason()
    )
    resp = make_response(jsonify(items))
    if reason:
        resp.headers["X-Companies-Empty-Reason"] = reason
    return resp


@bp.route("/api/beehus/filters/groupings")
def filter_groupings():
    """Groupings registered for the given company.

    Filters directly on `groupings.companyId` (the field is stored on every
    grouping document). `trashed` groupings are excluded.
    """
    company_id = request.args.get("companyId", "")
    if not company_visible(company_id):
        return jsonify([])

    items = []
    for gid, g in get_grouping_index().items():
        if g.get("trashed"):
            continue
        if g.get("companyId") != company_id:
            continue
        items.append({
            "id":        gid,
            "name":      g["name"] or gid,
            "walletIds": list(g["walletIds"]),
        })
    items.sort(key=lambda x: x["name"].lower())
    return jsonify(items)


@bp.route("/api/beehus/filters/wallets")
def filter_wallets():
    """Wallets in the company, optionally narrowed by grouping."""
    company_id  = request.args.get("companyId", "")
    grouping_id = request.args.get("groupingId", "") or None
    if not company_visible(company_id):
        return jsonify([])

    # Resolve cached maps OUTSIDE the loader (no nested cache lock).
    wallet_names = get_wallet_names()
    entity_names = get_entity_names()
    gindex = get_grouping_index().get(grouping_id) if grouping_id else None

    def _build():
        allowed_ids = None
        if grouping_id:
            if not gindex:
                return []
            allowed_ids = {str(w) for w in gindex["walletIds"]}

        items = []
        for w in beehus_catalog.wallets_in_company(company_id):
            wid = str(w["_id"])
            if allowed_ids is not None and wid not in allowed_ids:
                continue
            eid = beehus_catalog.id_str(w["entityId"]) if w.get("entityId") else ""
            items.append({
                "id":         wid,
                "name":       wallet_names.get(wid, "") or wid,
                "entityId":   eid,
                "entityName": entity_names.get(eid, "") if eid else "",
            })
        items.sort(key=lambda x: x["name"].lower())
        return items

    # Per (company, grouping) result is stable within the 5-min TTL.
    return jsonify(_cached_ttl(("wallets_filter", company_id, grouping_id or ""), _build))


@bp.route("/api/beehus/filters/groupings-by-publish-state")
def filter_groupings_by_publish_state():
    """Groupings of `companyId` whose `navPackages` for `positionDate` match
    the requested `published` state.

    Used by the Publicar Agrupamentos / Despublicar Agrupamentos views to
    only let the user act on groupings that are actually in the right state.
    Each grouping on a given date is cleanly either fully published or fully
    unpublished (no mixed state observed in production), so a single
    `distinct(groupingId)` query is sufficient.
    """
    company_id    = request.args.get("companyId", "")
    position_date = request.args.get("positionDate", "")
    published_str = (request.args.get("published") or "").lower()
    if published_str not in ("true", "false"):
        return jsonify({"error": "published must be 'true' or 'false'"}), 400
    published = (published_str == "true")
    if not company_visible(company_id) or not position_date:
        return jsonify([])

    # navPackages de nível agrupamento via cache consolidado da empresa. Em
    # produção há 1 doc por agrupamento/data (validado), então `distinct` vira
    # um set sobre os docs filtrados pelo estado `published`. O cache já é
    # não-trashed. `published` casa estritamente (campo ausente/null não casa
    # nem true nem false, igual ao Mongo).
    eligible = set()
    for d in beehus_catalog.nav_grouping_docs(company_id, position_date):
        pub = d.get("published")
        if pub is None or bool(pub) != published:
            continue
        gid = beehus_catalog.id_str(d.get("groupingId"))
        if gid:
            eligible.add(gid)
    if not eligible:
        return jsonify([])

    gindex = get_grouping_index()
    items = []
    for gid in eligible:
        g = gindex.get(gid)
        if not g or g.get("trashed"):
            continue
        # Defense in depth: only return groupings that belong to the company.
        if g.get("companyId") and g["companyId"] != company_id:
            continue
        items.append({
            "id":        gid,
            "name":      g["name"] or gid,
            "walletIds": list(g["walletIds"]),
        })
    items.sort(key=lambda x: x["name"].lower())
    return jsonify(items)


@bp.route("/api/beehus/filters/grouping-return-deltas")
def filter_grouping_return_deltas():
    """Per-grouping `|returnNavPerShare - returnContribution|` from navPackages.

    Returns one row per grouping (matching the requested `published` state),
    aggregated as the **worst wallet** within that grouping — i.e., the
    navPackage doc whose `|returnNavPerShare - returnContribution|` is
    largest. The reported `returnNavPerShare` / `returnContribution` come
    from that worst-wallet doc; this is the most actionable signal for
    deciding whether the grouping is safe to publish.

    Query params:
        companyId    (str, required)
        positionDate (str, required, YYYY-MM-DD)
        published    (str, optional — 'true'/'false', default 'false')

    Returns: [{groupingId, groupingName,
               returnNavPerShare, returnContribution, deltaAbs}]
    Sorted by `deltaAbs` desc (None values last).
    """
    company_id    = request.args.get("companyId", "")
    position_date = request.args.get("positionDate", "")
    published_str = (request.args.get("published") or "false").lower()
    if published_str not in ("true", "false", "all"):
        return jsonify({"error": "published must be 'true', 'false', or 'all'"}), 400
    fetch_all = (published_str == "all")
    published = (published_str == "true") if not fetch_all else None
    if not company_visible(company_id) or not position_date:
        return jsonify([])

    # navPackages de nível agrupamento via cache consolidado da empresa (1 doc
    # por agrupamento/data em produção, validado). Não-trashed já garantido;
    # quando não é `all`, filtra pelo estado `published` (estrito como o Mongo).
    cursor = []
    for d in beehus_catalog.nav_grouping_docs(company_id, position_date):
        if not fetch_all:
            pub = d.get("published")
            if pub is None or bool(pub) != published:
                continue
        cursor.append(d)

    # For each grouping, keep the wallet doc with the largest |Δ|. A grouping
    # with at least one numeric pair wins over one with only None pairs (so
    # the response prefers actionable rows when both exist). When fetch_all,
    # a grouping is considered published if ANY of its navPackage docs is.
    by_grouping: dict[str, dict] = {}
    for d in cursor:
        gid  = str(d.get("groupingId") or "")
        if not gid:
            continue
        rnps = d.get("returnNavPerShare")
        rc   = d.get("returnContribution")
        pub  = bool(d.get("published")) if fetch_all else None
        delta_abs = None
        if isinstance(rnps, (int, float)) and isinstance(rc, (int, float)):
            delta_abs = abs(float(rnps) - float(rc))

        cur = by_grouping.get(gid)
        if cur is None:
            by_grouping[gid] = {"rnps": rnps, "rc": rc, "deltaAbs": delta_abs, "published": pub}
            continue
        # any published wallet → grouping is published
        if fetch_all and pub:
            cur["published"] = True
        # Prefer rows with a known delta; among those, keep the largest.
        cur_delta = cur["deltaAbs"]
        if delta_abs is None:
            continue
        if cur_delta is None or delta_abs > cur_delta:
            cur["rnps"], cur["rc"], cur["deltaAbs"] = rnps, rc, delta_abs

    gindex = get_grouping_index()
    items = []
    for gid, info in by_grouping.items():
        item: dict = {
            "groupingId":         gid,
            "groupingName":       (gindex.get(gid) or {}).get("name", "") or gid,
            "returnNavPerShare":  info["rnps"],
            "returnContribution": info["rc"],
            "deltaAbs":           info["deltaAbs"],
        }
        if fetch_all:
            item["published"] = bool(info.get("published"))
        items.append(item)
    items.sort(key=lambda x: (x["deltaAbs"] is None, -(x["deltaAbs"] or 0.0)))
    return jsonify(items)


@bp.route("/api/beehus/filters/wallets-with-position")
def filter_wallets_with_position():
    """Wallets in the company that have a processed position for the given date.

    Used by the Excluir Posições view to limit the Available pane to wallets
    that actually have a processed position to delete on that date.

    Source: pre-processing endpoint E (`processedWalletsDetailed`) via
    `beehus_catalog.wallets_with_position` — the company-wide set of wallets
    processed on that date. (Replaces the old direct `processedPosition` read.)
    """
    company_id    = request.args.get("companyId", "")
    position_date = request.args.get("positionDate", "")
    if not company_visible(company_id) or not position_date:
        return jsonify([])

    eligible = beehus_catalog.wallets_with_position(company_id, position_date)
    if not eligible:
        return jsonify([])

    # Prefer the UI's cached wallet name (keeps labels consistent with the rest
    # of the console); fall back to the name the endpoint returned, then the id.
    wallet_names = get_wallet_names()
    items = [
        {"id": w["id"], "name": wallet_names.get(w["id"], "") or w.get("name") or w["id"]}
        for w in eligible
    ]
    items.sort(key=lambda x: x["name"].lower())
    return jsonify(items)


@bp.route("/api/beehus/filters/entities")
def filter_entities():
    """Entities reachable through this company's wallets."""
    company_id = request.args.get("companyId", "")
    if not company_visible(company_id):
        return jsonify([])

    # Resolve the cached name map OUTSIDE the loader (no nested cache lock).
    enames = get_entity_names()

    def _build():
        eids = {
            beehus_catalog.id_str(w["entityId"])
            for w in beehus_catalog.wallets_in_company(company_id)
            if w.get("entityId")
        }
        items = [{"id": eid, "name": enames.get(eid, "") or eid} for eid in eids]
        items.sort(key=lambda x: x["name"].lower())
        return items

    # Per-company result is stable within the 5-min TTL; cache it so reopening
    # the same company doesn't re-scan its wallets + rebuild + sort each time.
    return jsonify(_cached_ttl(("entities_filter", company_id), _build))


@bp.route("/api/beehus/filters/securities")
def filter_securities():
    # The response is a sorted list of ALL ~16k securities — identical for
    # every caller and changing only as often as the underlying catalog
    # (5-min TTL). Building + sorting + serialising 16k dicts per request cost
    # ~1s; cache the finished list so that work runs at most once per TTL.
    # Each item carries {id, name, mainId}: the security-edit modal's free-text
    # search shows the mainId column and a manual pick keeps its mainId (see
    # pickSecForEdit), so the suggestion cell renders "mainId · name" instead of
    # the bare beehusName. The index read uses beehus_catalog's own cache (a
    # different lock than this `_cached_ttl`), so no lock nesting occurs.
    def _build():
        idx = beehus_catalog.securities_index()
        items = [
            {"id": sid,
             "name": d.get("beehusName") or sid,
             "mainId": d.get("mainId") or ""}
            for sid, d in idx.items()
        ]
        items.sort(key=lambda x: (x["name"] or "").lower())
        return items

    return jsonify(_cached_ttl("securities_filter_list", _build))


# ── Transactions ──────────────────────────────────────────────────────────────

@bp.route("/api/beehus/transactions", methods=["POST"])
def transactions_create():
    data = request.get_json(silent=True) or {}
    required = ("companyId", "entityId", "walletId", "balance",
                "operationDate", "liquidationDate")
    missing = [k for k in required if data.get(k) in (None, "")]
    if missing:
        return jsonify({"error": f"missing fields: {', '.join(missing)}"}), 400
    try:
        result = create_transaction(
            company_id=data["companyId"],
            entity_id=data["entityId"],
            wallet_id=data["walletId"],
            balance=float(data["balance"]),
            operation_date=data["operationDate"],
            liquidation_date=data["liquidationDate"],
            currency_id=data.get("currencyId", "BRL"),
            transaction_type=data.get("beehusTransactionType", "withdrawalDeposit"),
            description=data.get("description", ""),
            comment=data.get("comment", ""),
            hide=bool(data.get("hide", False)),
        )
    except BeehusAPIError as e:
        return _api_error_response(e)
    return jsonify(result), 201


@bp.route("/api/beehus/transactions/<txn_id>", methods=["DELETE"])
def transactions_delete(txn_id):
    try:
        result = delete_transaction(txn_id)
    except BeehusAPIError as e:
        return _api_error_response(e)
    return jsonify({"deleted": txn_id, "response": result})


_TXN_PATCHABLE = {
    "balance",
    "beehusTransactionType",
    "currencyId",
    "description",
    "entityId",
    "liquidationDate",
    "operationDate",
    "securityId",
}


@bp.route("/api/beehus/transactions/<txn_id>", methods=["PATCH"])
def transactions_patch(txn_id):
    """Forward a partial PATCH to the upstream Beehus API.

    Body is a JSON object containing any subset of the patchable fields
    (balance, beehusTransactionType, currencyId, description, entityId,
    liquidationDate, operationDate, securityId). Unknown keys are dropped
    so that a typo in the UI does not poison the upstream request."""
    data = request.get_json(silent=True) or {}
    patch = {k: v for k, v in data.items() if k in _TXN_PATCHABLE}
    if not patch:
        return jsonify({"error": "no patchable fields in body"}), 400

    try:
        result = update_transaction(txn_id, patch)
    except BeehusAPIError as e:
        return _api_error_response(e)

    # A API (upstream) é a fonte de verdade. Não há mais espelho no Mongo local
    # — o cache de leitura saiu; os listings vêm do endpoint G via beehus_catalog.
    return jsonify({
        "updated":  txn_id,
        "patch":    patch,
        "response": result,
    })


# ── Positions ─────────────────────────────────────────────────────────────────

@bp.route("/api/beehus/positions/process", methods=["POST"])
def positions_process():
    """Trigger processed-position processing for a company on a given date.

    Body:
        companyId    (str, required)
        positionDate (str, required, YYYY-MM-DD)
        wallets      (list[str], optional) — restricts processing to these
                     wallet ids. Empty list means "all wallets in the company"
                     per the upstream API contract.

    Antes de disparar o processamento, cada carteira pedida é expandida com a
    sua CADEIA DE EXPLOSÃO (carteiras que ela arrasta em todos os níveis) — sem
    isso o processamento não conclui. A expansão é feita aqui (ponto único por
    onde todo `process` passa), então vale para data única, faixa e demais
    chamadores. `wallets` vazio ("todas") não precisa expansão.
    """
    data = request.get_json(silent=True) or {}
    company_id    = (data.get("companyId") or "").strip()
    position_date = (data.get("positionDate") or "").strip()
    wallets       = data.get("wallets") or []

    if not company_id or not position_date:
        return jsonify({"error": "companyId and positionDate are required"}), 400
    if not company_visible(company_id):
        return jsonify({"error": "company is not visible to this user"}), 403
    if not isinstance(wallets, list) or not all(isinstance(w, str) for w in wallets):
        return jsonify({"error": "wallets must be a list of strings"}), 400

    # Arrasta a cadeia de explosão de cada carteira pedida (no-op quando vazio).
    effective = beehus_catalog.expand_wallets_with_explosion(company_id, wallets)
    dragged = [w for w in effective if w not in set(wallets)]

    try:
        result = process_processed_position(
            company_id=company_id,
            position_date=position_date,
            wallets=effective,
        )
    except BeehusAPIError as e:
        return _api_error_response(e)
    out = result if isinstance(result, dict) else {"ok": True}
    if dragged:
        out["draggedWallets"] = dragged
    return jsonify(out)


@bp.route("/api/beehus/positions/explosion-chain", methods=["GET"])
def positions_explosion_chain():
    """Preview das carteiras arrastadas pela cadeia de explosão de uma carteira.

    Query:
        companyId (str, required)
        walletId  (str, required)

    Devolve `{"chain": [{securityId, walletId, name, level, viaWalletId}, ...]}`
    — lista achatada, em todos os níveis, vazia quando não há explosão. Alimenta
    o modal de confirmação antes de processar (Painel de Controle > Processar).
    """
    company_id = (request.args.get("companyId") or "").strip()
    wallet_id  = (request.args.get("walletId") or "").strip()
    if not company_id or not wallet_id:
        return jsonify({"error": "companyId and walletId are required"}), 400
    if not company_visible(company_id):
        return jsonify({"error": "company is not visible to this user"}), 403
    try:
        chain = beehus_catalog.explosion_chain(company_id, wallet_id)
    except BeehusAPIError as e:
        return _api_error_response(e)
    return jsonify({"chain": chain})


@bp.route("/api/beehus/positions/delete", methods=["POST"])
def positions_delete():
    """Delete processed positions for a company on a given date.

    Body:
        companyId    (str, required)
        positionDate (str, required, YYYY-MM-DD)
        walletIds    (list[str], optional) — restricts deletion to these
                     wallet ids. Empty list means "all wallets in the company"
                     per the upstream API contract.

    Mounted as POST locally because DELETE requests with a JSON body are
    awkward for the browser fetch API; the upstream call is still a real
    DELETE on the Beehus API.
    """
    data = request.get_json(silent=True) or {}
    company_id    = (data.get("companyId") or "").strip()
    position_date = (data.get("positionDate") or "").strip()
    wallet_ids    = data.get("walletIds") or []

    if not company_id or not position_date:
        return jsonify({"error": "companyId and positionDate are required"}), 400
    if not company_visible(company_id):
        return jsonify({"error": "company is not visible to this user"}), 403
    if not isinstance(wallet_ids, list) or not all(isinstance(w, str) for w in wallet_ids):
        return jsonify({"error": "walletIds must be a list of strings"}), 400

    try:
        result = delete_processed_position(
            company_id=company_id,
            position_date=position_date,
            wallet_ids=wallet_ids,
        )
    except BeehusAPIError as e:
        return _api_error_response(e)
    return jsonify(result if result is not None else {"ok": True})


# ── NAV calculation ───────────────────────────────────────────────────────────

@bp.route("/api/beehus/nav/calculate-wallets", methods=["POST"])
def nav_calculate_wallets():
    """Trigger NAV-contribution recalculation for a company on a given date.

    Body:
        companyId    (str, required)
        positionDate (str, required, YYYY-MM-DD)
        wallets      (list[str], optional) — restricts calculation to these
                     wallet ids. Empty list means "all wallets in the company"
                     per the upstream API contract.
    """
    data = request.get_json(silent=True) or {}
    company_id    = (data.get("companyId") or "").strip()
    position_date = (data.get("positionDate") or "").strip()
    wallets       = data.get("wallets") or []

    if not company_id or not position_date:
        return jsonify({"error": "companyId and positionDate are required"}), 400
    if not company_visible(company_id):
        return jsonify({"error": "company is not visible to this user"}), 403
    if not isinstance(wallets, list) or not all(isinstance(w, str) for w in wallets):
        return jsonify({"error": "wallets must be a list of strings"}), 400

    try:
        result = calculate_nav_wallets(
            company_id=company_id,
            position_date=position_date,
            wallets=wallets,
        )
    except BeehusAPIError as e:
        return _api_error_response(e)
    # NAV mudou upstream → invalida o cache consolidado da empresa (a grade de
    # divergência e os filtros de publicação leem de beehus_catalog.nav_packages).
    beehus_catalog.invalidate_nav(company_id)
    return jsonify(result if result is not None else {"ok": True})


@bp.route("/api/beehus/nav/explosion-proportions", methods=["POST"])
def nav_explosion_proportions():
    """Trigger recalculation of explosion proportions for groupings.

    Body:
        companyId    (str, required)
        positionDate (str, required, YYYY-MM-DD)
        groupings    (list[str], optional) — restricts calculation to these
                     grouping ids. Empty list means "all groupings in the
                     company" per the upstream API contract.
    """
    data = request.get_json(silent=True) or {}
    company_id    = (data.get("companyId") or "").strip()
    position_date = (data.get("positionDate") or "").strip()
    groupings     = data.get("groupings") or []

    if not company_id or not position_date:
        return jsonify({"error": "companyId and positionDate are required"}), 400
    if not company_visible(company_id):
        return jsonify({"error": "company is not visible to this user"}), 403
    if not isinstance(groupings, list) or not all(isinstance(g, str) for g in groupings):
        return jsonify({"error": "groupings must be a list of strings"}), 400

    try:
        result = proportion_explosion(
            company_id=company_id,
            position_date=position_date,
            groupings=groupings,
        )
    except BeehusAPIError as e:
        return _api_error_response(e)
    beehus_catalog.invalidate_nav(company_id)
    return jsonify(result if result is not None else {"ok": True})


@bp.route("/api/beehus/nav/calculate-groupings", methods=["POST"])
def nav_calculate_groupings():
    """Trigger NAV-contribution recalculation at the grouping level.

    Body:
        companyId    (str, required)
        positionDate (str, required, YYYY-MM-DD)
        groupings    (list[str], optional) — restricts calculation to these
                     grouping ids. Empty list means "all groupings in the
                     company" per the upstream API contract.
    """
    data = request.get_json(silent=True) or {}
    company_id    = (data.get("companyId") or "").strip()
    position_date = (data.get("positionDate") or "").strip()
    groupings     = data.get("groupings") or []

    if not company_id or not position_date:
        return jsonify({"error": "companyId and positionDate are required"}), 400
    if not company_visible(company_id):
        return jsonify({"error": "company is not visible to this user"}), 403
    if not isinstance(groupings, list) or not all(isinstance(g, str) for g in groupings):
        return jsonify({"error": "groupings must be a list of strings"}), 400

    try:
        result = calculate_nav_groupings(
            company_id=company_id,
            position_date=position_date,
            groupings=groupings,
        )
    except BeehusAPIError as e:
        return _api_error_response(e)
    beehus_catalog.invalidate_nav(company_id)
    return jsonify(result if result is not None else {"ok": True})


# Upstream's publish/unpublish endpoints take a JSON body, so a single PATCH
# can technically carry an unbounded `groupingIds` array. We still chunk for
# two reasons that survived the URL-length era: (1) partial-success
# granularity — if one batch fails, the operator sees exactly how many
# groupings made it through and can retry from the failed batch (publish /
# unpublish are idempotent per grouping); (2) bounded per-call latency so the
# upstream timeout never trips on long lists.
_PUBLISH_CHUNK_SIZE = 50


def _run_publish_in_chunks(fn, *, company_id, position_date, grouping_ids):
    """Call `fn` (publish_nav or unpublish_nav) with `grouping_ids` split into
    chunks of `_PUBLISH_CHUNK_SIZE`. Aggregates per-chunk outcomes into a
    response dict the front-end can render.

    Empty `grouping_ids` is treated as a single "all groupings of the company"
    call (upstream contract).

    On the first chunk that raises `BeehusAPIError` the loop stops and we
    return what's been done so far plus the error — publish/unpublish are
    idempotent at the grouping level, so retrying from the failed chunk is
    safe; surfacing the boundary helps the operator do that.
    """
    if not grouping_ids:
        result = fn(company_id=company_id, position_date=position_date, grouping_ids=[])
        return {
            "ok":               True,
            "totalGroupings":   0,
            "chunkSize":        _PUBLISH_CHUNK_SIZE,
            "chunkCount":       1,
            "chunksSucceeded":  1,
            "chunkResults":     [result if result is not None else {"ok": True}],
        }

    chunks = [
        grouping_ids[i:i + _PUBLISH_CHUNK_SIZE]
        for i in range(0, len(grouping_ids), _PUBLISH_CHUNK_SIZE)
    ]
    results = []
    for idx, ch in enumerate(chunks):
        try:
            r = fn(company_id=company_id, position_date=position_date, grouping_ids=ch)
            results.append(r if r is not None else {"ok": True})
        except BeehusAPIError as e:
            return {
                "ok":               False,
                "totalGroupings":   len(grouping_ids),
                "chunkSize":        _PUBLISH_CHUNK_SIZE,
                "chunkCount":       len(chunks),
                "chunksSucceeded":  idx,
                "failedChunkIndex": idx,
                "failedChunkSize":  len(ch),
                "error":            str(e),
                "upstream_status":  e.status,
                "upstream_body":    e.body,
                "chunkResults":     results,
            }
    return {
        "ok":               True,
        "totalGroupings":   len(grouping_ids),
        "chunkSize":        _PUBLISH_CHUNK_SIZE,
        "chunkCount":       len(chunks),
        "chunksSucceeded":  len(chunks),
        "chunkResults":     results,
    }


@bp.route("/api/beehus/nav/publish", methods=["POST"])
def nav_publish():
    """Publish NAV-contribution results for the listed groupings.

    Body (local):
        companyId    (str, required)
        positionDate (str, required, YYYY-MM-DD)
        groupingIds  (list[str], optional) — restricts publication to
                     these grouping ids. Empty list means "all groupings
                     in the company" per the upstream API contract.

    The upstream API expects a PATCH with a JSON body
    (`companyId`, `positionDate`, `groupingIds[]`); this route forwards the
    same shape and chunks long `groupingIds` lists to preserve partial-
    success granularity and bound per-call latency (see
    `_run_publish_in_chunks`).
    """
    data = request.get_json(silent=True) or {}
    company_id    = (data.get("companyId") or "").strip()
    position_date = (data.get("positionDate") or "").strip()
    grouping_ids  = data.get("groupingIds") or []

    if not company_id or not position_date:
        return jsonify({"error": "companyId and positionDate are required"}), 400
    if not company_visible(company_id):
        return jsonify({"error": "company is not visible to this user"}), 403
    if not isinstance(grouping_ids, list) or not all(isinstance(g, str) for g in grouping_ids):
        return jsonify({"error": "groupingIds must be a list of strings"}), 400

    summary = _run_publish_in_chunks(
        publish_nav,
        company_id=company_id,
        position_date=position_date,
        grouping_ids=grouping_ids,
    )
    # Mesmo em sucesso parcial, chunks publicados mudaram o estado `published`
    # → invalida o cache (filtros de publicação leem de nav_packages).
    beehus_catalog.invalidate_nav(company_id)
    if summary["ok"]:
        return jsonify(summary), 200
    # Match the rest of the routes: 401 for auth errors, 502 for upstream
    # failures. The chunk summary preserves upstream_status/body for the log.
    code = 401 if summary.get("upstream_status") == 401 else 502
    return jsonify(summary), code


@bp.route("/api/beehus/nav/unpublish", methods=["POST"])
def nav_unpublish():
    """Unpublish NAV-contribution results for the listed groupings.

    Mirrors `nav_publish` — same body shape forwarded upstream, same
    chunking strategy, just the inverse endpoint.
    """
    data = request.get_json(silent=True) or {}
    company_id    = (data.get("companyId") or "").strip()
    position_date = (data.get("positionDate") or "").strip()
    grouping_ids  = data.get("groupingIds") or []

    if not company_id or not position_date:
        return jsonify({"error": "companyId and positionDate are required"}), 400
    if not company_visible(company_id):
        return jsonify({"error": "company is not visible to this user"}), 403
    if not isinstance(grouping_ids, list) or not all(isinstance(g, str) for g in grouping_ids):
        return jsonify({"error": "groupingIds must be a list of strings"}), 400

    summary = _run_publish_in_chunks(
        unpublish_nav,
        company_id=company_id,
        position_date=position_date,
        grouping_ids=grouping_ids,
    )
    beehus_catalog.invalidate_nav(company_id)
    if summary["ok"]:
        return jsonify(summary), 200
    code = 401 if summary.get("upstream_status") == 401 else 502
    return jsonify(summary), code


@bp.route("/api/beehus/transactions/search", methods=["POST"])
def transactions_search():
    """Query local `db.transactions` using the filters from the UI.

    Mandatory: companyId, initialDate, finalDate (YYYY-MM-DD).
    Optional:  groupingId (str, legacy), groupingIds[] (preferred),
               walletIds[], beehusTransactionTypes[], securityIds[],
               entityIds[],
               identified (str: 'true' / 'false'/ ''=both) — when set,
               restricts to rows whose `beehusTransactionType` is filled
               ('true') or empty/missing ('false').

    Selecting groupings widens the scope rather than narrowing it: a row
    matches if its `walletId` is in the (grouping-narrowed) wallet set OR its
    own `groupingId` is one of the selected groupings — so transactions
    attached directly to a grouping show up next to the wallet-level ones.
    """
    data = request.get_json(silent=True) or {}

    company_id   = data.get("companyId") or ""
    initial_date = data.get("initialDate") or ""
    final_date   = data.get("finalDate") or ""
    if not company_id or not initial_date or not final_date:
        return jsonify({"error": "companyId, initialDate and finalDate are required"}), 400
    if not company_visible(company_id):
        return jsonify({"transactions": []})

    # Validate list[str] inputs up-front. The values flow into Mongo `$in`
    # filters and ObjectId() conversions; passing through arbitrary types
    # would either crash with TypeError (ObjectId) or bypass the project's
    # "list of strings" invariant, allowing operator-controlled query shapes.
    for _list_key in ("walletIds", "groupingIds", "entityIds",
                      "beehusTransactionTypes", "securityIds"):
        _v = data.get(_list_key)
        if _v is None:
            continue
        if not isinstance(_v, list) or not all(isinstance(x, str) for x in _v):
            return jsonify({"error": f"{_list_key} must be a list of strings"}), 400

    # Resolve the search scope (company → grouping(s) → explicit wallets).
    # `groupingIds[]` is the new shape (Editar Transações multi-upload);
    # `groupingId` (str) is kept for backwards compatibility and treated as a
    # singleton list.
    #
    # A selected grouping plays two roles: it narrows the wallet candidate set
    # to its member wallets, AND a transaction can be attached to a grouping
    # *directly* via its own `groupingId` field — without belonging to any of
    # the grouping's wallets. Those grouping-tagged transactions must show up
    # alongside the wallet ones, so we keep the selected groupings and union a
    # `groupingId` branch into the final query below. Only groupings that
    # belong to the visible company count (the gindex carries `companyId`,
    # same guard the classify route uses).
    grouping_ids = list(data.get("groupingIds") or [])
    legacy_grouping = (data.get("groupingId") or "").strip()
    if legacy_grouping and legacy_grouping not in grouping_ids:
        grouping_ids.append(legacy_grouping)
    valid_grouping_ids: list[str] = []
    # When groupings are selected the wallet branch is narrowed to their member
    # wallets (intersected with the company's wallets); otherwise it is every
    # wallet of the company. `allowed_wallet_ids is None` means "no grouping
    # narrowing"; an (possibly empty) set means "restrict to these".
    allowed_wallet_ids: set[str] | None = None
    if grouping_ids:
        gindex = get_grouping_index()
        wallet_ids: set[str] = set()
        for gid in grouping_ids:
            g = gindex.get(gid)
            if not g:
                continue
            if g.get("companyId") and g["companyId"] != company_id:
                continue
            valid_grouping_ids.append(gid)
            wallet_ids.update(str(w) for w in (g.get("walletIds") or []))
        # Narrow the wallet branch to the groupings' members. If none carry
        # wallets the branch is simply empty — the groupingId branch below can
        # still surface grouping-tagged transactions.
        allowed_wallet_ids = wallet_ids

    # Candidate wallets = the company's wallets (visibility!), optionally
    # narrowed to the selected groupings' members. Same result set as the
    # former db.wallets.find({companyId[, _id:$in]}, {_id:1}), via the index.
    candidate_wallets = {
        str(w["_id"]) for w in beehus_catalog.wallets_in_company(company_id)
        if allowed_wallet_ids is None or str(w["_id"]) in allowed_wallet_ids
    }
    explicit_wallets  = set(data.get("walletIds") or [])
    if explicit_wallets:
        candidate_wallets &= explicit_wallets

    # The transaction is in scope if it lives in a candidate wallet OR (when
    # groupings are selected) carries one of the selected groupingIds. The
    # `groupingId` field is stored as ObjectId in production, so the IN clause
    # spans both string and ObjectId representations (mirrors classify/probe).
    scope_branches: list[dict] = []
    if candidate_wallets:
        scope_branches.append({"walletId": {"$in": list(candidate_wallets)}})
    if valid_grouping_ids:
        grouping_in: list = list(valid_grouping_ids)
        for gid in valid_grouping_ids:
            try:
                grouping_in.append(ObjectId(gid))
            except (InvalidId, TypeError):
                pass
        scope_branches.append({"groupingId": {"$in": grouping_in}})
    if not scope_branches:
        return jsonify({"transactions": []})

    txn_query = {
        "liquidationDate": {"$gte": initial_date, "$lte": final_date},
        "trashed": {"$ne": True},
    }
    if len(scope_branches) == 1:
        txn_query.update(scope_branches[0])
    else:
        txn_query["$or"] = scope_branches
    types = data.get("beehusTransactionTypes") or []
    if types:
        txn_query["beehusTransactionType"] = {"$in": types}
    sec_ids = data.get("securityIds") or []
    if sec_ids:
        try:
            txn_query["securityId"] = {"$in": [ObjectId(s) for s in sec_ids]}
        except (InvalidId, TypeError):
            return jsonify({"transactions": []})
    entity_ids = data.get("entityIds") or []
    if entity_ids:
        txn_query["entityId"] = {"$in": entity_ids}

    # Optional `identified` filter: only used by the Identificar Transações view.
    # 'true'  → rows with a non-empty `beehusTransactionType`
    # 'false' → rows missing the field or with an empty string
    # anything else (including absent) → no filter (both buckets returned).
    # When `beehusTransactionTypes` is also set, the type filter wins for
    # 'true' (non-empty implied) and for 'false' we ignore the type filter
    # since a missing field can't be in any list anyway.
    identified_str = (data.get("identified") or "").strip().lower()
    if identified_str == "true":
        # Only rows whose `beehusTransactionType` is set and non-empty. If a
        # types list was already added above, leave it; otherwise add the
        # "non-empty" predicate.
        if "beehusTransactionType" not in txn_query:
            txn_query["beehusTransactionType"] = {"$nin": [None, ""]}
    elif identified_str == "false":
        # Override any user-provided types filter — a row without the field
        # cannot match a list.
        txn_query["beehusTransactionType"] = {"$in": [None, ""]}

    wallet_names   = get_wallet_names()
    entity_names   = get_entity_names()
    security_names = get_security_names()

    out = []
    # Cap the result set so a broad company-wide range can't return an
    # unbounded list; the `truncated` flag surfaces over-cap cases. 10k
    # handles realistic company-wide ranges. Sorted by `liquidationDate`
    # desc (most recent settlement first).
    _TXN_SEARCH_CAP = 10_000

    # Migrated from `db.transactions.find` → Beehus endpoint G via
    # beehus_catalog.transactions_search. The endpoint handles the
    # liquidationDate range + company + wallet/security/entity/grouping
    # filters; everything the API can't express (the wallet⋁grouping OR,
    # the type/identified predicates, the `trashed` guard, the
    # liquidationDate-desc sort and the cap) is reapplied client-side.
    #
    # The original scope was `walletId ∈ candidate_wallets OR groupingId ∈
    # valid_grouping_ids`. The endpoint ANDs its filters, so when both
    # branches exist we run two searches (one per branch) and union the
    # docs by `_id`; with a single branch we run one search.
    sec_filter = [str(s) for s in sec_ids] if sec_ids else None
    ent_filter = [str(e) for e in entity_ids] if entity_ids else None

    def _search(*, wallet_ids=None, grouping_ids=None):
        return beehus_catalog.transactions_search(
            company_id, initial_date=initial_date, final_date=final_date,
            wallet_ids=wallet_ids, grouping_ids=grouping_ids,
            security_ids=sec_filter, entity_ids=ent_filter,
            date_type="liquidation")

    raw_docs: list[dict] = []
    seen_ids: set[str] = set()
    branch_calls: list[dict] = []
    if candidate_wallets:
        branch_calls.append({"wallet_ids": list(candidate_wallets)})
    if valid_grouping_ids:
        branch_calls.append({"grouping_ids": list(valid_grouping_ids)})
    for _call in branch_calls:
        for d in _search(**_call):
            _key = str(d.get("_id") or d.get("id") or "")
            if _key and _key in seen_ids:
                continue
            if _key:
                seen_ids.add(_key)
            raw_docs.append(d)

    # Reapply filters the endpoint G doesn't cover, then sort desc by
    # liquidationDate and cap. `trashed` guard mirrors the original
    # "trashed": {"$ne": True}.
    def _type_ok(d):
        bt = d.get("beehusTransactionType")
        # 'false' overrides any user-provided types filter (mirrors the
        # original `txn_query["beehusTransactionType"] = {"$in": [None, ""]}`
        # which *replaced* the types `$in`). A missing/empty field can't be in
        # any types list anyway, so the types predicate is intentionally skipped
        # here — applying it first would make 'false' + non-empty types match
        # nothing (the two sets are disjoint).
        if identified_str == "false":
            return bt in (None, "")
        if types:
            if bt not in types:
                return False
        if identified_str == "true":
            if not types and (bt is None or bt == ""):
                return False
        return True

    filtered = [d for d in raw_docs if not d.get("trashed") and _type_ok(d)]
    filtered.sort(key=lambda d: str(d.get("liquidationDate") or ""), reverse=True)
    docs = filtered[:_TXN_SEARCH_CAP]

    for d in docs:
        wid = str(d.get("walletId") or "")
        eid = str(d.get("entityId") or "")
        sid = str(d.get("securityId") or "") if d.get("securityId") else ""
        out.append({
            "id":              str(d.get("_id") or d.get("id") or ""),
            "walletId":        wid,
            "walletName":      wallet_names.get(wid, "") or wid,
            "entityId":        eid,
            "entityName":      entity_names.get(eid, "") or eid,
            "securityId":      sid,
            "securityName":    security_names.get(sid, "") if sid else "",
            "balance":         d.get("balance"),
            "quantity":        d.get("quantity"),
            "price":           d.get("price"),
            "type":            d.get("beehusTransactionType") or "",
            "currencyId":      d.get("currencyId") or "",
            "operationDate":   _date_str(d.get("operationDate")),
            "liquidationDate": _date_str(d.get("liquidationDate")),
            "description":     d.get("description") or "",
            "comment":         d.get("comment") or "",
        })

    return jsonify({"transactions": out, "truncated": len(out) >= _TXN_SEARCH_CAP})


# ── Identificar Transações (config + identification stub) ─────────────────────

# All beehusTransactionType values currently observed in production. Kept in
# sync with the TYPES list in templates/beehus_console.html (DeleteTxn module).
_IDENTIFY_TXN_ALL_TYPES = [
    "amortization",
    "brokerageFee",
    "buySell",
    "bzFundTaxes",
    "contributionAdjustment",
    "coupon",
    "dividend",
    "dividendOnboarding",
    "gainsExpenses",
    "interestOnEquity",
    "managementFee",
    "maturity",
    "other",
    "otherFee",
    "performanceFee",
    "rebate",
    "securityContributionAdjustment",
    "securityTransfer",
    "taxes",
    "withdrawalDeposit",
    "withdrawalDepositAdjustment",
]

# Default list of types whose identification step should also resolve a
# `securityId`. Other types only get a `beehusTransactionType` suggestion.
_IDENTIFY_TXN_DEFAULT_TYPES_NEED_SECURITY = [
    "amortization",
    "buySell",
    "coupon",
    "dividend",
    "dividendOnboarding",
    "interestOnEquity",
    "maturity",
    "securityContributionAdjustment",
    "securityTransfer",
    "taxes",
]

_IDENTIFY_TXN_CONFIG_FILE = os.path.join(
    os.path.dirname(__file__), "..", "data", "identify_transactions_config.json",
)

_IDENTIFY_TXN_REINFORCEMENTS_FILE = os.path.join(
    os.path.dirname(__file__), "..", "data", "identify_transactions_reinforcements.json",
)


def _normalize_reinforcement_key(desc):
    """Thin wrapper around :func:`reinforcement_keys.normalize_reinforcement_key`.

    Kept as a private re-export so existing call sites in this module
    don't need to learn about the new module. The actual logic
    (normalize → mask transaction codes) now lives in the shared
    `reinforcement_keys` module so the build / migration scripts in
    `scripts/` use the exact same pipeline."""
    from reinforcement_keys import normalize_reinforcement_key as _normalize
    return _normalize(desc)


# mtime-keyed cache for the reinforcement table. The previous version
# of `_load_reinforcements` re-read the JSON on every `_lookup_reinforcement`
# call — fine on day-1 when the file was tiny, but with 10k+ rules and
# the file living on OneDrive we observed intermittent parse failures
# (the sync agent rewrites the file in place, and a reader landing mid-
# rewrite gets a truncated/malformed view). Even with atomic_write_json
# guarding *our* writes, OneDrive's own pull-sync from another machine
# still introduces transient inconsistency.
#
# Strategy: read once, cache; invalidate when `os.stat().st_mtime`
# changes. On a parse error, prefer the previous cached value (if we
# have one) over returning empty — empty would silently disable every
# reinforcement until the next valid save, which is much worse than
# serving slightly-stale rules for a few seconds while OneDrive settles.
_REINFORCEMENTS_CACHE = {
    "mtime":      None,  # last seen st_mtime
    "data":       None,  # last successful {"rules": {...}}
    "eligible":   None,  # precomputed Tier-2 substring-eligible [(key, rule)]
    "warned_at":  None,  # last mtime we logged a parse warning for (de-dup)
}


def _load_reinforcements():
    """Return ``{rules: {normalized_desc: {...}}}`` from the cache,
    refreshing it only when the file's mtime changed.

    Resilience contract:
      • Missing file → empty rules (file will be created on next save).
      • Parse error AND we have a previous successful read → return the
        cached value and log once per mtime; the read is almost
        certainly mid-OneDrive-sync and the file becomes valid again
        on the next mtime tick.
      • Parse error AND no prior cache → empty rules; the live UI
        gracefully degrades to "no reinforcement matches" for the
        next call instead of crashing.
    """
    if not os.path.exists(_IDENTIFY_TXN_REINFORCEMENTS_FILE):
        return {"rules": {}}
    try:
        mtime = os.stat(_IDENTIFY_TXN_REINFORCEMENTS_FILE).st_mtime
    except OSError:
        # Concurrent rename — fall through to a fresh read attempt.
        mtime = None

    cached_data = _REINFORCEMENTS_CACHE["data"]
    if mtime is not None and mtime == _REINFORCEMENTS_CACHE["mtime"] and cached_data is not None:
        return cached_data

    try:
        with open(_IDENTIFY_TXN_REINFORCEMENTS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f) or {}
        rules = data.get("rules")
        if not isinstance(rules, dict):
            rules = {}
        new_data = {"rules": rules}
        # Precompute the Tier-2 substring-eligible keys (len ≥ threshold) once
        # per file-load. `_match_against_rules` previously re-filtered all
        # ~958 rules on every lookup — and runs up to twice per transaction
        # (inner negating-prefix key + full key). Caching turns that O(rules)
        # filter into a single pass per file mtime. Stored as a SIBLING cache
        # field (not inside `new_data`) so it never leaks into the persisted
        # JSON when a caller writes the returned state back to disk.
        _REINFORCEMENTS_CACHE["mtime"] = mtime
        _REINFORCEMENTS_CACHE["data"] = new_data
        _REINFORCEMENTS_CACHE["eligible"] = [
            (k, v) for k, v in rules.items() if _is_tier2_eligible(k)
        ]
        return new_data
    except Exception as exc:
        log = logging.getLogger(__name__)
        # De-duplicate the warning per mtime so a busy server doesn't
        # spam the log when OneDrive holds the file in a bad state for
        # a few seconds.
        if _REINFORCEMENTS_CACHE.get("warned_at") != mtime:
            log.warning(
                "identify-transactions reinforcements unreadable (mtime=%s): %s",
                mtime, exc,
            )
            _REINFORCEMENTS_CACHE["warned_at"] = mtime
        if cached_data is not None:
            # Prefer the last good snapshot — way better than serving
            # an empty rule set just because OneDrive is mid-sync.
            return cached_data
        return {"rules": {}}


def _eligible_rules(state):
    """Tier-2 substring-eligible ``(key, rule)`` pairs for ``state``.

    Returns the precomputed sibling-cache list ONLY when ``state`` is the
    object currently held in the cache (identity check) — guaranteeing the
    eligible list matches that exact rules dict. For the fresh
    ``{"rules": {}}`` literals returned on the missing-file / no-prior-cache
    paths, identity won't match, so we return ``None`` and let
    ``_match_against_rules`` derive an (empty) eligible set from ``rules``.
    This prevents a deleted-file scenario from matching against a stale
    eligible list."""
    if (_REINFORCEMENTS_CACHE.get("eligible") is not None
            and _REINFORCEMENTS_CACHE.get("data") is state):
        return _REINFORCEMENTS_CACHE["eligible"]
    return None


def _record_reinforcement(description, beehus_txn_type, security_id,
                          security_name, security_main_id, no_security=None):
    """Add (or update) a reinforcement entry keyed by normalised description.

    Called after a manual user PATCH succeeds. Subsequent identify runs match
    on the normalised description and short-circuit the classifier — exact
    match only, never partial. A repeat of the same description bumps the
    `hits` counter and updates `lastSeenAt` so stale rules can be pruned later
    if needed.

    `no_security` is the explicit "this transaction has no security" flag.
    When ``True`` the stored rule carries ``noSecurity: true`` and future
    identifies skip the security cascade for it (see `_suggest_for_transaction`).
    Pass ``None`` (the default) to leave any existing flag untouched on an
    update; pass ``False`` to clear it. The key is only written when truthy so
    the file stays lean — absence reads as "needs a security normally"."""
    if not description:
        return
    key = _normalize_reinforcement_key(description)
    if not key:
        return
    if not (beehus_txn_type or security_id):
        return  # nothing useful to remember

    state = _load_reinforcements()
    existing = state["rules"].get(key) or {}
    now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    if no_security is None:
        no_security = bool(existing.get("noSecurity"))
    entry = {
        "beehusTransactionType": beehus_txn_type or existing.get("beehusTransactionType") or "",
        "securityId":            security_id or existing.get("securityId") or "",
        "securityName":          security_name or existing.get("securityName") or "",
        "securityMainId":        security_main_id or existing.get("securityMainId") or "",
        "lastDescriptionRaw":    description,
        "addedAt":               existing.get("addedAt") or now,
        "lastSeenAt":            now,
        "hits":                  int(existing.get("hits") or 0) + 1,
    }
    if no_security:
        entry["noSecurity"] = True
    state["rules"][key] = entry
    try:
        atomic_write_json(_IDENTIFY_TXN_REINFORCEMENTS_FILE, state)
        # We mutated the cached rules dict in place; its sibling `eligible`
        # index is now stale. Force the next _load_reinforcements to re-read
        # the file we just wrote and rebuild rules + eligible together
        # (coarse mtime resolution could otherwise keep serving the stale
        # index if the read lands in the same filesystem mtime tick).
        _REINFORCEMENTS_CACHE["mtime"] = None
    except Exception as exc:
        # Don't fail the PATCH because the reinforcement file couldn't be
        # written — the upstream change already succeeded.
        logging.getLogger(__name__).error(
            "failed to save identify-transactions reinforcement: %s", exc,
        )


def _lookup_reinforcement(description):
    """Match a transaction description against the reinforcement table.

    Returns ``None`` if nothing matches, otherwise a dict::

        {"rule":     <stored reinforcement entry>,
         "score":    float in [0.70, 1.00],
         "exact":    bool,                         # True ⇔ score == 1.0
         "matched":  <normalised key that matched>}

    Matching is three-tier:

    0. **Negating-prefix override**: descriptions that start with a tax
       / fee prefix (`IR -`, `IRRF`, `IOF`, `DEBITO IOF`, `DEBITO CBLC
       IRRF`) refer to the *tax on* an underlying transaction, not the
       transaction itself. We strip the prefix, recurse-lookup the
       inner description, and — when an inner match is found — force
       the returned `beehusTransactionType` to `taxes` while keeping
       the inner rule's `securityId` / `securityName` (the tax is on
       that security). The returned `score` is the inner score capped
       at 0.99 (Tier 0 hits never reach 1.0 because the prefix itself
       was synthetic information added to the underlying rule).
    1. **Exact match** on the normalised description → score 1.0.
    2. **Substring match**: any stored key that is a substring of the
       normalised description matches. The longest such key wins (more
       specific). The score reflects coverage —
       ``len(matched_key) / len(description_key)`` — clamped to
       ``[0.70, 0.99]`` so substring hits never reach 1.0 (1.0 is reserved
       for exact matches alone).

    This lets the operator save a short, generic snippet (e.g. ``AQUISICAO
    DE COTAS NO FUNDO TESOURO SIMPL RF PCO``) once and have it reuse on
    every transaction that contains that snippet — without false-positive
    risk from fuzzy / character-level matching."""
    if not description:
        return None
    key = _normalize_reinforcement_key(description)
    if not key:
        return None
    state = _load_reinforcements()
    rules = state["rules"]

    # ── Tier 0 — negating-prefix override ─────────────────────────────────
    # The normalised key keeps the prefix (e.g. "IRRF - …"), so a regular
    # Tier 1/2 lookup would either miss or — worse — match the prefix-free
    # rule via substring and return the wrong type. Strip the prefix
    # first, look up the inner key, and overwrite the type on the way
    # back. We don't recurse through `_lookup_reinforcement` itself to
    # avoid infinite loops when a stripped key still looks like a
    # negating prefix (defensive: shouldn't happen, but cheap to bound).
    from reinforcement_keys import strip_negating_prefix
    eligible = _eligible_rules(state)
    inner_key = strip_negating_prefix(key)
    if inner_key:
        inner_rule, inner_match, inner_score = _match_against_rules(inner_key, rules, eligible)
        if inner_rule is not None:
            taxed_rule = dict(inner_rule)
            taxed_rule["beehusTransactionType"] = "taxes"
            return {
                "rule":    taxed_rule,
                # Tier 0 hits never reach 1.0 — the prefix is a synthetic
                # override, not direct evidence the stored key matches.
                "score":   round(min(inner_score, 0.99), 4),
                "exact":   False,
                "matched": inner_match,
            }

    # ── Tier 1 / 2 — direct lookup on the full key ────────────────────────
    rule, match_key, score = _match_against_rules(key, rules, eligible)
    if rule is None:
        return None
    is_exact = (match_key == key) and (score == 1.0)
    return {"rule": dict(rule),
            "score": round(score, 4) if not is_exact else 1.0,
            "exact": is_exact,
            "matched": match_key}


# Minimum length for a key to be eligible as a Tier-2 substring match.
# Single-word keys ("SAQUE", "CUPOM") and one-char keys ("A") would
# otherwise gain massive false-positive surface by matching as substring
# inside any longer description that happens to contain those letters.
# Exact (Tier-1) matches stay allowed at any length — a description that
# is literally "SAQUE" should still hit the rule keyed "SAQUE".
_TIER2_MIN_KEY_LEN = 10

# A MULTI-word key (contains a space) is inherently far more specific than a
# bare word — two adjacent tokens must appear in sequence — so it qualifies
# for Tier-2 at a lower floor. This lets short-but-specific snippets like
# "COR JSCP" (8 chars) match "COR JSCP ITUB4" without lowering the bar for
# dangerous single words. The single-word protection above is unchanged.
_TIER2_MIN_MULTIWORD_LEN = 6


def _is_tier2_eligible(key):
    """True if `key` may match as a Tier-2 substring. Long keys (>= 10) always
    qualify; multi-word keys qualify from `_TIER2_MIN_MULTIWORD_LEN`; short
    single words never do (they stay exact-only)."""
    if not key:
        return False
    if len(key) >= _TIER2_MIN_KEY_LEN:
        return True
    return (" " in key) and len(key) >= _TIER2_MIN_MULTIWORD_LEN


def _match_against_rules(key, rules, eligible=None):
    """Tier-1 / Tier-2 lookup helper.

    Returns ``(rule_dict_or_None, matched_key_or_None, score)``. Score
    is 1.0 for an exact hit, ``coverage`` clamped to ``[0.70, 0.99]``
    for a substring hit, or ``0.0`` when nothing matches (paired with
    ``None`` rule / matched values so callers can short-circuit on the
    rule check).

    ``eligible`` is the precomputed list of ``(rule_key, rule)`` pairs whose
    key is long enough for a Tier-2 substring match (see
    ``_load_reinforcements``). When omitted, it is derived from ``rules`` on
    the fly so direct callers / tests keep working."""
    exact = rules.get(key)
    if exact:
        return exact, key, 1.0
    if eligible is None:
        eligible = [(k, v) for k, v in rules.items() if _is_tier2_eligible(k)]
    best_key = None
    best_rule = None
    for rule_key, rule in eligible:
        if rule_key in key:
            if best_key is None or len(rule_key) > len(best_key):
                best_key = rule_key
                best_rule = rule
    if best_rule is None:
        return None, None, 0.0
    coverage = len(best_key) / max(len(key), 1)
    return best_rule, best_key, max(0.70, min(0.99, coverage))


def _load_identify_txn_config():
    """Return `{typesNeedingSecurity: [str]}`. Falls back to the built-in
    defaults when the file is missing or corrupt. Unknown type names that
    sneak in are kept as-is — surfacing them lets the operator notice when
    the `_IDENTIFY_TXN_ALL_TYPES` list drifts from production."""
    defaults = {"typesNeedingSecurity": list(_IDENTIFY_TXN_DEFAULT_TYPES_NEED_SECURITY)}
    if not os.path.exists(_IDENTIFY_TXN_CONFIG_FILE):
        return defaults
    try:
        with open(_IDENTIFY_TXN_CONFIG_FILE, "r", encoding="utf-8") as f:
            data = json.load(f) or {}
        types = data.get("typesNeedingSecurity")
        if not isinstance(types, list) or not all(isinstance(t, str) for t in types):
            return defaults
        # Drop blanks/dupes while preserving insertion order.
        seen = set()
        cleaned = []
        for t in types:
            if t and t not in seen:
                seen.add(t)
                cleaned.append(t)
        return {"typesNeedingSecurity": cleaned}
    except Exception:
        return defaults


def _save_identify_txn_config(cfg):
    try:
        atomic_write_json(_IDENTIFY_TXN_CONFIG_FILE, cfg)
        return True, ""
    except Exception as exc:
        # Match the conciliacao_config redaction policy: don't leak a
        # filesystem path back to the client.
        logging.getLogger(__name__).error(
            "failed to save identify-transactions config: %s", exc,
        )
        return False, "verifique sincronização do OneDrive e tente novamente"


@bp.route("/api/beehus/identify-transactions/config", methods=["GET"])
def identify_txn_config_get():
    """Return current Identificar Transações config plus the full list of
    known beehusTransactionType values so the UI can render every checkbox
    without a second round-trip."""
    cfg = _load_identify_txn_config()
    return jsonify({
        "typesNeedingSecurity": cfg["typesNeedingSecurity"],
        "allTypes":             list(_IDENTIFY_TXN_ALL_TYPES),
    })


@bp.route("/api/beehus/identify-transactions/config", methods=["PUT"])
def identify_txn_config_put():
    """Replace the `typesNeedingSecurity` list. Body: ``{typesNeedingSecurity: [str]}``."""
    data = request.get_json(silent=True) or {}
    types = data.get("typesNeedingSecurity")
    if not isinstance(types, list) or not all(isinstance(t, str) for t in types):
        return jsonify({"error": "typesNeedingSecurity must be a list of strings"}), 400
    seen = set()
    cleaned = []
    for t in types:
        t = t.strip()
        if t and t not in seen:
            seen.add(t)
            cleaned.append(t)
    cfg = {"typesNeedingSecurity": cleaned}
    ok, err = _save_identify_txn_config(cfg)
    if not ok:
        return jsonify({"error": err}), 500
    return jsonify({
        "typesNeedingSecurity": cfg["typesNeedingSecurity"],
        "allTypes":             list(_IDENTIFY_TXN_ALL_TYPES),
    })


@bp.route("/api/beehus/identify-transactions/reinforcements", methods=["GET"])
def identify_txn_reinforcements_list():
    """Return all reinforcement rules, sorted by most-recently-used first.

    Each entry carries the normalised key (the substring that triggers the
    match) plus the saved values, so the Configurações modal can render a
    full CRUD table without follow-up lookups."""
    state = _load_reinforcements()
    rows = []
    for key, rule in (state.get("rules") or {}).items():
        rows.append({
            "key":                   key,
            "beehusTransactionType": rule.get("beehusTransactionType", "") or "",
            "securityId":            rule.get("securityId", "") or "",
            "securityName":          rule.get("securityName", "") or "",
            "securityMainId":        rule.get("securityMainId", "") or "",
            "lastDescriptionRaw":    rule.get("lastDescriptionRaw", "") or "",
            "addedAt":               rule.get("addedAt", "") or "",
            "lastSeenAt":            rule.get("lastSeenAt", "") or "",
            "hits":                  int(rule.get("hits") or 0),
            "noSecurity":            bool(rule.get("noSecurity")),
        })
    rows.sort(key=lambda r: r["lastSeenAt"], reverse=True)
    return jsonify({"reinforcements": rows})


@bp.route("/api/beehus/identify-transactions/reinforcement", methods=["DELETE"])
def identify_txn_reinforcement_delete():
    """Remove a single reinforcement rule by its normalised key.

    Body: ``{key: <normalised key>}``. Soft 404 if the key is unknown
    (idempotent — caller can safely re-issue)."""
    data = request.get_json(silent=True) or {}
    key = str(data.get("key") or "")
    if not key:
        return jsonify({"error": "key is required"}), 400
    state = _load_reinforcements()
    rules = state.get("rules") or {}
    existed = key in rules
    if existed:
        del rules[key]
        state["rules"] = rules
        try:
            atomic_write_json(_IDENTIFY_TXN_REINFORCEMENTS_FILE, state)
            # Stale eligible index after the in-place delete — force a rebuild
            # on the next read so a removed rule can't linger in the Tier-2
            # substring path. (See the matching note in _record_reinforcement.)
            _REINFORCEMENTS_CACHE["mtime"] = None
        except Exception as exc:
            logging.getLogger(__name__).error(
                "failed to delete identify-transactions reinforcement: %s", exc,
            )
            return jsonify({"error": "save failed"}), 500
    return jsonify({"deleted": existed, "key": key})


@bp.route("/api/beehus/identify-transactions/reinforcement", methods=["POST"])
def identify_txn_reinforcement_save():
    """Manually record an identification reinforcement.

    Replaces the previous auto-recording on every successful manual PATCH —
    operators now opt in explicitly via the security-edit modal (type +
    security) or the type-edit modal (type-only) so the reinforcement set
    stays curated. The body must carry the original transaction description
    (used as the lookup key after normalisation) plus at least one of
    beehusTransactionType or securityId. A type-only rule (securityId empty)
    is valid: future identifies inherit the type and skip the security
    cascade only when the type itself doesn't require a security — unless the
    rule is explicitly flagged ``noSecurity: true``, in which case the security
    cascade is always skipped (the transaction has no security).
    """
    data = request.get_json(silent=True) or {}
    description = str(data.get("description") or "")
    if not description.strip():
        return jsonify({"error": "description is required"}), 400
    btt = str(data.get("beehusTransactionType") or "")
    sid = str(data.get("securityId") or "")
    if not btt and not sid:
        return jsonify({
            "error": "at least one of beehusTransactionType or securityId is required",
        }), 400
    sname = str(data.get("securityName") or "")
    smain = str(data.get("securityMainId") or "")
    # `noSecurity` absent ⇒ None (leave any existing flag untouched on update);
    # present ⇒ explicit True/False from the operator's checkbox.
    no_security = None if data.get("noSecurity") is None else bool(data.get("noSecurity"))

    _record_reinforcement(
        description     = description,
        beehus_txn_type = btt,
        security_id     = sid,
        security_name   = sname,
        security_main_id= smain,
        no_security     = no_security,
    )
    # Echo back the saved entry so the UI can confirm what was stored.
    state = _load_reinforcements()
    key = _normalize_reinforcement_key(description)
    return jsonify({
        "saved":      True,
        "key":        key,
        "rule":       state["rules"].get(key) or {},
    })


@bp.route("/api/beehus/identify-transactions/reinforcement/normalize", methods=["POST"])
def identify_txn_reinforcement_normalize():
    """Normalize a raw description into its reinforcement lookup key.

    Single source of truth for the key (uppercase → accent strip → variable-
    token masking, e.g. <OPCODE>/CDB<CODE>). Powers the edit modal's live key
    preview and the key-collision check so the UI never drifts from the real
    server-side normalisation (the JS-only NFD/uppercase approximation misses
    token masking). Body: ``{description}``. Returns ``{key}`` ('' when blank).
    """
    data = request.get_json(silent=True) or {}
    description = str(data.get("description") or "")
    key = _normalize_reinforcement_key(description) if description.strip() else ""
    return jsonify({"key": key})


@bp.route("/api/beehus/identify-transactions/wallet-securities", methods=["GET"])
def identify_txn_wallet_securities():
    """Return the union of securityIds held by `walletId` at the FIXED dates
    D, D-1 and D-2 business days (D = `liquidationDate`), fetched via the
    Beehus API at each EXACT date — no backward scan over sparse dates.

    Used by the security-edit modal to restrict the free-text search to a
    universe of securities the wallet actually holds — matches the L1∪L2
    candidate pool that `TransactionSecurityClassifier` scores against.
    """
    wallet_id        = (request.args.get("walletId") or "").strip()
    liquidation_date = (request.args.get("liquidationDate") or "").strip()
    if not wallet_id or not liquidation_date:
        return jsonify({"error": "walletId and liquidationDate are required"}), 400

    # Authorization: resolve the wallet's company and gate against the
    # caller's visible-companies set (consistent with the rest of the page).
    try:
        ObjectId(wallet_id)
    except (InvalidId, TypeError):
        return jsonify({"error": "invalid walletId"}), 400
    wallet_doc = resolve_wallet(wallet_id, {"companyId": 1})
    if not wallet_doc:
        return jsonify({"securityIds": []})
    company_id = str(wallet_doc.get("companyId") or "")
    if not company_visible(company_id):
        return jsonify({"securityIds": []}), 403

    # Fixed dates: D (liquidationDate), D-1 and D-2 business days. One exact-
    # date API call each; a date with no position simply contributes nothing.
    d = str(liquidation_date)[:10]
    dates = [d, business_days_before(d, 1), business_days_before(d, 2)]
    sec_ids = set()
    for dt in dates:
        if not dt:
            continue
        doc = beehus_catalog.processed_doc(wallet_id, dt, company_id)
        if not doc:
            continue
        for s in (doc.get("securities") or []):
            sid = beehus_catalog.id_str(s.get("securityId")) if s.get("securityId") else ""
            if sid:
                sec_ids.add(sid)
    return jsonify({"securityIds": sorted(sec_ids)})


@bp.route("/api/beehus/identify-transactions/wallet-position-detail", methods=["GET"])
def identify_txn_wallet_position_detail():
    """Return the wallet's position snapshot at level L1 or L2, enriched with
    security metadata so the security-edit modal can render the full position
    (not just the matcher's top alternatives). Prefers the processedPosition;
    falls back to the raw (unprocessed) position of the SAME date when there is
    none (see `_position_rows`).

    Query params:
      walletId          required
      liquidationDate   required, YYYY-MM-DD (D — the txn liquidation date)
      level             optional, one of {'l1', 'l2'}. Default 'l1'.
                          l1 = position at D (the liquidationDate)
                          l2 = position at D-1 ∪ D-2 business days; L1 ids
                               are excluded so the two groups are disjoint.

    Returns:
      {
        level: 'l1' | 'l2',
        positionDate: 'YYYY-MM-DD' (L1 only),
        positionDates: ['YYYY-MM-DD', 'YYYY-MM-DD'] (L2 only),
        securities: [
          {securityId, beehusName, mainId, ticker, securityType,
           maturityDate, quantity, pu}
        ]
      }

    Fixed-offset, EXACT-date lookups via the Beehus API (no backward scan over
    sparse dates): for each target date the processedPosition is tried first and,
    when absent, the raw unprocessed position of that same date (reading the
    resolved `preProcessingData.securityId`). A date with neither contributes
    nothing. The "expand L1/L2" affordance in the modal calls this and replaces
    the short matcher-top-5 list with the wallet's complete position — useful
    when the matcher missed the right security (descrição com nomenclatura
    diferente do beehusName) but the operator can recognise it by glance, and so
    that carteiras em tombamento (só foto-base processada) still show the
    holdings em D.
    """
    wallet_id        = (request.args.get("walletId") or "").strip()
    liquidation_date = (request.args.get("liquidationDate") or "").strip()
    level            = (request.args.get("level") or "l1").strip().lower()
    if not wallet_id or not liquidation_date:
        return jsonify({"error": "walletId and liquidationDate are required"}), 400
    if level not in ("l1", "l2"):
        return jsonify({"error": "level must be 'l1' or 'l2'"}), 400

    # Authorization (same gate as the wallet-securities endpoint).
    try:
        ObjectId(wallet_id)
    except (InvalidId, TypeError):
        return jsonify({"error": "invalid walletId"}), 400
    wallet_doc = resolve_wallet(wallet_id, {"companyId": 1})
    if not wallet_doc:
        return jsonify({"securities": [], "level": level})
    company_id = str(wallet_doc.get("companyId") or "")
    if not company_visible(company_id):
        return jsonify({"securities": [], "level": level}), 403

    # Fixed dates: D (liquidationDate), D-1 and D-2 business days.
    d0 = str(liquidation_date)[:10]
    d1 = business_days_before(d0, 1)
    d2 = business_days_before(d0, 2)
    if level == "l1":
        target_dates = [d0]
    else:
        target_dates = [x for x in (d1, d2) if x]   # D-1, D-2 (most recent first)
        if not target_dates:
            return jsonify({"securities": [], "level": level, "positionDates": []})

    def _position_rows(dt):
        """Security rows held by the wallet on the EXACT `dt`, as
        ``[{securityId, quantity, pu, pricingType}]``. Prefers the PROCESSED
        position; when the wallet has no processed position that date (carteira
        em tombamento, que só tem a foto-base processada) falls back to the RAW
        (unprocessed) position of the SAME date, reading each item's already
        resolved ``preProcessingData.securityId``. Unresolved raw items (no
        securityId yet) are skipped — they aren't pickable. Same exact-date
        semantics as the L1/L2 classifier cascade (see _WalletCandidatePool)."""
        if not dt:
            return []
        pdoc = beehus_catalog.processed_doc(wallet_id, dt, company_id)
        if pdoc and (pdoc.get("securities") or []):
            rows = []
            for s in (pdoc.get("securities") or []):
                sid = beehus_catalog.id_str(s.get("securityId")) if s.get("securityId") else ""
                if sid:
                    rows.append({"securityId": sid, "quantity": s.get("quantity"),
                                 "pu": s.get("pu"), "pricingType": s.get("pricingType")})
            return rows
        udoc = beehus_catalog.unprocessed_doc(wallet_id, dt, company_id)
        rows = []
        if udoc:
            for s in (udoc.get("securities") or []):
                ppd = s.get("preProcessingData") if isinstance(s, dict) else None
                raw = ppd.get("securityId") if isinstance(ppd, dict) else None
                sid = beehus_catalog.id_str(raw) if raw else ""
                if sid:
                    rows.append({"securityId": sid, "quantity": s.get("quantity"),
                                 "pu": s.get("pu"),
                                 "pricingType": ppd.get("pricingType") if isinstance(ppd, dict) else None})
        return rows

    # Collect securities across the target dates (exact-date API). For L2 we
    # may see the same securityId on D-1 and D-2 — dedupe by id, keeping the
    # entry from the most recent date (target_dates is sorted DESC).
    by_sid = {}
    for dt in target_dates:
        for r in _position_rows(dt):
            sid = r["securityId"]
            if sid in by_sid:
                continue
            by_sid[sid] = {
                "securityId":  sid,
                "quantity":    r["quantity"],
                "pu":          r["pu"],
                "pricingType": r["pricingType"],
            }

    # For L2, exclude anything already in L1 so the two groups are disjoint
    # (matches the classifier's L2 definition: D-1 ∪ D-2 \ D).
    if level == "l2":
        l1_ids = {r["securityId"] for r in _position_rows(d0)}
        for sid in list(by_sid.keys()):
            if sid in l1_ids:
                by_sid.pop(sid, None)

    if not by_sid:
        out = []
    else:
        # Enrich via securities collection. We do one batch find here rather
        # than reach into the SecurityCache because this endpoint is called
        # on demand (operator click) and a fresh read avoids stale-cache
        # surprises for recently-cadastred securities.
        oids = []
        for sid in by_sid:
            try:
                oids.append(ObjectId(sid))
            except (TypeError, ValueError):
                continue
        meta_by_id = {}
        if oids:
            for s in beehus_catalog.securities_by_ids(oids).values():
                meta_by_id[str(s["_id"])] = s

        out = []
        for sid, info in by_sid.items():
            meta = meta_by_id.get(sid, {})
            out.append({
                "securityId":   sid,
                "beehusName":   meta.get("beehusName", "") or "",
                "mainId":       meta.get("mainId", "") or "",
                "ticker":       meta.get("ticker", "") or "",
                "taxId":        meta.get("taxId", "") or "",
                "isIn":         meta.get("isIn", "") or "",
                "selicCode":    meta.get("selicCode", "") or "",
                "securityType": meta.get("securityType", "") or "",
                "maturityDate": str(meta.get("maturityDate") or "")[:10],
                "quantity":     info["quantity"],
                "pu":           info["pu"],
                "pricingType":  info["pricingType"],
            })
        # Stable order: enriched names first, then alphabetical. Orphaned
        # securities (no entry in `securities` collection) drop to the end.
        out.sort(key=lambda r: (not r["beehusName"], r["beehusName"].lower(), r["mainId"]))

    payload = {"level": level, "securities": out}
    if level == "l1":
        payload["positionDate"] = d0
    else:
        payload["positionDates"] = target_dates
    return jsonify(payload)


def _num(v):
    """Coerce to float, returning None for None / non-numeric / NaN."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return None if f != f else f  # drop NaN


def _compute_execution_extras(txn_docs_by_id, suggestions_by_id):
    """Per-transaction executionPrice + IRRF for the Identificar Transações grid.

    Returns ``{transactionId: {...}}`` with keys ``executionPrice``, ``irrf``,
    ``pu``, ``amountDifference``, ``securityType``, ``formerDate``,
    ``withinGate``, ``execGroupKey``. Each value is ``None`` when not
    applicable. See docs/BEEHUS_CONSOLE.md for the rules. Summary:

    * Effective type/security = suggestion value, falling back to the txn's
      current value.
    * Gate: ``biz_days_between(operationDate, liquidationDate) < 3``. Outside
      the window executionPrice/IRRF are left ``None`` (operator sees "—").
    * ``amountDifference`` = the txn's own signed ``quantity`` (+ buy / − sell);
      when absent it falls back to the Δquantity between the previous and the
      most-recent processedPosition (≤ liquidationDate).
    * ``executionPrice = -Σbalance / Σquantity`` for buySell, grouped by
      (walletId, securityId, liquidationDate) over the identified batch so
      multiple partial fills on the same security/date share one price.
      ``maturity`` is computed per-row. Requires a securityId.
    * ``IRRF = balance + amountDifference × PU`` (≈ −tax, negative) for
      buySell + brazilianFund + balance > 0.
    * PU = ``processedPosition.securities.pu`` of the latest position
      (≤ liquidationDate) holding the security.
    """
    out = {}
    if not txn_docs_by_id:
        return out

    # ── 1. Effective (type, securityId) + raw fields per transaction ──────────
    eff = {}
    sec_ids = set()
    for tid, doc in txn_docs_by_id.items():
        if not doc:
            continue
        sug = suggestions_by_id.get(tid, {})
        etype = sug.get("beehusTransactionType") or doc.get("beehusTransactionType") or ""
        sid = str(sug.get("securityId") or doc.get("securityId") or "")
        eff[tid] = {
            "type":     etype,
            "sid":      sid,
            "wallet":   str(doc.get("walletId") or ""),
            "liq":      _date_str(doc.get("liquidationDate")),
            "op":       _date_str(doc.get("operationDate")),
            "balance":  _num(doc.get("balance")),
            "quantity": _num(doc.get("quantity")),
        }
        if sid:
            sec_ids.add(sid)

    # ── 2. securityType per securityId (batch) ────────────────────────────────
    sec_type = {}
    oids = []
    for sid in sec_ids:
        try:
            oids.append(ObjectId(sid))
        except (InvalidId, TypeError):
            pass
    if oids:
        for s in beehus_catalog.securities_by_ids(oids).values():
            sec_type[str(s["_id"])] = s.get("securityType") or ""

    # ── 3. processedPositions per wallet (PU + Δqty), via endpoint A (no Mongo) ──
    # Endpoint A is EXACT-date only (no range / no "≤ date"), so we walk business
    # days back from the liquidationDate to the latest position at/-before it (d0,
    # for PU) and — lazily, only when a Δqty fallback is actually needed — to the
    # position before that (d1). Daily-position wallets resolve in 1-2 calls;
    # sparser wallets walk back up to _POS_MAXBACK business days, then degrade
    # gracefully (PU None / Δqty = absolute qty), as the old 40-doc window did
    # for liq dates beyond its reach. Each (wallet, date) is fetched once (cached).
    _POS_MAXBACK = 25  # business days — covers monthly positions + holiday gaps
    _wallets = {e["wallet"] for e in eff.values() if e["wallet"]}
    _wallet_company = {
        wid: str(cid or "")
        for wid, cid in beehus_catalog.wallet_company_map(list(_wallets)).items()
    } if _wallets else {}

    _secs_cache = {}   # (wallet, date) -> {sid: {pu, qty}} | None  (None = no position)

    def _secs_on(wallet, date):
        key = (wallet, date)
        if key in _secs_cache:
            return _secs_cache[key]
        cid = _wallet_company.get(wallet)
        doc = beehus_catalog.processed_doc(wallet, date, cid) if (wallet and date and cid) else None
        secs = None
        if doc:
            secs = {}
            for s in (doc.get("securities") or []):
                s_sid = str(s.get("securityId") or "")
                if s_sid:
                    secs[s_sid] = {"pu": s.get("pu"), "qty": s.get("quantity")}
        _secs_cache[key] = secs
        return secs

    def _walk_back(wallet, start_date, predicate):
        """First (date, secs) at/-before `start_date` (business days) where
        `predicate(secs)` holds, within _POS_MAXBACK; else (None, None)."""
        d = str(start_date)[:10]
        for _ in range(_POS_MAXBACK + 1):
            if not d:
                break
            secs = _secs_on(wallet, d)
            if secs is not None and predicate(secs):
                return d, secs
            d = business_days_before(d, 1)
        return None, None

    _d0_cache = {}   # (wallet, liq) -> (date0, secs0)  — latest position ≤ liq
    def _d0(wallet, liq):
        key = (wallet, liq)
        if key not in _d0_cache:
            _d0_cache[key] = _walk_back(wallet, liq, lambda s: True)
        return _d0_cache[key]

    _d1_cache = {}   # (wallet, liq) -> secs1  — position immediately before d0
    def _d1(wallet, liq):
        key = (wallet, liq)
        if key not in _d1_cache:
            d0_date, _ = _d0(wallet, liq)
            secs1 = None
            if d0_date:
                _, secs1 = _walk_back(wallet, business_days_before(d0_date, 1), lambda s: True)
            _d1_cache[key] = secs1
        return _d1_cache[key]

    def _pu(wallet, sid, liq):
        if not (wallet and sid and liq):
            return None
        d0_date, secs0 = _d0(wallet, liq)
        if secs0 is not None and sid in secs0:
            return secs0[sid].get("pu")
        # sid absent from the position ≤ liq: latest older position containing it
        # (mirrors the original "most recent position ≤ liq that holds sid").
        if not d0_date:
            return None
        _, secs = _walk_back(wallet, business_days_before(d0_date, 1), lambda s: sid in s)
        return (secs.get(sid) or {}).get("pu") if secs else None

    def _pos_delta(wallet, sid, liq):
        if not (wallet and sid and liq):
            return None
        _, secs0 = _d0(wallet, liq)
        if secs0 is None:
            return None
        secs1 = _d1(wallet, liq)   # lazy: triggers the longer walk only here
        cur = (secs0.get(sid) or {}).get("qty")
        prev = (secs1.get(sid) or {}).get("qty") if secs1 else None
        if cur is None and prev is None:
            return None
        delta = (_num(cur) or 0.0) - (_num(prev) or 0.0)
        return delta or None

    def _amount_diff(e):
        """Signed quantity for the txn: own quantity, else position Δqty."""
        q = e["quantity"]
        if q is not None and q != 0:
            return q
        return _pos_delta(e["wallet"], e["sid"], e["liq"])

    # ── 4. buySell group prices: -Σbalance / Σquantity by (wallet, sid, liq) ──
    groups = {}
    for tid, e in eff.items():
        if e["type"] != "buySell" or not e["sid"] or not e["liq"]:
            continue
        key = (e["wallet"], e["sid"], e["liq"])
        g = groups.setdefault(key, {"balance": 0.0, "qty": 0.0, "all_qty": True})
        g["balance"] += (e["balance"] or 0.0)
        if e["quantity"] is not None:
            g["qty"] += e["quantity"]
        else:
            g["all_qty"] = False

    group_price = {}
    for key, g in groups.items():
        wallet, sid, liq = key
        denom = g["qty"] if (g["all_qty"] and g["qty"]) else _pos_delta(wallet, sid, liq)
        group_price[key] = round(-g["balance"] / denom, 6) if denom else None

    # ── 5. Per-transaction assembly ───────────────────────────────────────────
    for tid, e in eff.items():
        etype, sid, wallet, liq, op = e["type"], e["sid"], e["wallet"], e["liq"], e["op"]
        within_gate = bool(op and liq and biz_days_between(op, liq) < 3)
        pu = _pu(wallet, sid, liq) if sid else None
        amt = _amount_diff(e)
        stype = sec_type.get(sid, "")

        exec_price = None
        group_key = None
        # brazilianFund: execution price is not meaningful (the cash already
        # reflects NAV/quotas) — only the IRRF applies. Skip executionPrice for
        # both buySell and maturity when the security is a brazilianFund.
        if (within_gate and sid and etype in ("buySell", "maturity")
                and stype != "brazilianFund"):
            if etype == "buySell":
                key = (wallet, sid, liq)
                group_key = "|".join(key)
                exec_price = group_price.get(key)
            else:  # maturity — per row
                group_key = "|".join((wallet, sid, liq))
                denom = e["quantity"] if (e["quantity"] not in (None, 0)) else amt
                if denom:
                    exec_price = round(-(e["balance"] or 0.0) / denom, 6)

        irrf = None
        if (within_gate and etype == "buySell" and stype == "brazilianFund"
                and e["balance"] is not None and e["balance"] > 0
                and amt is not None and pu is not None):
            irrf = round(e["balance"] + amt * pu, 2)

        out[tid] = {
            "executionPrice":  exec_price,
            "irrf":            irrf,
            "pu":              pu,
            "amountDifference": amt,
            "securityType":    stype,
            "formerDate":      op,
            "withinGate":      within_gate,
            "execGroupKey":    group_key,
        }
    return out


@bp.route("/api/beehus/identify-transactions/identify", methods=["POST"])
def identify_txn_identify():
    """Suggest a `beehusTransactionType` and (when needed) a `securityId`
    for each transaction in the body.

    The transaction data comes FROM THE CLIENT — the `/search` response
    (`this._txns`) already carries every field the classifier and the
    execution-extras pass need — so this route does **not** read the
    transactions back from Mongo (no by-id round-trip).

    Body: ``{transactions: [{id, description, walletId, entityId, securityId,
            balance, quantity, price, operationDate, liquidationDate,
            beehusTransactionType, currencyId}]}``  (max 5000).
      - `beehusTransactionType` may also arrive as `type` (the `/search` alias).
      - Only `id` is required per item; missing fields just weaken the result.
    """
    data = request.get_json(silent=True) or {}
    txns = data.get("transactions")
    if not isinstance(txns, list) or not all(isinstance(x, dict) for x in txns):
        return jsonify({"error": "transactions must be a list of objects "
                                 "({id, description, walletId, liquidationDate, "
                                 "balance, beehusTransactionType, ...})"}), 400
    if not txns:
        return jsonify({"suggestions": []})
    if len(txns) > 5000:
        return jsonify({"error": "too many transactions (max 5000)"}), 400

    cfg = _load_identify_txn_config()
    types_need_security = set(cfg["typesNeedingSecurity"])

    # Re-hydrate the txn docs from the client payload (which came from /search)
    # — no Mongo round-trip. Mirrors the field set the old by-id projection
    # loaded, so `_suggest_for_transaction` and `_compute_execution_extras` see
    # the same shape. `type` is the /search alias for `beehusTransactionType`.
    txn_by_id = {}
    for t in txns:
        tid = str(t.get("id") or t.get("transactionId") or t.get("_id") or "")
        if not tid:
            continue
        txn_by_id[tid] = {
            "walletId":              t.get("walletId") or "",
            "entityId":              t.get("entityId") or "",
            "securityId":            t.get("securityId") or "",
            "balance":               t.get("balance"),
            "quantity":              t.get("quantity"),
            "price":                 t.get("price"),
            "beehusTransactionType": t.get("beehusTransactionType") or t.get("type") or "",
            "operationDate":         t.get("operationDate") or "",
            "liquidationDate":       t.get("liquidationDate") or "",
            "description":           t.get("description") or "",
            "comment":               t.get("comment") or "",
            "currencyId":            t.get("currencyId") or "",
        }

    # Authorization: drop any transaction whose wallet's companyId is outside
    # the caller's company_filter. Transactions don't carry companyId directly;
    # we resolve it via wallets in a single batch query.
    visible = get_company_filter()
    if visible and txn_by_id:
        wallet_ids_to_check = {
            str(d.get("walletId") or "") for d in txn_by_id.values() if d.get("walletId")
        }
        # walletId → companyId via the cached wallets index (O(1) lookups, not N+1).
        wallet_company = {
            wid: str(cid or "")
            for wid, cid in beehus_catalog.wallet_company_map(
                list(wallet_ids_to_check)
            ).items()
        }
        txn_by_id = {
            tid: d for tid, d in txn_by_id.items()
            if wallet_company.get(str(d.get("walletId") or "")) in visible
        }

    # Reset per-batch cache on the security classifier so processedPosition
    # data is fresh for this request (within the request, results are cached
    # per wallet to avoid re-querying for every transaction).
    sec_clf = _get_security_classifier()
    if sec_clf is not None:
        sec_clf.reset_request_cache()

    suggestions = []
    for t in txns:
        tid = str(t.get("id") or t.get("transactionId") or t.get("_id") or "")
        s = _suggest_for_transaction(txn_by_id.get(tid), types_need_security)
        suggestions.append({
            "transactionId":         tid,
            "beehusTransactionType": s["beehusTransactionType"],
            "securityId":            s["securityId"],
            "needsSecurity":         s["needsSecurity"],
            "confidence":            s["confidence"],
            "source":                s["source"],
            "needsReview":           s["needsReview"],
            # Security-id suggestion fields (only populated when needsSecurity):
            "securityName":          s["securityName"],
            "securityMainId":        s["securityMainId"],
            "securityConfidence":    s["securityConfidence"],
            "securitySource":        s["securitySource"],
            "securityAmbiguous":     s["securityAmbiguous"],
            "securityAlternatives":  s["securityAlternatives"],
            "securityScore":         s["securityScore"],
            "securityTiebreak":      s["securityTiebreak"],
        })

    # Enrich with executionPrice / IRRF (and supporting PU / amountDifference /
    # gate fields) so the grid can render the two derived columns and the
    # Implementar flow knows what to push upstream. Computed in one batch pass
    # so buySell fills on the same security/date share a single price.
    extras = _compute_execution_extras(
        txn_by_id, {s["transactionId"]: s for s in suggestions}
    )
    for s in suggestions:
        s.update(extras.get(s["transactionId"], {}))

    return jsonify({"suggestions": suggestions})


# Singleton classifier — trained on first request (5190 samples ≈ 2-3 s).
# A double-trained instance on a concurrent first hit is harmless; the lock
# only prevents the actual fitting from running twice.
_classifier        = None
_classifier_lock   = threading.Lock()
_classifier_failed = False  # set True if training has failed once, to avoid retrying every request


def _get_classifier():
    global _classifier, _classifier_failed
    if _classifier is not None or _classifier_failed:
        return _classifier
    with _classifier_lock:
        if _classifier is not None or _classifier_failed:
            return _classifier
        try:
            from transaction_type_classifier import TransactionTypeClassifier
            clf = TransactionTypeClassifier()
            clf.train()
            _classifier = clf
        except Exception as exc:
            _classifier_failed = True
            logging.getLogger(__name__).exception(
                "Failed to train transaction-type classifier: %s. "
                "Identify-transactions will return no suggestions.", exc,
            )
    return _classifier


# Singleton transaction-security classifier. Cheap to construct (no model
# training). The internal pool cache is reset between identify_txn_identify
# requests so processedPositions data stays current per batch.
_sec_classifier        = None
_sec_classifier_lock   = threading.Lock()
_sec_classifier_failed = False


def _get_security_classifier():
    global _sec_classifier, _sec_classifier_failed
    if _sec_classifier is not None or _sec_classifier_failed:
        return _sec_classifier
    with _sec_classifier_lock:
        if _sec_classifier is not None or _sec_classifier_failed:
            return _sec_classifier
        try:
            from transaction_security_classifier import TransactionSecurityClassifier
            _sec_classifier = TransactionSecurityClassifier(db)
        except Exception as exc:
            _sec_classifier_failed = True
            logging.getLogger(__name__).exception(
                "Failed to construct transaction-security classifier: %s. "
                "Identify-transactions will skip securityId suggestions.", exc,
            )
    return _sec_classifier


def _suggest_for_transaction(txn_doc, types_need_security):
    """Suggest beehusTransactionType (and securityId, when required) for one txn.

    Type pipeline: rules → ML fallback (transaction_type_classifier).

    Security pipeline (only when the predicted/current type appears in
    types_need_security): cascade L1 (current processedPosition) → L2 (T-1, T-2)
    → L3 (full securities collection), via transaction_security_classifier.
    For ambiguous buySell rows, an amountDifference + offset tie-breaker runs
    against processedPositions.

    The response always includes the security fields; they are blank when
    needsSecurity is false (skipping the relatively expensive Mongo lookups).
    """
    suggested_type     = ""
    suggested_security = ""
    confidence         = 0.0
    source             = ""
    needs_review       = False

    current_type = ""
    description  = ""
    if txn_doc:
        current_type = txn_doc.get("beehusTransactionType") or ""
        description  = txn_doc.get("description") or ""

    # ── Reinforcements: first lookup, before rules + ML ──────────────────────
    # The lookup is two-tier: exact match → score 1.0, otherwise the longest
    # stored snippet that is a substring of the description wins, with a
    # coverage-based score in [0.70, 0.99]. Source is reported as
    # 'reinforcement' (exact) or 'reinforcement_partial' (substring) so the UI
    # badge distinguishes the two and operators can audit which rule fired.
    reinforced_match = _lookup_reinforcement(description) if description else None
    reinforced = (reinforced_match or {}).get("rule")
    sec_name = sec_main = sec_source = ""
    sec_conf = 0.0
    sec_ambiguous = False
    sec_alts = []
    sec_score = None
    sec_tiebreak = None
    # Set when the matched reinforcement is explicitly flagged "no security"
    # (`noSecurity: true`) — the operator's deliberate, unambiguous signal
    # that the transaction has no security at all (e.g. "AJ POS DE FUT"),
    # distinct from a type-only rule that simply hasn't pinned a security yet
    # and still wants the ML cascade to classify one. When set, the L1/L2/L3
    # cascade is skipped and the row is not flagged pending-security even if
    # its type is in `typesNeedingSecurity`.
    reinforced_no_security = False
    if reinforced:
        score = float(reinforced_match.get("score") or 1.0)
        is_exact = bool(reinforced_match.get("exact"))
        if reinforced.get("beehusTransactionType"):
            suggested_type = reinforced["beehusTransactionType"]
        if reinforced.get("securityId"):
            suggested_security = reinforced["securityId"]
            sec_name   = reinforced.get("securityName", "") or ""
            sec_main   = reinforced.get("securityMainId", "") or ""
            sec_conf   = score
            sec_source = "reinforcement" if is_exact else "reinforcement_partial"
        elif reinforced.get("noSecurity"):
            reinforced_no_security = True
        confidence = score
        source = "reinforcement" if is_exact else "reinforcement_partial"
        needs_review = not is_exact and score < 0.85
    elif description:
        clf = _get_classifier()
        if clf is not None:
            try:
                result = clf.predict(description)
                if result["type"]:
                    suggested_type = result["type"]
                confidence  = float(result["confidence"])
                source      = str(result["source"])
                needs_review = bool(result["needs_review"])
            except Exception as exc:
                logging.getLogger(__name__).warning(
                    "Classifier prediction failed for txn description %r: %s",
                    description[:80], exc,
                )

    effective_type = suggested_type or current_type
    # A reinforcement flagged `noSecurity` overrides the type-based default:
    # the txn neither gets a cascade lookup below nor is reported as
    # pending-security in the response.
    needs_security = (effective_type in types_need_security) and not reinforced_no_security

    # ── Security suggestion (only when the type requires one) ─────────────────
    # Skipped entirely when the reinforcement already filled it — the operator
    # has already chosen, so the L1/L2/L3 cascade would just waste cycles.
    if needs_security and description and txn_doc and not (reinforced and reinforced.get("securityId")):
        sec_clf = _get_security_classifier()
        if sec_clf is not None:
            try:
                liq = txn_doc.get("liquidationDate")
                if liq is not None and not isinstance(liq, str):
                    liq = _date_str(liq)
                wallet = str(txn_doc.get("walletId") or "")
                bal = txn_doc.get("balance")
                sec_result = sec_clf.predict(
                    description=description,
                    wallet_id=wallet or None,
                    liquidation_date=liq,
                    transaction_type=effective_type,
                    balance=float(bal) if isinstance(bal, (int, float)) else None,
                )
                suggested_security = sec_result.get("securityId") or ""
                sec_name      = sec_result.get("beehusName", "") or ""
                sec_main      = sec_result.get("mainId", "") or ""
                sec_conf      = float(sec_result.get("confidence", 0.0) or 0.0)
                sec_source    = str(sec_result.get("source", ""))
                sec_ambiguous = bool(sec_result.get("ambiguous", False))
                sec_alts      = sec_result.get("alternatives") or []
                sec_score     = sec_result.get("score")
                sec_tiebreak  = sec_result.get("tiebreak")
            except Exception as exc:
                logging.getLogger(__name__).warning(
                    "Security classifier failed for txn %s: %s",
                    description[:80], exc,
                )

    return {
        "beehusTransactionType": suggested_type,
        "securityId":            suggested_security,
        "needsSecurity":         bool(needs_security),
        "confidence":            confidence,
        "source":                source,
        "needsReview":           bool(needs_review),
        "securityName":          sec_name,
        "securityMainId":        sec_main,
        "securityConfidence":    sec_conf,
        "securitySource":        sec_source,
        "securityAmbiguous":     sec_ambiguous,
        "securityAlternatives":  sec_alts,
        "securityScore":         sec_score,
        "securityTiebreak":      sec_tiebreak,
    }


# ── Provisions ────────────────────────────────────────────────────────────────

@bp.route("/api/beehus/provisions", methods=["POST"])
def provisions_create():
    data = request.get_json(silent=True) or {}
    required = ("companyId", "walletId", "balance",
                "initialDate", "liquidationDate",
                "provisionType", "provisionSource")
    missing = [k for k in required if data.get(k) in (None, "")]
    if missing:
        return jsonify({"error": f"missing fields: {', '.join(missing)}"}), 400
    if not company_visible(data["companyId"]):
        return jsonify({"error": "company is not visible to this user"}), 403
    try:
        result = create_provision(
            company_id=data["companyId"],
            wallet_id=data["walletId"],
            balance=float(data["balance"]),
            initial_date=data["initialDate"],
            liquidation_date=data["liquidationDate"],
            provision_type=data["provisionType"],
            provision_source=data["provisionSource"],
            currency_id=data.get("currencyId", "BRL"),
            description=data.get("description", ""),
            security_id=data.get("securityId") or None,
        )
    except BeehusAPIError as e:
        return _api_error_response(e)
    return jsonify(result), 201


@bp.route("/api/beehus/provisions/<prov_id>", methods=["DELETE"])
def provisions_delete(prov_id):
    """Forward DELETE /beehus/provisions/<id> to upstream Beehus API."""
    try:
        result = delete_provision(prov_id)
    except BeehusAPIError as e:
        return _api_error_response(e)
    return jsonify({"deleted": prov_id, "response": result})


@bp.route("/api/beehus/provisions/search", methods=["POST"])
def provisions_search():
    """Query local `db.provisions` for the company by date-range OVERLAP
    or by single-point ACTIVE-ON-DATE coverage.

    Mandatory: companyId, *and either* (initialDate + finalDate) *or* coverDate.
    Optional: walletId — quando presente, restringe a busca à wallet
    indicada. Sem ele, busca todas as wallets da company (legado,
    usado pelo Beehus Console). Filtrar no backend evita carregar
    provisões de outras wallets que o cliente descartaria de qualquer
    jeito.

    **Dois modos de filtro mutuamente exclusivos**:

    1. **`coverDate`** (ponto único — usado pela prévia de Posição
       Projetada). Mesma régua estrita de `_provisions_detail` em
       `pages/repetir_posicoes.py`:

           provision.initialDate     <= coverDate
           provision.liquidationDate >  coverDate

       Captura **exatamente** as provisões ativas em `coverDate` e que
       **ainda não foram liquidadas** nessa data. Garante que o painel
       de provisões e o cálculo do NAV projetado mostrem o mesmo
       conjunto de docs.

    2. **`initialDate` + `finalDate`** (overlap — usado pelo Beehus
       Console em buscas por janela). Régua legada, intervalo fechado
       nas duas pontas:

           provision.initialDate     <= user.finalDate
           provision.liquidationDate >= user.initialDate

       Captura toda provisão que esteve **ativa em qualquer momento**
       do range — incluindo as que liquidam no último dia.

    Quando `coverDate` é informado, `initialDate`/`finalDate` são
    ignorados. Trashed provisions são excluídas em ambos os modos.
    """
    data = request.get_json(silent=True) or {}
    company_id   = data.get("companyId") or ""
    initial_date = data.get("initialDate") or ""
    final_date   = data.get("finalDate") or ""
    cover_date   = (data.get("coverDate") or "").strip()
    wallet_id    = (data.get("walletId") or "").strip()
    if not company_id:
        return jsonify({"error": "companyId is required"}), 400
    if not cover_date and (not initial_date or not final_date):
        return jsonify({
            "error": "either coverDate or (initialDate + finalDate) is required"
        }), 400
    if not company_visible(company_id):
        return jsonify({"provisions": []})

    _wids = [wallet_id] if wallet_id else None
    if cover_date:
        # Modo "ativo em coverDate" — régua estrita da prévia.
        _provs = beehus_catalog.provisions_active(company_id, cover_date, wallet_ids=_wids)
    else:
        # Modo overlap (legado).
        _provs = beehus_catalog.provisions_overlapping(
            company_id, initial_date, final_date, wallet_ids=_wids)
    _provs.sort(key=lambda p: str(p.get("liquidationDate") or "")[:10], reverse=True)

    wallet_names   = get_wallet_names()
    security_names = get_security_names()

    out = []
    cursor = _provs[:1000]
    for d in cursor:
        wid = beehus_catalog.id_str(d.get("walletId"))
        sid = str(d.get("securityId") or "") if d.get("securityId") else ""
        out.append({
            "id":              str(d["_id"]),
            "walletId":        wid,
            "walletName":      wallet_names.get(wid, "") or wid,
            "securityId":      sid,
            "securityName":    security_names.get(sid, "") if sid else "",
            "balance":         d.get("balance"),
            "provisionType":   d.get("provisionType") or "",
            "currencyId":      d.get("currencyId") or "",
            "provisionSource": d.get("provisionSource") or "",
            "initialDate":     _date_str(d.get("initialDate")),
            "liquidationDate": _date_str(d.get("liquidationDate")),
            "description":     d.get("description") or "",
        })

    return jsonify({"provisions": out, "truncated": len(out) >= 1000})


# ── Execution prices ──────────────────────────────────────────────────────────

@bp.route("/api/beehus/execution-prices", methods=["POST"])
def execution_prices_create():
    """Create an execution price for a security on a given position date.

    Body:
        companyId       (str, required)
        walletId        (str, required)
        securityId      (str, required)
        positionDate    (str, required, YYYY-MM-DD)
        executionPrice  (number, required)
    """
    data = request.get_json(silent=True) or {}
    required = ("companyId", "walletId", "securityId",
                "positionDate", "executionPrice")
    missing = [k for k in required if data.get(k) in (None, "")]
    if missing:
        return jsonify({"error": f"missing fields: {', '.join(missing)}"}), 400
    if not company_visible(data["companyId"]):
        return jsonify({"error": "company is not visible to this user"}), 403
    try:
        execution_price = float(data["executionPrice"])
    except (TypeError, ValueError):
        return jsonify({"error": "executionPrice must be numeric"}), 400
    try:
        result = create_execution_price(
            company_id=data["companyId"],
            wallet_id=data["walletId"],
            security_id=data["securityId"],
            position_date=data["positionDate"],
            execution_price=execution_price,
        )
    except BeehusAPIError as e:
        return _api_error_response(e)
    return jsonify(result), 201


@bp.route("/api/beehus/identify-transactions/execution-extras", methods=["POST"])
def identify_txn_execution_extras():
    """Push executionPrices and create IRRF `taxes` transactions for the rows
    the operator just implemented in Identificar Transações.

    Body:
        executionPrices: [{walletId, securityId, positionDate, executionPrice}]
                         (already deduped client-side by wallet/security/date)
        taxes:           [{sourceTransactionId, balance}]

    companyId is resolved server-side from the wallet (transactions don't carry
    it). The new `taxes` transaction copies entity/wallet/security/dates/currency
    from the source transaction. Returns a per-bucket summary; an upstream 401
    aborts the remaining sends so the operator can re-login.
    """
    data        = request.get_json(silent=True) or {}
    exec_prices = data.get("executionPrices") or []
    taxes       = data.get("taxes") or []

    summary = {"execOk": 0, "execFail": 0, "taxOk": 0, "taxFail": 0, "errors": []}

    def _company_for(wallet_id):
        w = resolve_wallet(wallet_id, {"companyId": 1})
        return str(w.get("companyId")) if w else ""

    # ── executionPrices ───────────────────────────────────────────────────────
    for item in exec_prices:
        wallet_id = str(item.get("walletId") or "")
        sec_id    = str(item.get("securityId") or "")
        pos_date  = str(item.get("positionDate") or "")
        price     = _num(item.get("executionPrice"))
        if not (wallet_id and sec_id and pos_date) or price is None:
            summary["execFail"] += 1
            summary["errors"].append(f"preço de execução inválido: {item}")
            continue
        company_id = _company_for(wallet_id)
        if not company_id or not company_visible(company_id):
            summary["execFail"] += 1
            summary["errors"].append(f"wallet {wallet_id} sem company visível")
            continue
        try:
            create_execution_price(
                company_id=company_id, wallet_id=wallet_id, security_id=sec_id,
                position_date=pos_date, execution_price=price,
            )
            summary["execOk"] += 1
        except BeehusAuthError as e:
            return _api_error_response(e)
        except BeehusAPIError as e:
            summary["execFail"] += 1
            summary["errors"].append(f"preço de execução {sec_id}@{pos_date}: {e}")

    # ── IRRF → new `taxes` transactions ───────────────────────────────────────
    for item in taxes:
        src_id  = str(item.get("sourceTransactionId") or "")
        balance = _num(item.get("balance"))
        if not src_id or balance is None:
            summary["taxFail"] += 1
            summary["errors"].append(f"IRRF inválido: {item}")
            continue
        # Source-transaction fields come FROM THE CLIENT (the Identificar grid
        # already loaded them via /search) — no by-id Mongo round-trip, mirroring
        # the /identify route. companyId is still resolved server-side from the
        # wallet and gated by company visibility (authorization by wallet
        # visibility, not by-id ownership — same trade-off as the conciliação guards).
        wallet_id = str(item.get("walletId") or "")
        if not wallet_id:
            summary["taxFail"] += 1
            summary["errors"].append(f"IRRF sem walletId (origem {src_id})")
            continue
        company_id = _company_for(wallet_id)
        if not company_id or not company_visible(company_id):
            summary["taxFail"] += 1
            summary["errors"].append(f"IRRF wallet {wallet_id} sem company visível")
            continue
        desc = item.get("description") or ""
        sec_id = str(item.get("securityId") or "")
        try:
            create_transaction(
                company_id=company_id,
                entity_id=str(item.get("entityId") or ""),
                wallet_id=wallet_id,
                balance=balance,
                operation_date=_date_str(item.get("operationDate")),
                liquidation_date=_date_str(item.get("liquidationDate")),
                currency_id=item.get("currencyId") or "BRL",
                transaction_type="taxes",
                description=(f"IRRF — {desc}" if desc else "IRRF"),
                security_id=(sec_id or None),
            )
            summary["taxOk"] += 1
        except BeehusAuthError as e:
            return _api_error_response(e)
        except BeehusAPIError as e:
            summary["taxFail"] += 1
            summary["errors"].append(f"IRRF {src_id}: {e}")

    return jsonify(summary)


# ── Util ──────────────────────────────────────────────────────────────────────

@bp.route("/api/beehus/util/parse-strings-excel", methods=["POST"])
def util_parse_strings_excel():
    """Extract a deduped list of non-empty cell values from an uploaded .xlsx.

    The sheet has no header — every non-empty cell is stringified, trimmed,
    and added to the result. Used by the "por datas" pickers to bulk-load
    a list of grouping ids (or any other opaque string id) from a spreadsheet.

    Returns the values in the same order they were first seen on the sheet,
    not sorted — IDs have no inherent ordering and the caller may want to
    preserve the user's original sort.
    """
    f = request.files.get("file")
    if f is None:
        return jsonify({"error": "no file uploaded (expected multipart 'file')"}), 400

    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({"error": "openpyxl is not installed on the server"}), 500

    try:
        wb = load_workbook(f, read_only=True, data_only=True)
    except Exception as e:
        return jsonify({"error": f"could not read workbook: {e}"}), 400

    seen = set()
    values = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is None:
                    continue
                # Numeric ids land as int/float in openpyxl; everything else
                # comes through as a string. Trim and skip empties.
                s = str(v).strip()
                if not s or s in seen:
                    continue
                seen.add(s)
                values.append(s)

    return jsonify({"values": values, "count": len(values)})


@bp.route("/api/beehus/util/parse-dates-excel", methods=["POST"])
def util_parse_dates_excel():
    """Extract a sorted, deduped list of YYYY-MM-DD dates from an uploaded
    .xlsx. The sheet has no header — every cell that looks like a date is
    collected. Used by the "por datas" pickers (Processar / NAV Wallets / NAV
    Groupings / Publicar) to bulk-load a custom list of business days.
    """
    f = request.files.get("file")
    if f is None:
        return jsonify({"error": "no file uploaded (expected multipart 'file')"}), 400

    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({"error": "openpyxl is not installed on the server"}), 500

    try:
        wb = load_workbook(f, read_only=True, data_only=True)
    except Exception as e:
        return jsonify({"error": f"could not read workbook: {e}"}), 400

    dates = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is None:
                    continue
                # openpyxl returns datetime/date objects for date-typed cells.
                if hasattr(v, "strftime"):
                    dates.add(v.strftime("%Y-%m-%d"))
                    continue
                # Fall back to common string formats so people pasting plain
                # text don't get silently dropped.
                if isinstance(v, str):
                    s = v.strip()
                    if not s:
                        continue
                    parsed = None
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                        try:
                            parsed = datetime.strptime(s, fmt).date()
                            break
                        except ValueError:
                            pass
                    if parsed is not None:
                        dates.add(parsed.strftime("%Y-%m-%d"))

    return jsonify({"dates": sorted(dates), "count": len(dates)})


def _date_str(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10]

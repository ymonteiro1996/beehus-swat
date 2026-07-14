import csv
import io
import json
import os
import threading
from datetime import date

from bson import ObjectId
from flask import Blueprint, render_template, jsonify, make_response, request
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from beehus_api import (
    BeehusAPIError,
    BeehusAuthError,
    update_security_mappings,
)
from db import (db, get_biz_dates, load_config_delays, get_company_filter,
                company_visible, get_company_names, get_wallet_names,
                get_grouping_index, atomic_write_json, today_in_brt, _cached_ttl,
                invalidate_cache, business_days_before, IDENTIFICAR_ENABLED)
import beehus_catalog
from security_type_classifier import SecurityTypeClassifier, JSON_PATH
from security_matcher import (
    SecurityMatcher, get_cache, get_mapping_cache, _score_breakdown, _confidence_label,
    extract_features,
)

bp = Blueprint("controlpanel", __name__)

# Number of business-day cards rendered in the date strip at the top of /controlpanel.
# Kept in sync with the inline grid (`grid-template-columns: repeat(N, ...)`)
# in templates/controlpanel.html — change both together. Matches the value used by
# sibling pages (painel, posicoes, caixa, conciliacao, etc.).
_NUM_DATES = 10

# Guards lazy init of _clf and _matcher under multi-threaded WSGI: without
# this, concurrent first-loads each retrain a fresh sklearn model and
# refresh_cache/rebuild reassignments race with in-flight match requests.
_init_lock = threading.Lock()
# Serialises read-modify-write on classifier_overrides.json. Dropdown JS
# fires fire-and-forget POSTs per row change, so several land on different
# WSGI threads and clobber each other's saved overrides.
_overrides_lock = threading.Lock()

# ── Lazy classifier singleton ─────────────────────────────────────────────────
_clf = None

def _get_classifier():
    global _clf
    if _clf is not None:
        return _clf
    with _init_lock:
        if _clf is None and os.path.exists(JSON_PATH):
            c = SecurityTypeClassifier()
            c.train()
            _clf = c          # only cache after successful training
        return _clf

def reset_classifier():
    """Force retrain on next call (used after rebuild_mapping)."""
    global _clf
    with _init_lock:
        _clf = None

# ── Lazy matcher singleton ───────────────────────────────────────────────────
_matcher = None

def _get_matcher():
    global _matcher
    if _matcher is not None:
        return _matcher
    with _init_lock:
        if _matcher is None:
            # SecurityMatcher is API/file-backed: its cache loads from the daily
            # JSON or beehus_catalog (API), never Mongo, and `db` is passed only
            # for signature compatibility (never dereferenced). So the matcher
            # builds even with Mongo disconnected — do NOT gate on db._ready(),
            # or POST /api/controlpanel/match 500s on a Mongo-free instance.
            _matcher = SecurityMatcher(db, classifier=_get_classifier())
        return _matcher

def _reset_matcher():
    """Drop the cached matcher so the next call rebuilds it under the lock."""
    global _matcher
    with _init_lock:
        _matcher = None

# Each entry: (type key in MongoDB, column label shown in the table)
ISSUE_TYPES = [
    ("missing_wallet",                       "Carteira"),
    ("missing_unprocessed_position",         "Posição"),
    ("security_unmapped",                    "Mapeamento"),
    ("security_missing_classification",      "Classificação"),
    ("security_missing_price",               "Registro de Preço"),
    ("security_missing_history_price",       "Preço para o dia"),
]
# NOTA: os tipos de pós-processamento `missing_fund_position_for_explosion` e
# `explosion_error` foram REMOVIDOS do Painel. O endpoint de pre-processing (E)
# — fonte das contagens — não os expõe de forma confiável: as carteiras
# bloqueadas por explosão nem aparecem na resposta, então contá-los exigiria
# voltar a varrer a coleção `issues` no Mongo. Decisão: não exibir explosão aqui.

# Columns appended to the right of ISSUE_TYPES, separated by a visual divider
# in the UI. They surface day-by-day pipeline state per company, not pending
# issues — so the count/total fraction is the natural display format.
EXTRA_COLS = [
    ("processed",     "Posições Processadas"),
    ("nav_wallet",    "NAV Wallet"),
    ("gap",           "GAP"),
    ("nav_grouping",  "NAV Grouping"),
    ("published",     "Published"),
]

# ── Threshold helpers (mirrors pages.conciliacao) ────────────────────────────
# Painel de Controle's GAP column shares the same |returnNavPerShare - returnContribution|
# threshold as the conciliação page, persisted in the same JSON file so both
# UIs stay in sync when an operator tweaks the value. We re-implement the
# tiny read/write helpers here instead of importing from pages.conciliacao to
# avoid coupling page modules; the file format (single key `diffThresholdPct`,
# unit = percent) is the contract.
_CONCILIACAO_CONFIG_FILE = os.path.join(os.path.dirname(__file__), "..", "data",
                                        "conciliacao_config.json")
_DEFAULT_DIFF_THRESHOLD_PCT = 0.01  # 0.01% = 1 basis point

# Per-securityType field configuration for the "Cadastrar ativos" modal
# (templates/controlpanel.html → #registration-modal). Defines which editable
# fields the bottom line of each asset shows, keyed by securityType. The
# frontend renders the inputs and builds the registration JSON straight from
# this file, so adding fields for a new type is a JSON edit — no code change.
# See docs/CONTROLPANEL.md → "Cadastrar ativos".
_SECURITY_TYPE_FIELDS_FILE = os.path.join(os.path.dirname(__file__), "..", "data",
                                          "security_type_fields.json")


def _load_threshold_config():
    defaults = {"diffThresholdPct": _DEFAULT_DIFF_THRESHOLD_PCT}
    if not os.path.exists(_CONCILIACAO_CONFIG_FILE):
        return defaults
    try:
        with open(_CONCILIACAO_CONFIG_FILE, "r", encoding="utf-8") as f:
            return {**defaults, **(json.load(f) or {})}
    except Exception:
        return defaults


def _save_threshold_config(cfg):
    """Persist via atomic_write_json. Returns (ok, friendly_error_message)."""
    try:
        atomic_write_json(_CONCILIACAO_CONFIG_FILE, cfg)
        return True, ""
    except Exception as exc:
        import logging
        logging.getLogger(__name__).error("save threshold cfg failed: %s", exc)
        # Don't leak the full filesystem path in the error.
        return False, "verifique sincronização do OneDrive e tente novamente"


def _diff_threshold_decimal(req=None):
    """Resolve threshold in DECIMAL form (UI/storage unit is percent).

    Priority: explicit ?threshold=<pct> query param → config file → default.
    """
    pct = None
    if req is not None:
        raw = (req.args.get("threshold") or "").strip()
        if raw:
            try:
                pct = float(raw)
            except ValueError:
                pct = None
    if pct is None:
        pct = float(_load_threshold_config().get("diffThresholdPct",
                                                _DEFAULT_DIFF_THRESHOLD_PCT))
    return max(0.0, pct / 100.0)


# ── Per-company / per-date counters for the extra columns ────────────────────

def _wallets_by_company():
    """Returns ({companyId: total_wallets}, {walletIdStr: companyIdStr}).

    processedPosition has no companyId field — we resolve company through
    walletId. Doing the join in Python (one full wallets scan) avoids a
    MongoDB $lookup on the hot path. Cached 5 min via _cached_ttl: the
    wallets collection rarely changes (new wallet registrations) and this
    function runs on every /api/controlpanel/rows call. Callers must NOT mutate.
    """
    def _load():
        by_company = {}
        wallet_to_company = {}
        for w in beehus_catalog.wallets_index().values():
            cid = str(w.get("companyId") or "")
            wid = beehus_catalog.id_str(w.get("_id"))
            if cid and wid:
                by_company[cid] = by_company.get(cid, 0) + 1
                wallet_to_company[wid] = cid
        return (by_company, wallet_to_company)
    result = _cached_ttl("controlpanel.wallets_by_company", _load)
    if not result[0]:
        # Não deixe um índice vazio TRANSITÓRIO (token ainda não pronto / hiccup
        # na API) grudar pelos 5 min do TTL: ele zera o denominador (`total`) das
        # colunas de progresso do Painel e elas viram "—" mesmo com dado vivo
        # disponível. Mesmo guard que beehus_catalog.wallets_index() já aplica na
        # fonte — aqui protege a cache DERIVADA, que senão mascararia o self-heal.
        invalidate_cache("controlpanel.wallets_by_company")
    return result


def _groupings_by_company():
    """{companyId: total_untrashed_groupings}. Cached 5 min via _cached_ttl —
    callers must NOT mutate the returned dict.

    Derivado de `get_grouping_index()` (API + fallback Mongo) em vez de um
    aggregate direto no Mongo: conta os agrupamentos não-trashed por empresa."""
    def _load():
        out = {}
        for g in get_grouping_index().values():
            if g.get("trashed"):
                continue
            cid = g.get("companyId") or ""
            if cid:
                out[cid] = out.get(cid, 0) + 1
        return out
    result = _cached_ttl("controlpanel.groupings_by_company", _load)
    if not result:
        # Mesmo guard de _wallets_by_company: um vazio transitório não pode
        # grudar pelo TTL e zerar os totais das colunas NAV Grouping / Published.
        invalidate_cache("controlpanel.groupings_by_company")
    return result


def _navpackage_counts_by_company(date, threshold):
    """Quatro dicts paralelos (por companyId) na data:

      - nav_wallet:   carteiras com NAV calculado
      - gap:          carteiras com |returnNavPerShare - returnContribution| > threshold
      - nav_grouping: agrupamentos com NAV calculado
      - published:    agrupamentos publicados

    Via endpoint consolidado `/results` (1 chamada por empresa, AO VIVO,
    paralelizado), com fallback Mongo por empresa. `walletsWithNav`/
    `totalGroupings`/`publishedGroupings` já vêm prontos; o `gap` é contado sobre
    `walletsWithNavDetailed` com a MESMA regra do mismatch da Conciliação. Só
    entram empresas com atividade na data (espelha o $group por companyId)."""
    nav_wallet, gap, nav_grouping, published = {}, {}, {}, {}
    results = beehus_catalog.nav_results_many(beehus_catalog.all_company_ids(), date)
    for cid, res in results.items():
        nw = res.get("walletsWithNav") or 0
        ng = res.get("totalGroupings") or 0
        pub = res.get("publishedGroupings") or 0
        g = sum(1 for w in res.get("walletsWithNavDetailed", [])
                if beehus_catalog._nav_results_is_gap(w, threshold))
        if nw:
            nav_wallet[cid] = nw
        if g:
            gap[cid] = g
        if ng:
            nav_grouping[cid] = ng
        if pub:
            published[cid] = pub
    return nav_wallet, gap, nav_grouping, published


def _extra_cell(count, total, *, mode="ratio"):
    """Format an extras cell. mode: 'ratio' shows X/Y with colour by completeness;
    'count' shows just X with red if > 0 (for GAP)."""
    if mode == "count":
        if count > 0:
            cls = "bg-red-100 text-red-700 font-medium"
        else:
            cls = "text-gray-400"
        return {"count": count, "total": None, "label": str(count), "cls": cls}

    # ratio mode
    if total <= 0:
        cls = "text-gray-300"
        label = "—"
    elif count >= total:
        cls = "bg-green-100 text-green-700 font-medium"
        label = f"{count}/{total}"
    elif count == 0:
        cls = "bg-red-50 text-red-600"
        label = f"{count}/{total}"
    else:
        cls = "bg-amber-50 text-amber-700"
        label = f"{count}/{total}"
    return {"count": count, "total": total, "label": label, "cls": cls}

# ── Helpers ────────────────────────────────────────────────────────────────────

def _cell_cls(count):
    return "bg-green-100 text-green-700 font-medium" if count > 0 else "text-gray-300"


# Como derivar cada contagem do endpoint E (pre-processing) p/ bater 1:1 com o
# Mongo `issues` (validado ao vivo). `(tipo_painel, chave_top_level, chave_Detailed)`:
#  - missing_wallet / missing_unprocessed_position → contagem TOP-LEVEL do E
#    (1 issue por carteira; top-level == len(Detailed) == Mongo).
#  - security_* → contar via *Detailed, pois o shape difere por tipo e o Mongo
#    conta OCORRÊNCIAS (security × carteira):
#      · unmapped / missing_price / missing_classification: 1 item por security
#        DISTINTA, com `affectedWallets[]` → ocorrências = soma de affectedWallets.
#      · missing_history_price: 1 item por carteira (tem `walletId`, SEM
#        affectedWallets) → ocorrências = len(Detailed).
#    Regra única que cobre os dois: len(affectedWallets) se presente, senão 1.
#    Validado == Mongo: unmapped 29, missing_price 13, history_price 31.
_PREPROC_SPEC = [
    ("missing_wallet",                  "missingWallet",  None),
    ("missing_unprocessed_position",    "missingPosition", None),
    ("security_unmapped",               None, "securityUnmappedDetailed"),
    ("security_missing_classification", None, "securityMissingClassificationDetailed"),
    ("security_missing_price",          None, "securityMissingPriceDetailed"),
    ("security_missing_history_price",  None, "securityMissingHistoryPriceDetailed"),
]


def _detailed_occurrences(lst):
    """Ocorrências (security × carteira) numa lista *Detailed do E: soma
    len(affectedWallets) quando o item tem essa chave; senão conta o item como 1
    (itens por-carteira, ex.: securityMissingHistoryPriceDetailed)."""
    total = 0
    for it in (lst or []):
        if not isinstance(it, dict):
            continue
        aw = it.get("affectedWallets")
        total += len(aw) if isinstance(aw, list) else 1
    return total


def _counts_from_status(st):
    """`{tkey: count}` para UMA empresa, a partir do status do endpoint E,
    aplicando a regra `_PREPROC_SPEC` (top-level p/ missing_*, soma de
    ocorrências nas `*Detailed` p/ security_*). Mesma contagem usada no grid do
    Painel — garante que summary e grid não divergem."""
    out = {}
    for tkey, top_key, det_key in _PREPROC_SPEC:
        n = (st.get(top_key) or 0) if top_key else _detailed_occurrences(st.get(det_key))
        if n:
            out[tkey] = n
    return out


def _preproc_counts(date):
    """Via endpoint E (pre-processing), fan-out por empresa (paralelo, como
    `nav_results_many`): retorna ({(companyId, type): count} p/ os 6 tipos do
    Painel, {companyId: processedWallets}). Substitui os reads Mongo de `issues`
    (contagem por tipo) e `processedPosition` (contagem) do grid numa só passada."""
    counts, processed = {}, {}
    statuses = beehus_catalog.preprocessing_status_many(
        beehus_catalog.all_company_ids(), date)
    for cid, st in statuses.items():
        for tkey, n in _counts_from_status(st).items():
            counts[(cid, tkey)] = n
        pw = st.get("processedWallets") or 0
        if pw:
            processed[cid] = pw
    return counts, processed


def _format_issue(issue):
    return {
        "type":                  issue.get("type", ""),
        "description":           issue.get("description", ""),
        "walletId":              str(issue.get("walletId", "") or ""),
        "externalId":            str(issue.get("externalId", "") or ""),
        "externalOrigin":        str(issue.get("externalOrigin", "") or ""),
        "securityId":            str(issue.get("securityId", "") or ""),
        "unprocessedSecurityId": str(issue.get("unprocessedSecurityId", "") or ""),
        "createdAt":             issue["createdAt"].strftime("%Y-%m-%d %H:%M") if issue.get("createdAt") else "",
    }


# ── Routes ─────────────────────────────────────────────────────────────────────

def _date_cards(dates):
    """Lista de {date} para os cards do seletor (SEM contador de issues). As datas
    são os `_NUM_DATES` dias ÚTEIS de `get_biz_dates` (já exclui fins de semana);
    o seletor é pura navegação de datas e o grid (/rows) preenche ao clicar.
    Mongo-free — o contador de pendências por data foi removido."""
    return [{"date": d} for d in dates]


@bp.route("/controlpanel")
def index():
    dates   = get_biz_dates(_NUM_DATES)
    cards   = _date_cards(dates)
    delays  = load_config_delays()
    default_delay = min(delays.values(), default=1) if delays else 1
    threshold_pct = float(_load_threshold_config().get("diffThresholdPct",
                                                      _DEFAULT_DIFF_THRESHOLD_PCT))
    return render_template(
        "controlpanel.html",
        cards=cards,
        types=ISSUE_TYPES,
        extra_cols=EXTRA_COLS,
        default_delay=default_delay,
        threshold_pct=threshold_pct,
        identificar_enabled=IDENTIFICAR_ENABLED,
    )


@bp.route("/api/controlpanel/date-cards")
def date_cards():
    """Return `_NUM_DATES` business-day cards ending on the given date.

    endDate must be ISO YYYY-MM-DD; anything else falls back to today so a
    malformed query param never reaches get_biz_dates / datetime parsing.
    """
    import re
    end_date = (request.args.get("endDate") or "").strip() or None
    if end_date and not re.match(r"^\d{4}-\d{2}-\d{2}$", end_date):
        end_date = None
    try:
        dates = get_biz_dates(_NUM_DATES, end_date=end_date)
    except Exception:
        dates = get_biz_dates(_NUM_DATES)
    return jsonify({"cards": _date_cards(dates), "dates": dates})


def _unidentified_txn_count_by_company(date):
    """Count transactions still missing a `beehusTransactionType` per company
    for `date`, via endpoint G (`list_transactions`) — fan-out por empresa.

    Espelha a busca padrão de Identificar Transações: `dateType=liquidation` na
    data, "não identificada" = `beehusTransactionType` null/'' , `trashed`
    excluído. O endpoint G já escopa por empresa (dispensa o mapa
    walletId→companyId); os filtros de tipo/trashed são aplicados no cliente
    sobre o resultado. Returns {companyId: count}."""
    counts = {}
    by_company = beehus_catalog.transactions_search_many(
        beehus_catalog.all_company_ids(),
        initial_date=date, final_date=date, date_type="liquidation")
    for cid, txns in by_company.items():
        n = sum(1 for t in txns
                if not t.get("trashed")
                and (t.get("beehusTransactionType") or "") == "")
        if n:
            counts[cid] = n
    return counts


@bp.route("/api/controlpanel/rows")
def get_rows():
    date          = request.args.get("date", get_biz_dates(1)[0])
    company_names = get_company_names()
    threshold     = _diff_threshold_decimal(request)

    # Issues (6 tipos) + posições processadas saem do endpoint E (pre-processing),
    # 1 chamada por empresa em paralelo (o fan-out já usa 10 workers internos) —
    # substitui os reads Mongo de `issues` e `processedPosition`. Inclui empresas
    # com atividade no pipeline mesmo sem issues pendentes (senão as colunas
    # processed/NAV/published sumiriam quando o backlog zera).
    #
    # A coluna TXN (transações não identificadas, endpoint G) NÃO é carregada aqui:
    # medição mostrou que G é o endpoint mais caro (~60% do tempo de /rows) e
    # bloqueava a renderização da tela inteira. Ela é buscada à parte por
    # GET /api/controlpanel/txn-counts depois que a grade renderiza (ver template).
    # Por isso o set de empresas abaixo NÃO inclui mais txn — empresas que só têm
    # transações pendentes (sem issue/posição/NAV) são anexadas pelo front quando
    # o /txn-counts retorna.
    counts, processed = _preproc_counts(date)
    wallets_total, _ = _wallets_by_company()
    groupings_total = _groupings_by_company()
    nav_wallet, gap, nav_grouping, published = _navpackage_counts_by_company(date, threshold)

    company_ids = (
        {cid for (cid, _) in counts}
        | set(processed.keys())
        | set(nav_wallet.keys())
        | set(nav_grouping.keys())
        | set(published.keys())
    )
    cf = get_company_filter()
    if cf:
        company_ids = company_ids & cf

    rows = []
    for cid in sorted(company_ids, key=lambda c: company_names.get(c, c)):
        cells = []
        for key, _label in ISSUE_TYPES:
            count = counts.get((cid, key), 0)
            cells.append({
                "type":  key,
                "count": count,
                "label": str(count) if count > 0 else "—",
                "cls":   _cell_cls(count),
            })

        wt = wallets_total.get(cid, 0)
        gt = groupings_total.get(cid, 0)
        extras = [
            {"key": "processed",
             **_extra_cell(processed.get(cid, 0),    wt)},
            {"key": "nav_wallet",
             **_extra_cell(nav_wallet.get(cid, 0),   wt)},
            {"key": "gap",
             **_extra_cell(gap.get(cid, 0),          wt, mode="count")},
            {"key": "nav_grouping",
             **_extra_cell(nav_grouping.get(cid, 0), gt)},
            {"key": "published",
             **_extra_cell(published.get(cid, 0),    gt)},
        ]

        rows.append({
            "companyId": cid,
            "company":   company_names.get(cid, cid),
            "cells":     cells,
            "extras":    extras,
            # A coluna TXN é preenchida assíncronamente via /txn-counts depois que
            # a grade renderiza (G é o endpoint mais caro). Aqui vai só um marcador
            # "pendente" — o front renderiza a célula como "…" e troca quando o
            # /txn-counts chega.
            "txn": {"pending": True},
        })

    return jsonify({"rows": rows, "date": date})


@bp.route("/api/controlpanel/txn-counts")
def txn_counts():
    """Contagem de transações não identificadas por empresa (coluna TXN) na data.

    Servido SEPARADO de /rows porque o endpoint G (transactions) é, de longe, o
    mais caro do grid (~60% do tempo medido) e bloqueava a renderização da tela.
    O fluxo é: /rows pinta a grade com E + /results; o front então busca esta rota
    e preenche a coluna TXN — além de ANEXAR linhas de empresas que só têm
    transações pendentes (sem issue/posição/NAV), preservando o conjunto de
    empresas que o /rows antigo cobria.

    Cada item já vem com a célula pronta (`{count,label,cls}`) para o front não
    precisar replicar a regra de cor (`_cell_cls`). Só empresas com count > 0
    aparecem (o fan-out de G só conta positivos)."""
    date          = request.args.get("date", get_biz_dates(1)[0])
    company_names = get_company_names()
    txn_unident   = _unidentified_txn_count_by_company(date)
    cf = get_company_filter()
    items = []
    for cid, n in txn_unident.items():
        if cf and cid not in cf:
            continue
        items.append({
            "companyId": cid,
            "company":   company_names.get(cid, cid),
            "txn": {"count": n, "label": str(n) if n > 0 else "—", "cls": _cell_cls(n)},
        })
    return jsonify({"date": date, "items": items})


# ── Per-company issue summary (used by Fluxo apontamentos) ────────────────────

@bp.route("/api/controlpanel/issues-summary")
def issues_summary():
    """Pending-issue counts for a single (company, date), narrowed to the
    types the caller cares about.

    Query params:
        companyId  required
        date       required, YYYY-MM-DD
        types      optional, comma-separated. Defaults to all ISSUE_TYPES keys.

    The Fluxo apontamento for each step asks for a focused subset (e.g. just
    the four "post-process" issue types). Returning one type at a time would
    multiply round-trips for no benefit, so the endpoint accepts a list and
    aggregates once. Labels come from ISSUE_TYPES so the UI doesn't need to
    keep its own copy in sync.
    """
    cid  = (request.args.get("companyId") or "").strip()
    date = (request.args.get("date") or "").strip()
    if not cid or not date:
        return jsonify({"error": "companyId and date are required"}), 400
    if not company_visible(cid):
        return jsonify({"error": "company not visible"}), 403

    valid_types = {k for k, _ in ISSUE_TYPES}
    label_by_type = dict(ISSUE_TYPES)
    raw_types = (request.args.get("types") or "").strip()
    if raw_types:
        requested = [t.strip() for t in raw_types.split(",") if t.strip()]
        types = [t for t in requested if t in valid_types]
    else:
        types = [k for k, _ in ISSUE_TYPES]
    if not types:
        return jsonify({"companyId": cid, "date": date, "types": [], "total": 0})

    # Counts come from the pre-processing endpoint (E) — the SAME source the
    # Painel grid uses (`_preproc_counts`), so the apontamento summary and the
    # grid can never drift. Replaces the direct `db.issues.aggregate`. One E
    # call per (company, date); empty/failed → all-zero counts.
    statuses = beehus_catalog.preprocessing_status_many([cid], date)
    st = statuses.get(cid)
    counts = _counts_from_status(st) if st else {}

    # Preserve the order requested by the caller so the UI can render the
    # apontamento list deterministically — order in ISSUE_TYPES matches the
    # left-to-right column order of the Painel de Controle table.
    items = [
        {"type": t, "label": label_by_type[t], "count": counts.get(t, 0)}
        for t in types
    ]
    total = sum(it["count"] for it in items)
    return jsonify({"companyId": cid, "date": date, "types": items, "total": total})


# ── Threshold endpoints (shared with conciliação) ─────────────────────────────

@bp.route("/api/controlpanel/threshold", methods=["GET"])
def get_threshold():
    """Return the current diffThresholdPct (percent units, e.g. 0.01)."""
    cfg = _load_threshold_config()
    return jsonify({"diffThresholdPct": cfg.get("diffThresholdPct",
                                               _DEFAULT_DIFF_THRESHOLD_PCT)})


@bp.route("/api/controlpanel/threshold", methods=["PUT", "POST"])
def set_threshold():
    """Update the diffThresholdPct. Writes to the same config file as
    conciliação so both pages stay in sync."""
    body = request.get_json(force=True, silent=True) or {}
    try:
        pct = float(body.get("diffThresholdPct"))
    except (TypeError, ValueError):
        return jsonify({"error": "diffThresholdPct inválido"}), 400
    if pct < 0 or pct > 10:
        return jsonify({"error": "diffThresholdPct fora do intervalo [0, 10]"}), 400
    cfg = _load_threshold_config()
    cfg["diffThresholdPct"] = pct
    ok, err = _save_threshold_config(cfg)
    if not ok:
        return jsonify({"error": f"falha ao salvar: {err}"}), 500
    return jsonify(cfg)


def _enrich_issue_securities(issues):
    """Anexa `beehusName`/`mainId` a cada issue (in-place) cruzando `securityId`
    com a coleção de securities via `beehus_catalog.securities_by_ids`. Uma só
    busca em lote para toda a lista. Compartilhado por `/detail` e
    `/wallet-issues` para não duplicar a regra de enriquecimento."""
    def _to_oid(val):
        try:
            return ObjectId(val)
        except (TypeError, ValueError):
            return None

    sec_ids = [_to_oid(i["securityId"]) for i in issues if i.get("securityId")]
    sec_ids = [s for s in sec_ids if s]
    sec_map = {}
    if sec_ids:
        sec_map = {
            str(s["_id"]): s
            for s in beehus_catalog.securities_by_ids(sec_ids).values()
        }
    for issue in issues:
        sec = sec_map.get(issue.get("securityId", "")) or {}
        issue["beehusName"] = sec.get("beehusName", "") or ""
        issue["mainId"]     = sec.get("mainId", "") or ""
    return issues


@bp.route("/api/controlpanel/detail")
def get_detail():
    cid  = request.args.get("companyId")
    date = request.args.get("date")
    typ  = request.args.get("type")

    if cid and not company_visible(cid):
        return jsonify({"issues": [], "date": date, "type": typ}), 403

    wallet_names = get_wallet_names()

    # Issues vêm do pre-processing (E) via `beehus_catalog.issues_detail` — os
    # mesmos arrays `*Detailed` que o grid já consome, expandidos por
    # affectedWallets (1 linha por security×carteira). Drop-in do antigo
    # `db.issues.find({companyId, status:'pending', date, type})`; validado ao
    # vivo == Mongo nos 5 tipos. O enriquecimento (walletName/beehusName/mainId)
    # segue abaixo, igual a antes.
    issues = sorted([
        {**_format_issue(issue),
         "walletName": wallet_names.get(str(issue.get("walletId", "") or ""), "")}
        for issue in beehus_catalog.issues_detail(cid, date, typ)
    ], key=lambda x: x["createdAt"])

    _enrich_issue_securities(issues)

    return jsonify({"issues": issues, "date": date, "type": typ})


@bp.route("/api/controlpanel/wallet-issues")
def wallet_issues():
    """Verificação POR CARTEIRA (modal aberto ao clicar no nome da empresa).

    Dois modos, mesma fonte única (pre-processing E, 1 chamada por
    company+date — os mesmos arrays `*Detailed` do grid, via
    `beehus_catalog.issues_by_wallet_detail`):

      • SEM `walletId` — lista as carteiras da empresa com o contador de issues
        pendentes na data (`wallets: [{walletId, walletName, issueCount}]`),
        carteiras com issue primeiro. Popula o seletor do modal.
      • COM `walletId` — issues daquela carteira na data, agrupadas pelos 6
        tipos do Painel em ordem canônica (`types: [{type, label, issues[]}]`),
        cada issue enriquecida com walletName/beehusName/mainId (igual ao
        `/detail`).

    Query params: companyId (obrig.), date (obrig., YYYY-MM-DD), walletId (opc.).
    """
    cid  = (request.args.get("companyId") or "").strip()
    date = (request.args.get("date") or "").strip()
    wid  = (request.args.get("walletId") or "").strip()
    if not cid or not date:
        return jsonify({"error": "companyId and date are required"}), 400
    if not company_visible(cid):
        return jsonify({"error": "company not visible"}), 403

    company_names = get_company_names()
    company_name  = company_names.get(cid) or cid
    wallets       = beehus_catalog.wallets_for_company(cid)  # {walletId: name}
    by_wallet     = beehus_catalog.issues_by_wallet_detail(cid, date)

    # ── Modo 1: lista de carteiras para o seletor ─────────────────────────
    if not wid:
        items = [
            {"walletId": w, "walletName": nm or w,
             "issueCount": len(by_wallet.get(w, []))}
            for w, nm in wallets.items()
        ]
        # Defensivo: carteiras que só aparecem nas issues (fora do índice
        # partner_wallets) não podem ficar órfãs — o operador precisa poder
        # abri-las. Improvável, mas o índice pode estar desatualizado.
        known = set(wallets)
        for owid, rows in by_wallet.items():
            if owid and owid not in known:
                items.append({"walletId": owid, "walletName": owid,
                              "issueCount": len(rows)})
        # Carteiras com issue no topo (mais issues primeiro), depois alfabético.
        items.sort(key=lambda it: (-it["issueCount"], (it["walletName"] or "").lower()))
        return jsonify({
            "companyId": cid, "companyName": company_name, "date": date,
            "wallets": items,
            "totalWithIssues": sum(1 for it in items if it["issueCount"]),
        })

    # ── Modo 2: issues de uma carteira específica ─────────────────────────
    wallet_name = wallets.get(wid) or get_wallet_names().get(wid, "") or wid
    rows = sorted(
        [{**_format_issue(r), "walletName": wallet_name}
         for r in by_wallet.get(wid, [])],
        key=lambda x: (x["type"], x["createdAt"]),
    )
    _enrich_issue_securities(rows)

    label_by_type = dict(ISSUE_TYPES)
    groups = []
    for tkey, _label in ISSUE_TYPES:
        tissues = [r for r in rows if r["type"] == tkey]
        if tissues:
            groups.append({"type": tkey, "label": label_by_type[tkey],
                           "issues": tissues})
    return jsonify({
        "companyId": cid, "companyName": company_name, "date": date,
        "walletId": wid, "walletName": wallet_name,
        "types": groups, "total": len(rows),
    })


# ── Cell drill-down (Posições Processadas / NAV Wallet / NAV Grouping / Published)
#
# The 4 right-most columns in /controlpanel show pipeline progress per
# (company, date). Operators want to click the count and see which wallets
# (or groupings) are actually done vs pending. This endpoint backs the
# `#cell-detail-modal` on the Home view — one row per wallet (for the
# wallet-level columns) or per grouping (for the grouping-level columns),
# each with a boolean `done` flag the UI renders as a status badge.
#
# Column → meaning of `done`:
#   processed     — wallet has a `processedPosition` for (positionDate)
#   nav_wallet    — wallet has a NAV `navPackages` doc (walletId set)
#   nav_grouping  — grouping has a NAV `navPackages` doc (groupingId set)
#   published     — grouping has a NAV `navPackages` doc with published=true
#
# All navPackages queries respect `trashed != true` and filter by companyId
# to match the count rendered in the table.

_WALLET_COLUMNS    = {"processed", "nav_wallet"}
_GROUPING_COLUMNS  = {"nav_grouping", "published"}
_CELL_DETAIL_COLS  = _WALLET_COLUMNS | _GROUPING_COLUMNS

# Issue types that can block a wallet from being processed. Surfaced in
# the "Posições Processadas" drill-down so the operator can tell at a
# glance why an unprocessed wallet hasn't reached `processedPosition`
# yet. Post-processing types (`missing_fund_position_for_explosion`,
# `explosion_error`) are deliberately excluded — they're symptoms of a
# *later* stage and would only add noise to a "why isn't this wallet
# processed?" view. Order mirrors the canonical ISSUE_TYPES list so the
# chips render in pipeline order.
_PROCESSING_BLOCKING_ISSUE_TYPES = (
    "missing_wallet",
    "missing_unprocessed_position",
    "security_unmapped",
    "security_missing_classification",
    "security_missing_price",
    "security_missing_history_price",
)


def _wallets_for_company(company_id):
    """List of {id, name} for the given company, sorted by name."""
    out = []
    for wid, nm in beehus_catalog.wallets_for_company(company_id).items():
        out.append({"id": wid, "name": (nm or "")})
    out.sort(key=lambda w: (w["name"] or w["id"]).lower())
    return out


def _untrashed_groupings_for_company(company_id):
    """List of {id, name, walletIds} from `get_grouping_index()`. Filtered
    to untrashed and matching company. We copy the wallet-id list rather
    than aliasing the cache entry so callers can mutate freely."""
    gindex = get_grouping_index()
    out = []
    for gid, info in gindex.items():
        if info.get("trashed"):
            continue
        if info.get("companyId") != company_id:
            continue
        out.append({
            "id":        gid,
            "name":      info.get("name") or gid,
            "walletIds": list(info.get("walletIds") or []),
        })
    out.sort(key=lambda g: (g["name"] or g["id"]).lower())
    return out


def _processed_done_wallets(company_id, date, wallet_ids):
    """Set of walletIds (str) with a processedPosition for the date.

    Via endpoint A (`processed_existing_wallets`, walletIds plural numa data),
    com fallback Mongo."""
    if not wallet_ids:
        return set()
    return beehus_catalog.processed_existing_wallets(company_id, date, wallet_ids)


def _unprocessed_existing_wallets(company_id, date, wallet_ids):
    """Set of walletIds (str) that have at least one
    `unprocessedSecurityPositions` doc for the date. Surfaces in the
    "Posições Processadas" drill-down so operators can tell whether a
    pending wallet is waiting on raw positions arriving from upstream
    (no unprocessed doc yet) vs waiting on the processing step itself
    (unprocessed doc exists but no processedPosition yet).

    Via endpoint B (`unprocessed_existing_wallets`), com fallback Mongo."""
    if not wallet_ids:
        return set()
    return beehus_catalog.unprocessed_existing_wallets(company_id, date, wallet_ids)


def _blocking_issues_by_wallet(company_id, date, wallet_ids):
    """{walletId: [{type, label, count}, ...]} for pending issues whose
    type is in `_PROCESSING_BLOCKING_ISSUE_TYPES`. One aggregation
    covers every wallet in the company so the per-wallet rendering on
    the frontend is just a dict lookup. Issues without a walletId
    (company-level rows) are skipped — they don't belong next to a
    specific wallet in the UI."""
    if not wallet_ids:
        return {}
    # Sourced from the pre-processing endpoint (get_preprocessing_status) via
    # beehus_catalog, which returns {walletId: [{type, label, count}, ...]} in
    # the canonical pipeline order ({} if the endpoint shape isn't recognised).
    # As contagens do grid (/rows) também vêm do E (via _preproc_counts) e os
    # date-cards não têm mais contador (puro get_biz_dates, sem DB). Ainda no
    # Mongo só os drill-downs (on-click, não a tela inicial): issues-summary e
    # detail (lista de docs individuais).
    return beehus_catalog.blocking_issues_by_wallet(
        company_id, date, wallet_ids,
        _PROCESSING_BLOCKING_ISSUE_TYPES, dict(ISSUE_TYPES))


def _nav_done_wallets(company_id, date):
    """Set of walletIds (str) com NAV calculado em (companyId, positionDate).
    Via endpoint consolidado `/results` (1 chamada, ao vivo), fallback Mongo."""
    res = beehus_catalog.nav_results(company_id, date)
    return {beehus_catalog.id_str(w.get("walletId"))
            for w in res.get("walletsWithNavDetailed", [])
            if beehus_catalog.id_str(w.get("walletId"))}


def _nav_done_groupings(company_id, date, *, only_published=False):
    """Set of groupingIds (str) com NAV calculado em (companyId, positionDate).
    `only_published=True` exige `published == true` (coluna "Published").
    Via endpoint consolidado `/results`, fallback Mongo."""
    res = beehus_catalog.nav_results(company_id, date)
    return {beehus_catalog.id_str(g.get("groupingId"))
            for g in res.get("groupingsDetailed", [])
            if beehus_catalog.id_str(g.get("groupingId"))
            and (not only_published or g.get("published"))}


@bp.route("/api/controlpanel/cell-detail")
def cell_detail():
    """Drill-down detail for the 4 progress columns on the Home view.

    Query params:
        companyId  required
        date       required, YYYY-MM-DD
        column     required, one of {processed, nav_wallet, nav_grouping, published}

    For wallet-level columns the response groups wallets under their
    grouping (with an `orphanWallets` bucket for wallets that aren't part
    of any grouping for this company). For grouping-level columns the
    response is a flat sorted list of groupings.
    """
    company_id = (request.args.get("companyId") or "").strip()
    date       = (request.args.get("date") or "").strip()
    column     = (request.args.get("column") or "").strip()

    if not company_id or not date or not column:
        return jsonify({"error": "companyId, date and column are required"}), 400
    if column not in _CELL_DETAIL_COLS:
        return jsonify({"error": f"invalid column: {column}"}), 400
    if not company_visible(company_id):
        return jsonify({"error": "company not visible"}), 403

    company_names = get_company_names()
    company_name  = company_names.get(company_id) or company_id

    groupings = _untrashed_groupings_for_company(company_id)

    # ── Grouping-level columns ────────────────────────────────────────────
    if column in _GROUPING_COLUMNS:
        done = _nav_done_groupings(company_id, date,
                                   only_published=(column == "published"))
        items = [{
            "groupingId":   g["id"],
            "groupingName": g["name"],
            "done":         g["id"] in done,
        } for g in groupings]
        return jsonify({
            "column":          column,
            "level":           "grouping",
            "companyId":       company_id,
            "companyName":     company_name,
            "date":            date,
            "totalGroupings":  len(groupings),
            "doneGroupings":   sum(1 for it in items if it["done"]),
            "groupings":       items,
        })

    # ── Wallet-level columns (processed / nav_wallet) ─────────────────────
    wallets = _wallets_for_company(company_id)
    wallet_ids = {w["id"] for w in wallets}
    wallet_names = {w["id"]: (w["name"] or w["id"]) for w in wallets}

    if column == "processed":
        done = _processed_done_wallets(company_id, date, wallet_ids)
        # Side-channel flag: tells the operator whether the wallet is
        # pending because raw positions haven't arrived (no unprocessed
        # doc) or because processing simply hasn't run yet (unprocessed
        # doc exists). Only relevant for the "Posições Processadas"
        # drill-down, so we skip the query for `nav_wallet`.
        unprocessed_set = _unprocessed_existing_wallets(company_id, date, wallet_ids)
        # Blocking issues per wallet — surfaced only for the "Posições
        # Processadas" view so the operator can tell *why* an
        # unprocessed wallet hasn't reached `processedPosition`.
        issues_by_wallet = _blocking_issues_by_wallet(
            company_id, date, wallet_ids
        )
    else:  # nav_wallet
        done = _nav_done_wallets(company_id, date)
        unprocessed_set    = None
        issues_by_wallet   = None

    # Build per-grouping wallet lists; track which company wallets get
    # captured so we can surface the leftovers as `orphanWallets`.
    seen = set()
    groupings_out = []
    for g in groupings:
        # Restrict to wallets that actually belong to this company — the
        # cached grouping index may carry stale ids if a wallet was moved.
        wallets_out = []
        for wid in g["walletIds"]:
            if wid not in wallet_ids:
                continue
            entry = {
                "walletId":   wid,
                "walletName": wallet_names.get(wid) or wid,
                "done":       wid in done,
            }
            if unprocessed_set is not None:
                entry["hasUnprocessed"] = wid in unprocessed_set
            if issues_by_wallet is not None:
                entry["issues"] = issues_by_wallet.get(wid, [])
            wallets_out.append(entry)
            seen.add(wid)
        wallets_out.sort(key=lambda w: w["walletName"].lower())
        if not wallets_out:
            # Skip empty groupings — they would render an empty section with
            # zero wallets and only add noise to the modal.
            continue
        groupings_out.append({
            "groupingId":   g["id"],
            "groupingName": g["name"],
            "wallets":      wallets_out,
            "doneCount":    sum(1 for w in wallets_out if w["done"]),
            "totalCount":   len(wallets_out),
        })

    orphan_wallets = []
    for w in wallets:
        if w["id"] in seen:
            continue
        entry = {
            "walletId":   w["id"],
            "walletName": w["name"] or w["id"],
            "done":       w["id"] in done,
        }
        if unprocessed_set is not None:
            entry["hasUnprocessed"] = w["id"] in unprocessed_set
        if issues_by_wallet is not None:
            entry["issues"] = issues_by_wallet.get(w["id"], [])
        orphan_wallets.append(entry)

    payload = {
        "column":         column,
        "level":          "wallet",
        "companyId":      company_id,
        "companyName":    company_name,
        "date":           date,
        "totalWallets":   len(wallets),
        "doneWallets":    sum(1 for w in wallets if w["id"] in done),
        "groupings":      groupings_out,
        "orphanWallets":  orphan_wallets,
    }
    if unprocessed_set is not None:
        payload["totalUnprocessed"] = sum(
            1 for w in wallets if w["id"] in unprocessed_set
        )
    return jsonify(payload)


# ── Classifier endpoints ──────────────────────────────────────────────────────

@bp.route("/api/controlpanel/classify/override", methods=["POST"])
def classify_override():
    """Save a user correction: {unprocessedId, securityType}."""
    body = request.get_json(force=True, silent=True) or {}
    uid  = body.get("unprocessedId", "").strip()
    stype = body.get("securityType", "").strip()
    if not uid or not stype:
        return jsonify({"ok": False, "error": "missing fields"}), 400

    overrides_path = os.path.join(os.path.dirname(JSON_PATH), "classifier_overrides.json")
    with _overrides_lock:
        overrides = {}
        if os.path.exists(overrides_path):
            try:
                with open(overrides_path, "r", encoding="utf-8") as f:
                    overrides = json.load(f)
            except json.JSONDecodeError:
                overrides = {}
        overrides[uid] = stype
        atomic_write_json(overrides_path, overrides)
    return jsonify({"ok": True})


@bp.route("/api/controlpanel/security-types")
def security_types():
    """Return the list of known securityType values for the dropdown."""
    types = set()
    if os.path.exists(JSON_PATH):
        with open(JSON_PATH, encoding="utf-8") as f:
            for row in json.load(f):
                if row.get("securityType"):
                    types.add(row["securityType"])
    return jsonify({"types": sorted(types)})


@bp.route("/api/controlpanel/security-type-fields")
def security_type_fields():
    """Per-securityType field config for the "Cadastrar ativos" modal.

    Returns the contents of data/security_type_fields.json:
      { "types": { "<securityType>": { "label": str, "fields": [ {key,...} ] } } }

    The frontend uses `.types` to render the editable (bottom) line of each
    asset and to build the registration JSON. Degrades gracefully to an empty
    map when the file is missing/corrupt — the modal then shows only the
    structural fields (beehusName + securityType)."""
    try:
        if os.path.exists(_SECURITY_TYPE_FIELDS_FILE):
            with open(_SECURITY_TYPE_FIELDS_FILE, encoding="utf-8") as f:
                data = json.load(f)
            return jsonify({"types": data.get("types", {})})
    except Exception:
        import traceback, logging
        logging.error("security_type_fields error: %s", traceback.format_exc())
    return jsonify({"types": {}})


def _batch_pu(company_id, date, unprocessed_ids):
    """Return {unprocessedId: pu} from `unprocessedSecurityPositions` on the
    selected `date`, via the Beehus API (endpoint B) — not Mongo.

    Lê via `beehus_catalog.unprocessed_docs_map` (→ unprocessed-security-positions,
    range de 1 dia = a data exata), escopado às carteiras da empresa, e varre os
    `securities[]` retornados pelo PU de cada `unprocessedId`. Só a **data
    selecionada** é consultada (não a posição mais recente ≤ data). `{}` quando
    falta company/date/ids ou a API falha (sem Mongo).
    """
    if not company_id or not unprocessed_ids or not date:
        return {}
    wallet_ids = [w.get("_id") for w in beehus_catalog.wallets_in_company(company_id)
                  if w.get("_id")]
    if not wallet_ids:
        return {}
    docs_map = beehus_catalog.unprocessed_docs_map(company_id, wallet_ids, [date])

    pu_map = {}
    remaining = set(unprocessed_ids)
    for doc in docs_map.values():
        for sec in doc.get("securities") or []:
            uid = sec.get("unprocessedId")
            if uid and uid in remaining:
                pu_map[uid] = sec.get("pu")
                remaining.discard(uid)
        if not remaining:
            break
    return pu_map


def _batch_last_price(security_ids, target_date=None, *, company_id=None,
                      entity_id=None, wallet_id=None):
    """`{securityId: {value, date}}` do record de preço RESOLVIDO (escopo
    company/entity/wallet → C3→C2→C1→B2→B1) via seam (`filtered-security-price`;
    só `status=="approved"`/não-trashed). Sem `target_date`: última entrada. Com
    `target_date`: a entrada exata, senão a mais próxima por |Δdias|. Sem contexto
    de carteira, resolve o B1 global. Substitui os reads diretos de
    `db.securityPrices`."""
    price_map = {}
    if not security_ids:
        return price_map

    hp_map = beehus_catalog.security_prices_resolved(
        list(security_ids), company_id=company_id, entity_id=entity_id,
        wallet_id=wallet_id)

    if not target_date:
        for sid, hp in hp_map.items():
            if not hp:
                continue
            last = max(hp, key=lambda e: str(e.get("date", "")))
            price_map[sid] = {"value": last.get("value"),
                              "date": str(last.get("date", ""))[:10]}
        return price_map

    from datetime import datetime
    try:
        target_dt = datetime.strptime(target_date[:10], "%Y-%m-%d")
    except (ValueError, TypeError):
        # target malformado → cai no comportamento "última entrada".
        return _batch_last_price(security_ids, company_id=company_id,
                                 entity_id=entity_id, wallet_id=wallet_id)

    for sid, hp in hp_map.items():
        exact = None
        best = None
        best_diff = None
        for entry in (hp or []):
            d_str = str(entry.get("date", ""))[:10]
            if not d_str:
                continue
            if d_str == target_date[:10]:
                exact = entry
                break
            try:
                entry_dt = datetime.strptime(d_str, "%Y-%m-%d")
            except ValueError:
                continue
            diff = abs((entry_dt - target_dt).days)
            if best_diff is None or diff < best_diff:
                best = entry
                best_diff = diff

        chosen = exact or best
        if chosen:
            price_map[sid] = {
                "value": chosen.get("value"),
                "date":  str(chosen.get("date", ""))[:10],
            }
    return price_map


def _price_agreement(pu, last_price):
    """Score contribution from comparing the unprocessed position's ``pu``
    against the matched security's ``lastPrice`` (securityPrices.historyPrice).

    Extra evidence available ONLY in the mapping flow (the identification flow
    has no PU to compare). A tight agreement confirms the match; a large
    divergence is evidence against it. Returns ``(points, label)`` — points are
    0 when either side is missing/non-positive or the gap is in the neutral
    band, so a missing price never penalises a candidate."""
    try:
        pu_f = float(pu)
        lp_f = float(last_price)
    except (TypeError, ValueError):
        return 0, None
    if pu_f <= 0 or lp_f <= 0:
        return 0, None
    rel = abs(pu_f - lp_f) / max(abs(pu_f), abs(lp_f))
    if rel <= 0.001:
        return 30, "Preço idêntico ao PU (Δ ≤ 0,1%)"
    if rel <= 0.01:
        return 20, "Preço bate com o PU (Δ ≤ 1%)"
    if rel <= 0.05:
        return 8,  "Preço próximo do PU (Δ ≤ 5%)"
    if rel >= 0.50:
        return -25, "Preço diverge muito do PU (Δ > 50%)"
    if rel >= 0.20:
        return -10, "Preço diverge do PU (Δ > 20%)"
    return 0, None


@bp.route("/api/controlpanel/match", methods=["POST"])
def match_securities():
    """
    Match unprocessedIds against the securities collection.

    Body: { "items": [...], "companyId": str (optional), "date": str (optional) }
    Returns: { "results": [ { unprocessedId, predicted_type, type_confidence,
                               candidate, pu, lastPrice }, ... ] }
    """
    try:
        body  = request.get_json(force=True, silent=True) or {}
        items = body.get("items", [])
        company_id          = body.get("companyId", "")
        date                = body.get("date", "")
        complement_priority = body.get("complement_priority", "uid")
        # Build complement lookup: normalised-uid → complement string
        complement_map = {}
        for entry in (body.get("complements") or []):
            k = str(entry.get("uid", "")).strip().lower()
            v = str(entry.get("complement", "")).strip()
            if k and v:
                complement_map[k] = v
        if not items:
            return jsonify({"results": []})
        if company_id and not company_visible(company_id):
            return jsonify({"results": [], "error": "company not visible"}), 403

        matcher = _get_matcher()
        if matcher is None:
            return jsonify({"results": [], "error": "matcher not available"}), 500

        # Collect uids, resolving item format
        uids = []
        item_types = {}  # uid → user-supplied securityType (if any)
        for item in items:
            uid   = item.get("unprocessedId", "") if isinstance(item, dict) else str(item)
            stype = item.get("securityType") if isinstance(item, dict) else None
            if uid:
                uids.append(uid)
                if stype:
                    item_types[uid] = stype

        if not uids:
            return jsonify({"results": []})

        # Step 1: batch-classify types for all uids that need it
        uids_needing_clf = [u for u in uids if u not in item_types]
        clf = _get_classifier()
        type_map = {}  # uid → {type, confidence}
        if clf and uids_needing_clf:
            preds = clf.predict_batch(uids_needing_clf)
            for p in preds:
                type_map[p["unprocessedId"]] = {
                    "type": p["type"], "confidence": p["confidence"]
                }
        for uid, stype in item_types.items():
            type_map[uid] = {"type": stype, "confidence": None}

        # Step 2: match each uid (classifier step already done)
        results = []
        sec_id_set = set()
        for uid in uids:
            tm = type_map.get(uid, {})
            stype = tm.get("type")
            sconf = tm.get("confidence")

            # Enrich uid with complement when available, controlling order by priority.
            comp = complement_map.get(uid.strip().lower(), "")
            if comp:
                effective_uid = (comp + " " + uid) if complement_priority == "complement" else (uid + " " + comp)
            else:
                effective_uid = uid

            # Pre-compute raw complement tokens so they can be injected into
            # matcher.match() BEFORE search/scoring run internally. All tokens are
            # injected (no filtering yet); display filtering happens after the merge.
            raw_comp_tokens = [t for t in comp.split() if t] if comp else []
            inject = {f"complement_{i}": t for i, t in enumerate(raw_comp_tokens[:3], 1)}

            match = matcher.match(effective_uid, security_type=stype, type_confidence=sconf, limit=3,
                                  inject_features=inject or None)
            top   = match["candidates"][0] if match["candidates"] else None
            if top:
                sec_id_set.add(top["securityId"])

            # Feature merge (3 steps when complement present):
            #   1. Extract features from original uid only (clean baseline)
            #   2. Extract features from effective_uid (already in match["extracted"])
            #   3. Merge: priority source wins per-field; the other fills blanks
            if comp:
                detected_type = match.get("predicted_type") or stype or ""
                uid_feats      = extract_features(uid, detected_type)
                combined_feats = match.get("extracted") or {}
                if complement_priority == "uid":
                    feats = {k: uid_feats.get(k) or combined_feats.get(k)
                             for k in set(uid_feats) | set(combined_feats)}
                else:
                    feats = {k: combined_feats.get(k) or uid_feats.get(k)
                             for k in set(uid_feats) | set(combined_feats)}
            else:
                feats = match.get("extracted") or {}
            # Complement display tokens: remove injected complement_N (they came back
            # via combined_feats) then recalculate keeping only tokens not already
            # captured as a specific feature (exact string match).
            for _k in ("complement_1", "complement_2", "complement_3"):
                feats.pop(_k, None)
            if raw_comp_tokens:
                non_comp_values = {str(v) for v in feats.values() if v}
                display_tokens  = [t for t in raw_comp_tokens if t not in non_comp_values]
                for _ci, _ct in enumerate(display_tokens[:3], 1):
                    feats[f"complement_{_ci}"] = _ct
            results.append({
                "unprocessedId":  uid,
                "predicted_type": match["predicted_type"],
                "type_confidence": match["type_confidence"],
                "extracted": {
                    # Keys used by the registration modal — do not rename
                    "isin":          feats.get("isin", ""),
                    "ticker":        feats.get("ticker", ""),
                    "taxId":         feats.get("cnpj", ""),
                    "type":          feats.get("bond_type") or feats.get("instrument", ""),
                    # Diagnostic fields — shown in the uid tooltip
                    "instrument":    feats.get("instrument", ""),
                    "bond_type":     feats.get("bond_type", ""),
                    "coupon":        feats.get("coupon", ""),
                    "fund_code":     feats.get("fund_code", ""),
                    "issuer":        feats.get("issuer", ""),
                    "indexer":       feats.get("indexer", ""),
                    "rate":          feats.get("rate", ""),
                    "maturity_date": feats.get("maturity_date", ""),
                    "selic_code":    feats.get("selic_code", ""),
                    "cetip_code":    feats.get("cetip_code", ""),
                    "internal_code": feats.get("internal_code", ""),
                    "external_code": feats.get("external_code", ""),
                    "underlying":    feats.get("underlying", ""),
                    "option_type":   feats.get("option_type", ""),
                    "strike":        feats.get("strike", ""),
                    "expiry":        feats.get("expiry", ""),
                    "expiry_month":  feats.get("expiry_month", ""),
                    "expiry_year":   feats.get("expiry_year", ""),
                    "name":          feats.get("name", ""),
                    "contract":      feats.get("contract", ""),
                    "generic_code_1": feats.get("generic_code_1", ""),
                    "generic_code_2": feats.get("generic_code_2", ""),
                    "generic_code_3": feats.get("generic_code_3", ""),
                    "complement_1":  feats.get("complement_1", ""),
                    "complement_2":  feats.get("complement_2", ""),
                    "complement_3":  feats.get("complement_3", ""),
                },
                "candidate": {
                    "securityId": top["securityId"],
                    "mainId":     top["mainId"],
                    "beehusName": top["beehusName"],
                    "indexer":    top.get("indexer", ""),
                    "score":      top["score"],
                    "confidence": top["confidence"],
                    "matched_on": top["matched_on"],
                } if top else None,
            })

        # Step 3: batch enrich with PU and lastPrice (2 queries total)
        uid_set   = set(uids)
        pu_map    = _batch_pu(company_id, date, uid_set)
        price_map = _batch_last_price(sec_id_set, target_date=date)

        for r in results:
            r["pu"] = pu_map.get(r["unprocessedId"])
            cand = r.get("candidate")
            r["lastPrice"] = price_map.get(cand["securityId"]) if cand else None
            if not cand:
                continue
            # Score decomposition for the "como o score foi calculado" tooltip
            # (same logic as Identificar Transações; here it's L3-only — full
            # cadastro sweep — since there's no wallet position to scope to).
            base_score = cand.get("score") or 0
            breakdown = _score_breakdown(cand.get("matched_on") or [], base_score)
            # Price-vs-PU verification: only confirms/penalises a candidate that
            # already has a real match signal (base_score > 0), so a price
            # coincidence can never lift an otherwise-unmatched security.
            if base_score > 0:
                lp = r["lastPrice"]
                pts, label = _price_agreement(r["pu"], lp.get("value") if lp else None)
                if pts:
                    # Clamp so a penalty can't push the score below 0; keep the
                    # breakdown honest by recording the *applied* delta.
                    applied = pts if base_score + pts >= 0 else -base_score
                    if applied:
                        breakdown.append({"code": "price", "points": applied, "label": label})
                        new_score = base_score + applied
                        cand["score"] = new_score
                        cand["confidence"] = _confidence_label(new_score)
            cand["breakdown"] = breakdown

        return jsonify({"results": results})
    except Exception:
        import traceback, logging
        logging.error("match_securities error: %s", traceback.format_exc())
        return jsonify({"results": [], "error": "internal error"}), 500


@bp.route("/api/controlpanel/security-mapping-id")
def security_mapping_id():
    """Return the securityMappings _id for a given companyId."""
    company_id = request.args.get("companyId", "").strip()
    if not company_id:
        return jsonify({"securityMappingId": None, "error": "missing companyId"}), 400
    if not company_visible(company_id):
        return jsonify({"securityMappingId": None, "error": "company not visible"}), 403
    smid = beehus_catalog.security_mapping_id(company_id)
    return jsonify({"securityMappingId": smid})


@bp.route("/api/controlpanel/apply-mapping", methods=["POST"])
def apply_mapping():
    """Apply selected security mappings directly into Beehus via upstream PATCH.

    Body: { "companyId": str, "mappingsToInclude": [{from, to}, ...] }

    Looks up the `securityMappings._id` for the company server-side (so the
    client cannot tamper with it) and forwards the include list upstream.
    Only `mappingsToInclude` is accepted here by design — exclusions are out
    of scope for the Painel de Controle page.
    """
    body = request.get_json(silent=True) or {}
    company_id = (body.get("companyId") or "").strip()
    includes   = body.get("mappingsToInclude") or []

    if not company_id:
        return jsonify({"error": "companyId is required"}), 400
    if not company_visible(company_id):
        return jsonify({"error": "company is not visible to this user"}), 403
    if not isinstance(includes, list) or not includes:
        return jsonify({"error": "mappingsToInclude must be a non-empty list"}), 400

    cleaned = []
    for m in includes:
        if not isinstance(m, dict):
            return jsonify({"error": "each mapping must be an object {from, to}"}), 400
        frm = (m.get("from") or "").strip()
        to  = (m.get("to") or "").strip()
        if not frm or not to:
            return jsonify({"error": "each mapping requires non-empty 'from' and 'to'"}), 400
        cleaned.append({"from": frm, "to": to})

    mapping_id = beehus_catalog.security_mapping_id(company_id)
    if not mapping_id:
        return jsonify({"error": "securityMappingId not found for this company"}), 404

    try:
        result = update_security_mappings(
            mapping_id,
            mappings_to_include=cleaned,
            mappings_to_exclude=[],
        )
    except BeehusAuthError as e:
        return jsonify({
            "error": str(e),
            "upstream_status": e.status,
            "upstream_body": e.body,
        }), 401
    except BeehusAPIError as e:
        return jsonify({
            "error": str(e),
            "upstream_status": e.status,
            "upstream_body": e.body,
        }), 502

    return jsonify({
        "ok": True,
        "securityMappingId": mapping_id,
        "applied": len(cleaned),
        "response": result if result is not None else {},
    })


@bp.route("/api/controlpanel/search-securities")
def search_securities():
    """
    Free-text search over the in-memory securities cache.

    Query params:
      - q:    search term (matched against beehusName, mainId, ticker, taxId,
              isIn, selicCode — case/accent-insensitive, substring).
      - type: optional securityType filter. If provided but yields no hits,
              falls back to the full cache so the user is never stuck.
      - limit: max results (default 50, max 200).

    Returns: { "results": [...], "cacheCount": N, "poolCount": M, "filteredByType": bool }
    """
    import logging
    log = logging.getLogger(__name__)
    try:
        q        = (request.args.get("q") or "").strip()
        stype    = (request.args.get("type") or "").strip()
        try:
            limit = min(int(request.args.get("limit", 50)), 200)
        except ValueError:
            limit = 50

        cache = get_cache()
        if not cache.is_loaded:
            # Prefer the on-disk snapshot (sub-second); only fall back to a
            # full MongoDB scan when no file exists. This keeps the search
            # endpoint responsive on cold start.
            if not cache.load_from_file():
                cache.load_from_db(db)

        cache_count = cache.count
        # Access the raw list through the private attribute (cache has no public iterator).
        full = cache._securities
        pool = cache.get_by_type(stype) if stype else full
        filtered_by_type = bool(stype)
        # Fallback: type filter matched nothing → search whole cache.
        if stype and not pool:
            pool = full
            filtered_by_type = False

        if not pool:
            return jsonify({
                "results": [], "cacheCount": cache_count, "poolCount": 0,
                "filteredByType": False,
            })

        results = []
        if not q:
            results = pool[:limit]
        else:
            from security_matcher import _strip_accents
            needle = _strip_accents(q.lower())
            search_fields = ("beehusName", "mainId", "ticker", "taxId", "isIn", "selicCode")
            for sec in pool:
                for field in search_fields:
                    val = sec.get(field) or ""
                    if val and needle in _strip_accents(str(val).lower()):
                        results.append(sec)
                        break
                if len(results) >= limit:
                    break

        out = []
        for s in results:
            out.append({
                "securityId":   s.get("_id", ""),
                "beehusName":   s.get("beehusName", ""),
                "mainId":       s.get("mainId", ""),
                "ticker":       s.get("ticker", ""),
                "taxId":        s.get("taxId", ""),
                "isIn":         s.get("isIn", ""),
                "selicCode":    s.get("selicCode", ""),
                "indexer":      s.get("indexer", ""),
                "securityType": s.get("securityType", ""),
                "maturityDate": s.get("maturityDate", ""),
            })
        return jsonify({
            "results": out,
            "cacheCount": cache_count,
            "poolCount": len(pool),
            "filteredByType": filtered_by_type,
        })
    except Exception:
        import traceback
        log.error("search_securities error: %s", traceback.format_exc())
        return jsonify({"results": [], "error": "internal error"}), 500


# Max business days to walk back when a wallet has no processedPosition on the
# requested date. The processed-position API is single-date (no server-side
# "most recent <= date"), so we reproduce the old Mongo `.sort(positionDate,-1)`
# fallback with a bounded backward scan — gaps are almost always weekends or a
# single holiday, so a week of business days covers the realistic cases without
# fanning out (one API call per company per day, only for wallets still missing).
_WALLET_POS_MAX_BACK_DAYS = 7


def _latest_positions_by_wallet(company_to_wids, end_date):
    """`{walletId_str: securities_list}` for the most recent processedPosition
    with positionDate <= end_date, per wallet.

    API-only via `beehus_catalog.processed_positions_map` (the processed-position
    route), grouped by company. Replaces the direct
    `db.processedPosition.find({walletId:$in[, positionDate<=date]}).sort(positionDate,-1)`
    that picked the latest snapshot per wallet. Walks back at most
    `_WALLET_POS_MAX_BACK_DAYS` business days for wallets without a snapshot on
    `end_date`; a wallet still missing after that is simply absent from the map."""
    out = {}
    for cid, wids in company_to_wids.items():
        if not cid:
            continue
        remaining = set(wids)
        d = end_date
        tries = 0
        while remaining and tries <= _WALLET_POS_MAX_BACK_DAYS:
            # Single date per call keeps the map robust: processed_positions_map
            # returns {} only if THIS date's fetch errors, so a bad day just
            # falls through to the prior one instead of nuking earlier hits.
            pos_map = beehus_catalog.processed_positions_map(cid, list(remaining), [d])
            for (wid, _pd), secs in pos_map.items():
                if wid in remaining:
                    out[wid] = secs or []
                    remaining.discard(wid)
            d = business_days_before(d, 1)
            tries += 1
    return out


@bp.route("/api/controlpanel/wallet-positions")
def wallet_positions():
    """Suggest securities by reading the most recent processedPosition for the
    given wallets.

    The "Identificar" modal lets the operator pick a security for an
    unprocessedId. Besides the global securities cache search, this endpoint
    powers a second source: the securities that already live in the wallet's
    most recent processed snapshot. That snapshot is the strongest signal that
    a given security is the right counterpart for an unmapped position —
    operators almost always pick from this short list when it exists.

    Query params:
      walletIds  required, comma-separated wallet ids. Multiple wallets are
                 supported because issues are grouped by unprocessedId across
                 every wallet that exhibits them.
      date       optional YYYY-MM-DD. We look up the most recent
                 processedPosition with positionDate <= date (defaults to the
                 latest available when omitted).
      q          optional free-text filter, applied after enrichment with the
                 same fields as /search-securities (beehusName, mainId, …).
      limit      max enriched rows (default 100, max 500).

    Returns: { results: [...], walletsScanned: N, securityCount: M }
    where each result mirrors the /search-securities shape and adds
    `walletIds` / `walletNames` describing which wallets carry the security.
    """
    import logging
    log = logging.getLogger(__name__)
    # Identification is an instance-role feature, hidden in the UI when off
    # (templates/controlpanel.html). The data itself now comes from the
    # processed-position API (no Mongo), so this gate only mirrors the
    # instance role; the modal's "Cadastro" source (global securities cache,
    # API-backed) remains available regardless.
    if not IDENTIFICAR_ENABLED:
        return jsonify({"results": [], "walletsScanned": 0, "securityCount": 0})
    try:
        raw_ids = (request.args.get("walletIds") or "").strip()
        date    = (request.args.get("date") or "").strip() or None
        q       = (request.args.get("q") or "").strip()
        try:
            limit = min(int(request.args.get("limit", 100)), 500)
        except ValueError:
            limit = 100

        wallet_ids = [w.strip() for w in raw_ids.split(",") if w.strip()]
        if not wallet_ids:
            return jsonify({"results": [], "walletsScanned": 0, "securityCount": 0})

        # Resolve wallet names + enforce company visibility. Without this an
        # operator could exfiltrate positions from a wallet the company filter
        # would otherwise hide.
        cf = get_company_filter()
        visible_wallets = {}
        company_to_wids = {}   # companyId -> [walletId, ...] (for the API fetch)
        for w in wallet_ids:
            if not ObjectId.is_valid(w):
                continue
            doc = beehus_catalog.wallet_doc(w)
            if not doc:
                continue
            wid = beehus_catalog.id_str(doc.get("_id")) or w
            cid = str(doc.get("companyId") or "")
            if cf and cid not in cf:
                continue
            visible_wallets[wid] = doc.get("name") or wid
            company_to_wids.setdefault(cid, []).append(wid)
        if not visible_wallets:
            return jsonify({"results": [], "walletsScanned": 0, "securityCount": 0})

        # Latest processedPosition per wallet, via the processed-position API
        # (no Mongo). The API is single-date, so `_latest_positions_by_wallet`
        # reproduces the old `.sort(positionDate,-1)` "most recent <= date"
        # fallback with a bounded backward scan. Default to today (BRT) when no
        # date is supplied — the old "latest available" needs a concrete date
        # without a server-side sort.
        end_date = date or today_in_brt().isoformat()
        positions = _latest_positions_by_wallet(company_to_wids, end_date)

        # {securityId: {pu, quantity, pricingType, walletIds[], walletNames[]}}
        by_security = {}
        wallets_with_snapshot = set()
        for wid, securities in positions.items():
            wallets_with_snapshot.add(wid)
            wname = visible_wallets.get(wid, wid)
            for s in securities or []:
                sid = beehus_catalog.id_str(s.get("securityId")) or ""
                if not sid:
                    continue
                entry = by_security.setdefault(sid, {
                    "pu": s.get("pu"),
                    "quantity": s.get("quantity"),
                    "pricingType": s.get("pricingType"),
                    "walletIds": [],
                    "walletNames": [],
                })
                if wid not in entry["walletIds"]:
                    entry["walletIds"].append(wid)
                    entry["walletNames"].append(wname)

        if not by_security:
            return jsonify({
                "results": [], "walletsScanned": len(wallets_with_snapshot),
                "securityCount": 0,
            })

        # Enrich with security metadata. Prefer the in-memory cache (already
        # used by /search-securities) to avoid hitting MongoDB on every
        # modal-open — the cache is warmed by /warmup on page load.
        cache = get_cache()
        cache_by_id = {s.get("_id"): s for s in cache._securities} if cache.is_loaded else {}

        # Fetch metadata for any sids missing from the cache (the cache is
        # rebuilt daily, but a brand-new security registered today might not
        # be there yet).
        missing_ids = [sid for sid in by_security if sid not in cache_by_id]
        if missing_ids:
            oids = []
            for sid in missing_ids:
                try:
                    oids.append(ObjectId(sid))
                except (TypeError, ValueError):
                    continue
            if oids:
                for s in beehus_catalog.securities_by_ids(oids).values():
                    cache_by_id[str(s["_id"])] = {
                        "_id":          str(s["_id"]),
                        "beehusName":   s.get("beehusName", ""),
                        "mainId":       s.get("mainId", ""),
                        "ticker":       s.get("ticker", ""),
                        "taxId":        s.get("taxId", ""),
                        "isIn":         s.get("isIn", ""),
                        "selicCode":    s.get("selicCode", ""),
                        "indexer":      s.get("indexer", ""),
                        "securityType": s.get("securityType", ""),
                        "maturityDate": str(s.get("maturityDate") or ""),
                    }

        out = []
        for sid, info in by_security.items():
            meta = cache_by_id.get(sid, {})
            out.append({
                "securityId":   sid,
                "beehusName":   meta.get("beehusName", ""),
                "mainId":       meta.get("mainId", ""),
                "ticker":       meta.get("ticker", ""),
                "taxId":        meta.get("taxId", ""),
                "isIn":         meta.get("isIn", ""),
                "selicCode":    meta.get("selicCode", ""),
                "indexer":      meta.get("indexer", ""),
                "securityType": meta.get("securityType", ""),
                "maturityDate": str(meta.get("maturityDate", ""))[:10],
                "pu":           info["pu"],
                "quantity":     info["quantity"],
                "pricingType":  info["pricingType"],
                "walletIds":    info["walletIds"],
                "walletNames":  info["walletNames"],
            })

        # Optional free-text filter (same fields as /search-securities). We
        # apply it after enrichment so the operator can search by ticker /
        # mainId / name on the wallet's positions just like on the global
        # cadastro.
        if q:
            from security_matcher import _strip_accents
            needle = _strip_accents(q.lower())
            search_fields = ("beehusName", "mainId", "ticker", "taxId", "isIn", "selicCode")
            filtered = []
            for r in out:
                for f in search_fields:
                    val = r.get(f) or ""
                    if val and needle in _strip_accents(str(val).lower()):
                        filtered.append(r)
                        break
            out = filtered

        # Stable order: prefer those with a beehusName (i.e. enriched), then
        # sort alphabetically. Securities still pending registration would
        # otherwise pollute the top of the list with empty names.
        out.sort(key=lambda r: (not r["beehusName"], r["beehusName"].lower(), r["mainId"]))

        return jsonify({
            "results":        out[:limit],
            "walletsScanned": len(wallets_with_snapshot),
            "securityCount":  len(by_security),
        })
    except Exception:
        import traceback
        log.error("wallet_positions error: %s", traceback.format_exc())
        return jsonify({"results": [], "error": "internal error"}), 500


@bp.route("/api/controlpanel/last-price")
def last_price():
    """Return historyPrice for a single securityId.

    If `date` (YYYY-MM-DD) is provided, returns the price on that date or the
    nearest available; otherwise returns the last entry.
    """
    sid  = (request.args.get("securityId") or "").strip()
    date = (request.args.get("date") or "").strip() or None
    if not sid:
        return jsonify({"lastPrice": None})
    price_map = _batch_last_price([sid], target_date=date)
    return jsonify({"lastPrice": price_map.get(sid)})


# ── Cache endpoints ──────────────────────────────────────────────────────────

@bp.route("/api/controlpanel/cache-status")
def cache_status():
    """Return current cache state so the frontend can show refresh prompts."""
    cache = get_cache()
    mcache = get_mapping_cache()
    return jsonify({
        "loaded":     cache.is_loaded,
        "stale":      cache.is_stale,
        "loadedDate": cache.loaded_date,
        "count":      cache.count,
        "classifierReady": _clf is not None,
        "mappingLoaded": mcache.is_loaded,
        "mappingCount":  mcache.count,
    })


@bp.route("/api/controlpanel/warmup", methods=["POST"])
def warmup():
    """Pre-load classifier and cache so the first match request is fast."""
    try:
        actions = []
        # Load cache from file if today's, else from DB
        cache = get_cache()
        if not cache.is_loaded or cache.is_stale:
            if cache.load_from_file():
                actions.append("cache_from_file")
            else:
                cache.load_from_db(db)
                actions.append("cache_from_db")

        # Warm the security-mappings cache (drives Recalcular) — today's file or
        # API. Best-effort: a failure here must never break the warmup.
        try:
            mcache = get_mapping_cache()
            if mcache.ensure_loaded():
                actions.append("mapping_cache_from_api")
            elif mcache.is_loaded:
                actions.append("mapping_cache_ready")
        except Exception:
            import logging
            logging.warning("warmup: mapping cache warm failed", exc_info=True)

        # Train classifier if not ready
        if _get_classifier() is not None:
            actions.append("classifier_ready")

        return jsonify({"ok": True, "actions": actions, "count": cache.count})
    except Exception:
        import traceback, logging
        logging.error("warmup error: %s", traceback.format_exc())
        return jsonify({"ok": False, "error": "internal error"}), 500


@bp.route("/api/controlpanel/refresh-cache", methods=["POST"])
def refresh_cache():
    """Force-reload the securities cache from MongoDB."""
    cache = get_cache()
    cache.load_from_db(db)
    _reset_matcher()  # force re-init with fresh cache
    return jsonify({
        "ok":    True,
        "count": cache.count,
        "date":  cache.loaded_date,
    })


@bp.route("/api/controlpanel/parse-complement", methods=["POST"])
def parse_complement():
    """Parse a complement file and return a preview of {uid, complement} rows."""
    file = request.files.get("file")
    if not file:
        return jsonify({"ok": False, "error": "Nenhum arquivo enviado."}), 400

    filename = (file.filename or "").lower()
    ext = filename.rsplit(".", 1)[-1] if "." in filename else ""
    if ext not in ("xlsx", "csv", "json", "txt"):
        return jsonify({"ok": False, "error": f"Formato não suportado: .{ext}"}), 400

    try:
        raw = file.read()
        rows = []  # list of {"uid": str, "complement": str}

        if ext == "xlsx":
            wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:
                    first = False
                    continue
                uid = str(row[0]).strip() if row[0] is not None else ""
                comp = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                if uid:
                    rows.append({"uid": uid, "complement": comp})
            wb.close()

        elif ext in ("csv", "txt"):
            text = raw.decode("utf-8-sig", errors="replace")
            try:
                dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(io.StringIO(text), dialect)
            first = True
            for row in reader:
                if first:
                    first = False
                    continue
                if not row:
                    continue
                uid = row[0].strip() if row[0] else ""
                comp = row[1].strip() if len(row) > 1 else ""
                if uid:
                    rows.append({"uid": uid, "complement": comp})

        elif ext == "json":
            data = json.loads(raw.decode("utf-8", errors="replace"))
            if isinstance(data, dict):
                rows = [
                    {"uid": str(k).strip(), "complement": str(v).strip()}
                    for k, v in data.items() if str(k).strip()
                ]
            elif isinstance(data, list):
                for item in data:
                    if isinstance(item, (list, tuple)) and len(item) >= 2:
                        uid = str(item[0]).strip()
                        comp = str(item[1]).strip()
                    elif isinstance(item, dict):
                        uid = str(item.get("uid") or item.get("unprocessedId") or "").strip()
                        comp = str(item.get("complement") or item.get("complemento") or "").strip()
                    else:
                        continue
                    if uid:
                        rows.append({"uid": uid, "complement": comp})

        return jsonify({
            "ok":      True,
            "total":   len(rows),
            "rows":    rows,
            "preview": rows[:10],
        })

    except Exception:
        import traceback
        import logging
        logging.error("parse_complement error: %s", traceback.format_exc())
        return jsonify({"ok": False, "error": "Erro interno ao processar o arquivo."}), 500


@bp.route("/api/controlpanel/export-c3", methods=["POST"])
def export_c3():
    """Generate Excel for C3 assets selected from security_missing_price issues."""
    body  = request.get_json(force=True, silent=True) or {}
    items = body.get("items", [])

    # The "Data" column is what gets pasted into the C3 system. The frontend
    # sends the currently selected date pill (YYYY-MM-DD). We write it as a real
    # Excel date (a datetime.date cell displayed as DD/MM/YYYY) instead of text,
    # so downstream tools see an actual date value. Fall back to today's BRT
    # date so a missing field never silently writes a stale year.
    raw_date = (body.get("date") or "").strip()
    try:
        date_brt = date.fromisoformat(raw_date) if raw_date else today_in_brt()
    except ValueError:
        date_brt = today_in_brt()

    cf = get_company_filter()
    if cf:
        items = [it for it in items if str(it.get("companyId", "")) in cf]

    wb = Workbook()
    ws = wb.active
    ws.title = "C3"

    headers = [
        "Data", "SecurityId", "EntityId",
        "CompanyId", "WalletId", "PU", "C3Automatico", "BeehusName",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in items:
        ws.append([
            date_brt,
            item.get("securityId", ""),
            item.get("entityId", ""),
            item.get("companyId", ""),
            item.get("walletId", ""),
            item.get("pu", 0),
            "V" if item.get("consumoAutomatico") else "",
            item.get("beehusName", ""),
        ])

    # Force the "Data" column (A) to display as DD/MM/YYYY. openpyxl already
    # stores the datetime.date as a real Excel date serial; this just fixes the
    # display so it never falls back to Excel's default date mask.
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).number_format = "DD/MM/YYYY"

    # Highlight the BeehusName column (last column) — header keeps the bold
    # font from the loop above; we just add a red fill to the whole column
    # (header + every data row) so it stands out as the "informational only"
    # field that doesn't get pasted into the C3 system.
    beehus_col = ws.cell(row=1, column=len(headers)).column_letter
    red_fill = PatternFill("solid", fgColor="FFC7CE")  # soft red, Excel-native
    red_font = Font(color="9C0006", bold=False)
    for row_idx in range(1, ws.max_row + 1):
        cell = ws[f"{beehus_col}{row_idx}"]
        cell.fill = red_fill
        # Preserve bold on header; only restyle data rows.
        if row_idx == 1:
            cell.font = Font(bold=True, color="9C0006")
        else:
            cell.font = red_font

    buf = io.BytesIO()
    wb.save(buf)

    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp = make_response(buf.getvalue())
    resp.headers["Content-Type"] = mime
    resp.headers["Content-Disposition"] = "attachment; filename=c3_registro_preco.xlsx"
    return resp


@bp.route("/api/controlpanel/rebuild-mapping", methods=["POST"])
def rebuild():
    from security_type_classifier import rebuild_mapping
    # The operator clicks Recalcular right after labelling new mappings, so pull
    # a fresh copy from upstream before regenerating the corpus: drop the
    # all_mappings TTL and force the MappingCache off its daily file.
    beehus_catalog.invalidate("all_mappings")
    get_mapping_cache().refresh()
    total, mapped = rebuild_mapping()
    reset_classifier()
    _reset_matcher()
    # Also refresh cache since securities may have changed
    get_cache().load_from_db(db)
    return jsonify({"total": total, "mapped": mapped})

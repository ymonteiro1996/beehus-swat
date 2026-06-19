"""Correções — single source of truth for generated correction files.

Stores accepted transaction/provision rows as JSON on disk, one file per
companyId+date+walletId. The CRUD page lets users view, edit, add, and
delete individual rows. The export endpoint is the ONLY place in the
project that emits downloadable JSON / XLSX files.

Storage layout:
    data/correcoes/<companyId>/<date>/<walletId>.json
    {
      "companyId":       "...",
      "walletId":        "...",
      "date":            "YYYY-MM-DD",
      "updatedAt":       "ISO-8601",
      "transactions":    [ {"id": "uuid", ...row fields...}, ... ],
      "provisions":      [ {"id": "uuid", ...row fields...}, ... ],
      "deletions":       [ {"id": "uuid", "originalId": "...", ...}, ... ],
      "executionPrices": [ {"id": "uuid", "inputed": false, ...}, ... ]
    }

The `deletions` bucket is internal-only bookkeeping for MISCLASSIFIED
acceptance: each row points at a DB transaction that should be ignored by
the reconciliation pipeline (because a replacement was appended to
`transactions`). See docs/CONCILIACAO_BAYESIAN.md for the full flow.
Deletions are NOT emitted in the export endpoint.

The `executionPrices` bucket holds accepted MISSING_EXECUTION_PRICE
corrections. Unlike transactions/provisions, these are NOT exported as
files — they are pushed to the upstream Beehus API via
`POST /beehus/financial/execution-prices`. The UI in /correcoes provides
an "Enviar via API" button per row; on success the row is marked
`inputed=true` (with `inputedAt` and `beehusId`). Inputed rows are
considered "applied upstream" and are skipped by any local gap recalc
that would otherwise double-count them.

`transactions` and `provisions` rows can additionally be pushed individually
via `POST /api/correcoes/transactions/submit` and
`POST /api/correcoes/provisions/submit` (mirroring the *Funções > Criar
Transação* and *Criar Provisão* forms on /beehus). The same `inputed`/
`inputedAt`/`beehusId` triplet is set on success and the same exclusion
applies — `load_corrections_for_wallet` filters `inputed=True` rows from
the diagnose injection so they are not double-counted locally.
"""
import io
import json
import os
import re
import uuid
from datetime import datetime

from flask import Blueprint, jsonify, render_template, request, send_file
from openpyxl import Workbook

from db import (db, get_biz_dates, get_company_filter, company_visible, get_company_names,
                resolve_wallet, today_in_brt, atomic_write_json)
from beehus_api import (create_execution_price, create_provision, create_transaction,
                        BeehusAPIError, BeehusAuthError)


def _wallet_in_company(wallet_id, expected_company_id):
    """Verify the wallet's stored companyId matches `expected_company_id`.

    Closes the cross-company write-leak: a user whose company_filter allows
    company A could otherwise POST `{companyId: A, walletId: <wallet from B>}`
    and pollute A's correction store with B's wallet IDs.
    """
    if not wallet_id or not expected_company_id:
        return False
    w = resolve_wallet(wallet_id, {"companyId": 1})
    return bool(w) and str(w.get("companyId", "")) == str(expected_company_id)


# `currencyId` on persisted transactions/provisions is authoritative-from-wallet:
# we look it up in `db.wallets` instead of trusting whatever the upstream caller
# (Painel, Conciliação, manual Add modal) passed. Otherwise a stale "BRL" default
# in the front-end could overwrite a wallet whose true currency is USD/EUR.
#
# Note on field names: the wallets collection stores the currency code under
# `currency` (e.g. "USD", "BRL"). We also read `currencyId` as a defensive
# fallback in case any wallet doc was migrated to that name. The earlier
# version of this helper only looked at `currencyId` and silently defaulted to
# "BRL" for every wallet — meaning USD/EUR wallets were being submitted to the
# upstream API as BRL.
def _wallet_currency(wallet_id, cache=None):
    """Return the wallet's currency code, falling back to "BRL" when missing.

    Pass a dict in `cache` to memoize across rows in the same request — bulk
    appends can touch the same wallet many times and we don't want one DB
    round-trip per row."""
    if cache is not None and wallet_id in cache:
        return cache[wallet_id]
    w = resolve_wallet(wallet_id, {"currency": 1, "currencyId": 1}) or {}
    cur = str(w.get("currency") or w.get("currencyId") or "BRL")
    if cache is not None:
        cache[wallet_id] = cur
    return cur


def _visible_company_ids():
    """Return the list of companyId strings currently visible to the user.

    Takes the union of ids in `db.companies` and directories actually present
    under `data/correcoes/`. This way a correction persisted against a
    companyId that has no matching Mongo document (test data, orphans after
    renames, etc.) is still listed. Empty `company_filter` means "all".
    """
    cf = get_company_filter()  # set() means show all
    ids = set(get_company_names().keys())
    if os.path.isdir(_ROOT):
        try:
            for name in os.listdir(_ROOT):
                if _SAFE_ID_RE.match(name) and os.path.isdir(os.path.join(_ROOT, name)):
                    ids.add(name)
        except OSError:
            pass
    if not cf:
        return sorted(ids)
    return sorted(i for i in ids if i in cf)

bp = Blueprint("correcoes", __name__)

_NUM_DATES = 10
_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data", "correcoes"))

_KINDS = ("transactions", "provisions", "deletions", "executionPrices")

# Field whitelists — anything outside these is dropped on write.
_TXN_FIELDS = [
    "companyId", "entityId", "walletId", "currencyId", "securityId",
    "operationDate", "liquidationDate", "balance", "description",
    "inputType", "beehusTransactionType", "hide", "comment",
]
_PROV_FIELDS = [
    "walletId", "initialDate", "liquidationDate", "provisionType",
    "securityId", "balance", "description", "provisionSource", "currencyId",
]
# `deletions` — pointers to DB transactions the user asked to disregard.
# `originalId` is the Mongo _id (str) of the txn to exclude from the gap
# calculation. `reason` is a flag tag (usually "MISCLASSIFIED"). The other
# fields are snapshots of the original txn so the UI can render the row
# without another DB lookup.
_DEL_FIELDS = [
    "companyId", "walletId", "originalId", "securityId", "balance",
    "operationDate", "liquidationDate", "beehusTransactionType",
    "description", "reason",
]
# `executionPrices` — accepted MISSING_EXECUTION_PRICE corrections.
# `executionPrice` is the user-confirmed price; `expectedExecPrice` /
# `pu` / `priorExecutionPrice` are diagnostic snapshots so the row can be
# rendered without re-running the diagnose. `inputed`/`inputedAt`/
# `beehusId` are populated AFTER the user pushes the row to the upstream
# API via /api/correcoes/execution-prices/submit.
_EXECPRICE_FIELDS = [
    "companyId", "walletId", "securityId", "securityName", "positionDate",
    "executionPrice", "expectedExecPrice", "pu", "priorExecutionPrice",
    "amountDiff", "actualBalance", "expectedValue", "description",
    "inputed", "inputedAt", "beehusId",
]

_FIELDS_BY_KIND = {
    "transactions":    _TXN_FIELDS,
    "provisions":      _PROV_FIELDS,
    "deletions":       _DEL_FIELDS,
    "executionPrices": _EXECPRICE_FIELDS,
}

# Internal metadata preserved on disk but NEVER included in export output.
# `sourceAnomalyKey` links a correction row back to the anomaly that spawned it
# (e.g. a Painel "Anomalias em ativos" row). Upstream pages read these keys to
# mark already-moved anomalies and prevent re-adding.
# `inputed`/`inputedAt`/`beehusId` mark provisions and executionPrices that
# were already pushed to the upstream Beehus API by the per-row "Enviar via
# API" buttons in /correcoes. Pipelines must skip inputed rows when applying
# corrections locally (already baked-in upstream — would double-count).
_META_FIELDS = ("sourceAnomalyKey", "inputed", "inputedAt", "beehusId")

_SAFE_ID_RE = re.compile(r"^[A-Za-z0-9_-]+$")
_SAFE_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


# ── Path / IO helpers ─────────────────────────────────────────────────────────

def _safe(segment):
    """Reject any path segment that isn't alphanumeric/dash/underscore."""
    if not segment or not _SAFE_ID_RE.match(str(segment)):
        raise ValueError(f"invalid segment: {segment!r}")
    return str(segment)


def _safe_date(date):
    if not date or not _SAFE_DATE_RE.match(str(date)):
        raise ValueError(f"invalid date: {date!r}")
    return str(date)


def _wallet_file(company_id, date, wallet_id):
    return os.path.join(_ROOT, _safe(company_id), _safe_date(date), f"{_safe(wallet_id)}.json")


def _company_dir(company_id):
    return os.path.join(_ROOT, _safe(company_id))


def _date_dir(company_id, date):
    return os.path.join(_ROOT, _safe(company_id), _safe_date(date))


def _load(company_id, date, wallet_id):
    path = _wallet_file(company_id, date, wallet_id)
    if not os.path.isfile(path):
        return {
            "companyId":       company_id,
            "walletId":        wallet_id,
            "date":            date,
            "updatedAt":       None,
            "transactions":    [],
            "provisions":      [],
            "deletions":       [],
            "executionPrices": [],
        }
    with open(path, "r", encoding="utf-8") as f:
        blob = json.load(f)
    blob.setdefault("transactions", [])
    blob.setdefault("provisions", [])
    blob.setdefault("deletions", [])
    blob.setdefault("executionPrices", [])
    return blob


def _save(blob):
    company_id = _safe(blob["companyId"])
    date       = _safe_date(blob["date"])
    wallet_id  = _safe(blob["walletId"])
    os.makedirs(_date_dir(company_id, date), exist_ok=True)
    blob["updatedAt"] = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    path = _wallet_file(company_id, date, wallet_id)
    # atomic_write_json uses tempfile.mkstemp — without it, two concurrent
    # WSGI threads writing the same wallet file race on a fixed `.tmp` name,
    # truncating each other's in-flight writes on OneDrive paths.
    atomic_write_json(path, blob, indent=2)


def _delete_if_empty(blob):
    """Remove the wallet file if all arrays are empty. Also prune empty dirs."""
    if (blob["transactions"] or blob["provisions"]
            or blob.get("deletions") or blob.get("executionPrices")):
        return
    company_id = _safe(blob["companyId"])
    date       = _safe_date(blob["date"])
    wallet_id  = _safe(blob["walletId"])
    path = _wallet_file(company_id, date, wallet_id)
    if os.path.isfile(path):
        os.remove(path)
    # Prune date dir and company dir if now empty
    for d in (_date_dir(company_id, date), _company_dir(company_id)):
        try:
            os.rmdir(d)
        except OSError:
            break


def _iter_wallet_files(company_id, date):
    d = _date_dir(company_id, date)
    if not os.path.isdir(d):
        return
    for name in os.listdir(d):
        if name.endswith(".json"):
            yield os.path.join(d, name)


def append_rows_for_wallet(company_id, date, wallet_id, *,
                           transactions=None, provisions=None, deletions=None,
                           execution_prices=None):
    """Server-side helper mirroring /api/correcoes/bulk for one wallet.

    Used when another endpoint wants to persist corrections without a round-trip
    through HTTP (e.g. the refinement-accept flow). Dedup is keyed on
    `sourceAnomalyKey` so repeated calls are idempotent. Returns
    (added_counts, skipped_counts)."""
    buckets = {
        "transactions":    transactions     or [],
        "provisions":      provisions       or [],
        "deletions":       deletions        or [],
        "executionPrices": execution_prices or [],
    }
    try:
        _safe(company_id); _safe_date(date); _safe(wallet_id)
    except ValueError:
        return ({k: 0 for k in _KINDS}, {k: 0 for k in _KINDS})

    blob = _load(company_id, date, wallet_id)
    existing_keys = {k: set() for k in _KINDS}
    for k in _KINDS:
        for r in blob[k]:
            sk = r.get("sourceAnomalyKey")
            if sk:
                existing_keys[k].add(sk)

    added   = {k: 0 for k in _KINDS}
    skipped = {k: 0 for k in _KINDS}
    cur_cache = {}
    for kind in _KINDS:
        fields = _FIELDS_BY_KIND[kind]
        for row in buckets[kind]:
            clean = _pick(row, fields)
            sk = clean.get("sourceAnomalyKey")
            if sk and sk in existing_keys[kind]:
                skipped[kind] += 1
                continue
            clean["id"] = str(uuid.uuid4())
            clean["walletId"] = wallet_id
            if "balance" in clean:
                clean["balance"] = _coerce_balance(clean["balance"])
            if "currencyId" in fields:
                clean["currencyId"] = _wallet_currency(wallet_id, cur_cache)
            blob[kind].append(clean)
            if sk:
                existing_keys[kind].add(sk)
            added[kind] += 1
    if any(added.values()):
        _save(blob)
    return added, skipped


def load_all_pending_provisions_by_wallet(company_id):
    """Single-pass scan of the entire company's correcoes tree, returning a
    dict mapping `wallet_id → [provisions]`. Use this when iterating many
    wallets for the same company (e.g., `/api/conciliacao/rows`) to avoid
    an O(wallets × dateFolders) file-open explosion.

    Each provision row carries an added `acceptanceDate` field. Returns `{}`
    on any I/O or path validation error.

    Provisions marked `inputed=True` are excluded — they were already pushed
    to the upstream Beehus API via /api/correcoes/provisions/submit and are
    considered baked-in upstream. Applying them again locally would
    double-count against the now-upstream document."""
    if not company_id:
        return {}
    try:
        _safe(company_id)
    except ValueError:
        return {}
    comp_dir = _company_dir(company_id)
    if not os.path.isdir(comp_dir):
        return {}

    try:
        date_names = os.listdir(comp_dir)
    except OSError:
        return {}

    out = {}
    for d in date_names:
        if not _SAFE_DATE_RE.match(d):
            continue
        date_dir = os.path.join(comp_dir, d)
        try:
            files = os.listdir(date_dir)
        except OSError:
            continue
        for name in files:
            if not name.endswith(".json"):
                continue
            wid = name[:-len(".json")]
            if not _SAFE_ID_RE.match(wid):
                continue
            try:
                with open(os.path.join(date_dir, name), "r", encoding="utf-8") as f:
                    blob = json.load(f)
            except (OSError, json.JSONDecodeError):
                continue
            bucket = out.setdefault(wid, [])
            for p in (blob.get("provisions") or []):
                if p.get("inputed"):
                    continue
                row = dict(p)
                row["acceptanceDate"] = d
                bucket.append(row)
    return out


def load_all_pending_provisions(company_id, wallet_id):
    """Return every pending provision for a wallet across all acceptance-date
    folders (no date filtering). Each row carries an added `acceptanceDate`
    field. Use this when the caller wants to apply its own active-window /
    lifecycle filter — e.g., the diagnose pipeline needs both (active window
    injects into `prov_map`; lifecycle flag injects into
    `prov_lifecycle_sids`).

    Provisions marked `inputed=True` are excluded — see the docstring of
    `load_all_pending_provisions_by_wallet` for the rationale."""
    if not company_id or not wallet_id:
        return []
    try:
        _safe(company_id); _safe(wallet_id)
    except ValueError:
        return []
    comp_dir = _company_dir(company_id)
    if not os.path.isdir(comp_dir):
        return []

    results = []
    try:
        date_names = os.listdir(comp_dir)
    except OSError:
        return []
    for d in date_names:
        if not _SAFE_DATE_RE.match(d):
            continue
        wallet_file = os.path.join(comp_dir, d, f"{_safe(wallet_id)}.json")
        if not os.path.isfile(wallet_file):
            continue
        try:
            with open(wallet_file, "r", encoding="utf-8") as f:
                blob = json.load(f)
        except (OSError, json.JSONDecodeError):
            continue
        for p in (blob.get("provisions") or []):
            if p.get("inputed"):
                continue
            row = dict(p)
            row["acceptanceDate"] = d
            results.append(row)
    return results


def load_active_pending_provisions(company_id, wallet_id, target_date):
    """Return pending provisions for a wallet whose active window includes
    `target_date`, irrespective of which acceptance-date folder they live in.

    Thin filter over `load_all_pending_provisions` enforcing
    `initialDate ≤ target_date < liquidationDate` (upper-bound exclusive —
    same rule used by `db.provisions.find` in the NAV aggregation).

    Use this in UI lookups that ask "what pending corrections are *active* on
    this date?" — not in pipelines that ask "what corrections should be
    applied *to this reconciliation run*?" (the latter is what
    `load_corrections_for_wallet` is for — keyed by acceptance date)."""
    if not target_date:
        return []
    try:
        _safe_date(target_date)
    except ValueError:
        return []
    td = str(target_date)[:10]
    def _active(p):
        init = str(p.get("initialDate") or "")[:10]
        liq  = str(p.get("liquidationDate") or "")[:10]
        return bool(init) and bool(liq) and init <= td < liq
    return [p for p in load_all_pending_provisions(company_id, wallet_id) if _active(p)]


def load_corrections_for_wallet(company_id, date, wallet_id):
    """Public helper: return (transactions, provisions, deletions) stored for a
    (companyId, date, walletId) triple. Rows preserve every field on disk
    (including `id`, `originalId` and `sourceAnomalyKey`). Returns
    ([], [], []) on any missing input, invalid path segment, or unreadable file.

    Used by `pages/conciliacao.py` to inject pending corrections into the
    diagnostic pipeline so that gaps, flags and listings reflect the
    "post-correction" view without mutating the database. The `deletions`
    bucket carries `originalId` pointers — callers must filter DB
    transactions whose `_id` appears in any deletion row.

    `executionPrices` is intentionally NOT returned here — that bucket is
    consumed via `load_pending_execution_prices` (cross-folder, like
    provisions) and the gap-recalc semantics are still being defined."""
    if not company_id or not date or not wallet_id:
        return [], [], []
    try:
        _safe(company_id); _safe_date(date); _safe(wallet_id)
    except ValueError:
        return [], [], []
    path = _wallet_file(company_id, date, wallet_id)
    if not os.path.isfile(path):
        return [], [], []
    try:
        with open(path, "r", encoding="utf-8") as f:
            blob = json.load(f)
    except (OSError, json.JSONDecodeError):
        return [], [], []
    # Provisions and transactions marked `inputed=True` were already pushed to
    # the upstream Beehus API via /api/correcoes/<bucket>/submit and must be
    # excluded from local diagnose injection — they're baked-in upstream.
    txns_pending  = [t for t in (blob.get("transactions") or []) if not t.get("inputed")]
    provs_pending = [p for p in (blob.get("provisions") or []) if not p.get("inputed")]
    return (
        txns_pending,
        provs_pending,
        blob.get("deletions") or [],
    )


def load_pending_execution_prices(company_id, wallet_id):
    """Return all `executionPrices` rows accepted for a wallet across every
    acceptance-date folder. Each row carries an added `acceptanceDate`.

    Pipelines that want to honour user-accepted price corrections should
    filter the result further by `inputed == False` — rows already pushed to
    the upstream Beehus API are considered baked-in upstream and applying
    them again locally would double-count."""
    if not company_id or not wallet_id:
        return []
    try:
        _safe(company_id); _safe(wallet_id)
    except ValueError:
        return []
    comp_dir = _company_dir(company_id)
    if not os.path.isdir(comp_dir):
        return []
    results = []
    try:
        date_names = os.listdir(comp_dir)
    except OSError:
        return []
    for d in date_names:
        if not _SAFE_DATE_RE.match(d):
            continue
        wallet_file = os.path.join(comp_dir, d, f"{_safe(wallet_id)}.json")
        if not os.path.isfile(wallet_file):
            continue
        try:
            with open(wallet_file, "r", encoding="utf-8") as f:
                blob = json.load(f)
        except (OSError, json.JSONDecodeError):
            continue
        for ep in (blob.get("executionPrices") or []):
            row = dict(ep)
            row["acceptanceDate"] = d
            results.append(row)
    return results


def _latest_date_with_data(company_ids):
    """Return the most recent YYYY-MM-DD directory (across given companies)
    that contains at least one correction row, or None."""
    candidates = set()
    for cid in company_ids:
        try:
            safe_cid = _safe(cid)
        except ValueError:
            continue
        comp_dir = os.path.join(_ROOT, safe_cid)
        if not os.path.isdir(comp_dir):
            continue
        try:
            for d in os.listdir(comp_dir):
                if _SAFE_DATE_RE.match(d):
                    candidates.add((d, safe_cid))
        except OSError:
            continue

    for d, safe_cid in sorted(candidates, reverse=True):
        date_dir = os.path.join(_ROOT, safe_cid, d)
        try:
            names = os.listdir(date_dir)
        except OSError:
            continue
        for name in names:
            if not name.endswith(".json"):
                continue
            try:
                with open(os.path.join(date_dir, name), "r", encoding="utf-8") as f:
                    blob = json.load(f)
            except (OSError, json.JSONDecodeError):
                continue
            if (blob.get("transactions") or blob.get("provisions")
                    or blob.get("deletions") or blob.get("executionPrices")):
                return d
    return None


def _pick(row, fields):
    """Return a new dict with only the whitelisted fields plus the meta fields.
    The caller is always responsible for assigning `id` (uuid for create, the
    path row_id for update); we don't echo `row.id` from the client to avoid
    confusion about which side owns identity."""
    out = {k: row.get(k) for k in fields if k in row}
    for m in _META_FIELDS:
        if row.get(m):
            out[m] = row[m]
    return out


def _coerce_balance(v):
    if v is None or v == "":
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


# ── Page route ────────────────────────────────────────────────────────────────

@bp.route("/correcoes")
def index():
    companies = sorted(
        [{"id": cid, "name": name or cid} for cid, name in get_company_names().items()],
        key=lambda c: c["name"],
    )
    cf = get_company_filter()
    if cf:
        companies = [c for c in companies if c["id"] in cf]
    return render_template("correcoes.html", companies=companies)


# ── Date pills ────────────────────────────────────────────────────────────────

@bp.route("/api/correcoes/dates")
def get_dates():
    """Date pills.

    If `companyId` query param is provided, counts rows only for that company.
    Otherwise aggregates across every company visible to the user — this is the
    mode used by the Correções page after the company filter was removed.
    """
    company_id = request.args.get("companyId", "")
    end_date_param = request.args.get("endDate") or None

    if company_id:
        if not company_visible(company_id):
            return jsonify({"cards": []})
        try:
            _safe(company_id)
        except ValueError:
            end_date = end_date_param or today_in_brt().isoformat()
            dates = get_biz_dates(_NUM_DATES, end_date)
            return jsonify({"cards": [{"date": d, "total": 0} for d in dates]})
        company_ids = [company_id]
    else:
        company_ids = _visible_company_ids()

    # Default `endDate` to the latest date that actually has data; fall back to
    # today when nothing is persisted yet. When the caller supplies `endDate`
    # explicitly, respect it.
    if end_date_param:
        end_date = end_date_param
    else:
        end_date = _latest_date_with_data(company_ids) or today_in_brt().isoformat()
    dates = get_biz_dates(_NUM_DATES, end_date)

    totals = {}
    for d in dates:
        total = 0
        for cid in company_ids:
            try:
                safe_cid = _safe(cid)
            except ValueError:
                continue
            date_dir = os.path.join(_ROOT, safe_cid, d)
            if not os.path.isdir(date_dir):
                continue
            for name in os.listdir(date_dir):
                if not name.endswith(".json"):
                    continue
                try:
                    with open(os.path.join(date_dir, name), "r", encoding="utf-8") as f:
                        blob = json.load(f)
                    total += (
                        len(blob.get("transactions") or [])
                        + len(blob.get("provisions") or [])
                        + len(blob.get("deletions") or [])
                        + len(blob.get("executionPrices") or [])
                    )
                except (OSError, json.JSONDecodeError):
                    continue
        totals[d] = total

    cards = [{"date": d, "total": totals.get(d, 0)} for d in dates]
    return jsonify({"cards": cards})


# ── Read rows for a company+date (all wallets combined) ───────────────────────

@bp.route("/api/correcoes")
def list_items():
    """List rows for a given date.

    If `companyId` is provided, scopes to that one company and returns the
    wallet map for it. Otherwise returns rows across every visible company,
    with `companyId` attached to each row plus a `companies` map `{id: name}`
    and a `wallets` map `{walletId: name}` spanning all of them.
    """
    company_id = request.args.get("companyId", "")
    date       = request.args.get("date", "")
    if not company_visible(company_id):
        return jsonify({"transactions": [], "provisions": [], "deletions": [],
                        "executionPrices": [], "wallets": {}, "companies": {}})

    try:
        _safe_date(date)
        if company_id:
            _safe(company_id)
    except ValueError:
        return jsonify({"error": "invalid companyId or date"}), 400

    company_ids = [company_id] if company_id else _visible_company_ids()

    # Company name map (id -> name). Fetch all (cached) then intersect.
    wanted_ids    = set(company_ids)
    _names        = get_company_names()
    companies_map = {cid: (_names[cid] or cid) for cid in wanted_ids if cid in _names}
    # Also include orphans (ids on disk without a companies document).
    for cid in company_ids:
        companies_map.setdefault(cid, cid)

    # Wallet name map across all requested companies, plus a grouping by company
    # so the Add modal can cascade company → wallet dropdown. `wallet_currencies`
    # surfaces each wallet's authoritative `currencyId` so the UI can show it
    # alongside the wallet name — the submit endpoints already resolve currency
    # from `db.wallets`, so the displayed value matches what the API will receive.
    wallet_names = {}
    wallet_currencies = {}
    wallets_by_company = {}
    for w in db.wallets.find(
        {"companyId": {"$in": company_ids}},
        {"name": 1, "companyId": 1, "currency": 1, "currencyId": 1},
    ):
        wid = str(w["_id"])
        wname = w.get("name", wid)
        wcid = str(w.get("companyId", ""))
        wallet_names[wid] = wname
        # `currency` is the field name actually stored on wallet documents
        # (e.g. "USD", "BRL"). `currencyId` is read as a defensive fallback
        # only — see the note on `_wallet_currency` above.
        wallet_currencies[wid] = str(w.get("currency") or w.get("currencyId") or "")
        wallets_by_company.setdefault(wcid, {})[wid] = wname

    transactions, provisions, deletions, execution_prices = [], [], [], []
    for cid in company_ids:
        for path in _iter_wallet_files(cid, date):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    blob = json.load(f)
            except (OSError, json.JSONDecodeError):
                continue
            wid = blob.get("walletId", "")
            bcid = blob.get("companyId") or cid
            for r in blob.get("transactions", []):
                transactions.append({**r, "walletId": r.get("walletId") or wid, "companyId": bcid})
            for r in blob.get("provisions", []):
                provisions.append({**r, "walletId": r.get("walletId") or wid, "companyId": bcid})
            for r in blob.get("deletions", []):
                deletions.append({**r, "walletId": r.get("walletId") or wid, "companyId": bcid})
            for r in blob.get("executionPrices", []):
                execution_prices.append({**r, "walletId": r.get("walletId") or wid, "companyId": bcid})

    return jsonify({
        "transactions":      transactions,
        "provisions":        provisions,
        "deletions":         deletions,
        "executionPrices":   execution_prices,
        "wallets":           wallet_names,
        "walletCurrencies":  wallet_currencies,
        "walletsByCompany":  wallets_by_company,
        "companies":         companies_map,
    })


# ── Anomaly keys for upstream (which painel rows are already moved) ───────────

@bp.route("/api/correcoes/anomaly-keys")
def anomaly_keys():
    """Return the unique sourceAnomalyKey values for a (company, date).

    Upstream pages (e.g. Painel) call this to gray out rows whose corrections
    already live in the correcoes store.
    """
    company_id = request.args.get("companyId", "")
    date       = request.args.get("date", "")
    if not company_visible(company_id):
        return jsonify({"keys": []})
    try:
        _safe(company_id); _safe_date(date)
    except ValueError:
        return jsonify({"keys": []})

    keys = set()
    for path in _iter_wallet_files(company_id, date):
        try:
            with open(path, "r", encoding="utf-8") as f:
                blob = json.load(f)
        except (OSError, json.JSONDecodeError):
            continue
        # Every bucket that a Painel/Conciliação accept can write into is
        # surveyed — otherwise the front-end's `_currentAnomalyKeys` set is
        # incomplete and the Aceitar button never flips to "✓ Aceito" for
        # rows that landed in the missing bucket (e.g. executionPrices).
        for bucket in ("transactions", "provisions", "executionPrices"):
            for r in blob.get(bucket) or []:
                k = r.get("sourceAnomalyKey")
                if k:
                    keys.add(k)
    return jsonify({"keys": sorted(keys)})


# ── Create / Update / Delete ──────────────────────────────────────────────────

@bp.route("/api/correcoes/items", methods=["POST"])
def create_item():
    body = request.get_json() or {}
    company_id = body.get("companyId", "")
    date       = body.get("date", "")
    wallet_id  = body.get("walletId", "")
    kind       = body.get("kind", "")
    row        = body.get("row") or {}

    if not company_visible(company_id):
        return jsonify({"error": "forbidden"}), 403
    if kind not in _KINDS:
        return jsonify({"error": f"kind must be one of {list(_KINDS)}"}), 400

    try:
        _safe(company_id); _safe_date(date); _safe(wallet_id)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    if not _wallet_in_company(wallet_id, company_id):
        return jsonify({"error": "wallet/company mismatch"}), 403

    fields = _FIELDS_BY_KIND[kind]
    clean = _pick(row, fields)
    clean["id"] = str(uuid.uuid4())
    clean["walletId"] = wallet_id
    if "balance" in clean:
        clean["balance"] = _coerce_balance(clean["balance"])
    if "currencyId" in fields:
        clean["currencyId"] = _wallet_currency(wallet_id)

    blob = _load(company_id, date, wallet_id)
    blob[kind].append(clean)
    _save(blob)
    return jsonify({"id": clean["id"], "row": clean})


@bp.route("/api/correcoes/items", methods=["PUT"])
def update_item():
    body = request.get_json() or {}
    company_id = body.get("companyId", "")
    date       = body.get("date", "")
    wallet_id  = body.get("walletId", "")
    kind       = body.get("kind", "")
    row        = body.get("row") or {}
    row_id     = row.get("id", "")

    if not company_visible(company_id):
        return jsonify({"error": "forbidden"}), 403
    if kind not in _KINDS:
        return jsonify({"error": f"kind must be one of {list(_KINDS)}"}), 400
    if not row_id:
        return jsonify({"error": "row.id required"}), 400

    try:
        _safe(company_id); _safe_date(date); _safe(wallet_id)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    if not _wallet_in_company(wallet_id, company_id):
        return jsonify({"error": "wallet/company mismatch"}), 403

    fields = _FIELDS_BY_KIND[kind]
    clean = _pick(row, fields)
    clean["id"] = row_id
    clean["walletId"] = wallet_id
    if "balance" in clean:
        clean["balance"] = _coerce_balance(clean["balance"])
    if "currencyId" in fields:
        clean["currencyId"] = _wallet_currency(wallet_id)

    blob = _load(company_id, date, wallet_id)
    found = False
    for i, r in enumerate(blob[kind]):
        if r.get("id") == row_id:
            blob[kind][i] = clean
            found = True
            break
    if not found:
        return jsonify({"error": "row not found"}), 404
    _save(blob)
    return jsonify({"row": clean})


@bp.route("/api/correcoes/items", methods=["DELETE"])
def delete_item():
    body = request.get_json() or {}
    company_id = body.get("companyId", "")
    date       = body.get("date", "")
    wallet_id  = body.get("walletId", "")
    kind       = body.get("kind", "")
    row_id     = body.get("id", "")

    if not company_visible(company_id):
        return jsonify({"error": "forbidden"}), 403
    if kind not in _KINDS:
        return jsonify({"error": f"kind must be one of {list(_KINDS)}"}), 400
    if not row_id:
        return jsonify({"error": "id required"}), 400

    try:
        _safe(company_id); _safe_date(date); _safe(wallet_id)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    if not _wallet_in_company(wallet_id, company_id):
        return jsonify({"error": "wallet/company mismatch"}), 403

    blob = _load(company_id, date, wallet_id)
    before = len(blob[kind])
    blob[kind] = [r for r in blob[kind] if r.get("id") != row_id]
    if len(blob[kind]) == before:
        return jsonify({"error": "row not found"}), 404
    _save(blob)
    _delete_if_empty(blob)
    return jsonify({"ok": True})


# ── Bulk append (called by upstream pages like painel/conciliacao) ────────────

@bp.route("/api/correcoes/bulk", methods=["POST"])
def bulk_append():
    """Append many rows at once, grouped by wallet.

    Body: {
      companyId, date,
      wallets: {
        walletId: { transactions: [...], provisions: [...] }
      }
    }
    """
    body = request.get_json() or {}
    company_id = body.get("companyId", "")
    date       = body.get("date", "")
    wallets    = body.get("wallets") or {}

    if not company_visible(company_id):
        return jsonify({"error": "forbidden"}), 403

    try:
        _safe(company_id); _safe_date(date)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    added    = {k: 0 for k in _KINDS}
    skipped  = {k: 0 for k in _KINDS}
    rejected = []   # wallets dropped due to id-validation or company mismatch
    for wallet_id, buckets in (wallets or {}).items():
        try:
            _safe(wallet_id)
        except ValueError:
            rejected.append({"walletId": wallet_id, "reason": "invalid_wallet_id"})
            continue
        if not _wallet_in_company(wallet_id, company_id):
            # Used to silently skip — that masked a real issue (front-end
            # calling with a stale `_selectedCompany` that doesn't match the
            # diagnose-time wallet) where users saw "Aceitar" succeed in the
            # network tab but no rows landed on disk. Surface it instead.
            rejected.append({"walletId": wallet_id, "reason": "wallet_company_mismatch"})
            continue
        blob = _load(company_id, date, wallet_id)
        # Build an index of existing sourceAnomalyKeys to make re-sends idempotent.
        existing_keys = {k: set() for k in _KINDS}
        for k in _KINDS:
            for r in blob[k]:
                sk = r.get("sourceAnomalyKey")
                if sk:
                    existing_keys[k].add(sk)

        wallet_currency = _wallet_currency(wallet_id)
        for kind in _KINDS:
            fields = _FIELDS_BY_KIND[kind]
            for row in (buckets or {}).get(kind, []) or []:
                clean = _pick(row, fields)
                sk = clean.get("sourceAnomalyKey")
                if sk and sk in existing_keys[kind]:
                    skipped[kind] += 1
                    continue
                clean["id"] = str(uuid.uuid4())
                clean["walletId"] = wallet_id
                if "balance" in clean:
                    clean["balance"] = _coerce_balance(clean["balance"])
                if "currencyId" in fields:
                    clean["currencyId"] = wallet_currency
                blob[kind].append(clean)
                if sk:
                    existing_keys[kind].add(sk)
                added[kind] += 1
        _save(blob)
    # If every requested wallet was rejected and nothing landed, fail loudly so
    # the front-end can show the user a real error instead of pretending it
    # worked. A partial reject (some wallets ok, some not) still returns 200
    # but the `rejected[]` payload tells the caller which ones failed.
    body_resp = {"added": added, "skipped": skipped, "rejected": rejected}
    if rejected and not any(added.values()) and not any(skipped.values()):
        return jsonify({**body_resp, "error": "todas as carteiras foram rejeitadas — verifique companyId × walletId"}), 422
    return jsonify(body_resp)


# ── Export ────────────────────────────────────────────────────────────────────

_INTERNAL_FIELDS = ("id",) + _META_FIELDS

def _strip_internal(rows):
    return [{k: v for k, v in r.items() if k not in _INTERNAL_FIELDS} for r in rows]


def _collect(company_id, date):
    """Return (transactions, provisions) — ids + meta stripped."""
    txns, provs = [], []
    for path in _iter_wallet_files(company_id, date):
        try:
            with open(path, "r", encoding="utf-8") as f:
                blob = json.load(f)
        except (OSError, json.JSONDecodeError):
            continue
        txns.extend(blob.get("transactions") or [])
        provs.extend(blob.get("provisions") or [])
    return _strip_internal(txns), _strip_internal(provs)


@bp.route("/api/correcoes/export", methods=["POST"])
def export_files():
    body = request.get_json() or {}
    company_id = body.get("companyId", "")
    date       = body.get("date", "")
    fmt        = (body.get("format") or "json").lower()

    if not company_visible(company_id):
        return jsonify({"error": "forbidden"}), 403
    if fmt not in ("json", "xlsx"):
        return jsonify({"error": "format must be 'json' or 'xlsx'"}), 400

    try:
        _safe(company_id); _safe_date(date)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    txns, provs = _collect(company_id, date)
    if not txns and not provs:
        return jsonify({"error": "nenhuma correcao para exportar"}), 400

    if fmt == "json":
        payload = {}
        if txns:
            payload["transactions"] = {"companyId": company_id, "transactions": txns}
        if provs:
            payload["provisions"] = {"provisions": provs}
        buf = io.BytesIO(json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"))
        return send_file(
            buf, mimetype="application/json", as_attachment=True,
            download_name=f"correcoes_{company_id}_{date}.json",
        )

    # XLSX: one sheet per kind
    wb = Workbook()
    wb.remove(wb.active)
    if txns:
        ws = wb.create_sheet("Transactions")
        ws.append(_TXN_FIELDS)
        for r in txns:
            ws.append([r.get(k) for k in _TXN_FIELDS])
    if provs:
        ws = wb.create_sheet("Provisions")
        ws.append(_PROV_FIELDS)
        for r in provs:
            ws.append([r.get(k) for k in _PROV_FIELDS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"correcoes_{company_id}_{date}.xlsx",
    )


# ── Upstream push: shared per-row cores ───────────────────────────────────────
#
# Each `_push_<kind>_row` validates one persisted row, forwards it to the
# Beehus API and, on success, stamps `inputed`/`inputedAt`/`beehusId` on the
# row in place. They DO NOT persist the blob — the caller saves once. They
# return (ok: bool, info: dict); on failure `info` carries `error` and,
# optionally, `auth=True` (→ 401) or `upstream_status`/`upstream_body` (→ 502).
# Both the single-row /submit endpoints and the batch /bulk-submit endpoint
# go through these, so upstream-submission logic lives in exactly one place.

def _stamp_inputed(target, result):
    target["inputed"]   = True
    target["inputedAt"] = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    target["beehusId"]  = (result or {}).get("_id") or (result or {}).get("id")


def _push_transaction_row(company_id, wallet_id, target):
    if target.get("inputed"):
        return True, {"alreadyInputed": True}
    operation_date   = target.get("operationDate") or ""
    liquidation_date = target.get("liquidationDate") or ""
    txn_type         = target.get("beehusTransactionType") or ""
    balance          = target.get("balance")
    missing = []
    if not operation_date:    missing.append("operationDate")
    if not liquidation_date:  missing.append("liquidationDate")
    if not txn_type:          missing.append("beehusTransactionType")
    if balance in (None, ""): missing.append("balance")
    if missing:
        return False, {"error": f"row missing required fields: {', '.join(missing)}"}
    try:
        balance_num = float(balance)
    except (TypeError, ValueError):
        return False, {"error": "balance not numeric"}

    # currencyId is ALWAYS resolved from the wallet (authoritative); entityId
    # falls back from the wallet only when missing on the row.
    entity_id   = target.get("entityId") or ""
    currency_id = _wallet_currency(wallet_id)
    if not entity_id:
        w = resolve_wallet(wallet_id, {"entityId": 1}) or {}
        entity_id = str(w.get("entityId") or "")
    if not entity_id:
        return False, {"error": "row missing entityId and wallet has no entityId"}
    target["currencyId"] = currency_id

    try:
        result = create_transaction(
            company_id=company_id, entity_id=entity_id, wallet_id=wallet_id,
            balance=balance_num, operation_date=operation_date,
            liquidation_date=liquidation_date, currency_id=currency_id,
            transaction_type=txn_type, description=target.get("description") or "",
            comment=target.get("comment") or "", hide=bool(target.get("hide", False)),
            input_type=target.get("inputType") or "sheets",
            security_id=target.get("securityId") or None,
        )
    except BeehusAuthError as e:
        return False, {"error": str(e), "auth": True,
                       "upstream_status": getattr(e, "status", None),
                       "upstream_body": getattr(e, "body", None)}
    except BeehusAPIError as e:
        return False, {"error": str(e), "upstream_status": getattr(e, "status", None),
                       "upstream_body": getattr(e, "body", None)}
    _stamp_inputed(target, result)
    return True, {"upstream": result}


def _push_provision_row(company_id, wallet_id, target):
    if target.get("inputed"):
        return True, {"alreadyInputed": True}
    initial_date     = target.get("initialDate") or ""
    liquidation_date = target.get("liquidationDate") or ""
    provision_type   = target.get("provisionType") or ""
    provision_source = target.get("provisionSource") or "adjustments"
    balance          = target.get("balance")
    missing = []
    if not initial_date:      missing.append("initialDate")
    if not liquidation_date:  missing.append("liquidationDate")
    if not provision_type:    missing.append("provisionType")
    if balance in (None, ""): missing.append("balance")
    if missing:
        return False, {"error": f"row missing required fields: {', '.join(missing)}"}
    try:
        balance_num = float(balance)
    except (TypeError, ValueError):
        return False, {"error": "balance not numeric"}

    currency_id = _wallet_currency(wallet_id)
    target["currencyId"] = currency_id
    try:
        result = create_provision(
            company_id=company_id, wallet_id=wallet_id, balance=balance_num,
            initial_date=initial_date, liquidation_date=liquidation_date,
            provision_type=provision_type, provision_source=provision_source,
            currency_id=currency_id, description=target.get("description") or "",
            security_id=target.get("securityId") or None,
        )
    except BeehusAuthError as e:
        return False, {"error": str(e), "auth": True,
                       "upstream_status": getattr(e, "status", None),
                       "upstream_body": getattr(e, "body", None)}
    except BeehusAPIError as e:
        return False, {"error": str(e), "upstream_status": getattr(e, "status", None),
                       "upstream_body": getattr(e, "body", None)}
    _stamp_inputed(target, result)
    return True, {"upstream": result}


def _push_execution_price_row(company_id, wallet_id, date, target):
    if target.get("inputed"):
        return True, {"alreadyInputed": True}
    sec_id        = target.get("securityId") or ""
    position_date = target.get("positionDate") or date
    exec_price    = target.get("executionPrice")
    if not sec_id or exec_price in (None, ""):
        return False, {"error": "row missing securityId or executionPrice"}
    try:
        exec_price_num = float(exec_price)
    except (TypeError, ValueError):
        return False, {"error": "executionPrice not numeric"}
    try:
        result = create_execution_price(
            company_id=company_id, wallet_id=wallet_id, security_id=sec_id,
            position_date=position_date, execution_price=exec_price_num,
        )
    except BeehusAuthError as e:
        return False, {"error": str(e), "auth": True,
                       "upstream_status": getattr(e, "status", None),
                       "upstream_body": getattr(e, "body", None)}
    except BeehusAPIError as e:
        return False, {"error": str(e), "upstream_status": getattr(e, "status", None),
                       "upstream_body": getattr(e, "body", None)}
    _stamp_inputed(target, result)
    return True, {"upstream": result}


def _push_info_to_response(ok, info, target):
    """Map a (ok, info) push result to the (json, http_status) the single-row
    /submit endpoints return. Assumes the blob has already been saved by the
    caller on success."""
    if ok:
        return jsonify({"ok": True, "row": target, **info}), 200
    if info.get("auth"):
        return jsonify({"error": info["error"], "upstream_status": info.get("upstream_status"),
                        "upstream_body": info.get("upstream_body")}), 401
    if info.get("upstream_status") is not None or "upstream_body" in info:
        return jsonify({"error": info["error"], "upstream_status": info.get("upstream_status"),
                        "upstream_body": info.get("upstream_body")}), 502
    return jsonify({"error": info["error"]}), 400


# ── Execution prices: push to upstream API and mark row as inputed ────────────

@bp.route("/api/correcoes/execution-prices/submit", methods=["POST"])
def submit_execution_price():
    """Push a single `executionPrices` row to the Beehus API and mark it as
    `inputed=true` on success.

    Body: { companyId, date, walletId, id }  — `id` is the row uuid as
    persisted in `data/correcoes/.../<wallet>.json`. The row's
    `executionPrice`, `securityId`, `walletId` and `positionDate` are read
    from disk and forwarded to `POST /beehus/financial/execution-prices`.

    On HTTP 2xx, the row is updated in place: `inputed=true`,
    `inputedAt=<utc-iso8601>`, `beehusId=<upstream _id>`. The diagnose
    pipeline excludes inputed rows from any local gap recalc (they are
    considered baked into upstream)."""
    body = request.get_json() or {}
    company_id = body.get("companyId", "")
    date       = body.get("date", "")
    wallet_id  = body.get("walletId", "")
    row_id     = body.get("id", "")

    if not company_visible(company_id):
        return jsonify({"error": "forbidden"}), 403
    if not row_id:
        return jsonify({"error": "id required"}), 400
    try:
        _safe(company_id); _safe_date(date); _safe(wallet_id)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if not _wallet_in_company(wallet_id, company_id):
        return jsonify({"error": "wallet/company mismatch"}), 403

    blob = _load(company_id, date, wallet_id)
    target = next((r for r in blob.get("executionPrices") or [] if r.get("id") == row_id), None)
    if not target:
        return jsonify({"error": "row not found"}), 404

    ok, info = _push_execution_price_row(company_id, wallet_id, date, target)
    if ok and not info.get("alreadyInputed"):
        _save(blob)
    return _push_info_to_response(ok, info, target)


# ── Transactions: push to upstream API and mark row as inputed ────────────────

@bp.route("/api/correcoes/transactions/submit", methods=["POST"])
def submit_transaction():
    """Push a single `transactions` row to the Beehus API and mark it as
    `inputed=true` on success.

    Body: { companyId, date, walletId, id }  — `id` is the row uuid as
    persisted in `data/correcoes/.../<wallet>.json`. Forwards the row's
    `entityId`, `balance`, `operationDate`, `liquidationDate`, `currencyId`,
    `beehusTransactionType`, `description`, `comment`, `hide`, `inputType`
    and `securityId` to `POST /beehus/financial/transactions` (same upstream
    call used by /beehus *Funções > Criar Transação*).

    On HTTP 2xx, the row is updated in place: `inputed=true`,
    `inputedAt=<utc-iso8601>`, `beehusId=<upstream _id>`. The diagnose
    pipeline (`load_corrections_for_wallet`) excludes inputed transactions
    from local injection — they're considered baked into upstream.

    `entityId` falls back to the wallet's stored `entityId` when the row
    has none. `currencyId` likewise falls back to the wallet's currency,
    matching the server-side enforcement on write."""
    body = request.get_json() or {}
    company_id = body.get("companyId", "")
    date       = body.get("date", "")
    wallet_id  = body.get("walletId", "")
    row_id     = body.get("id", "")

    if not company_visible(company_id):
        return jsonify({"error": "forbidden"}), 403
    if not row_id:
        return jsonify({"error": "id required"}), 400
    try:
        _safe(company_id); _safe_date(date); _safe(wallet_id)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if not _wallet_in_company(wallet_id, company_id):
        return jsonify({"error": "wallet/company mismatch"}), 403

    blob = _load(company_id, date, wallet_id)
    target = next((r for r in blob.get("transactions") or [] if r.get("id") == row_id), None)
    if not target:
        return jsonify({"error": "row not found"}), 404

    ok, info = _push_transaction_row(company_id, wallet_id, target)
    if ok and not info.get("alreadyInputed"):
        _save(blob)
    return _push_info_to_response(ok, info, target)


# ── Provisions: push to upstream API and mark row as inputed ──────────────────

@bp.route("/api/correcoes/provisions/submit", methods=["POST"])
def submit_provision():
    """Push a single `provisions` row to the Beehus API and mark it as
    `inputed=true` on success.

    Body: { companyId, date, walletId, id }  — `id` is the row uuid as
    persisted in `data/correcoes/.../<wallet>.json`. Forwards the row's
    `provisionType`, `provisionSource`, `balance`, `initialDate`,
    `liquidationDate`, `currencyId`, `description`, `securityId` to
    `POST /beehus/provisions` (same upstream call used by /beehus
    *Funções > Criar Provisão*).

    On HTTP 2xx, the row is updated in place: `inputed=true`,
    `inputedAt=<utc-iso8601>`, `beehusId=<upstream _id>`. Pipelines that
    apply pending provisions locally (`load_active_pending_provisions`,
    `load_all_pending_provisions`) should filter `inputed=False` to avoid
    double-counting against the now-upstream document."""
    body = request.get_json() or {}
    company_id = body.get("companyId", "")
    date       = body.get("date", "")
    wallet_id  = body.get("walletId", "")
    row_id     = body.get("id", "")

    if not company_visible(company_id):
        return jsonify({"error": "forbidden"}), 403
    if not row_id:
        return jsonify({"error": "id required"}), 400
    try:
        _safe(company_id); _safe_date(date); _safe(wallet_id)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if not _wallet_in_company(wallet_id, company_id):
        return jsonify({"error": "wallet/company mismatch"}), 403

    blob = _load(company_id, date, wallet_id)
    target = next((r for r in blob.get("provisions") or [] if r.get("id") == row_id), None)
    if not target:
        return jsonify({"error": "row not found"}), 404

    ok, info = _push_provision_row(company_id, wallet_id, target)
    if ok and not info.get("alreadyInputed"):
        _save(blob)
    return _push_info_to_response(ok, info, target)


# ── Bulk: persist AND push to upstream in one shot ────────────────────────────

@bp.route("/api/correcoes/bulk-submit", methods=["POST"])
def bulk_submit():
    """Persist correction rows (idempotent, like /bulk) AND immediately push
    them to the upstream Beehus API, stamping `inputed=true` on success.

    Used by the Conciliação > Diagnóstico "Aceitar" flow: after the user
    reviews/edits the generated rows in the confirmation modal, the rows are
    written to `data/correcoes/...` (audit trail) and sent straight to the
    API — no manual trip through the /correcoes page.

    Body: same shape as `/api/correcoes/bulk`:
      { companyId, date, wallets: { walletId: {transactions, provisions,
        deletions, executionPrices} } }

    `deletions` are persisted (internal bookkeeping) but NEVER pushed upstream.
    Submission targets are the persisted rows of the submittable kinds whose
    `sourceAnomalyKey` is in the request and that are not yet `inputed` — so a
    re-accept after a partial failure retries only the unsent rows, and a
    re-accept after full success is a no-op.

    Returns 200 with:
      { added, skipped, rejected, submitted, failed, authFailed, results: [
          {walletId, kind, id, securityId, ok, beehusId?, error?, upstream_status?}
      ] }
    Callers must inspect `failed`/`authFailed`/`rejected` — a 200 does not mean
    every row was accepted upstream."""
    body = request.get_json() or {}
    company_id = body.get("companyId", "")
    date       = body.get("date", "")
    wallets    = body.get("wallets") or {}

    if not company_visible(company_id):
        return jsonify({"error": "forbidden"}), 403
    try:
        _safe(company_id); _safe_date(date)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    added    = {k: 0 for k in _KINDS}
    skipped  = {k: 0 for k in _KINDS}
    rejected = []
    results  = []
    submitted = 0
    failed    = 0
    auth_failed = False

    # Only these kinds are pushed upstream — deletions stay local.
    submit_kinds = ("transactions", "provisions", "executionPrices")

    for wallet_id, buckets in (wallets or {}).items():
        try:
            _safe(wallet_id)
        except ValueError:
            rejected.append({"walletId": wallet_id, "reason": "invalid_wallet_id"})
            continue
        if not _wallet_in_company(wallet_id, company_id):
            rejected.append({"walletId": wallet_id, "reason": "wallet_company_mismatch"})
            continue

        buckets = buckets or {}
        a, s = append_rows_for_wallet(
            company_id, date, wallet_id,
            transactions=buckets.get("transactions"),
            provisions=buckets.get("provisions"),
            deletions=buckets.get("deletions"),
            execution_prices=buckets.get("executionPrices"),
        )
        for k in _KINDS:
            added[k]   += a.get(k, 0)
            skipped[k] += s.get(k, 0)

        # Keys present in THIS request, per submittable kind.
        req_keys = {kind: {r.get("sourceAnomalyKey")
                           for r in (buckets.get(kind) or []) if r.get("sourceAnomalyKey")}
                    for kind in submit_kinds}

        blob  = _load(company_id, date, wallet_id)
        dirty = False
        for kind in submit_kinds:
            keys = req_keys[kind]
            if not keys:
                continue
            for target in blob.get(kind) or []:
                sk = target.get("sourceAnomalyKey")
                if not sk or sk not in keys or target.get("inputed"):
                    continue
                if kind == "transactions":
                    ok, info = _push_transaction_row(company_id, wallet_id, target)
                elif kind == "provisions":
                    ok, info = _push_provision_row(company_id, wallet_id, target)
                else:
                    ok, info = _push_execution_price_row(company_id, wallet_id, date, target)

                res = {"walletId": wallet_id, "kind": kind, "id": target.get("id"),
                       "securityId": target.get("securityId"), "ok": ok}
                if ok:
                    submitted += 1
                    if not info.get("alreadyInputed"):
                        dirty = True
                    res["beehusId"] = target.get("beehusId")
                else:
                    failed += 1
                    if info.get("auth"):
                        auth_failed = True
                    res["error"] = info.get("error")
                    res["upstream_status"] = info.get("upstream_status")
                results.append(res)
        if dirty:
            _save(blob)

    return jsonify({"added": added, "skipped": skipped, "rejected": rejected,
                    "submitted": submitted, "failed": failed,
                    "authFailed": auth_failed, "results": results})

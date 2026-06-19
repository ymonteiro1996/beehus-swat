"""Routes for the Parceiro (Partner) console page.

Layout:
    /parceiro                                — page
    /api/parceiro/token         GET/POST/DELETE — separate from /api/beehus/token
    /api/parceiro/companies     GET            — visible companies
    /api/parceiro/groupings     GET ?companyId — groupings of the company
    /api/parceiro/users         POST           — create one user
    /api/parceiro/users/<id>    PATCH          — associate groupings
    /api/parceiro/util/parse-users-excel       POST — bulk-cadastro rows
    /api/parceiro/util/parse-groupings-excel   POST — bulk-association rows
"""
from flask import Blueprint, render_template, jsonify, request

from db import (
    company_visible,
    get_company_filter,
    get_company_names,
    get_grouping_index,
)
from partner_api import (
    PartnerAPIError,
    PartnerAuthError,
    create_user,
    update_user,
    set_token,
    clear_token,
    token_status,
)

bp = Blueprint("parceiro", __name__)


# ── Page ──────────────────────────────────────────────────────────────────────

@bp.route("/parceiro")
def index():
    return render_template("parceiro.html")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _api_error_response(e: PartnerAPIError):
    status = 401 if isinstance(e, PartnerAuthError) else 502
    return jsonify({
        "error":           str(e),
        "upstream_status": e.status,
        "upstream_body":   e.body,
    }), status


# ── Token ─────────────────────────────────────────────────────────────────────

@bp.route("/api/parceiro/token", methods=["GET"])
def token_get():
    return jsonify(token_status())


@bp.route("/api/parceiro/token", methods=["POST"])
def token_set():
    data = request.get_json(silent=True) or {}
    try:
        set_token(data.get("token", ""))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify(token_status())


@bp.route("/api/parceiro/token", methods=["DELETE"])
def token_clear():
    clear_token()
    return jsonify(token_status())


# ── Filters ───────────────────────────────────────────────────────────────────

@bp.route("/api/parceiro/companies")
def filter_companies():
    cf = get_company_filter()
    items = []
    for cid, name in get_company_names().items():
        if cf and cid not in cf:
            continue
        items.append({"id": cid, "name": name or cid})
    items.sort(key=lambda x: x["name"].lower())
    return jsonify(items)


@bp.route("/api/parceiro/groupings")
def filter_groupings():
    """Groupings registered for the given company. Trashed ones excluded."""
    company_id = request.args.get("companyId", "")
    if not company_visible(company_id):
        return jsonify([])
    items = []
    for gid, g in get_grouping_index().items():
        if g.get("trashed"):
            continue
        if g.get("companyId") != company_id:
            continue
        items.append({
            "id":   gid,
            "name": g["name"] or gid,
        })
    items.sort(key=lambda x: x["name"].lower())
    return jsonify(items)


# ── Users ─────────────────────────────────────────────────────────────────────

def _validate_user_fields(data):
    """Return (cleaned_dict, error_message)."""
    name  = (data.get("name")  or "").strip()
    email = (data.get("email") or "").strip()
    phone = (data.get("phone") or "").strip()
    if not name:
        return None, "name is required"
    if not email:
        return None, "email is required"
    if not phone:
        return None, "phone is required"
    return {"name": name, "email": email, "phone": phone}, None


@bp.route("/api/parceiro/users", methods=["POST"])
def users_create():
    data = request.get_json(silent=True) or {}
    company_id = (data.get("companyId") or "").strip()
    if not company_id:
        return jsonify({"error": "companyId is required"}), 400
    if not company_visible(company_id):
        return jsonify({"error": "company is not visible to this user"}), 403
    fields, err = _validate_user_fields(data)
    if err:
        return jsonify({"error": err}), 400
    try:
        result = create_user(
            company_id=company_id,
            name=fields["name"],
            email=fields["email"],
            phone=fields["phone"],
        )
    except PartnerAPIError as e:
        return _api_error_response(e)
    return jsonify(result), 201


@bp.route("/api/parceiro/users/<user_id>", methods=["PATCH"])
def users_patch(user_id):
    """Replace the partner user's groupings (and identity fields).

    Body:
        name, email, phone (str, required) — must be re-sent as upstream
            replaces the whole document.
        groupings (list[dict], required) — each item has {_id, description, isOwn}.
    """
    data = request.get_json(silent=True) or {}
    fields, err = _validate_user_fields(data)
    if err:
        return jsonify({"error": err}), 400
    raw_groupings = data.get("groupings")
    if not isinstance(raw_groupings, list):
        return jsonify({"error": "groupings must be a list"}), 400

    cleaned: list[dict] = []
    for g in raw_groupings:
        if not isinstance(g, dict):
            return jsonify({"error": "each grouping must be an object"}), 400
        gid = (g.get("_id") or "").strip()
        if not gid:
            return jsonify({"error": "grouping _id is required"}), 400
        cleaned.append({
            "_id":         gid,
            "description": (g.get("description") or "").strip(),
            "isOwn":       bool(g.get("isOwn", False)),
        })
    try:
        result = update_user(
            user_id,
            name=fields["name"],
            email=fields["email"],
            phone=fields["phone"],
            groupings=cleaned,
        )
    except PartnerAPIError as e:
        return _api_error_response(e)
    return jsonify({"updated": user_id, "response": result})


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _open_uploaded_workbook():
    """Return (workbook, error_response_tuple). Caller checks `error` first."""
    f = request.files.get("file")
    if f is None:
        return None, (jsonify({"error": "no file uploaded (expected multipart 'file')"}), 400)
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None, (jsonify({"error": "openpyxl is not installed on the server"}), 500)
    try:
        wb = load_workbook(f, read_only=True, data_only=True)
    except Exception as e:
        return None, (jsonify({"error": f"could not read workbook: {e}"}), 400)
    return wb, None


def _stringify(v):
    if v is None:
        return ""
    return str(v).strip()


@bp.route("/api/parceiro/util/parse-users-excel", methods=["POST"])
def util_parse_users_excel():
    """Parse a workbook with one row per user.

    Expected header row (case-insensitive, lowercased): name, email, phone.
    Returns an ordered list of {name, email, phone} dicts (deduped by email)
    so the UI can preview them before invoking POST per row.
    """
    wb, err = _open_uploaded_workbook()
    if err:
        return err

    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return jsonify({"users": [], "count": 0, "warnings": ["empty sheet"]})

    cols = {(_stringify(h).lower()): i for i, h in enumerate(header) if h is not None}
    required = ("name", "email", "phone")
    missing = [c for c in required if c not in cols]
    if missing:
        return jsonify({
            "error": f"missing columns: {', '.join(missing)}",
            "expectedHeader": list(required),
            "foundHeader":    [_stringify(h) for h in header],
        }), 400

    out, seen = [], set()
    for r in rows:
        if r is None:
            continue
        name  = _stringify(r[cols["name"]])  if cols["name"]  < len(r) else ""
        email = _stringify(r[cols["email"]]) if cols["email"] < len(r) else ""
        phone = _stringify(r[cols["phone"]]) if cols["phone"] < len(r) else ""
        if not (name or email or phone):
            continue  # blank line
        if email and email.lower() in seen:
            continue
        if email:
            seen.add(email.lower())
        out.append({"name": name, "email": email, "phone": phone})
    return jsonify({"users": out, "count": len(out)})


@bp.route("/api/parceiro/util/parse-groupings-excel", methods=["POST"])
def util_parse_groupings_excel():
    """Parse a workbook with one row per (user, grouping) pair.

    Expected header row (case-insensitive, lowercased): userId, name, email,
    phone, groupingId. Optional columns: description, isOwn.

    Rows sharing a userId are aggregated into one association request: name,
    email, phone are taken from the FIRST occurrence of that userId; the
    remaining rows only contribute their grouping entry. The UI uses the
    aggregated structure to drive one PATCH per user.

    Missing `description` falls back to the local grouping name (resolved
    on the server using `get_grouping_index`). `isOwn` defaults to false;
    truthy values are 'true', '1', 'yes', 'sim', 'verdadeiro'.
    """
    wb, err = _open_uploaded_workbook()
    if err:
        return err

    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return jsonify({"users": [], "count": 0, "warnings": ["empty sheet"]})

    cols = {(_stringify(h).lower()): i for i, h in enumerate(header) if h is not None}
    required = ("userid", "name", "email", "phone", "groupingid")
    missing = [c for c in required if c not in cols]
    if missing:
        return jsonify({
            "error": f"missing columns: {', '.join(missing)}",
            "expectedHeader": ["userId", "name", "email", "phone", "groupingId",
                               "description", "isOwn"],
            "foundHeader":    [_stringify(h) for h in header],
        }), 400

    gindex = get_grouping_index()
    truthy = {"true", "1", "yes", "y", "sim", "verdadeiro", "x"}

    by_user: dict[str, dict] = {}
    order: list[str] = []
    for r in rows:
        if r is None:
            continue

        def cell(key):
            i = cols[key]
            return _stringify(r[i]) if i < len(r) else ""

        uid = cell("userid")
        gid = cell("groupingid")
        if not uid and not gid:
            continue  # blank line
        if not uid:
            continue  # row without user is unusable
        if not gid:
            continue  # row without grouping is unusable

        if uid not in by_user:
            by_user[uid] = {
                "userId":    uid,
                "name":      cell("name"),
                "email":     cell("email"),
                "phone":     cell("phone"),
                "groupings": [],
                "_seenGids": set(),
            }
            order.append(uid)

        bucket = by_user[uid]
        if gid in bucket["_seenGids"]:
            continue
        bucket["_seenGids"].add(gid)

        if "description" in cols:
            i = cols["description"]
            description = _stringify(r[i]) if i < len(r) else ""
        else:
            description = ""
        if not description:
            local = gindex.get(gid)
            if local:
                description = local.get("name") or ""

        is_own = False
        if "isown" in cols:
            i = cols["isown"]
            cellval = _stringify(r[i]).lower() if i < len(r) else ""
            is_own = cellval in truthy

        bucket["groupings"].append({
            "_id":         gid,
            "description": description,
            "isOwn":       is_own,
        })

    users = []
    for uid in order:
        b = by_user[uid]
        b.pop("_seenGids", None)
        users.append(b)
    return jsonify({"users": users, "count": len(users)})

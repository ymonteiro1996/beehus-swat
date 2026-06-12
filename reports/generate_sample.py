"""
Generate sample Excel file for Blue3 Wealth "Relatório Patrimonial de Revisão".
Each sheet = one MongoDB collection used to populate the report.
All data is fictitious.
"""

import os
import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Script sits in reports/; anchor the output path to the project-root data/ dir
# so running from any CWD writes to the same place.
_HERE = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_HERE)

# ── Styling ────────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
CATEGORY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CATEGORY_FONT = Font(name="Calibri", bold=True, size=10)
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D0D0D0"),
)
BRL_FMT = '#,##0.00'
PCT_FMT = '0.00%'
DATE_FMT = 'DD/MM/YYYY'

# ── Sample IDs (mimicking MongoDB ObjectIds) ───────────────────────────────────

CLIENT_ID = "67f4a1b2c3d4e5f607890abc"
CLIENT_NAME = "CARLOS EDUARDO MENDES"
COMPANY_ID = "10987654321000"
ENTITY_ID = "67c0a6b471f5e8c88f76044b"
WALLET_BRL = "68bb268b9a9a11e087ee53de"
WALLET_USD = "68bb268b9a9a11e087ee53df"
REFERENCE_MONTH = "2025-07"
REPORT_DATE = "2025-07-31"
EMISSION_DATE = "2025-08-15"
FX_RATE = 5.6021


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_body(ws, start_row, end_row, num_cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def apply_fmt(ws, col, start_row, end_row, fmt):
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=col).number_format = fmt


# ── Sheet 1: report_config ─────────────────────────────────────────────────────

def sheet_report_config(wb):
    ws = wb.create_sheet("report_config")
    headers = [
        "clientId", "clientName", "companyId", "entityId",
        "referenceMonth", "reportDate", "emissionDate",
        "reportType", "currencyId", "fxRate_BRL_USD",
        "walletId_BRL", "walletId_USD"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    row_mensal = [
        CLIENT_ID, CLIENT_NAME, COMPANY_ID, ENTITY_ID,
        REFERENCE_MONTH, REPORT_DATE, EMISSION_DATE,
        "mensal", "BRL", FX_RATE,
        WALLET_BRL, WALLET_USD
    ]
    row_tri = list(row_mensal)
    row_tri[7] = "trimestral"

    for c, v in enumerate(row_mensal, 1):
        ws.cell(row=2, column=c, value=v)
    for c, v in enumerate(row_tri, 1):
        ws.cell(row=3, column=c, value=v)

    style_body(ws, 2, 3, len(headers))
    auto_width(ws)


# ── Sheet 2: performance_summary ───────────────────────────────────────────────

def sheet_performance_summary(wb):
    ws = wb.create_sheet("performance_summary")
    headers = [
        "clientId", "referenceMonth", "competencia",
        "patrimonio_inicial", "movimentacoes", "ganho_financeiro",
        "cambio", "rebates", "taxa_gestao", "custos", "impostos",
        "patrimonio_final", "rentabilidade", "benchmark", "pct_benchmark"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    # Monthly row (Julho)
    pi = 4_850_000.00
    mov = 125_340.00
    ganho = 62_415.00
    cambio = -2_340.00
    rebates = 3_200.00
    taxa = -4_041.67
    custos = -185.00
    impostos = -8_922.00
    pf = pi + mov + ganho + cambio + rebates + taxa + custos + impostos
    rent = 0.0087
    bench = 0.0102
    pct_b = rent / bench

    row_jul = [
        CLIENT_ID, REFERENCE_MONTH, "Julho",
        pi, mov, ganho, cambio, rebates, taxa, custos, impostos,
        pf, rent, bench, pct_b
    ]
    for c, v in enumerate(row_jul, 1):
        ws.cell(row=2, column=c, value=v)

    # Year row
    pi_y = 4_200_000.00
    mov_y = 650_000.00
    ganho_y = 285_320.00
    cambio_y = -15_800.00
    rebates_y = 22_400.00
    taxa_y = -28_291.69
    custos_y = -1_295.00
    impostos_y = -62_480.00
    pf_y = pi_y + mov_y + ganho_y + cambio_y + rebates_y + taxa_y + custos_y + impostos_y
    rent_y = 0.0641
    bench_y = 0.0785
    pct_b_y = rent_y / bench_y

    row_yr = [
        CLIENT_ID, REFERENCE_MONTH, "2025",
        pi_y, mov_y, ganho_y, cambio_y, rebates_y, taxa_y, custos_y, impostos_y,
        pf_y, rent_y, bench_y, pct_b_y
    ]
    for c, v in enumerate(row_yr, 1):
        ws.cell(row=3, column=c, value=v)

    style_body(ws, 2, 3, len(headers))

    for col in [4, 5, 6, 7, 8, 9, 10, 11, 12]:
        apply_fmt(ws, col, 2, 3, BRL_FMT)
    for col in [13, 14, 15]:
        apply_fmt(ws, col, 2, 3, PCT_FMT)

    auto_width(ws)


# ── Sheet 3: references ────────────────────────────────────────────────────────

def sheet_references(wb):
    ws = wb.create_sheet("references")
    headers = ["clientId", "referenceMonth", "indicator", "mes", "ano", "12m", "24m", "inicio"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Portfolio",  0.0087, 0.0641, 0.1145, 0.2380, 0.1620],
        ["CDI",        0.0102, 0.0785, 0.1190, 0.2510, 0.1530],
        ["Ibovespa",   -0.0045, 0.1420, 0.0580, 0.1850, 0.1210],
        ["IPCA",       0.0038, 0.0290, 0.0485, 0.0910, 0.0650],
        ["Dolar",      0.0065, -0.0980, -0.0185, 0.0540, 0.0890],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        for c, v in enumerate(row_data, 3):
            ws.cell(row=r, column=c, value=v)
        for col in range(4, 9):
            ws.cell(row=r, column=col).number_format = PCT_FMT

    style_body(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


# ── Sheet 4: return_by_asset_class ─────────────────────────────────────────────

def sheet_return_by_asset_class(wb):
    ws = wb.create_sheet("return_by_asset_class")
    headers = ["clientId", "referenceMonth", "period", "asset_class", "return_pct"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    classes = ["Floating", "Short Duration", "Multimercado", "Long Duration",
               "Renda Variavel", "Alternativos", "Total"]
    returns_mes = [0.0098, 0.0045, 0.0210, -0.0115, -0.0032, 0.0145, 0.0087]
    returns_tri = [0.0310, 0.0135, 0.0620, -0.0085, 0.0210, 0.0380, 0.0350]
    returns_12m = [0.1120, 0.0890, 0.1450, 0.0620, 0.0850, 0.1080, 0.1145]

    r = 2
    for period, rets in [("mes", returns_mes), ("trimestre", returns_tri), ("12m", returns_12m)]:
        for cls, ret in zip(classes, rets):
            ws.cell(row=r, column=1, value=CLIENT_ID)
            ws.cell(row=r, column=2, value=REFERENCE_MONTH)
            ws.cell(row=r, column=3, value=period)
            ws.cell(row=r, column=4, value=cls)
            ws.cell(row=r, column=5, value=ret)
            ws.cell(row=r, column=5).number_format = PCT_FMT
            r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Sheet 5: allocation ────────────────────────────────────────────────────────

def sheet_allocation(wb):
    ws = wb.create_sheet("allocation")
    headers = ["clientId", "referenceMonth", "asset_class", "brasil_pct", "exterior_pct", "total_pct"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Fundos",                          0.0620, None,   0.0620],
        ["Renda Fixa",                      0.0380, None,   0.0380],
        ["Multimercado",                    0.0510, None,   0.0510],
        ["Pos-fixado (Floating)",           0.2750, None,   0.2750],
        ["Curta Duracao (Short Duration)",   0.1285, 0.0120, 0.1405],
        ["Multimercados (Hedge Funds)",     0.0830, None,   0.0830],
        ["Longa Duracao (Long Duration)",    0.1720, 0.0480, 0.2200],
        ["Renda Variavel (Equities)",       0.1050, None,   0.1050],
        ["Alternativos (Alternatives)",     0.0185, None,   0.0185],
        ["Caixa e Provisoes",               0.0540, 0.9200, 0.0070],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        ws.cell(row=r, column=3, value=row_data[0])
        for c, v in enumerate(row_data[1:], 4):
            ws.cell(row=r, column=c, value=v)
            if v is not None:
                ws.cell(row=r, column=c).number_format = PCT_FMT

    style_body(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


# ── Sheet 6: allocation_by_institution ─────────────────────────────────────────

def sheet_allocation_by_institution(wb):
    ws = wb.create_sheet("alloc_by_institution")
    headers = ["clientId", "referenceMonth", "institution", "saldo_bruto_BRL", "pct"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["BTG Pactual",       3_612_450.80, 0.7240],
        ["XP Investimentos",    910_325.15, 0.1825],
        ["Avenue Securities",   327_077.55, 0.0655],
        ["Itau Private",        140_000.00, 0.0280],
    ]
    total = sum(d[1] for d in data)
    data.append(["Total", total, 1.0000])

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        ws.cell(row=r, column=3, value=row_data[0])
        ws.cell(row=r, column=4, value=row_data[1])
        ws.cell(row=r, column=4).number_format = BRL_FMT
        ws.cell(row=r, column=5, value=row_data[2])
        ws.cell(row=r, column=5).number_format = PCT_FMT

    # Bold total row
    last = len(data) + 1
    for c in range(1, 6):
        ws.cell(row=last, column=c).font = Font(name="Calibri", bold=True, size=10)

    style_body(ws, 2, last, len(headers))
    auto_width(ws)


# ── Sheet 7: geographic_exposure ───────────────────────────────────────────────

def sheet_geographic_exposure(wb):
    ws = wb.create_sheet("geographic_exposure")
    headers = ["clientId", "referenceMonth", "region", "pct"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Brasil",  0.72],
        ["EUA",     0.18],
        ["Europa",  0.05],
        ["Asia",    0.02],
        ["Outros",  0.03],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        ws.cell(row=r, column=3, value=row_data[0])
        ws.cell(row=r, column=4, value=row_data[1])
        ws.cell(row=r, column=4).number_format = PCT_FMT

    style_body(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


# ── Sheet 8: transactions (movimentacoes) ──────────────────────────────────────

def sheet_transactions(wb):
    ws = wb.create_sheet("transactions")
    headers = [
        "clientId", "referenceMonth", "direction", "date",
        "asset_class", "product", "value_BRL"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    entradas = [
        ["2025-06-10", "Floating",       "TREND DI FIC FI RF",                            150_000.00],
        ["2025-06-10", "Short Duration",  "CDB BTG PACTUAL PRE 12.10% 15/03/2027",        350_000.00],
        ["2025-06-15", "Short Duration",  "LCA BANCO SAFRA CDI+0.8% 18/12/2026",          200_000.00],
        ["2025-06-22", "Long Duration",   "DEB EQUATORIAL IPCA+2.50% 15/06/2045",          85_000.00],
        ["2025-06-22", "Short Duration",  "CDB BTG PACTUAL PRE 12.10% 15/03/2027",        180_000.00],
    ]
    saidas = [
        ["2025-06-05", "Floating",       "KADIMA FIC FI RF CP LP",                        100_000.00],
        ["2025-06-12", "Short Duration",  "CDB BANCO PINE PRE 11.80% 20/09/2026",         250_000.00],
        ["2025-06-18", "Long Duration",   "NTN-B IPCA+5.80% 15/08/2035",                  120_000.00],
        ["2025-06-25", "Renda Variavel",  "VALE3 - VALE S.A.",                              75_340.00],
    ]

    r = 2
    for d, cls, prod, val in entradas:
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        ws.cell(row=r, column=3, value="entrada")
        ws.cell(row=r, column=4, value=d)
        ws.cell(row=r, column=5, value=cls)
        ws.cell(row=r, column=6, value=prod)
        ws.cell(row=r, column=7, value=val)
        ws.cell(row=r, column=7).number_format = BRL_FMT
        r += 1

    for d, cls, prod, val in saidas:
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        ws.cell(row=r, column=3, value="saida")
        ws.cell(row=r, column=4, value=d)
        ws.cell(row=r, column=5, value=cls)
        ws.cell(row=r, column=6, value=prod)
        ws.cell(row=r, column=7, value=val)
        ws.cell(row=r, column=7).number_format = BRL_FMT
        r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Sheet 9: monthly_flow (aplicacoes e resgates) ─────────────────────────────

def sheet_monthly_flow(wb):
    ws = wb.create_sheet("monthly_flow")
    headers = ["clientId", "referenceMonth", "mes_ano", "movimentacoes", "saldo_bruto"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["jan/2025",  1_230_500.00, 4_200_000.00],
        ["fev/2025",   -45_200.00,  4_385_300.00],
        ["mar/2025",   -18_750.00,  4_421_850.00],
        ["abr/2025",   850_000.00,  4_556_100.00],
        ["mai/2025",   -62_300.00,  4_694_800.00],
        ["jun/2025",   125_340.00,  4_850_000.00],
        ["jul/2025",   -22_034.40,  5_049_853.50],
    ]
    acum = sum(d[1] for d in data)
    data.append(["acum/2025", acum, 5_049_853.50])

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        ws.cell(row=r, column=3, value=row_data[0])
        ws.cell(row=r, column=4, value=row_data[1])
        ws.cell(row=r, column=4).number_format = BRL_FMT
        ws.cell(row=r, column=5, value=row_data[2])
        ws.cell(row=r, column=5).number_format = BRL_FMT

    style_body(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


# ── Sheet 10: period_flow (movimentacoes no periodo) ──────────────────────────

def sheet_period_flow(wb):
    ws = wb.create_sheet("period_flow")
    headers = ["clientId", "referenceMonth", "item", "value_BRL"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Saldo Inicial Periodo",  4_200_000.00],
        ["Movimentacoes",           -22_034.40],
        ["Impostos Pagos",        -140_020.18],
        ["Rendimento Bruto",       4_461_936.92],
        ["Saldo Final Periodo",    5_049_853.50],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        ws.cell(row=r, column=3, value=row_data[0])
        ws.cell(row=r, column=4, value=row_data[1])
        ws.cell(row=r, column=4).number_format = BRL_FMT

    style_body(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


# ── Sheet 11: risk_statistics ──────────────────────────────────────────────────

def sheet_risk_statistics(wb):
    ws = wb.create_sheet("risk_statistics")
    headers = [
        "clientId", "referenceMonth",
        "meses_negativos", "meses_positivos",
        "rentabilidade_mensal_maxima", "rentabilidade_mensal_minima",
        "meses_acima_cdi", "meses_abaixo_cdi",
        "volatilidade_12m", "volatilidade_inicio",
        "sharpe",
        "var_12m", "max_drawdown", "tempo_recuperacao_dias"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    row_data = [
        CLIENT_ID, REFERENCE_MONTH,
        2, 5,
        0.0125, 0.0032,
        4, 3,
        0.0285, 0.0310,
        1.45,
        0.8520, 0.0945, 195
    ]
    for c, v in enumerate(row_data, 1):
        ws.cell(row=2, column=c, value=v)

    for col in [5, 6, 9, 10, 12, 13]:
        ws.cell(row=2, column=col).number_format = PCT_FMT

    style_body(ws, 2, 2, len(headers))
    auto_width(ws)


# ── Sheet 12: risk_contribution_by_class ───────────────────────────────────────

def sheet_risk_contribution_by_class(wb):
    ws = wb.create_sheet("risk_contribution")
    headers = [
        "clientId", "referenceMonth", "asset_class",
        "peso_pct", "volatilidade", "corr_portfolio",
        "mrc_aa", "rc_absoluto_aa", "rc_pct"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Acoes Brasil",    0.35, 0.22, 0.82, 0.1600, 0.0560, 0.55],
        ["Credito Privado", 0.28, 0.08, 0.35, 0.0550, 0.0154, 0.15],
        ["FII",             0.22, 0.14, 0.45, 0.1350, 0.0297, 0.29],
        ["Caixa (risk-free)", 0.15, 0.02, 0.00, 0.0000, 0.0000, 0.00],
    ]
    total_rc_abs = sum(d[6] for d in data)
    data.append(["Total / Check", 1.00, None, None, None, total_rc_abs, 1.00])

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        ws.cell(row=r, column=3, value=row_data[0])
        for c, v in enumerate(row_data[1:], 4):
            ws.cell(row=r, column=c, value=v)

    for col in [4, 5, 7, 8, 9]:
        apply_fmt(ws, col, 2, len(data) + 1, PCT_FMT)

    style_body(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


# ── Sheet 13: liquidity ────────────────────────────────────────────────────────

def sheet_liquidity(wb):
    ws = wb.create_sheet("liquidity")
    headers = ["clientId", "referenceMonth", "time_bucket", "pct_portfolio"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Diaria",         0.18],
        ["Ate 1 Mes",      0.22],
        ["Ate 6 meses",    0.12],
        ["Ate 1 ano",      0.08],
        ["Ate 2 anos",     0.05],
        ["Acima de 2 anos", 0.35],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        ws.cell(row=r, column=3, value=row_data[0])
        ws.cell(row=r, column=4, value=row_data[1])
        ws.cell(row=r, column=4).number_format = PCT_FMT

    style_body(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


# ── Sheet 14: liquidity_profiling ──────────────────────────────────────────────

def sheet_liquidity_profiling(wb):
    ws = wb.create_sheet("liquidity_profiling")
    headers = [
        "clientId", "referenceMonth", "asset_class",
        "asset_class_alloc_pct", "investment_alloc_pct",
        "investment_vehicle",
        "highly_liquid_pct", "liquid_pct", "semi_liquid_pct", "illiquid_pct"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Cash",                     0.02, 0.02, "Separate Account", 1.00, 0.00, 0.00, 0.00],
        ["Fixed Income",             0.15, 0.06, "Separate Account", 1.00, 0.00, 0.00, 0.00],
        ["Fixed Income",             0.15, 0.07, "Commingled Fund",  1.00, 0.00, 0.00, 0.00],
        ["Fixed Income",             0.15, 0.02, "Futures",          1.00, 0.00, 0.00, 0.00],
        ["Domestic Equity",          0.20, 0.10, "Commingled Fund",  0.00, 0.50, 0.50, 0.00],
        ["Domestic Equity",          0.20, 0.08, "Separate Account", 0.00, 1.00, 0.00, 0.00],
        ["Domestic Equity",          0.20, 0.02, "Futures",          1.00, 0.00, 0.00, 0.00],
        ["Intl Developed Equity",    0.12, 0.08, "Commingled Fund",  0.00, 0.50, 0.30, 0.20],
        ["Intl Developed Equity",    0.12, 0.04, "Separate Account", 0.00, 0.80, 0.20, 0.00],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        ws.cell(row=r, column=3, value=row_data[0])
        for c, v in enumerate(row_data[1:], 4):
            ws.cell(row=r, column=c, value=v)

    pct_cols = [4, 5, 7, 8, 9, 10]
    for col in pct_cols:
        apply_fmt(ws, col, 2, len(data) + 1, PCT_FMT)

    style_body(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


# ── Sheet 15: liquidity_budget ─────────────────────────────────────────────────

def sheet_liquidity_budget(wb):
    ws = wb.create_sheet("liquidity_budget")
    headers = ["clientId", "referenceMonth", "time_to_cash", "liquidity_classification", "budget_pct"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["< 1 Week",    "Highly Liquid", "At Least 10%"],
        ["< 1 Quarter", "Liquid",        "At Least 35%"],
        ["< 1 Year",    "Semi-Liquid",   "At Least 50%"],
        ["> 1 Year",    "Illiquid",      "Up to 50%"],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        for c, v in enumerate(row_data, 3):
            ws.cell(row=r, column=c, value=v)

    style_body(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


# ── Sheet 16: asset_detail (ativos detalhados) ────────────────────────────────

def sheet_asset_detail(wb):
    ws = wb.create_sheet("asset_detail")
    headers = [
        "clientId", "referenceMonth", "category", "asset_name",
        "saldo_bruto", "ir_iof", "saldo_liquido", "pct_aloc",
        "rent_mes_pct", "rent_mes_cdi_pct",
        "rent_ano_pct", "rent_ano_cdi_pct",
        "rent_inicio_data", "rent_inicio_pct", "rent_inicio_cdi_pct"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    assets = [
        # category, name, saldo_bruto, ir_iof, saldo_liq, %aloc, r_mes%, r_mes_cdi%, r_ano%, r_ano_cdi%, data_inicio, r_inicio%, r_inicio_cdi%
        ["SD High Grade", "LF BTG PACTUAL IPCA+5.20% 23/06/2026",
         6_280_450.20, 385_120.50, 5_895_329.70, 0.0165, 0.0058, 0.4820, 0.0735, 0.7910, "2023-08-31", 0.3850, 1.5240],
        ["SD High Grade", "CRI REDE D'OR IPCA+4.50% 15/10/2031",
         5_120_340.00, None, 5_120_340.00, 0.0135, 0.0042, 0.3280, 0.0685, 0.7410, "2023-08-31", 0.2010, 0.8120],
        ["SD High Grade", "LF BTG PACTUAL 11.50% PRE 21/12/2026",
         1_340_200.00, 64_500.00, 1_275_700.00, 0.0035, 0.0092, 0.7820, 0.0745, 0.8120, "2023-08-31", 0.2380, 0.9350],

        ["LD High Grade", "DEB LOCALIZA IPCA+5.60% 15/03/2031",
         8_420_600.00, 310_200.50, 8_110_399.50, 0.0225, 0.0050, 0.4120, 0.0750, 0.8250, "2023-08-31", 0.2280, 0.8920],
        ["LD High Grade", "DEB SAO MARTINHO IPCA+6.20% 15/01/2037",
         7_015_400.00, 162_800.00, 6_852_600.00, 0.0190, 0.0054, 0.4510, 0.0790, 0.8650, "2023-08-31", 0.2420, 0.9450],
        ["LD High Grade", "LF ITAU IPCA+5.40% 13/08/2031",
         5_180_320.00, 280_450.00, 4_899_870.00, 0.0140, 0.0070, 0.5940, 0.0780, 0.8520, "2023-08-31", 0.2250, 0.8810],
        ["LD High Grade", "CDB BANCO CONTINENTAL 14.80% PRE 16/11/2027",
         5_120_000.00, 238_400.00, 4_881_600.00, 0.0138, 0.0118, 0.9820, 0.0960, 1.0450, "2023-08-31", 0.3140, 1.2280],
        ["LD High Grade", "CRA M. DIAS BRANCO IPCA+4.20% 15/03/2031",
         4_950_200.00, 186_200.00, 4_764_000.00, 0.0134, 0.0061, 0.5120, 0.0700, 0.7650, "2023-08-31", 0.1970, 0.7710],
        ["LD High Grade", "DEB SABESP IPCA+5.90% 15/12/2031",
         4_310_500.00, 118_300.00, 4_192_200.00, 0.0118, 0.0050, 0.4250, 0.0770, 0.8410, "2023-08-31", 0.2350, 0.9200],
        ["LD High Grade", "NTN-B IPCA+5.20% 15/08/2030",
         2_070_400.00, 21_600.00, 2_048_800.00, 0.0056, 0.0048, 0.3940, 0.0725, 0.7920, "2023-08-31", 0.2200, 0.8610],

        ["Public Equity", "PETR4 - PETROLEO BRASILEIRO S.A.",
         142_500.00, None, 142_500.00, 0.0004, 0.2250, 2.4500, 0.0650, 0.5380, "2024-07-26", -0.8520, None],
        ["Public Equity", "ITUB4 - ITAU UNIBANCO",
         118_200.00, None, 118_200.00, 0.0003, 0.0320, 0.2810, 0.0480, 0.3950, "2024-07-26", -0.1250, None],

        ["Overnight", "LFT SELIC+0.14 01/03/2028",
         88_420_500.00, 3_340_200.00, 85_080_300.00, 0.2380, 0.0122, 1.0420, 0.0950, 1.0350, "2024-08-30", 0.1380, 1.0480],
        ["Overnight", "LFT SELIC+0.15 01/03/2028",
         88_200_000.00, 3_318_500.00, 84_881_500.00, 0.2375, 0.0122, 1.0420, 0.0950, 1.0350, "2024-08-30", 0.1380, 1.0480],
        ["Overnight", "CDB BTG 102.00% CDI 25/09/2025",
         2_800_400.00, 96_500.00, 2_703_900.00, 0.0075, 0.0124, 1.0580, 0.0955, 1.0420, "2024-05-29", 0.1680, 1.0580],
        ["Overnight", "TREND CASH",
         312_000.00, 2_600.00, 309_400.00, 0.0009, 0.0120, 1.0180, 0.0930, 1.0150, "2024-04-16", 0.1750, 1.0120],

        ["FL High Grade", "BRADESCO FIC FIM CP UPPER",
         45_520_300.00, 245_800.00, 45_274_500.00, 0.1265, 0.0122, 1.0380, 0.0985, 1.0780, "2023-08-31", 0.2940, 1.1510],
    ]

    r = 2
    for row_data in assets:
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        for c, v in enumerate(row_data, 3):
            ws.cell(row=r, column=c, value=v)
        # Formats
        for col in [5, 6, 7]:
            ws.cell(row=r, column=col).number_format = BRL_FMT
        for col in [8, 9, 10, 11, 12, 14, 15]:
            cell = ws.cell(row=r, column=col)
            if cell.value is not None:
                cell.number_format = PCT_FMT
        r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Sheet 17: portfolio_timeseries ─────────────────────────────────────────────

def sheet_portfolio_timeseries(wb):
    ws = wb.create_sheet("portfolio_timeseries")
    headers = ["clientId", "date", "portfolio_value", "benchmark_value"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    # Generate monthly points from Jan 2024 to Jul 2025
    base_port = 1000.0
    base_bench = 1000.0
    start = date(2024, 1, 1)
    r = 2
    for m in range(19):  # 19 months
        d = date(2024 + (m // 12), ((m % 12) + 1), 1) - timedelta(days=1)
        if d > date(2025, 7, 31):
            break
        d_str = d.strftime("%Y-%m-%d")
        port_ret = random.uniform(0.004, 0.015) * (1 if random.random() > 0.15 else -1)
        bench_ret = random.uniform(0.006, 0.012)
        base_port *= (1 + port_ret)
        base_bench *= (1 + bench_ret)
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=d_str)
        ws.cell(row=r, column=3, value=round(base_port, 2))
        ws.cell(row=r, column=4, value=round(base_bench, 2))
        r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Sheet 18: historical_performance (trimestral table) ────────────────────────

def sheet_historical_performance(wb):
    ws = wb.create_sheet("historical_performance")
    headers = [
        "clientId", "referenceMonth", "indicator",
        "mes", "ano", "12m", "24m", "inicio"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Rentabilidade",  0.0087,  0.0641,  0.1145,  0.2380, 0.1620],
        ["% CDI",          0.8529,  0.8166,  0.9622,  0.9483, 1.0588],
        ["CDI",            0.0102,  0.0785,  0.1190,  0.2510, 0.1530],
        ["Dolar (PTAX)",  -0.0165, -0.1210, -0.0285,  0.0540, 0.0482],
        ["IPCA",           0.0038,  0.0290,  0.0485,  0.0910, 0.0650],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=CLIENT_ID)
        ws.cell(row=r, column=2, value=REFERENCE_MONTH)
        for c, v in enumerate(row_data, 3):
            ws.cell(row=r, column=c, value=v)
        for col in range(4, 9):
            ws.cell(row=r, column=col).number_format = PCT_FMT

    style_body(ws, 2, len(data) + 1, len(headers))
    auto_width(ws)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    random.seed(42)
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    sheet_report_config(wb)
    sheet_performance_summary(wb)
    sheet_references(wb)
    sheet_return_by_asset_class(wb)
    sheet_allocation(wb)
    sheet_allocation_by_institution(wb)
    sheet_geographic_exposure(wb)
    sheet_transactions(wb)
    sheet_monthly_flow(wb)
    sheet_period_flow(wb)
    sheet_risk_statistics(wb)
    sheet_risk_contribution_by_class(wb)
    sheet_liquidity(wb)
    sheet_liquidity_profiling(wb)
    sheet_liquidity_budget(wb)
    sheet_asset_detail(wb)
    sheet_portfolio_timeseries(wb)
    sheet_historical_performance(wb)

    output_path = os.path.join(_PROJECT_ROOT, "data", "sample_relatorio_patrimonial.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Sample Excel saved to: {output_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()

"""
Generate report Excel from real MongoDB data for 4 wallets.
Each sheet = one MongoDB collection used to populate the report.
Reference month: 2025-12 (latest available data).
"""

import os
import sys
import statistics

# Script sits in reports/; make the project root importable so `from db import db`
# resolves regardless of the caller's CWD.
_HERE = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_HERE)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from db import db
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── NAV-series helpers ─────────────────────────────────────────────────────────
# The navPackages collection stores `positionDate` sometimes as an ISO string and
# sometimes as a BSON Date (pipeline-vintage drift). These helpers normalize to
# "YYYY-MM"/"YYYY-MM-DD" strings and resample a possibly-daily series to
# one entry per month so that `rentabilidade_mes`, `volatilidade_anualizada`,
# and `meses_positivos` operate on true monthly returns — not daily data.

def _to_ym(d):
    """Normalize a positionDate (str or datetime) to 'YYYY-MM'."""
    if d is None:
        return ""
    if hasattr(d, "strftime"):
        return d.strftime("%Y-%m")
    return str(d)[:7]


def _month_end_navs(nav_list):
    """Return one entry per month (the last seen for each YYYY-MM), sorted
    ascending by month key. Survives BSON mixed-type sort (where Strings sort
    before Dates) because we re-sort by our own string key."""
    by_month = {}
    for n in nav_list:
        k = _to_ym(n.get("positionDate"))
        if k:
            by_month[k] = n
    return [by_month[k] for k in sorted(by_month.keys())]


def _prev_month_end(month_navs, ref_month):
    """Last month-end NAV strictly before `ref_month` (YYYY-MM)."""
    for n in reversed(month_navs):
        if _to_ym(n.get("positionDate")) < ref_month:
            return n
    return None


def _prev_year_end(month_navs, ref_year):
    """Last month-end NAV whose month is <= (ref_year-1)-12."""
    target = f"{int(ref_year) - 1}-12"
    for n in reversed(month_navs):
        ym = _to_ym(n.get("positionDate"))
        if ym and ym <= target:
            return n
    return None


def _ret(cur, base):
    """Safe return: None if either endpoint missing or base is 0."""
    if cur is None or not base:
        return None
    return (cur / base) - 1


def _monthly_returns(month_navs):
    """Compute month-over-month returns from a resampled month-end NAV series."""
    out = []
    for i in range(1, len(month_navs)):
        prev_nps = month_navs[i - 1].get("navPerShare")
        cur_nps  = month_navs[i].get("navPerShare")
        r = _ret(cur_nps, prev_nps)
        if r is not None:
            out.append(r)
    return out

# ── Config ─────────────────────────────────────────────────────────────────────

WALLET_IDS = [
    "68cabc8f62a3c4412e8d2121",
    "68cabd0162a3c4412e8d2145",
    "6931fe0c3cb026c5483de086",
    "68cb072562a3c4412e8d5741",
]
REFERENCE_MONTH = "2025-12"
POSITION_DATE = "2025-12-31"
COMPANY_ID = "55555555555555"

# ── Styling ────────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
CATEGORY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CATEGORY_FONT = Font(name="Calibri", bold=True, size=10)
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(bottom=Side(style="thin", color="D0D0D0"))
BRL_FMT = '#,##0.00'
PCT_FMT = '0.00%'


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_body(ws, start_row, end_row, num_cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def apply_fmt(ws, col, start_row, end_row, fmt):
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=col).number_format = fmt


def style_category_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = CATEGORY_FONT
        cell.fill = CATEGORY_FILL


# ── Data loading ───────────────────────────────────────────────────────────────

def load_wallets():
    wallets = {}
    for wid in WALLET_IDS:
        w = db.wallets.find_one({"_id": ObjectId(wid)})
        if w:
            wallets[wid] = w
    return wallets


def load_positions():
    positions = {}
    for wid in WALLET_IDS:
        pos = db.processedPosition.find_one(
            {"walletId": wid, "positionDate": POSITION_DATE},
            {"_id": 0}
        )
        if pos:
            positions[wid] = pos
    return positions


def load_nav_history():
    navs = {}
    for wid in WALLET_IDS:
        nav_list = list(db.navPackages.find(
            {"walletId": wid, "trashed": {"$ne": True}},
            {"_id": 0, "positionDate": 1, "nav": 1, "navPerShare": 1, "returnNavPerShare": 1}
        ).sort("positionDate", 1))
        navs[wid] = nav_list
    return navs


def load_transactions():
    txns = {}
    for wid in WALLET_IDS:
        txn_list = list(db.transactions.find(
            {"walletId": wid},
            {"_id": 0, "operationDate": 1, "liquidationDate": 1, "balance": 1,
             "description": 1, "beehusTransactionType": 1}
        ).sort("liquidationDate", -1))
        txns[wid] = txn_list
    return txns


def load_entity_name(entity_id):
    ent = db.entities.find_one({"_id": ObjectId(entity_id)})
    return ent.get("name", str(entity_id)) if ent else str(entity_id)


# ── Sheet 1: report_config ─────────────────────────────────────────────────────

def sheet_report_config(wb, wallets):
    ws = wb.create_sheet("report_config")
    headers = [
        "walletId", "walletName", "companyId", "companyName",
        "entityId", "entityName", "currency",
        "referenceMonth", "reportDate", "reportType",
        "startDateConsolidation", "startDateReturn"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    r = 2
    for wid, w in wallets.items():
        entity_name = load_entity_name(w.get("entityId", ""))
        ws.cell(row=r, column=1, value=wid)
        ws.cell(row=r, column=2, value=w.get("name", ""))
        ws.cell(row=r, column=3, value=w.get("companyId", ""))
        ws.cell(row=r, column=4, value="One Wealth")
        ws.cell(row=r, column=5, value=str(w.get("entityId", "")))
        ws.cell(row=r, column=6, value=entity_name)
        ws.cell(row=r, column=7, value=w.get("currency", ""))
        ws.cell(row=r, column=8, value=REFERENCE_MONTH)
        ws.cell(row=r, column=9, value=POSITION_DATE)
        ws.cell(row=r, column=10, value="mensal")
        ws.cell(row=r, column=11, value=w.get("startDateConsolidation", ""))
        ws.cell(row=r, column=12, value=w.get("startDateReturn", ""))
        r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Sheet 2: performance_summary ───────────────────────────────────────────────

def sheet_performance_summary(wb, wallets, navs):
    ws = wb.create_sheet("performance_summary")
    headers = [
        "walletId", "walletName", "currency", "referenceMonth",
        "patrimonio_inicial", "patrimonio_final",
        "rentabilidade_mes", "rentabilidade_ano", "rentabilidade_inicio"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    r = 2
    ref_year = REFERENCE_MONTH[:4]
    for wid, w in wallets.items():
        month_navs = _month_end_navs(navs.get(wid, []))
        if not month_navs:
            continue

        # Resolve anchors against the reference month — not the prior daily row.
        latest     = month_navs[-1]
        first      = month_navs[0]
        prev_month = _prev_month_end(month_navs, REFERENCE_MONTH)
        year_end   = _prev_year_end(month_navs, ref_year)

        nps_latest = latest.get("navPerShare")
        nps_prev   = prev_month.get("navPerShare") if prev_month else None
        nps_year   = year_end.get("navPerShare")   if year_end   else None
        nps_first  = first.get("navPerShare")

        patrimonio_inicial = prev_month["nav"] if prev_month else None
        patrimonio_final   = latest["nav"]
        rent_mes    = _ret(nps_latest, nps_prev)
        rent_ano    = _ret(nps_latest, nps_year)
        rent_inicio = _ret(nps_latest, nps_first)

        ws.cell(row=r, column=1, value=wid)
        ws.cell(row=r, column=2, value=w.get("name", ""))
        ws.cell(row=r, column=3, value=w.get("currency", ""))
        ws.cell(row=r, column=4, value=REFERENCE_MONTH)
        ws.cell(row=r, column=5, value=patrimonio_inicial)
        ws.cell(row=r, column=6, value=patrimonio_final)
        ws.cell(row=r, column=7, value=rent_mes)
        ws.cell(row=r, column=8, value=rent_ano)
        ws.cell(row=r, column=9, value=rent_inicio)

        for col in [5, 6]:
            ws.cell(row=r, column=col).number_format = BRL_FMT
        for col in [7, 8, 9]:
            ws.cell(row=r, column=col).number_format = PCT_FMT
        r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Sheet 3: nav_timeseries ────────────────────────────────────────────────────

def sheet_nav_timeseries(wb, wallets, navs):
    ws = wb.create_sheet("nav_timeseries")
    headers = [
        "walletId", "walletName", "currency", "positionDate",
        "nav", "navPerShare", "returnNavPerShare"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    r = 2
    for wid, w in wallets.items():
        for n in navs.get(wid, []):
            ws.cell(row=r, column=1, value=wid)
            ws.cell(row=r, column=2, value=w.get("name", ""))
            ws.cell(row=r, column=3, value=w.get("currency", ""))
            ws.cell(row=r, column=4, value=n["positionDate"])
            ws.cell(row=r, column=5, value=n["nav"])
            ws.cell(row=r, column=6, value=n.get("navPerShare", 0))
            ws.cell(row=r, column=7, value=n.get("returnNavPerShare", 0))
            ws.cell(row=r, column=5).number_format = BRL_FMT
            ws.cell(row=r, column=6).number_format = '0.000000'
            ws.cell(row=r, column=7).number_format = PCT_FMT
            r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Sheet 4: asset_detail ─────────────────────────────────────────────────────

def sheet_asset_detail(wb, wallets, positions):
    ws = wb.create_sheet("asset_detail")
    headers = [
        "walletId", "walletName", "currency", "positionDate",
        "category_var1", "category_var2", "beehusName",
        "quantity", "pu", "saldo_bruto",
        "formerPu", "formerQuantity",
        "dailyContribution", "totalContribution",
        "rent_mes_contribution"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    r = 2
    for wid, w in wallets.items():
        pos = positions.get(wid)
        if not pos:
            continue
        secs = pos.get("securities", [])
        # Sort by hierarchicalVariable.variable1 then balance desc
        secs_sorted = sorted(secs, key=lambda s: (
            s.get("hierarchicalVariable", {}).get("variable1", "ZZZ"),
            -(s.get("quantity", 0) * s.get("pu", 0))
        ))

        current_cat = None
        for s in secs_sorted:
            var1 = s.get("hierarchicalVariable", {}).get("variable1", "")
            var2 = s.get("hierarchicalVariable", {}).get("variable2", "")
            balance = s.get("quantity", 0) * s.get("pu", 0)

            # Category header row
            cat_key = var1
            if cat_key != current_cat:
                current_cat = cat_key
                ws.cell(row=r, column=1, value=wid)
                ws.cell(row=r, column=5, value=var1)
                style_category_row(ws, r, len(headers))
                r += 1

            ws.cell(row=r, column=1, value=wid)
            ws.cell(row=r, column=2, value=w.get("name", ""))
            ws.cell(row=r, column=3, value=w.get("currency", ""))
            ws.cell(row=r, column=4, value=pos.get("positionDate", ""))
            ws.cell(row=r, column=5, value=var1)
            ws.cell(row=r, column=6, value=var2)
            ws.cell(row=r, column=7, value=s.get("beehusName", ""))
            ws.cell(row=r, column=8, value=s.get("quantity", 0))
            ws.cell(row=r, column=9, value=s.get("pu", 0))
            ws.cell(row=r, column=10, value=balance)
            ws.cell(row=r, column=11, value=s.get("formerPu", 0))
            ws.cell(row=r, column=12, value=s.get("formerQuantity", 0))
            ws.cell(row=r, column=13, value=s.get("dailyContribution", 0))
            ws.cell(row=r, column=14, value=s.get("totalContribution", 0))
            # Monthly contribution = totalContribution (from the position)
            ws.cell(row=r, column=15, value=s.get("totalContribution", 0))

            for col in [8, 9, 10, 11, 12, 13, 14, 15]:
                ws.cell(row=r, column=col).number_format = BRL_FMT

            r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Sheet 5: allocation ────────────────────────────────────────────────────────

def sheet_allocation(wb, wallets, positions):
    ws = wb.create_sheet("allocation")
    headers = [
        "walletId", "walletName", "currency", "referenceMonth",
        "asset_class_var1", "asset_class_var2", "total_balance", "pct_allocation"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    r = 2
    for wid, w in wallets.items():
        pos = positions.get(wid)
        if not pos:
            continue
        secs = pos.get("securities", [])

        # Group by var1/var2
        alloc = {}
        total = 0
        for s in secs:
            var1 = s.get("hierarchicalVariable", {}).get("variable1", "Outros")
            var2 = s.get("hierarchicalVariable", {}).get("variable2", "") or ""
            balance = s.get("quantity", 0) * s.get("pu", 0)
            key = (var1, var2)
            alloc[key] = alloc.get(key, 0) + balance
            total += balance

        for (var1, var2), bal in sorted(alloc.items(), key=lambda x: -x[1]):
            pct = bal / total if total > 0 else 0
            ws.cell(row=r, column=1, value=wid)
            ws.cell(row=r, column=2, value=w.get("name", ""))
            ws.cell(row=r, column=3, value=w.get("currency", ""))
            ws.cell(row=r, column=4, value=REFERENCE_MONTH)
            ws.cell(row=r, column=5, value=var1)
            ws.cell(row=r, column=6, value=var2)
            ws.cell(row=r, column=7, value=bal)
            ws.cell(row=r, column=8, value=pct)
            ws.cell(row=r, column=7).number_format = BRL_FMT
            ws.cell(row=r, column=8).number_format = PCT_FMT
            r += 1

        # Total row
        ws.cell(row=r, column=1, value=wid)
        ws.cell(row=r, column=2, value=w.get("name", ""))
        ws.cell(row=r, column=5, value="TOTAL")
        ws.cell(row=r, column=7, value=total)
        ws.cell(row=r, column=8, value=1.0)
        ws.cell(row=r, column=7).number_format = BRL_FMT
        ws.cell(row=r, column=8).number_format = PCT_FMT
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).font = Font(name="Calibri", bold=True, size=10)
        r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Sheet 6: allocation_consolidated ───────────────────────────────────────────

def sheet_allocation_consolidated(wb, wallets, positions):
    """Consolidated allocation across all wallets (for the report overview)."""
    ws = wb.create_sheet("alloc_consolidated")
    headers = [
        "referenceMonth", "asset_class_var1",
        "total_BRL", "total_USD", "total_all_BRL",
        "pct_total"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    fx_rate = 6.19  # approximate BRL/USD for 2025-12

    alloc_brl = {}
    alloc_usd = {}
    for wid, w in wallets.items():
        pos = positions.get(wid)
        if not pos:
            continue
        currency = w.get("currency", "BRL")
        secs = pos.get("securities", [])
        for s in secs:
            var1 = s.get("hierarchicalVariable", {}).get("variable1", "Outros")
            balance = s.get("quantity", 0) * s.get("pu", 0)
            if currency == "BRL":
                alloc_brl[var1] = alloc_brl.get(var1, 0) + balance
            else:
                alloc_usd[var1] = alloc_usd.get(var1, 0) + balance

    all_classes = sorted(set(list(alloc_brl.keys()) + list(alloc_usd.keys())))
    grand_total = sum(alloc_brl.values()) + sum(v * fx_rate for v in alloc_usd.values())

    r = 2
    for cls in all_classes:
        brl_val = alloc_brl.get(cls, 0)
        usd_val = alloc_usd.get(cls, 0)
        total_brl = brl_val + (usd_val * fx_rate)
        pct = total_brl / grand_total if grand_total > 0 else 0

        ws.cell(row=r, column=1, value=REFERENCE_MONTH)
        ws.cell(row=r, column=2, value=cls)
        ws.cell(row=r, column=3, value=brl_val)
        ws.cell(row=r, column=4, value=usd_val)
        ws.cell(row=r, column=5, value=total_brl)
        ws.cell(row=r, column=6, value=pct)

        ws.cell(row=r, column=3).number_format = BRL_FMT
        ws.cell(row=r, column=4).number_format = BRL_FMT
        ws.cell(row=r, column=5).number_format = BRL_FMT
        ws.cell(row=r, column=6).number_format = PCT_FMT
        r += 1

    # Grand total
    ws.cell(row=r, column=1, value=REFERENCE_MONTH)
    ws.cell(row=r, column=2, value="TOTAL")
    ws.cell(row=r, column=3, value=sum(alloc_brl.values()))
    ws.cell(row=r, column=4, value=sum(alloc_usd.values()))
    ws.cell(row=r, column=5, value=grand_total)
    ws.cell(row=r, column=6, value=1.0)
    for col in [3, 4, 5]:
        ws.cell(row=r, column=col).number_format = BRL_FMT
    ws.cell(row=r, column=6).number_format = PCT_FMT
    for col in range(1, len(headers) + 1):
        ws.cell(row=r, column=col).font = Font(name="Calibri", bold=True, size=10)

    style_body(ws, 2, r, len(headers))
    auto_width(ws)


# ── Sheet 7: alloc_by_institution ──────────────────────────────────────────────

def sheet_alloc_by_institution(wb, wallets, navs):
    ws = wb.create_sheet("alloc_by_institution")
    headers = [
        "referenceMonth", "institution", "walletId", "walletName",
        "currency", "saldo_bruto", "saldo_bruto_BRL"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    fx_rate = 6.19

    r = 2
    total_brl = 0
    for wid, w in wallets.items():
        nav_list = navs.get(wid, [])
        latest_nav = nav_list[-1]["nav"] if nav_list else 0
        entity_name = load_entity_name(w.get("entityId", ""))
        currency = w.get("currency", "BRL")
        brl_val = latest_nav if currency == "BRL" else latest_nav * fx_rate

        ws.cell(row=r, column=1, value=REFERENCE_MONTH)
        ws.cell(row=r, column=2, value=entity_name)
        ws.cell(row=r, column=3, value=wid)
        ws.cell(row=r, column=4, value=w.get("name", ""))
        ws.cell(row=r, column=5, value=currency)
        ws.cell(row=r, column=6, value=latest_nav)
        ws.cell(row=r, column=7, value=brl_val)

        ws.cell(row=r, column=6).number_format = BRL_FMT
        ws.cell(row=r, column=7).number_format = BRL_FMT
        total_brl += brl_val
        r += 1

    # Total row
    ws.cell(row=r, column=1, value=REFERENCE_MONTH)
    ws.cell(row=r, column=2, value="TOTAL")
    ws.cell(row=r, column=7, value=total_brl)
    ws.cell(row=r, column=7).number_format = BRL_FMT
    for col in range(1, len(headers) + 1):
        ws.cell(row=r, column=col).font = Font(name="Calibri", bold=True, size=10)

    style_body(ws, 2, r, len(headers))
    auto_width(ws)


# ── Sheet 8: geographic_exposure ───────────────────────────────────────────────

def sheet_geographic_exposure(wb, wallets, positions):
    ws = wb.create_sheet("geographic_exposure")
    headers = ["referenceMonth", "region", "total_BRL", "pct"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    fx_rate = 6.19
    region_totals = {}
    grand_total = 0

    for wid, w in wallets.items():
        pos = positions.get(wid)
        if not pos:
            continue
        currency = w.get("currency", "BRL")
        secs = pos.get("securities", [])
        for s in secs:
            balance = s.get("quantity", 0) * s.get("pu", 0)
            brl_val = balance if currency == "BRL" else balance * fx_rate

            # Infer region from currency and asset class
            if currency == "BRL":
                region = "Brasil"
            else:
                var1 = s.get("hierarchicalVariable", {}).get("variable1", "")
                name = s.get("beehusName", "").upper()
                if "US" in name or "SPDR" in name or "RUSSELL" in name or "JPM" in name or "LORD" in name:
                    region = "EUA"
                elif "HEDGE" in var1.upper() or "PRIVATE" in var1.upper():
                    region = "Global"
                else:
                    region = "EUA"

            region_totals[region] = region_totals.get(region, 0) + brl_val
            grand_total += brl_val

    r = 2
    for region, total in sorted(region_totals.items(), key=lambda x: -x[1]):
        pct = total / grand_total if grand_total > 0 else 0
        ws.cell(row=r, column=1, value=REFERENCE_MONTH)
        ws.cell(row=r, column=2, value=region)
        ws.cell(row=r, column=3, value=total)
        ws.cell(row=r, column=4, value=pct)
        ws.cell(row=r, column=3).number_format = BRL_FMT
        ws.cell(row=r, column=4).number_format = PCT_FMT
        r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Sheet 9: transactions ─────────────────────────────────────────────────────

def sheet_transactions(wb, wallets, txns):
    ws = wb.create_sheet("transactions")
    headers = [
        "walletId", "walletName", "currency", "direction",
        "liquidationDate", "transactionType", "description", "balance"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    r = 2
    for wid, w in wallets.items():
        for t in txns.get(wid, []):
            # Filter to reference month only
            # Coerce: liquidationDate may be an ISO string or a BSON Date
            # (datetime); str() makes .startswith safe for both. A datetime
            # stringifies as "2025-12-15 00:00:00", which still matches.
            liq_date = str(t.get("liquidationDate", "") or "")
            if not liq_date.startswith("2025-12"):
                continue
            balance = t.get("balance", 0)
            direction = "entrada" if balance >= 0 else "saida"

            ws.cell(row=r, column=1, value=wid)
            ws.cell(row=r, column=2, value=w.get("name", ""))
            ws.cell(row=r, column=3, value=w.get("currency", ""))
            ws.cell(row=r, column=4, value=direction)
            ws.cell(row=r, column=5, value=liq_date)
            ws.cell(row=r, column=6, value=t.get("beehusTransactionType", ""))
            ws.cell(row=r, column=7, value=t.get("description", ""))
            ws.cell(row=r, column=8, value=balance)
            ws.cell(row=r, column=8).number_format = BRL_FMT
            r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Sheet 10: monthly_flow ─────────────────────────────────────────────────────

def sheet_monthly_flow(wb, wallets, navs):
    ws = wb.create_sheet("monthly_flow")
    headers = [
        "walletId", "walletName", "currency",
        "positionDate", "nav", "returnNavPerShare"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    r = 2
    for wid, w in wallets.items():
        for n in navs.get(wid, []):
            ws.cell(row=r, column=1, value=wid)
            ws.cell(row=r, column=2, value=w.get("name", ""))
            ws.cell(row=r, column=3, value=w.get("currency", ""))
            ws.cell(row=r, column=4, value=n["positionDate"])
            ws.cell(row=r, column=5, value=n["nav"])
            ws.cell(row=r, column=6, value=n.get("returnNavPerShare", 0))
            ws.cell(row=r, column=5).number_format = BRL_FMT
            ws.cell(row=r, column=6).number_format = PCT_FMT
            r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Sheet 11: risk_statistics ──────────────────────────────────────────────────

def sheet_risk_statistics(wb, wallets, navs):
    ws = wb.create_sheet("risk_statistics")
    headers = [
        "walletId", "walletName", "currency", "referenceMonth",
        "meses_positivos", "meses_negativos",
        "rentabilidade_mensal_maxima", "rentabilidade_mensal_minima",
        "volatilidade_anualizada"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    r = 2
    for wid, w in wallets.items():
        month_navs = _month_end_navs(navs.get(wid, []))
        returns = _monthly_returns(month_navs)
        if not returns:
            continue

        positive = sum(1 for ret in returns if ret > 0)
        negative = sum(1 for ret in returns if ret < 0)
        max_ret  = max(returns)
        min_ret  = min(returns)

        # Annualized vol on monthly returns: stdev * sqrt(12).
        vol = statistics.stdev(returns) * (12 ** 0.5) if len(returns) > 1 else 0

        ws.cell(row=r, column=1, value=wid)
        ws.cell(row=r, column=2, value=w.get("name", ""))
        ws.cell(row=r, column=3, value=w.get("currency", ""))
        ws.cell(row=r, column=4, value=REFERENCE_MONTH)
        ws.cell(row=r, column=5, value=positive)
        ws.cell(row=r, column=6, value=negative)
        ws.cell(row=r, column=7, value=max_ret)
        ws.cell(row=r, column=8, value=min_ret)
        ws.cell(row=r, column=9, value=vol)
        for col in [7, 8, 9]:
            ws.cell(row=r, column=col).number_format = PCT_FMT
        r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Sheet 12: liquidity ────────────────────────────────────────────────────────

def sheet_liquidity(wb, wallets, positions):
    ws = wb.create_sheet("liquidity")
    headers = [
        "walletId", "walletName", "currency", "referenceMonth",
        "time_bucket", "total_balance", "pct_portfolio"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    # Define liquidity buckets based on asset type heuristics
    def classify_liquidity(sec):
        name = sec.get("beehusName", "").upper()
        var1 = sec.get("hierarchicalVariable", {}).get("variable1", "")
        var2 = sec.get("hierarchicalVariable", {}).get("variable2", "") or ""

        if "SELIC" in name or "CASH" in name or "CDI" in name.upper() or "TESOURO" in name:
            return "Diaria"
        if "T BILL" in name:
            return "Ate 6 meses"
        if "FII" in name or "ETF" in name or "PETR" in name or "ITUB" in name or "SPDR" in name or "RUSSELL" in name:
            return "Ate 1 Mes"
        if "FICFI" in name or "FIC" in name or "LORD" in name or "JPM" in name:
            return "Ate 1 Mes"
        if "LCA" in name or "CDB" in name:
            return "Ate 6 meses"
        if "CRI" in name or "CRA" in name or "DEB" in name or "FIDC" in name or "FIAGRO" in name:
            return "Ate 2 anos"
        if "PRIVATE" in var1.upper() or "LEXINGTON" in name or "INSIGHT" in name:
            return "Acima de 2 anos"
        if "HEDGE" in var1.upper():
            return "Ate 1 ano"
        return "Ate 1 ano"

    r = 2
    for wid, w in wallets.items():
        pos = positions.get(wid)
        if not pos:
            continue
        secs = pos.get("securities", [])
        buckets = {}
        total = 0
        for s in secs:
            balance = s.get("quantity", 0) * s.get("pu", 0)
            bucket = classify_liquidity(s)
            buckets[bucket] = buckets.get(bucket, 0) + balance
            total += balance

        bucket_order = ["Diaria", "Ate 1 Mes", "Ate 6 meses", "Ate 1 ano", "Ate 2 anos", "Acima de 2 anos"]
        for bucket in bucket_order:
            bal = buckets.get(bucket, 0)
            if bal == 0:
                continue
            pct = bal / total if total > 0 else 0
            ws.cell(row=r, column=1, value=wid)
            ws.cell(row=r, column=2, value=w.get("name", ""))
            ws.cell(row=r, column=3, value=w.get("currency", ""))
            ws.cell(row=r, column=4, value=REFERENCE_MONTH)
            ws.cell(row=r, column=5, value=bucket)
            ws.cell(row=r, column=6, value=bal)
            ws.cell(row=r, column=7, value=pct)
            ws.cell(row=r, column=6).number_format = BRL_FMT
            ws.cell(row=r, column=7).number_format = PCT_FMT
            r += 1

    style_body(ws, 2, r - 1, len(headers))
    auto_width(ws)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("Loading data from MongoDB...")
    wallets = load_wallets()
    positions = load_positions()
    navs = load_nav_history()
    txns = load_transactions()

    print(f"Loaded {len(wallets)} wallets, {len(positions)} positions")
    for wid, w in wallets.items():
        n_secs = len(positions.get(wid, {}).get("securities", []))
        n_navs = len(navs.get(wid, []))
        n_txns = len(txns.get(wid, []))
        print(f"  {w['name']:40s}  secs={n_secs}  navs={n_navs}  txns={n_txns}")

    wb = Workbook()
    wb.remove(wb.active)

    print("\nGenerating sheets...")
    sheet_report_config(wb, wallets)
    print("  report_config")
    sheet_performance_summary(wb, wallets, navs)
    print("  performance_summary")
    sheet_nav_timeseries(wb, wallets, navs)
    print("  nav_timeseries")
    sheet_asset_detail(wb, wallets, positions)
    print("  asset_detail")
    sheet_allocation(wb, wallets, positions)
    print("  allocation")
    sheet_allocation_consolidated(wb, wallets, positions)
    print("  alloc_consolidated")
    sheet_alloc_by_institution(wb, wallets, navs)
    print("  alloc_by_institution")
    sheet_geographic_exposure(wb, wallets, positions)
    print("  geographic_exposure")
    sheet_transactions(wb, wallets, txns)
    print("  transactions")
    sheet_monthly_flow(wb, wallets, navs)
    print("  monthly_flow")
    sheet_risk_statistics(wb, wallets, navs)
    print("  risk_statistics")
    sheet_liquidity(wb, wallets, positions)
    print("  liquidity")

    output_path = os.path.join(_PROJECT_ROOT, "data", "relatorio_patrimonial_from_mongo.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"\nExcel saved to: {output_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()

"""Gera os arquivos de upload (positions + transactions) a partir de um
relatório mensal EMR (XP MB) .xlsm. Ver docs/mbonboarding.md.

Saída (em --out, default = mesma pasta do .xlsm):
  <stem>_positions.json     -> formato unprocessedSecurities (FILE_GENERATION.md #3)
  <stem>_transactions.json  -> formato transactions          (FILE_GENERATION.md #1)
  <stem>_positions.xlsx     -> idem + colunas de revisão (securityId resolvido, status)
  <stem>_transactions.xlsx  -> idem + colunas de revisão (match, motivo do tipo)

RASCUNHO: walletId = rótulo "Banco" do Excel (placeholder). Trocar pelos
walletId reais antes do upload. Security resolvida só por identificador EXATO
(CNPJ/ticker); bonds só-nome ficam sem securityId (status 'revisar').

Uso:
  python scripts/mb_generate.py data/.tmp/EMR_202507.xlsm
"""
import sys, os, re, json, argparse, warnings
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
warnings.filterwarnings("ignore")

import openpyxl
from openpyxl import Workbook

# ── Constantes de resolução (descobertas via Mongo em 2026-06) ────────────────
COMPANY_ID = "10000000000000"                      # Blue3
ENTITY_ID = {"XP": "67cf6a5c71f5e8c88f760505",     # entity XP
             "SAFRA": "67c0a80471f5e8c88f7604a1"}  # entity Safra
DEFAULT_CURRENCY = "BRL"

# DataWM (header r4, dados r5+) — índices 0-based na tupla values_only
D_DATE, D_REPORT, D_COD, D_BANCO = 2, 3, 4, 5
D_ATIVO, D_DVCTO, D_POSANT, D_POS = 8, 11, 12, 13
D_CLASSE, D_TIPO, D_VEIC, D_CNPJ = 14, 17, 18, 19

# MovWM (header r1, dados r2+)
M_DATA, M_REPORT, M_BANCO, M_COD, M_DESC, M_VALOR, M_ASSET = 2, 3, 4, 6, 7, 8, 9

_DIGITS = re.compile(r"\D")
_CODE = re.compile(r"^[A-Z0-9]{4,12}$")
_CASH_HINTS = ("caixa", "conta corrente", "conta-corrente")


def entity_for(banco):
    b = (banco or "").strip().lower()
    if b.startswith("xp"):
        return ENTITY_ID["XP"]
    if b.startswith("safra"):
        return ENTITY_ID["SAFRA"]
    return ""


def is_cash(cod, ativo, classe):
    blob = f"{cod} {ativo} {classe}".lower()
    if any(h in blob for h in _CASH_HINTS):
        return True
    # cod numérico curto + classe "Caixa ..."
    return classe.strip().lower().startswith("caixa")


def cnpj_digits(*texts):
    for t in texts:
        if not t:
            continue
        d = _DIGITS.sub("", str(t))
        if len(d) == 14:
            return d
    return None


def parse_mov_date(v):
    """Data do MovWM: datetime nativo OU texto 'dd/mm/aaaa'."""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # devolve cru se não parsear (flag visível no output)


def clean(x):
    """Normaliza string de identificador: '-'/'' viram '' (vazio)."""
    s = str(x or "").strip()
    return "" if s in ("", "-") else s


def _round(v, n=8):
    try:
        return round(float(v), n)
    except (TypeError, ValueError):
        return None


# ── Planilha auxiliar (de-para preenchido pelo usuário) ───────────────────────
# 3 abas: Wallets (walletName->walletId), Securities (ativo->securityId),
# PU (securityId->PU). Ver scripts/mb_aux.py e docs/mbonboarding.md.
def load_aux(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    def sheet_map(name, kcol, vcol):
        if name not in wb.sheetnames:
            return {}
        ws = wb[name]
        out = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # header
            if row is None or len(row) <= max(kcol, vcol):
                continue
            k, v = row[kcol], row[vcol]
            if k in (None, "") or v in (None, ""):
                continue
            out[clean(k)] = str(v).strip()
        return out

    wallet_map = sheet_map("Wallets", 0, 1)       # walletName -> walletId
    sec_map = sheet_map("Securities", 0, 1)       # ativo      -> securityId
    pu_raw = sheet_map("PU", 0, 1)                # securityId -> PU(str)
    pu_map = {k: _round(v) for k, v in pu_raw.items() if _round(v) is not None}
    wb.close()
    return wallet_map, sec_map, pu_map


# ── Resolução de security por identificador exato (via cache do projeto) ──────
class Resolver:
    def __init__(self, cache):
        self.cache = cache

    def _first(self, field, value):
        if not value:
            return None
        hits = self.cache.lookup(field, value)
        return hits[0] if hits else None, (len(hits) if hits else 0)

    def resolve(self, cod, ativo, cnpj_col=None):
        """Retorna (securityId, beehusName, status). status:
        'cnpj' | 'ticker' | 'ambiguo:<n>' | 'revisar' (não casou)."""
        cnpj = cnpj_digits(cnpj_col, cod, ativo)
        if cnpj:
            for f in ("taxId", "mainId"):
                hits = self.cache.lookup(f, cnpj)
                if hits:
                    st = "cnpj" if len(hits) == 1 else f"ambiguo:{len(hits)}"
                    return hits[0].get("_id"), hits[0].get("beehusName"), st
        token = (cod or "").strip().upper()
        if _CODE.match(token):
            for f in ("ticker", "mainId", "selicCode", "isIn"):
                hits = self.cache.lookup(f, token)
                if hits:
                    st = "ticker" if len(hits) == 1 else f"ambiguo:{len(hits)}"
                    return hits[0].get("_id"), hits[0].get("beehusName"), st
        return None, None, "revisar"


# ── Classificador de beehusTransactionType por palavra-chave ──────────────────
# Ordem importa: dividend (juros/rendimento) > buySell (aplicação/resgate) >
# gainsExpenses (taxa/ir) > withdrawalDeposit (ted/transferência).
_TYPE_RULES = [
    ("dividend",        ("dividendo", "rendiment", "juros", "jcp", "jscp",
                         "provento", "amortiz", "cupom", "remuneraç",
                         "reembolso de evento")),
    ("buySell",         ("aplicaç", "resgate", "compra", "venda", "subscri",
                         "integraliz", "conversão", "liquidação de cota",
                         "vencimento", "operações em bolsa", "operaçoes em bolsa",
                         "liquido das operaç", "líquido das operaç")),
    ("gainsExpenses",   ("taxa", "tarifa", "irrf", " ir ", "imposto", "iof",
                         "come-cotas", "comecotas", "emolument", "custód",
                         "intermediaç")),
    ("withdrawalDeposit", ("ted ", "doc ", "transferência", "saque",
                           "depósit", "deposit", "retirada", "pix")),
]


def classify_type(description):
    d = f" {(description or '').lower()} "
    for kind, kws in _TYPE_RULES:
        for kw in kws:
            if kw in d:
                return kind, kw.strip()
    return "", ""


# ── Geração ───────────────────────────────────────────────────────────────────
def build(xlsm_path, resolver=None, aux=None):
    wallet_map, sec_map, pu_map = aux if aux else ({}, {}, {})
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True,
                                keep_links=False)
    positions, pos_review = [], []
    tx, tx_review = [], []

    # Positions <- DataWM
    ws = wb["DataWM"]
    for r in ws.iter_rows(min_row=5, values_only=True):
        if r[D_DATE] is None:
            continue
        if not isinstance(r[D_POS], (int, float)):   # pula lixo sem saldo numérico
            continue
        date = r[D_DATE].strftime("%Y-%m-%d") if isinstance(r[D_DATE], datetime) else str(r[D_DATE])
        banco = str(r[D_BANCO] or "")
        cod = str(r[D_COD] or "")
        ativo = str(r[D_ATIVO] or "")
        classe = str(r[D_CLASSE] or "")
        cash = is_cash(cod, ativo, classe)
        sec_name = clean(cod) if not cash else (clean(ativo) or clean(classe) or "Caixa")
        # securityId: aux (fonte de verdade) > resolver Mongo > vazio
        if cash:
            sid, bname, status = None, None, "caixa"
        elif sec_map:
            sid = sec_map.get(sec_name) or None
            bname, status = None, ("aux" if sid else "revisar")
        elif resolver:
            sid, bname, status = resolver.resolve(cod, ativo, r[D_CNPJ])
        else:
            sid, bname, status = None, None, "revisar"
        pu = pu_map.get(sid) if sid else None
        bal = r[D_POS]
        qty = _round(bal / pu, 8) if (pu and pu != 0 and isinstance(bal, (int, float))) else None
        wallet_id = wallet_map.get(banco, banco)   # walletId real ou rótulo placeholder
        positions.append({
            "date": date,
            "walletId": wallet_id,
            "security": sec_name,
            "quantity": qty,                   # derivado de SaldoBruto/PU (se PU no aux)
            "pu": pu,                          # do aux (securityId->PU)
            "balance": bal,
            "currencyId": DEFAULT_CURRENCY,
            "cashAccount": "Sim" if cash else "Nao",
        })
        pos_review.append({
            "walletLabel": banco, "walletId": wallet_id, "date": date, "security": sec_name,
            "balance": bal, "pu": pu, "quantity": qty, "posAnterior": r[D_POSANT],
            "classe": classe, "tipo": str(r[D_TIPO] or ""),
            "veiculo": str(r[D_VEIC] or ""), "cnpj": str(r[D_CNPJ] or ""),
            "resolvedSecurityId": sid or "", "resolvedBeehusName": bname or "",
            "matchStatus": status,
        })

    # Transactions <- MovWM
    ws = wb["MovWM"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[M_DATA] is None:
            continue
        if not isinstance(r[M_VALOR], (int, float)):   # pula lixo sem valor numérico
            continue
        banco = str(r[M_BANCO] or "")
        desc = str(r[M_DESC] or "")
        asset = str(r[M_ASSET] or "")
        cod = str(r[M_COD] or "")
        date = parse_mov_date(r[M_DATA])
        cash = is_cash(cod, asset, "")
        if cash:
            sid, bname, mstatus = None, None, "caixa"
        elif sec_map:
            sid = sec_map.get(clean(asset)) or sec_map.get(clean(cod)) or None
            bname, mstatus = None, ("aux" if sid else "revisar")
        elif resolver:
            sid, bname, mstatus = resolver.resolve(cod, asset, cod)
        else:
            sid, bname, mstatus = None, None, "revisar"
        ttype, reason = classify_type(desc)
        doc = {
            "companyId": COMPANY_ID,
            "entityId": entity_for(banco),
            "walletId": banco,                 # PLACEHOLDER (rótulo Excel)
            "currencyId": DEFAULT_CURRENCY,
            "operationDate": date,
            "liquidationDate": date,
            "balance": r[M_VALOR],
            "description": desc,
            "inputType": "sheets",
            "beehusTransactionType": ttype,
            "hide": False,
            "comment": "",
        }
        if sid:                                # securityId omitido se não há
            doc["securityId"] = sid
        tx.append(doc)
        tx_review.append({
            "walletLabel": banco, "date": date, "balance": r[M_VALOR],
            "description": desc, "asset": asset,
            "beehusTransactionType": ttype, "typeReason": reason,
            "resolvedSecurityId": sid or "", "resolvedBeehusName": bname or "",
            "matchStatus": mstatus,
        })

    wb.close()
    return positions, pos_review, tx, tx_review


def write_xlsx(path, rows):
    wb = Workbook()
    ws = wb.active
    if not rows:
        wb.save(path)
        return
    cols = list(rows[0].keys())
    ws.append(cols)
    for row in rows:
        ws.append([row.get(c) for c in cols])
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsm")
    ap.add_argument("--aux", default=None,
                    help="planilha auxiliar de-para (Wallets/Securities/PU). "
                         "Quando passada, é a fonte de verdade e o Mongo não é consultado.")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    from db import atomic_write_json

    if args.aux:
        aux = load_aux(args.aux)
        print(f"Aux: wallets={len(aux[0])} securities={len(aux[1])} PUs={len(aux[2])}")
        positions, pos_review, tx, tx_review = build(args.xlsm, resolver=None, aux=aux)
    else:
        from db import db
        from security_matcher import SecurityCache
        if not db._ready():
            sys.exit("DB não conectado (registre em /setup).")
        cache = SecurityCache()
        cache.ensure_loaded(db)
        print(f"SecurityCache: {cache.count} securities")
        positions, pos_review, tx, tx_review = build(args.xlsm, resolver=Resolver(cache))

    out = args.out or os.path.dirname(args.xlsm)
    os.makedirs(out, exist_ok=True)
    stem = os.path.splitext(os.path.basename(args.xlsm))[0]
    p = lambda s: os.path.join(out, f"{stem}_{s}")

    atomic_write_json(p("positions.json"),
                      {"companyId": COMPANY_ID, "unprocessedSecurities": positions})
    atomic_write_json(p("transactions.json"),
                      {"companyId": COMPANY_ID, "transactions": tx})
    write_xlsx(p("positions.xlsx"), pos_review)
    write_xlsx(p("transactions.xlsx"), tx_review)

    # Stats
    import collections
    pmatch = collections.Counter(r["matchStatus"] for r in pos_review)
    tmatch = collections.Counter(r["matchStatus"] for r in tx_review)
    ttypes = collections.Counter(r["beehusTransactionType"] or "(vazio)" for r in tx_review)
    print(f"\nPOSITIONS: {len(positions)} linhas")
    print("  match:", dict(pmatch))
    print(f"TRANSACTIONS: {len(tx)} linhas")
    print("  match:", dict(tmatch))
    print("  beehusTransactionType:", dict(ttypes))
    print("\nArquivos:")
    for s in ("positions.json", "transactions.json", "positions.xlsx", "transactions.xlsx"):
        print("  ", p(s))


if __name__ == "__main__":
    main()

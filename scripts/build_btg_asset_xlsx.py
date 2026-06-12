"""Gera planilha Excel a partir do JSON consolidado de build_btg_asset_registration.py.

Cria 1 aba por categoria + 1 aba "Resumo".
"""
from __future__ import annotations

import io
import json
import sys
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if getattr(sys.stderr, "encoding", "").lower() != "utf-8":
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent))  # para importar build_btg_asset_registration

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

# Reaproveita o pipeline de extração: importa as funções do build_btg_asset_registration
from build_btg_asset_registration import (  # noqa: E402
    COMPANY_ID, CONSUME_DATE,
    _build_fi_name, _build_fund_name, _build_coe_name,
    _build_pension_name, _build_crypto_name, _short_emissor,
)
from db import db  # noqa: E402


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")
EXISTS_FILL = PatternFill("solid", fgColor="E2EFDA")  # verde claro: já cadastrado
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(italic=True, size=10, color="595959")


def _style_header_row(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row_idx].height = 28


def _autosize(ws, headers, data_rows, max_w=60, min_w=10):
    for i, h in enumerate(headers, 1):
        longest = len(str(h))
        for r in data_rows:
            v = r[i - 1] if i - 1 < len(r) else ""
            longest = max(longest, len(str(v)) if v is not None else 0)
        ws.column_dimensions[get_column_letter(i)].width = max(min_w, min(max_w, longest + 2))


def _write_sheet(wb, sheet_name, headers, rows, title=None):
    ws = wb.create_sheet(sheet_name)
    cur = 1
    if title:
        ws.cell(row=cur, column=1, value=title).font = TITLE_FONT
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=len(headers))
        cur += 1
        ws.cell(
            row=cur, column=1,
            value=f"companyId={COMPANY_ID} · consumeDate={CONSUME_DATE} · linhas={len(rows)}"
        ).font = SUBTITLE_FONT
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=len(headers))
        cur += 2

    header_row = cur
    for i, h in enumerate(headers, 1):
        ws.cell(row=cur, column=i, value=h)
    _style_header_row(ws, cur, len(headers))
    cur += 1

    # Pinta linha inteira em verde claro se a coluna B ("securityId (existente)")
    # estiver preenchida. Caso contrário, aplica zebra-stripe normal.
    existing_col_idx = (
        2 if len(headers) > 1 and "existente" in str(headers[1]).lower() else None
    )
    for ri, row in enumerate(rows):
        is_existing = (
            existing_col_idx is not None
            and ri < len(rows)
            and bool(row[existing_col_idx - 1])
        )
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=cur, column=ci, value=v)
            if is_existing:
                cell.fill = EXISTS_FILL
            elif ri % 2 == 1:
                cell.fill = ZEBRA_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        cur += 1

    _autosize(ws, headers, rows)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)  # congela header + 2 cols
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"
        if rows else f"A{header_row}:{get_column_letter(len(headers))}{header_row}"
    )


def collect_rows():
    from security_matcher import get_cache
    from btg_existing_matcher import find_existing

    coll = db["rawBTGPosition"]
    q = {"companyId": COMPANY_ID, "consumeDate": CONSUME_DATE}
    docs = list(coll.find(q))

    rows = OrderedDict()

    def add(key, payload, wallet):
        if not key:
            return
        if key not in rows:
            rows[key] = {"_wallets": set(), **payload}
        rows[key]["_wallets"].add(wallet)

    for d in docs:
        wallet = d.get("walletCode") or ""
        p = d.get("position") or {}

        for fi in (p.get("FixedIncome") or []):
            if not isinstance(fi, dict):
                continue
            name, jt, emissor, indexer, y, idx_pct, m_iso = _build_fi_name(fi)
            key = ("FI", fi.get("SecurityCode") or fi.get("Ticker") or name, jt)
            add(key, {
                "category": "FixedIncome",
                "beehusName": name,
                "type": jt,
                "ticker": fi.get("Ticker") or "",
                "isin": fi.get("ISIN") or "",
                "cetipCode": fi.get("CetipCode") or "",
                "selicCode": fi.get("SelicCode") or "",
                "securityCode": fi.get("SecurityCode") or "",
                "issuerRaw": fi.get("Issuer") or "",
                "issuerShort": emissor,
                "indexer": indexer,
                "indexerPercentual": idx_pct,
                "yield": y,
                "rawRate": fi.get("IndexYieldRate") or fi.get("ReferenceIndexValue") or "",
                "issueDate": fi.get("IssueDate", "")[:10] if isinstance(fi.get("IssueDate"), str) else "",
                "maturityDate": m_iso,
                "accountingGroupCode": fi.get("AccountingGroupCode") or "",
                "issuerType": fi.get("IssuerType") or "",
                "taxFree": fi.get("TaxFree") or "",
                "isLiquidity": fi.get("IsLiquidity") or "",
                "isRepo": fi.get("IsRepo") or "",
            }, wallet)

        for fnd in (p.get("InvestmentFund") or []):
            if not isinstance(fnd, dict):
                continue
            name, jt, cnpj, sec_code = _build_fund_name(fnd)
            fund = fnd.get("Fund") or {}
            key = ("FUND", sec_code or cnpj or name, jt)
            add(key, {
                "category": "InvestmentFund",
                "beehusName": name,
                "type": jt,
                "taxId": cnpj or "",
                "securityCode": sec_code or "",
                "managerName": fund.get("ManagerName") or "",
                "benchmark": fund.get("BenchMark") or "",
                "fundLiquidity": fund.get("FundLiquidity") or "",
                "tipoCvm": fund.get("TipoCvm") or "",
                "fundName": fund.get("FundName") or "",
            }, wallet)

        for coe in (p.get("FixedIncomeStructuredNote") or []):
            if not isinstance(coe, dict):
                continue
            name, jt, indexer, y, idx_pct, m_iso = _build_coe_name(coe)
            key = ("COE", coe.get("SecurityCode") or coe.get("Ticker") or name, jt)
            add(key, {
                "category": "FixedIncomeStructuredNote",
                "beehusName": name,
                "ticker": coe.get("Ticker") or "",
                "cetipCode": coe.get("CetipCode") or "",
                "securityCode": coe.get("SecurityCode") or "",
                "issuerRaw": coe.get("Issuer") or "",
                "issuerShort": _short_emissor(coe.get("Issuer") or ""),
                "maturityDate": m_iso,
                "fantasyName": coe.get("FantasyName") or "",
                "indexer": indexer,
                "indexerPercentual": idx_pct,
                "yield": y,
            }, wallet)

        for pen in (p.get("PensionInformations") or []):
            if not isinstance(pen, dict):
                continue
            for pos in (pen.get("Positions") or []):
                if not isinstance(pos, dict):
                    continue
                name, jt, fcode, fcge = _build_pension_name(pen, pos)
                key = ("PEN", fcode or fcge or name, jt)
                add(key, {
                    "category": "PensionInformations",
                    "beehusName": name,
                    "type": jt,
                    "securityCode": fcode or "",
                    "fundCgeCode": fcge or "",
                    "susepCode": pos.get("SusepCode") or pen.get("SusepCode") or "",
                    "taxId": pos.get("PensionCnpjCode") or "",
                    "fundType": pos.get("FundType") or pen.get("FundType") or "",
                    "taxRegime": pos.get("TaxRegime") or pen.get("TaxRegime") or "",
                }, wallet)

        for cr in (p.get("CryptoCoin") or []):
            if not isinstance(cr, dict):
                continue
            name, jt = _build_crypto_name(cr)
            asset = cr.get("Asset") or {}
            key = ("CRYPTO", asset.get("Code") or name, jt)
            add(key, {
                "category": "CryptoCoin",
                "beehusName": name,
                "type": jt,
                "securityCode": asset.get("Code") or "",
                "assetName": asset.get("Name") or "",
                "assetType": asset.get("Type") or "",
                "productCode": asset.get("ProductCode") or "",
            }, wallet)

    # Match contra securities (mesma lógica do securityMapping)
    cache = get_cache()
    cache.ensure_loaded(db)
    found = 0
    for r in rows.values():
        sid, nm, mo = find_existing(r, cache)
        r["existingSecurityId"] = sid or ""
        r["existingBeehusName"] = nm or ""
        r["matchedOn"] = mo or ""
        if sid:
            found += 1
    print(f"securities cache: {cache.count} | já cadastrados: {found}/{len(rows)} "
          f"({found*100/max(len(rows),1):.1f}%)")
    return rows


def main():
    print(f"Buscando rawBTGPosition({COMPANY_ID}, {CONSUME_DATE})...")
    rows = collect_rows()

    by_cat = OrderedDict()
    for v in rows.values():
        by_cat.setdefault(v["category"], []).append(v)

    print(f"Total ativos: {len(rows)} | categorias: {list(by_cat.keys())}")

    wb = Workbook()
    wb.remove(wb.active)  # remove o sheet default; recriaremos em ordem

    # ── Resumo ────────────────────────────────────────────────────────────
    resumo_rows = []
    for cat, items in by_cat.items():
        registered = sum(1 for r in items if r.get("existingSecurityId"))
        wallets = sum(len(r["_wallets"]) for r in items)
        resumo_rows.append([cat, len(items), registered, len(items) - registered, wallets])
    total_reg = sum(r[2] for r in resumo_rows)
    total_all = sum(r[1] for r in resumo_rows)
    resumo_rows.append([
        "TOTAL", total_all, total_reg, total_all - total_reg,
        sum(r[4] for r in resumo_rows)
    ])
    _write_sheet(
        wb, "Resumo",
        ["Categoria", "Ativos distintos", "Já cadastrados",
         "A cadastrar", "Soma de wallets"],
        resumo_rows,
        title="rawBTGPosition — cadastro de ativos (match vs securities collection)",
    )

    # ── FixedIncome ──────────────────────────────────────────────────────
    if "FixedIncome" in by_cat:
        items = by_cat["FixedIncome"]
        headers = [
            "#", "securityId (existente)", "beehusName existente", "matchedOn",
            "beehusName (padrão BTG)", "type", "ticker", "ISIN",
            "CetipCode", "SelicCode", "SecurityCode", "Issuer (raw)",
            "Emissor curto", "indexer", "idxPct", "yield", "rawRate",
            "IssueDate", "MaturityDate", "AccGroupCode", "IssuerType",
            "TaxFree", "IsLiquidity", "IsRepo", "wallets",
        ]
        rows_x = []
        for i, r in enumerate(items, 1):
            rows_x.append([
                i, r.get("existingSecurityId", ""), r.get("existingBeehusName", ""),
                r.get("matchedOn", ""),
                r["beehusName"], r["type"], r["ticker"], r["isin"], r["cetipCode"],
                r["selicCode"], r["securityCode"], r["issuerRaw"], r["issuerShort"],
                r["indexer"] or "", r["indexerPercentual"], r["yield"], r["rawRate"],
                r["issueDate"], r["maturityDate"], r["accountingGroupCode"],
                r["issuerType"], r["taxFree"], r["isLiquidity"], r["isRepo"],
                len(r["_wallets"]),
            ])
        _write_sheet(wb, "FixedIncome", headers, rows_x,
                     title="FixedIncome — CDB / LCA / LCI / Debênture / CRI / CRA / LF / CCB")

    # ── InvestmentFund ───────────────────────────────────────────────────
    if "InvestmentFund" in by_cat:
        items = by_cat["InvestmentFund"]
        headers = ["#", "securityId (existente)", "beehusName existente",
                   "matchedOn", "beehusName (padrão BTG)", "type", "CNPJ",
                   "SecurityCode", "Manager", "Benchmark",
                   "FundLiquidity (D+)", "TipoCvm", "wallets"]
        rows_x = []
        for i, r in enumerate(items, 1):
            rows_x.append([
                i, r.get("existingSecurityId", ""), r.get("existingBeehusName", ""),
                r.get("matchedOn", ""),
                r["beehusName"], r["type"], r["taxId"], r["securityCode"],
                r["managerName"], r["benchmark"], r["fundLiquidity"], r["tipoCvm"],
                len(r["_wallets"]),
            ])
        _write_sheet(wb, "InvestmentFund", headers, rows_x,
                     title="InvestmentFund — fundos (FIC, FIDC, FIM, FIA, FII, FIP)")

    # ── FixedIncomeStructuredNote (COE) ──────────────────────────────────
    if "FixedIncomeStructuredNote" in by_cat:
        items = by_cat["FixedIncomeStructuredNote"]
        headers = ["#", "securityId (existente)", "beehusName existente",
                   "matchedOn", "beehusName (padrão BTG)", "ticker", "CetipCode",
                   "SecurityCode", "Issuer (raw)", "Emissor curto",
                   "MaturityDate", "FantasyName", "indexer", "idxPct", "yield",
                   "wallets"]
        rows_x = []
        for i, r in enumerate(items, 1):
            rows_x.append([
                i, r.get("existingSecurityId", ""), r.get("existingBeehusName", ""),
                r.get("matchedOn", ""),
                r["beehusName"], r["ticker"], r["cetipCode"], r["securityCode"],
                r["issuerRaw"], r["issuerShort"], r["maturityDate"], r["fantasyName"],
                r["indexer"] or "", r["indexerPercentual"], r["yield"],
                len(r["_wallets"]),
            ])
        _write_sheet(wb, "COE", headers, rows_x,
                     title="FixedIncomeStructuredNote — COEs")

    # ── PensionInformations ──────────────────────────────────────────────
    if "PensionInformations" in by_cat:
        items = by_cat["PensionInformations"]
        headers = ["#", "securityId (existente)", "beehusName existente",
                   "matchedOn", "beehusName (padrão BTG)", "type", "FundCode",
                   "FundCGECode", "SusepCode", "CNPJ", "FundType",
                   "TaxRegime", "wallets"]
        rows_x = []
        for i, r in enumerate(items, 1):
            rows_x.append([
                i, r.get("existingSecurityId", ""), r.get("existingBeehusName", ""),
                r.get("matchedOn", ""),
                r["beehusName"], r["type"], r["securityCode"], r["fundCgeCode"],
                r["susepCode"], r["taxId"], r["fundType"], r["taxRegime"],
                len(r["_wallets"]),
            ])
        _write_sheet(wb, "Pension", headers, rows_x,
                     title="PensionInformations — VGBL / PGBL")

    # ── CryptoCoin ───────────────────────────────────────────────────────
    if "CryptoCoin" in by_cat:
        items = by_cat["CryptoCoin"]
        headers = ["#", "securityId (existente)", "beehusName existente",
                   "matchedOn", "beehusName (padrão BTG)", "type",
                   "SecurityCode", "AssetName", "AssetType", "ProductCode",
                   "wallets"]
        rows_x = []
        for i, r in enumerate(items, 1):
            rows_x.append([
                i, r.get("existingSecurityId", ""), r.get("existingBeehusName", ""),
                r.get("matchedOn", ""),
                r["beehusName"], r["type"], r["securityCode"], r["assetName"],
                r["assetType"], r["productCode"], len(r["_wallets"]),
            ])
        _write_sheet(wb, "CryptoCoin", headers, rows_x,
                     title="CryptoCoin")

    out_dir = Path(__file__).resolve().parent.parent / "data" / "temp"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"btg_registration_{COMPANY_ID}_{CONSUME_DATE}.xlsx"
    wb.save(out_path)
    print(f"Excel: {out_path}")


if __name__ == "__main__":
    main()

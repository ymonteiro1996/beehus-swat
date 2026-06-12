"""Gera o TEMPLATE da planilha auxiliar (de-para) que alimenta positions +
transactions. 3 abas (lidas por mb_generate.load_aux nas colunas A,B):

  Wallets     A=walletName   B=walletId          (+ nada)
  Securities  A=ativo        B=securityId        (+ beehusName, status — referência)
  PU          A=securityId   B=PU                (+ beehusName, ativo — referência)

Vem pré-preenchido com o que dá pra resolver via Mongo (securityId por CNPJ/
ticker exato). O USUÁRIO completa: walletId de cada carteira, securityId dos
ativos 'revisar', e o PU de cada securityId. Depois:

  python scripts/mb_generate.py data/.tmp/EMR_202507.xlsm --aux data/.tmp/EMR_202507_aux.xlsx
  python scripts/mb_to_beehus.py data/.tmp/EMR_202507_positions.json

Uso:
  python scripts/mb_aux.py data/.tmp/EMR_202507.xlsm
"""
import sys, os, argparse, warnings
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
warnings.filterwarnings("ignore")

import openpyxl
from openpyxl import Workbook

from mb_generate import (clean, is_cash, Resolver,
                         D_DATE, D_COD, D_BANCO, D_ATIVO, D_CLASSE, D_CNPJ,
                         M_DATA, M_BANCO, M_COD, M_DESC, M_ASSET)


def collect(xlsm_path, resolver):
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True, keep_links=False)
    wallets = []                  # ordem de aparição
    seen_w = set()
    keys = {}                     # ativo_key -> (cod, ativo, cnpj_hint)

    def add_wallet(b):
        b = clean(b)
        if b and b not in seen_w:
            seen_w.add(b); wallets.append(b)

    def add_key(key, cod, ativo, cnpj):
        key = clean(key)
        if key and key not in keys:
            keys[key] = (cod, ativo, cnpj)

    ws = wb["DataWM"]
    for r in ws.iter_rows(min_row=5, values_only=True):
        if r[D_DATE] is None:
            continue
        cod, ativo, classe = str(r[D_COD] or ""), str(r[D_ATIVO] or ""), str(r[D_CLASSE] or "")
        add_wallet(r[D_BANCO])
        if is_cash(cod, ativo, classe):
            continue
        add_key(cod, cod, ativo, str(r[D_CNPJ] or ""))   # positions: key = Código

    ws = wb["MovWM"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[M_DATA] is None:
            continue
        cod, asset = str(r[M_COD] or ""), str(r[M_ASSET] or "")
        add_wallet(r[M_BANCO])
        if is_cash(cod, asset, ""):
            continue
        add_key(asset, cod, asset, cod)                  # transactions: key = Asset
    wb.close()

    # resolve securityId por chave (mesma lógica do build sem --aux)
    securities = []   # (ativo, securityId, beehusName, status)
    by_sid = {}       # securityId -> beehusName (para a aba PU)
    for key, (cod, ativo, cnpj) in keys.items():
        sid, bname, status = resolver.resolve(cod, ativo, cnpj)
        securities.append((key, sid or "", bname or "", status))
        if sid and sid not in by_sid:
            by_sid[sid] = bname or ""
    return wallets, securities, by_sid


def write_aux(source, resolver, out_path):
    """Gera o template aux a partir de `source` (path ou file-like do .xlsm já
    descriptografado). Retorna dict de stats."""
    wallets, securities, by_sid = collect(source, resolver)

    wb = Workbook()
    ws = wb.active; ws.title = "Wallets"
    ws.append(["walletName", "walletId"])
    for w in wallets:
        ws.append([w, ""])               # usuário preenche walletId

    ws = wb.create_sheet("Securities")
    ws.append(["ativo", "securityId", "beehusName (ref)", "status (ref)"])
    for ativo, sid, bname, status in securities:
        ws.append([ativo, sid, bname, status])

    ws = wb.create_sheet("PU")
    ws.append(["securityId", "PU", "beehusName (ref)", "ativo (ref)"])
    ativo_by_sid = {}
    for ativo, sid, bname, status in securities:
        if sid and sid not in ativo_by_sid:
            ativo_by_sid[sid] = ativo
    for sid, bname in by_sid.items():
        ws.append([sid, "", bname, ativo_by_sid.get(sid, "")])

    wb.save(out_path)
    resolved = sum(1 for _, sid, _, _ in securities if sid)
    return {"wallets": wallets, "securities": len(securities),
            "resolved": resolved, "pu_ids": len(by_sid)}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsm")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    from db import db
    from security_matcher import SecurityCache
    if not db._ready():
        sys.exit("DB não conectado (registre em /setup).")
    cache = SecurityCache(); cache.ensure_loaded(db)
    print(f"SecurityCache: {cache.count} securities")
    resolver = Resolver(cache)

    out = args.out or (os.path.splitext(args.xlsm)[0] + "_aux.xlsx")
    st = write_aux(args.xlsm, resolver, out)
    print(f"\nWallets: {len(st['wallets'])} | Securities: {st['securities']} "
          f"(resolvidos {st['resolved']}, revisar {st['securities']-st['resolved']}) | "
          f"securityIds p/ PU: {st['pu_ids']}")
    print("Wallets a mapear:", st["wallets"])
    print("->", out)


if __name__ == "__main__":
    main()

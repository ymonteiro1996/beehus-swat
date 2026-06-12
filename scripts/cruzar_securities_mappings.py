"""Cruza securityMappings x securities para uma companyId e gera uma planilha .xlsx.

Filtra `securityMappings` por companyId (fixa abaixo), explode o array
`mappings[]` e, para cada par {from, to}, vincula `mappings.to` ->
`securities._id`. Uma linha por mapping.

Saída: unprocessedSecurityId (mappings.from), securityId (mappings.to) e os
campos do ativo: mainId, isIn, taxId, ticker, securityType, currency,
beehusName, maturityDate, indexer, indexerPercentual, yield,
subscriptionSettlementDays, subscriptionNAVDays, redemptionNAVDays,
redemptionSettlementDays, type.
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from db import db
from bson import ObjectId
from bson.errors import InvalidId
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

COMPANY_ID = "33333333333333"
OUT = os.path.join(os.path.expanduser("~"), "Downloads",
                   f"securities_x_mappings_{COMPANY_ID}.xlsx")

# Campos pedidos do ativo, na ordem de saída. A chave do dict é o nome da
# coluna; o valor é a lista de chaves aceitas no documento (1ª presente vence,
# cobrindo o casing NAV/Nav legado).
SEC_FIELDS = {
    "mainId":                    ["mainId"],
    "isIn":                      ["isIn"],
    "taxId":                     ["taxId"],
    "ticker":                    ["ticker"],
    "securityType":              ["securityType"],
    "currency":                  ["currency"],
    "beehusName":                ["beehusName"],
    "maturityDate":              ["maturityDate"],
    "indexer":                   ["indexer"],
    "indexerPercentual":         ["indexerPercentual"],
    "yield":                     ["yield"],
    "subscriptionSettlementDays":["subscriptionSettlementDays"],
    "subscriptionNAVDays":       ["subscriptionNAVDays", "subscriptionNavDays"],
    "redemptionNAVDays":         ["redemptionNAVDays", "redemptionNavDays"],
    "redemptionSettlementDays":  ["redemptionSettlementDays"],
    "type":                      ["type"],
}


def first_present(doc, keys):
    for k in keys:
        if k in doc and doc[k] is not None:
            return doc[k]
    return None


# ── 1. securityMappings da company ───────────────────────────────────────────
doc = db.securityMappings.find_one({"companyId": COMPANY_ID}, {"mappings": 1})
mappings = (doc or {}).get("mappings", []) or []
if not doc:
    print(f"AVISO: nenhum securityMappings para companyId={COMPANY_ID}")

# ── 2. resolve securities._id (misto string/ObjectId) em UMA query ───────────
to_values = [m.get("to") for m in mappings if m.get("to")]
or_ids = []
seen = set()
for v in to_values:
    s = str(v)
    if s in seen:
        continue
    seen.add(s)
    or_ids.append(s)
    try:
        or_ids.append(ObjectId(s))   # variante ObjectId p/ _id legado
    except (InvalidId, TypeError):
        pass

projection = {"_id": 1}
for keys in SEC_FIELDS.values():
    for k in keys:
        projection[k] = 1

sec_by_id = {}
if or_ids:
    for s in db.securities.find({"_id": {"$in": or_ids}}, projection):
        sec_by_id[str(s["_id"])] = s

# ── 3. monta linhas (uma por mapping) ─────────────────────────────────────────
rows = []
for m in mappings:
    frm = m.get("from")
    to = m.get("to")
    sec = sec_by_id.get(str(to)) if to else None
    row = {
        "unprocessedSecurityId": frm,
        "securityId": str(to) if to is not None else None,
    }
    for col, keys in SEC_FIELDS.items():
        row[col] = first_present(sec, keys) if sec else None
    rows.append(row)

# ── 4. escreve xlsx ───────────────────────────────────────────────────────────
headers = ["unprocessedSecurityId", "securityId"] + list(SEC_FIELDS.keys())

wb = Workbook()
ws = wb.active
ws.title = "securities_x_mappings"
ws.append(headers)
for c in ws[1]:
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center")

for r in rows:
    ws.append([r.get(h) for h in headers])

# Datas/ObjectId viram string para o Excel não quebrar com tipos não nativos.
for row in ws.iter_rows(min_row=2):
    for cell in row:
        v = cell.value
        if v is not None and not isinstance(v, (int, float, str)):
            cell.value = str(v)

ws.freeze_panes = "A2"
for i, h in enumerate(headers, start=1):
    col = ws.cell(row=1, column=i).column_letter
    ws.column_dimensions[col].width = max(14, min(40, len(h) + 4))

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)

matched = sum(1 for r in rows if r.get("beehusName") is not None
              or r.get("mainId") is not None or r.get("securityType") is not None)
print(f"companyId={COMPANY_ID}")
print(f"Mappings: {len(rows)} | Securities encontradas: {len(sec_by_id)} "
      f"| Linhas com ativo casado: {matched}")
print(f"Arquivo: {OUT}")

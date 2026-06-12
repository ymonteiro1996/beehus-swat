"""Ad-hoc simulation: IPCA + 7% a.a., base PU=100 on 2025-01-02.

Reproduces precificacao._calculate_curva_impl's inflacao_curva roll using the
REAL IPCA series from the DB:
    PU(t) = PU(t-1) * daily_factor * (1 + ipca_daily(t))
    daily_factor = (1 + yield/100) ** (1/252)
    ipca_daily(t) = value(t)/value(t-1) - 1     (no `rentability` field on IPCA)

Run from project root: python scripts/sim_ipca_curva.py
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from bson import ObjectId
from db import db

IPCA_ID  = "66fc4ab9e88f2f542b80563d"
BASE_DT  = "2025-01-02"
BASE_PU  = 100.0
YIELD    = 7.0          # % a.a.


def ipca_series():
    docs = []
    try:
        docs = list(db.securityPrices.find(
            {"securityId": ObjectId(IPCA_ID)}, {"historyPrice": 1}))
    except Exception:
        pass
    if not docs:
        docs = list(db.securityPrices.find(
            {"securityId": IPCA_ID}, {"historyPrice": 1}))
    hps = []
    for d in docs:
        raw = d.get("historyPrice")
        if isinstance(raw, dict):
            hps.append(raw)
        elif isinstance(raw, list):
            hps.extend(raw)
    seen, series = set(), []
    for hp in sorted(hps, key=lambda x: str(x.get("date", ""))):
        dt = str(hp.get("date", ""))[:10]
        if dt and dt not in seen and hp.get("value") is not None:
            seen.add(dt)
            series.append((dt, float(hp["value"])))
    return series


def export_xlsx(base, base_idx_val, daily_factor, daily, rows):
    """Write the full daily curve to an .xlsx (full precision stored, display
    formatted). `daily` and `rows` are parallel lists in the same date order."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    out_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "simulacao_curva_IPCA_mais_7.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Curva IPCA+7"

    bold   = Font(bold=True)
    hdr_fl = PatternFill("solid", fgColor="1F4E78")
    hdr_ft = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    thin   = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Parameters block
    ws["A1"] = "Simulação — Preço na Curva (Inflação HTM)"
    ws["A1"].font = Font(bold=True, size=13)
    params = [
        ("Taxa",                "IPCA + 7,00% a.a."),
        ("Data base",           base),
        ("PU base",             BASE_PU),
        ("IPCA índice base",    base_idx_val),
        ("Fator yield diário",  daily_factor),
        ("Fórmula",             "PU(t) = PU(t-1) × (1+yield)^(1/252) × (1 + IPCA_dia)"),
    ]
    r = 3
    for k, v in params:
        ws.cell(r, 1, k).font = bold
        ws.cell(r, 2, v)
        r += 1

    # Header
    head_row = r + 1
    headers = ["Dia", "Data", "IPCA índice", "IPCA dia (%)", "Fator IPCA dia",
               "Fator yield", "Fator combinado", "IPCA acum. (%)",
               "Yield real acum. (%)", "PU"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(head_row, c, h)
        cell.fill = hdr_fl; cell.font = hdr_ft; cell.alignment = center
        cell.border = border

    # Day-0 base row
    data_start = head_row + 1
    base_vals = [0, base, base_idx_val, None, None, None, None, 0.0, 0.0, BASE_PU]
    for c, v in enumerate(base_vals, start=1):
        ws.cell(data_start, c, v).border = border

    # Daily rows (full precision; display via number_format)
    rr = data_start + 1
    for i, ((dt, idx, ipd, comb, pu_v), (_, _, ia, ya)) in enumerate(zip(daily, rows), start=1):
        vals = [i, dt, idx, ipd * 100, 1 + ipd, daily_factor, comb,
                (ia - 1) * 100, (ya - 1) * 100, pu_v]
        for c, v in enumerate(vals, start=1):
            ws.cell(rr, c, v).border = border
        rr += 1
    last_row = rr - 1

    # Number formats per column
    fmts = {3: "0.00000", 4: "0.00000", 5: "0.00000000", 6: "0.00000000",
            7: "0.00000000", 8: "0.000000", 9: "0.000000", 10: "0.00000"}
    for col, fmt in fmts.items():
        for row in range(data_start, last_row + 1):
            ws.cell(row, col).number_format = fmt

    # Widths + freeze
    widths = [6, 12, 14, 13, 16, 14, 16, 15, 20, 14]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(data_start, 1)

    wb.save(out_path)
    print(f"\n>>> XLSX salvo: {out_path}")
    print(f"    {last_row - head_row} linhas de dados ({base} → {daily[-1][0]})")


def main():
    if not db._ready():
        print(">>> DB não conectado.")
        sys.exit(1)

    series = ipca_series()
    dates  = [d for d, _ in series]
    valmap = dict(series)

    if BASE_DT not in valmap:
        # first calendar date >= base
        ge = [d for d in dates if d >= BASE_DT]
        base = ge[0] if ge else None
        print(f"(02/01/2025 não está no calendário; usando primeira data ≥: {base})")
    else:
        base = BASE_DT
    base_i = dates.index(base)

    daily_factor = (1 + YIELD / 100) ** (1 / 252)
    print(f"Base: {base}  PU={BASE_PU:.2f}  | yield={YIELD}% a.a.  "
          f"daily_factor=(1.07)^(1/252)={daily_factor:.10f}")
    print(f"IPCA base value={valmap[base]:.6f}   última data disponível={dates[-1]}\n")

    pu = BASE_PU
    rows = []                       # (date, pu, infl_accum, yield_accum)
    daily = []                      # (date, ipca_daily, combined, pu)
    for i in range(base_i + 1, len(dates)):
        dt = dates[i]
        ipca_daily = valmap[dt] / valmap[dates[i-1]] - 1
        combined   = daily_factor * (1 + ipca_daily)     # "Fator Diário" da UI
        pu *= combined
        n = i - base_i
        infl_accum  = valmap[dt] / valmap[base]          # IPCA acumulado (índice)
        yield_accum = daily_factor ** n                  # rendimento real acumulado
        rows.append((dt, pu, infl_accum, yield_accum))
        daily.append((dt, valmap[dt], ipca_daily, combined, pu))

    # ── Detalhe diário: primeiros 30 dias úteis (5 casas decimais) ──────────
    print(f"\n{'='*86}\nDETALHE DIÁRIO — primeiros 30 dias úteis (5 casas decimais)\n{'='*86}")
    print(f"{'Dia':>3} {'Data':<12}{'IPCA índice':>14}{'IPCA dia %':>12}"
          f"{'Fator yield':>13}{'Fator comb.':>13}{'PU':>13}")
    print("-" * 86)
    print(f"{'0':>3} {base:<12}{valmap[base]:>14.5f}{'':>12}{'':>13}{'':>13}{BASE_PU:>13.5f}")
    for k, (dt, idx, ipd, comb, pu_v) in enumerate(daily[:30], start=1):
        print(f"{k:>3} {dt:<12}{idx:>14.5f}{ipd*100:>12.5f}"
              f"{daily_factor:>13.5f}{comb:>13.5f}{pu_v:>13.5f}")

    # checkpoints: last business day of each month + final
    print(f"{'Data':<12}{'PU':>14}{'IPCA acum':>12}{'Yield acum':>12}{'Verif.':>14}")
    print("-" * 64)
    last_by_month = {}
    for dt, pu_v, ia, ya in rows:
        last_by_month[dt[:7]] = (dt, pu_v, ia, ya)
    for ym in sorted(last_by_month):
        dt, pu_v, ia, ya = last_by_month[ym]
        verif = BASE_PU * ia * ya       # closed form must equal incremental pu
        print(f"{dt:<12}{pu_v:>14.6f}{ia:>12.6f}{ya:>12.6f}{verif:>14.6f}")

    # ── Export .xlsx do período completo ───────────────────────────────────
    export_xlsx(base, valmap[base], daily_factor, daily, rows)

    dt, pu_v, ia, ya = rows[-1]
    n = len(rows)
    eff_aa = (pu_v / BASE_PU) ** (252 / n) - 1
    print("-" * 64)
    print(f"\nFINAL {dt}:")
    print(f"  PU                 = {pu_v:.6f}")
    print(f"  IPCA acumulado     = {(ia-1)*100:.4f}%   (índice {valmap[base]:.4f} → {valmap[dt]:.4f})")
    print(f"  Yield real acum.   = {(ya-1)*100:.4f}%   ((1.07)^({n}/252))")
    print(f"  Retorno total PU   = {(pu_v/BASE_PU-1)*100:.4f}%")
    print(f"  Retorno efetivo    = {eff_aa*100:.4f}% a.a.  (≈ (1+IPCA)*(1.07)-1)")
    print(f"  dias úteis (n)     = {n}")


if __name__ == "__main__":
    main()

"""Exporta securities de unprocessedSecurityPositions para uma planilha .xlsx.

Carteira fixa, positionDate entre [START, END] (inclusive). Uma linha por
security de cada documento. Colunas: Data, Carteira, Ativo, Quant, PU, SaldoBruto.
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from db import db
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers

WALLET = "680a9ce53b2296d861271302"
START = "2026-04-30"
END = "2026-05-29"
OUT = os.path.join(os.path.expanduser("~"), "Downloads",
                   f"unproc_{WALLET}_{START}_a_{END}.xlsx")

cursor = db.unprocessedSecurityPositions.find(
    {"walletId": WALLET, "positionDate": {"$gte": START, "$lte": END}},
    {"positionDate": 1, "walletId": 1, "securities": 1},
).sort("positionDate", 1)

rows = []
for doc in cursor:
    data = doc.get("positionDate")
    wid = doc.get("walletId")
    for s in (doc.get("securities") or []):
        rows.append({
            "Data": data,
            "Carteira": wid,
            "Ativo": s.get("unprocessedId"),
            "Quant": s.get("quantity"),
            "PU": s.get("pu"),
            "SaldoBruto": s.get("balance"),
        })

wb = Workbook()
ws = wb.active
ws.title = "Posicoes"
headers = ["Data", "Carteira", "Ativo", "Quant", "PU", "SaldoBruto"]
ws.append(headers)
for c in ws[1]:
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center")

for r in rows:
    ws.append([r["Data"], r["Carteira"], r["Ativo"], r["Quant"], r["PU"], r["SaldoBruto"]])

# Formato numérico (pt-BR exibido pelo Excel conforme locale do usuário)
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):  # Quant
    row[0].number_format = "#,##0.0000000"
for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):  # PU
    row[0].number_format = "#,##0.00000000"
for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):  # SaldoBruto
    row[0].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

widths = {"A": 12, "B": 26, "C": 55, "D": 18, "E": 16, "F": 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)

dates = sorted({r["Data"] for r in rows})
print(f"Linhas: {len(rows)} | Datas distintas: {len(dates)} ({dates[0] if dates else '-'} .. {dates[-1] if dates else '-'})")
print(f"Arquivo: {OUT}")

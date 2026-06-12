"""Ad-hoc simulation: "Benchmark + spread" curve (inflacao_curva engine model).

Reproduces precificacao._calculate_curva_impl's roll using the REAL benchmark
series from the DB:
    PU(t) = PU(t-1) * daily_factor * (1 + bm_daily(t))
    daily_factor = (1 + YIELD/100) ** (1/252)
    bm_daily(t)  = value(t)/value(t-1) - 1     (value-ratio; no `rentability`)

Configure the block below and run from project root:
    python scripts/sim_curva.py
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from bson import ObjectId
from db import db

# ── CONFIG ────────────────────────────────────────────────────────────────────
BM_ID    = "66fc4a71e88f2f542b805639"   # CDI  (IPCA = 66fc4ab9e88f2f542b80563d)
BM_NAME  = "CDI"
YIELD    = 0.7                          # spread % a.a. sobre o benchmark
BASE_DT  = "2025-01-02"
BASE_PU  = 100.0
OUT_NAME = "simulacao_curva_CDI_mais_0_7.xlsx"
# ──────────────────────────────────────────────────────────────────────────────


def bm_series(sec_id):
    docs = []
    try:
        docs = list(db.securityPrices.find(
            {"securityId": ObjectId(sec_id)}, {"historyPrice": 1}))
    except Exception:
        pass
    if not docs:
        docs = list(db.securityPrices.find(
            {"securityId": sec_id}, {"historyPrice": 1}))
    hps = []
    for d in docs:
        raw = d.get("historyPrice")
        if isinstance(raw, dict):
            hps.append(raw)
        elif isinstance(raw, list):
            hps.extend(raw)
    seen, series = set(), []
    for hp in sorted(hps, key=lambda x: str(x.get("date", ""))):
        dt = str(hp.get("date", ""))[:10]
        if dt and dt not in seen and hp.get("value") is not None:
            seen.add(dt)
            series.append((dt, float(hp["value"])))
    return series


def export_xlsx(base, base_idx_val, daily_factor, daily, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    out_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), OUT_NAME)

    wb = Workbook(); ws = wb.active; ws.title = f"Curva {BM_NAME}+{YIELD:g}"
    bold   = Font(bold=True)
    hdr_fl = PatternFill("solid", fgColor="1F4E78")
    hdr_ft = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    thin   = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"Simulação — Preço na Curva ({BM_NAME} + {YIELD:g}% a.a.)"
    ws["A1"].font = Font(bold=True, size=13)
    params = [
        ("Taxa",                f"{BM_NAME} + {YIELD:.2f}% a.a."),
        ("Data base",           base),
        ("PU base",             BASE_PU),
        (f"{BM_NAME} índice base", base_idx_val),
        ("Fator spread diário", daily_factor),
        ("Fórmula",             f"PU(t) = PU(t-1) × (1+spread)^(1/252) × (1 + {BM_NAME}_dia)"),
    ]
    r = 3
    for k, v in params:
        ws.cell(r, 1, k).font = bold; ws.cell(r, 2, v); r += 1

    head_row = r + 1
    headers = ["Dia", "Data", f"{BM_NAME} índice", f"{BM_NAME} dia (%)",
               f"Fator {BM_NAME} dia", "Fator spread", "Fator combinado",
               f"{BM_NAME} acum. (%)", "Spread acum. (%)", "PU"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(head_row, c, h)
        cell.fill = hdr_fl; cell.font = hdr_ft; cell.alignment = center; cell.border = border

    data_start = head_row + 1
    for c, v in enumerate([0, base, base_idx_val, None, None, None, None, 0.0, 0.0, BASE_PU], start=1):
        ws.cell(data_start, c, v).border = border

    rr = data_start + 1
    for i, ((dt, idx, bmd, comb, pu_v), (_, _, bm_acc, sp_acc)) in enumerate(zip(daily, rows), start=1):
        for c, v in enumerate([i, dt, idx, bmd * 100, 1 + bmd, daily_factor, comb,
                               (bm_acc - 1) * 100, (sp_acc - 1) * 100, pu_v], start=1):
            ws.cell(rr, c, v).border = border
        rr += 1
    last_row = rr - 1

    fmts = {3: "0.00000", 4: "0.00000", 5: "0.00000000", 6: "0.00000000",
            7: "0.00000000", 8: "0.000000", 9: "0.000000", 10: "0.00000"}
    for col, fmt in fmts.items():
        for row in range(data_start, last_row + 1):
            ws.cell(row, col).number_format = fmt

    for c, w in enumerate([6, 12, 14, 13, 16, 14, 16, 15, 18, 14], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(data_start, 1)
    wb.save(out_path)
    print(f"\n>>> XLSX salvo: {out_path}")
    print(f"    {last_row - head_row} linhas de dados ({base} → {daily[-1][0]})")


def main():
    if not db._ready():
        print(">>> DB não conectado."); sys.exit(1)

    series = bm_series(BM_ID)
    dates  = [d for d, _ in series]
    valmap = dict(series)

    if BASE_DT not in valmap:
        ge = [d for d in dates if d >= BASE_DT]
        base = ge[0] if ge else None
        print(f"(02/01/2025 não está no calendário; usando primeira data ≥: {base})")
    else:
        base = BASE_DT
    base_i = dates.index(base)

    daily_factor = (1 + YIELD / 100) ** (1 / 252)
    print(f"Taxa: {BM_NAME} + {YIELD:g}% a.a.")
    print(f"Base: {base}  PU={BASE_PU:.2f}  | spread diário=(1+{YIELD/100:g})^(1/252)={daily_factor:.10f}")
    print(f"{BM_NAME} índice base={valmap[base]:.6f}   última data={dates[-1]}\n")

    pu = BASE_PU
    rows, daily = [], []
    for i in range(base_i + 1, len(dates)):
        dt = dates[i]
        bmd  = valmap[dt] / valmap[dates[i - 1]] - 1
        comb = daily_factor * (1 + bmd)
        pu  *= comb
        n = i - base_i
        rows.append((dt, pu, valmap[dt] / valmap[base], daily_factor ** n))
        daily.append((dt, valmap[dt], bmd, comb, pu))

    # detalhe diário — primeiros 30 dias úteis (5 casas)
    print(f"{'='*88}\nDETALHE DIÁRIO — primeiros 30 dias úteis (5 casas decimais)\n{'='*88}")
    print(f"{'Dia':>3} {'Data':<12}{BM_NAME+' índice':>14}{BM_NAME+' dia %':>12}"
          f"{'Fator spread':>14}{'Fator comb.':>13}{'PU':>13}")
    print("-" * 88)
    print(f"{'0':>3} {base:<12}{valmap[base]:>14.5f}{'':>12}{'':>14}{'':>13}{BASE_PU:>13.5f}")
    for k, (dt, idx, bmd, comb, pu_v) in enumerate(daily[:30], start=1):
        print(f"{k:>3} {dt:<12}{idx:>14.5f}{bmd*100:>12.5f}{daily_factor:>14.5f}{comb:>13.5f}{pu_v:>13.5f}")

    export_xlsx(base, valmap[base], daily_factor, daily, rows)

    dt, pu_v, bm_acc, sp_acc = rows[-1]
    n = len(rows)
    eff_aa = (pu_v / BASE_PU) ** (252 / n) - 1
    print(f"\nFINAL {dt}:")
    print(f"  PU                  = {pu_v:.6f}")
    print(f"  {BM_NAME} acumulado = {(bm_acc-1)*100:.4f}%   (índice {valmap[base]:.4f} → {valmap[dt]:.4f})")
    print(f"  Spread acumulado    = {(sp_acc-1)*100:.4f}%   ((1+{YIELD/100:g})^({n}/252))")
    print(f"  Retorno total PU    = {(pu_v/BASE_PU-1)*100:.4f}%")
    print(f"  Retorno efetivo     = {eff_aa*100:.4f}% a.a.")
    print(f"  dias úteis (n)      = {n}")


if __name__ == "__main__":
    main()

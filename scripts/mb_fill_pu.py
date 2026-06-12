"""Preenche a coluna PU da aba 'PU' de cada EMR_<mês>_aux.xlsx com o PU
(securityPrices.historyPrice[].value) do securityId na positionDate daquele mês.

Para cada aux: positionDate = a (primeira) data das posições do mês
(EMR_<mês>_positions.json). Para cada securityId já identificado na aba PU,
busca em db.securityPrices o value na data exata; se não houver, usa o último
disponível <= positionDate (último preço útil). Grava na coluna B (PU).

Uso:
  python scripts/mb_fill_pu.py "data/.tmp"
  python scripts/mb_fill_pu.py "data/.tmp" --only EMR_202507_aux.xlsx
"""
import sys, os, glob, json, argparse, bisect, warnings
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
warnings.filterwarnings("ignore")
import openpyxl


def _to_float(v):
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return None


class PriceBook:
    """Cache securityId -> (sorted_dates[], {date: value}). securityId pode estar
    gravado como str ou ObjectId em securityPrices; cobre os dois."""
    def __init__(self, db):
        self.db = db
        self._cache = {}
        try:
            from bson import ObjectId
            self._oid = ObjectId
        except Exception:
            self._oid = None

    def _load(self, sid):
        ids = [sid]
        if self._oid:
            try:
                ids.append(self._oid(sid))
            except Exception:
                pass
        merged = {}
        for doc in self.db.securityPrices.find({"securityId": {"$in": ids}},
                                               {"historyPrice": 1}):
            for e in doc.get("historyPrice") or []:
                d = str(e.get("date") or "")[:10]
                val = _to_float(e.get("value"))
                if d and val is not None:
                    merged[d] = val          # docs do mesmo sid devem concordar
        dates = sorted(merged)
        self._cache[sid] = (dates, merged)
        return self._cache[sid]

    def pu(self, sid, position_date):
        """Retorna (pu, kind): kind in {'exato','anterior','sem_preco'}."""
        dates, m = self._cache.get(sid) or self._load(sid)
        if not dates:
            return None, "sem_preco"
        if position_date in m:
            return m[position_date], "exato"
        i = bisect.bisect_right(dates, position_date)   # último <= position_date
        if i == 0:
            return None, "sem_preco"
        return m[dates[i - 1]], "anterior"


def position_date_of(aux_path):
    """A primeira data das posições do mês correspondente."""
    pos = aux_path.replace("_aux.xlsx", "_positions.json")
    if not os.path.exists(pos):
        return None
    data = json.load(open(pos, encoding="utf-8"))
    dates = [p.get("date") for p in (data.get("unprocessedSecurities") or []) if p.get("date")]
    return min(dates) if dates else None


def fill_one(aux_path, book):
    pdate = position_date_of(aux_path)
    if not pdate:
        return None
    wb = openpyxl.load_workbook(aux_path)
    if "PU" not in wb.sheetnames:
        wb.close(); return None
    ws = wb["PU"]
    stats = {"date": pdate, "exato": 0, "anterior": 0, "sem_preco": 0, "total": 0}
    for row in ws.iter_rows(min_row=2):
        sid = row[0].value
        if not sid:
            continue
        stats["total"] += 1
        pu, kind = book.pu(str(sid).strip(), pdate)
        stats[kind] += 1
        if pu is not None:
            row[1].value = pu          # coluna B = PU
    wb.save(aux_path)
    wb.close()
    return stats


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("folder")
    ap.add_argument("--only", default=None)
    args = ap.parse_args()

    auxes = sorted(glob.glob(os.path.join(args.folder, "EMR_*_aux.xlsx")))
    if args.only:
        auxes = [a for a in auxes if os.path.basename(a) == args.only]
    if not auxes:
        sys.exit("Nenhum EMR_*_aux.xlsx encontrado.")

    from db import db
    if not db._ready():
        sys.exit("DB não conectado (registre em /setup).")
    book = PriceBook(db)

    print(f"{len(auxes)} planilhas aux\n")
    for a in auxes:
        st = fill_one(a, book)
        name = os.path.basename(a)
        if not st:
            print(f"  {name}: SKIP (sem positions.json/aba PU)"); continue
        print(f"  {name} [{st['date']}]: PU preenchidos {st['exato']+st['anterior']}/{st['total']} "
              f"(exato {st['exato']}, anterior {st['anterior']}, sem preço {st['sem_preco']})")


if __name__ == "__main__":
    main()
